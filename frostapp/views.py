import hashlib
import json
import os
import logging
import requests
from requests.exceptions import RequestException
import time
from typing import Tuple, Optional, Any, Iterable
import random
import string
import ldap
from ldap.filter import escape_filter_chars
import datetime
import uuid
import pymysql
import hmac
from django.core.mail import EmailMultiAlternatives
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
import threading
from concurrent.futures import ThreadPoolExecutor
import xml.etree.ElementTree as ET
from xml.dom import minidom
from django.conf import settings
# import cx_Oracle
try:
    import cx_Oracle  # если вдруг где-то ещё стоит
except Exception:
    import oracledb as cx_Oracle

    # включаем thick-mode через Instant Client (у тебя он уже ставится)
    _ic_dir = os.getenv("ORACLE_IC_DIR", "/opt/oracle/instantclient_21_1")
    if os.path.isdir(_ic_dir):
        try:
            cx_Oracle.init_oracle_client(lib_dir=_ic_dir)
        except Exception:
            pass



from django.views.decorators.http import require_GET, require_POST
from django.utils.timezone import now as tz_now

import re   
from django.http import JsonResponse, HttpResponseForbidden
from django.http import StreamingHttpResponse, FileResponse, HttpResponse
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import ipaddress
import paramiko
import csv
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404

from django.db.models import Q
import math
from collections import defaultdict, Counter
from pathlib import Path
from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_http_methods
from django.views.decorators.cache import never_cache
from django.urls import reverse
from django.utils.http import urlencode
from django.views.decorators.csrf import csrf_protect
from django.middleware.csrf import get_token
from ldap.controls.libldap import SimplePagedResultsControl
from django.core.exceptions import FieldError
import base64
from urllib.parse import urlencode
from functools import wraps
import io
import mimetypes
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation
from django.db import connection
from django.core.files.base import ContentFile
from django.utils.html import escape

from .models import (
    Queue, 
    MODUL_logs, 
    User, 
    UKMUser, 
    OpenInSystem, 
    QRCode, 
    Department, 
    Position, 
    Store, 
    AuthSession,
    QRIssueLog, 
    VpnAccessSession, 
    AdminBadgeRequest, 
    VpnAccessBaseline, 
    VpnAccessLease,  
    MaxBotRole, 
    MaxBotEmployee, 
    MaxBotVehicle, 
    MaxBotScenario, 
    MaxBotQuestion, 
    MaxBotQuestionOption, 
    MaxBotRequest, 
    MaxBotRequestAnswer,
)

from .forms import (
    MaxBotRoleForm,
    MaxBotEmployeeForm,
    MaxBotVehicleForm,
    MaxBotScenarioForm,
    MaxBotQuestionForm,
    MaxBotQuestionOptionForm,
)


_HEX = set("0123456789abcdefABCDEF")
logger = logging.getLogger(__name__)
AGENT_API_TOKEN = os.getenv("AGENT_API_TOKEN", "zDFbCQWRzL7pKYxzpfSSLVdqCrAYsHiN7FORRUDt1hE")
MAX_BOT_INTERNAL_TOKEN = os.getenv("MAX_BOT_INTERNAL_TOKEN", "wc3wow")
UKM5_FULL_XML_STORE_ID = 2013

ONEC_WORKING_EMPLOYEES_AUTH_USER = os.getenv("ONEC_WORKING_EMPLOYEES_AUTH_USER", "")
ONEC_WORKING_EMPLOYEES_AUTH_PASSWORD = os.getenv("ONEC_WORKING_EMPLOYEES_AUTH_PASSWORD", "")

SYNC_DEFAULT_PAGE_SIZE = 100

def _parse_int_set_env(name: str, default_csv: str) -> set[int]:
    raw = os.getenv(name, default_csv) or ""
    out: set[int] = set()
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            out.add(int(part))
        except Exception:

            pass
    return out


def _oracle_connect(*, user: str, password: str, dsn: str, **kwargs):
    """
    Универсальный коннект: cx_Oracle + python-oracledb.
    Убирает encoding/nencoding если драйвер их не поддерживает.
    """
    try:
        return cx_Oracle.connect(user=user, password=password, dsn=dsn, **kwargs)
    except TypeError as e:
        s = str(e)
        if "unexpected keyword argument 'encoding'" in s or "unexpected keyword argument 'nencoding'" in s:
            kwargs.pop("encoding", None)
            kwargs.pop("nencoding", None)
            return cx_Oracle.connect(user=user, password=password, dsn=dsn, **kwargs)
        raise

INACTIVE_REPORT_TOKEN="wc3wow"
UKM5_FULL_XML_STORE_IDS: set[int] = _parse_int_set_env("UKM5_FULL_XML_STORE_IDS", "2013,9016,1003")
TRM_ID_MAX = 2147483647
_TELEGRAM_SESSION = requests.Session()
BADGE_REQ_TTL_MINUTES = int(os.getenv("BADGE_REQ_TTL_MINUTES", "10"))
TG_BOT_API_TOKEN = os.getenv("TG_BOT_API_TOKEN", "")

TRM_SMALL_MIN = int(os.getenv("TRM_SMALL_MIN", "10000"))   
TRM_SMALL_MAX = int(os.getenv("TRM_SMALL_MAX", "99999")) 
if TRM_SMALL_MIN < 4:
    TRM_SMALL_MIN = 4
if TRM_SMALL_MAX >= TRM_ID_MAX:
    TRM_SMALL_MAX = TRM_ID_MAX - 1


ONEC_WORKING_EMPLOYEES_URL = os.getenv(
    "ONEC_WORKING_EMPLOYEES_URL",
    "http://192.168.17.26/zupcorp_http/hs/API/Get_WorkingEmployees",
)

ALLOWED_WORKING_EMPLOYEE_POSITIONS = {
    "приемщик магазина",
    "администратор магазина",
    "директор магазина",
    "приемщик товара",
}

ONEC_WORKING_EMPLOYEES_TIMEOUT = int(os.getenv("ONEC_WORKING_EMPLOYEES_TIMEOUT", "180"))

ORACLE_SERVICES_ALL = [
    "BINUU00","BINUU01","BINUU02","BINUU03","BINUU04","BINUU5","BINUU05","BINUU06","BINUU07","BINUU08","BINUU09",
    "BINUU10","BINUU011","BINUU11","BINUU12","BINUU14","BINUU15","BINUU16","BINUU17","BINUU18","BINUU21","BINUU22",
    "BINUU23","BINUU24","BINUU25","BINUU26",
    "BINCH00","BINCH1","BINCH2","BINCH3","BINCH4","BINCH5","BINCH6","BINCH7","BINCH8","BINCH9","BINCH10","BINCH11",
    "BINCH12","BINCH13","BINCH14","BINCH15","BINCH16","BINCH17","BINCH18","BINCH19","BINCH20","BINCH21","BINCH22",
]

ORACLE_TNS_MAP = {
    "BINUU00": {"host": "192.168.17.239", "port": 1521, "service_name": "BINUU00"},
    "BINUU01": {"host": "omega1",         "port": 1521, "service_name": "BINUU01"},
    "BINUU02": {"host": "epsilon3",       "port": 1521, "service_name": "BINUU02"},
    "BINUU03": {"host": "omegadb3",       "port": 1521, "service_name": "BINUU03"},
    "BINUU04": {"host": "omega4",         "port": 1521, "service_name": "BINUU04"},
    "BINUU5":  {"host": "omega5",         "port": 1521, "service_name": "BINUU5"},
    "BINUU05": {"host": "delta1",         "port": 1521, "service_name": "BINUU05"},
    "BINUU06": {"host": "10.30.10.254",   "port": 1521, "service_name": "BINUU06"},
    "BINUU07": {"host": "10.30.1.254",    "port": 1521, "service_name": "BINUU07"},
    "BINUU08": {"host": "10.30.8.254",    "port": 1521, "service_name": "BINUU08"},
    "BINUU09": {"host": "omega9new",      "port": 1521, "service_name": "BINUU09"},
    "BINUU10": {"host": "omega14",        "port": 1521, "service_name": "BINUU10"},
    "BINUU011":{"host": "omega11",        "port": 1521, "service_name": "BINUU011"},
    "BINUU11": {"host": "omega33",        "port": 1521, "service_name": "BINUU11"},
    "BINUU12": {"host": "epsilon3",       "port": 1521, "service_name": "BINUU12"},
    "BINUU14": {"host": "epsilon3",       "port": 1521, "service_name": "BINUU14"},
    "BINUU15": {"host": "omega15",        "port": 1521, "service_name": "BINUU15"},
    "BINUU16": {"host": "omega16",        "port": 1521, "service_name": "BINUU16"},
    "BINUU17": {"host": "omega17",        "port": 1521, "service_name": "BINUU17"},


    "BINUU18": {"hosts": ["omega18", "omega18_"], "port": 1521, "service_name": "BINUU18"},

    "BINUU21": {"host": "omega21",        "port": 1521, "service_name": "BINUU21"},
    "BINUU22": {"host": "omega22",        "port": 1521, "service_name": "BINUU22"},
    "BINUU23": {"host": "omega42",        "port": 1521, "service_name": "BINUU23"},
    "BINUU24": {"host": "omega24",        "port": 1521, "service_name": "BINUU24"},
    "BINUU25": {"host": "omega25",        "port": 1521, "service_name": "BINUU25"},
    "BINUU26": {"host": "omega26",        "port": 1521, "service_name": "BINUU26"},

    "BINCH00": {"host": "192.168.202.253","port": 1521, "service_name": "BINCH00"},
    "BINCH1":  {"host": "10.50.50.254",   "port": 1521, "service_name": "BINCH1"},
    "BINCH2":  {"host": "192.168.202.253","port": 1521, "service_name": "BINCH2"},
    "BINCH3":  {"host": "192.168.202.238","port": 1521, "service_name": "BINCH3"},
    "BINCH4":  {"host": "192.168.202.249","port": 1521, "service_name": "BINCH4"},
    "BINCH5":  {"host": "192.168.202.7",  "port": 1521, "service_name": "BINCH5"},
    "BINCH6":  {"host": "192.168.202.12", "port": 1521, "service_name": "BINCH6"},
    "BINCH7":  {"host": "delta7",         "port": 1521, "service_name": "BINCH7"},
    "BINCH8":  {"host": "delta8",         "port": 1521, "service_name": "BINCH8"},
    "BINCH9":  {"host": "192.168.202.253","port": 1521, "service_name": "BINCH9"},
    "BINCH10": {"host": "192.168.202.8",  "port": 1521, "service_name": "BINCH10"},
    "BINCH11": {"host": "192.168.202.253","port": 1521, "service_name": "BINCH11"},
    "BINCH12": {"host": "192.168.202.249","port": 1521, "service_name": "BINCH12"},
    "BINCH13": {"host": "192.168.202.253","port": 1521, "service_name": "BINCH13"},
    "BINCH14": {"host": "192.168.202.13", "port": 1521, "service_name": "BINCH14"},
    "BINCH15": {"host": "delta15",        "port": 1521, "service_name": "BINCH15"},
    "BINCH16": {"host": "192.168.202.11", "port": 1521, "service_name": "BINCH16"},
    "BINCH17": {"host": "192.168.202.249","port": 1521, "service_name": "BINCH17"},
    "BINCH18": {"host": "192.168.202.249","port": 1521, "service_name": "BINCH18"},
    "BINCH19": {"host": "192.168.202.238","port": 1521, "service_name": "BINCH19"},
    "BINCH20": {"host": "192.168.202.249","port": 1521, "service_name": "BINCH20"},
    "BINCH21": {"host": "192.168.202.249","port": 1521, "service_name": "BINCH21"},
    "BINCH22": {"host": "192.168.202.253","port": 1521, "service_name": "BINCH22"},
}

# Транслитерация
_TRANSLIT = {
    "А":"A","Б":"B","В":"V","Г":"G","Д":"D","Е":"E","Ё":"YO","Ж":"ZH","З":"Z","И":"I","Й":"Y","К":"K","Л":"L",
    "М":"M","Н":"N","О":"O","П":"P","Р":"R","С":"S","Т":"T","У":"U","Ф":"F","Х":"KH","Ц":"TS","Ч":"CH","Ш":"SH",
    "Щ":"SHCH","Ъ":"","Ы":"Y","Ь":"","Э":"E","Ю":"YU","Я":"YA",
}

_LOGIN_SAFE_RE = re.compile(r"[^a-z0-9_]+", re.IGNORECASE)



INN_SYNC_PROGRESS_EVERY_ROWS = int(os.getenv("INN_SYNC_PROGRESS_EVERY_ROWS", "2000"))
INN_SYNC_HEARTBEAT_SEC = int(os.getenv("INN_SYNC_HEARTBEAT_SEC", "30"))
INN_SYNC_ORACLE_CALL_TIMEOUT_MS = int(os.getenv("INN_SYNC_ORACLE_CALL_TIMEOUT_MS", "120000"))  # 120s
INN_SYNC_SLOW_STEP_WARN_SEC = float(os.getenv("INN_SYNC_SLOW_STEP_WARN_SEC", "10"))

# Для MAX
MAX_BOT_TOKEN = os.getenv("MAX_BOT_TOKEN", "f9LHodD0cOLB5T7-pExgJ_hpfhRlo6c71eHipQzkJpONY4RosjHynb9NbY6bDk7ermUnV2IJnEbal0A-oRzl")
MAX_API_BASE = os.getenv("MAX_API_BASE", "https://platform-api.max.ru")
MAX_HTTP_TIMEOUT = int(os.getenv("MAX_HTTP_TIMEOUT", "10"))
MAX_ADMIN_LOG_USER_ID = int(os.getenv("MAX_ADMIN_LOG_USER_ID", "91759973"))


# Report endpoint config

INACTIVE_REPORT_TOKEN = os.getenv("INACTIVE_REPORT_TOKEN", "wc3wow").strip()

BITRIX_IM_DIALOG_GET_URL       = "https://gkbin.bitrix24.ru/rest/61518/u58sn3x77hzrm6d9/im.dialog.get.json"
BITRIX_IM_DISK_FOLDER_GET_URL  = "https://gkbin.bitrix24.ru/rest/61518/q4taty8uw51am09u/im.disk.folder.get.json"
BITRIX_DISK_FOLDER_UPLOAD_URL  = "https://gkbin.bitrix24.ru/rest/61518/p1rrvrb3ewm9mfx6/disk.folder.uploadfile.json"
BITRIX_IM_DISK_FILE_COMMIT_URL = "https://gkbin.bitrix24.ru/rest/61518/s0rg1x8c02bym84s/im.disk.file.commit.json"











# Config


BITRIX_DEPARTMENT_GET_URL = os.getenv(
    "BITRIX_DEPARTMENT_GET_URL",
    "https://gkbin.bitrix24.ru/rest/61518/df8m05y41a99szxh/department.get.json",
)
BITRIX_USER_GET_URL = os.getenv(
    "BITRIX_USER_GET_URL",
    "https://gkbin.bitrix24.ru/rest/61518/0ogeiqf5gdy3dot0/user.get.json",
)

BITRIX_USER_UPDATE_URL = os.getenv(
    "BITRIX_USER_UPDATE_URL",
    "https://gkbin.bitrix24.ru/rest/61518/xe5hh9jc83b1p7n4/user.update.json",
)

BITRIX_NOTIFY_URL = os.getenv(
    "BITRIX_NOTIFY_URL",
    "https://gkbin.bitrix24.ru/rest/61518/1ky2jzwneefj1aor/im.notify.personal.add.json",
)

BITRIX_INN_FIELD = os.getenv("BITRIX_INN_FIELD", "UF_USR_1761723694787")

AD_DOMAIN = os.getenv("AD_DOMAIN", "BINLTD")
AD_IP = os.getenv("AD_IP", "192.168.17.100")
AD_USERNAME = os.getenv("AD_USERNAME", "account_adm")
AD_PASSWORD = os.getenv("AD_PASSWORD", "BIN#FTyghu81@")
AD_SEARCH_BASE = os.getenv("AD_SEARCH_BASE", "OU=People,DC=binltd,DC=local")
AD_BASE_DN = os.getenv("AD_BASE_DN", "DC=binltd,DC=local")

VPN_GROUP_CN = os.getenv("VPN_GROUP_CN", "mikrotik_vpn")

VPN_PIN_TTL_MIN = int(os.getenv("VPN_PIN_TTL_MIN", "10"))
VPN_SESSION_TTL_MIN = int(os.getenv("VPN_SESSION_TTL_MIN", "60"))
VPN_MAX_PIN_ATTEMPTS = int(os.getenv("VPN_MAX_PIN_ATTEMPTS", "5"))
_INN_RE = re.compile(r"^\d{10}(\d{2})?$")
VPN_INSTRUCTION_URL = os.getenv("VPN_INSTRUCTION_URL", "https://gkbin.bitrix24.ru/bitrix/tools/disk/focus.php?folderId=10054660&action=openFolderList&ncc=1")
VPN_SCHEDULER_LOCK_KEY = int(os.getenv("VPN_SCHEDULER_LOCK_KEY", "778899"))

# UI: LDAP Tools (employees + sync employeeID from 1C)


LDAP_TOOLS_PAGE_SIZE = int(os.getenv("LDAP_TOOLS_PAGE_SIZE", "200"))
LDAP_EXPORT_PAGE_SIZE = int(os.getenv("LDAP_EXPORT_PAGE_SIZE", "500"))
LDAP_EXPORT_MAX_TOTAL = int(os.getenv("LDAP_EXPORT_MAX_TOTAL", "20000"))

# Helpers: hashing/pin/session


def _rand_pin_4() -> str:
    return f"{random.randint(0, 9999):04d}"

def _rand_salt(n: int = 16) -> str:
    # 16 bytes hex -> 32 chars
    return os.urandom(n).hex()

def _pin_hash(pin: str, salt: str) -> str:
    return hashlib.sha256((salt + ":" + pin).encode("utf-8")).hexdigest()

def _client_ip(request) -> str:
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "") or ""

def _expire_old_sessions():
    # мягкая чистка
    now = timezone.now()
    VpnAccessSession.objects.filter(expires_at__lt=now).exclude(status="EXPIRED").update(status="EXPIRED")


def _get_sid(request) -> str:
    return (request.GET.get("sid") or request.COOKIES.get("vpn_sid") or "").strip()

def _get_session_or_403(request, must_verified: bool) -> VpnAccessSession:
    sid = _get_sid(request)
    if not sid:
        raise PermissionError("NO_SESSION")
    try:
        sess = VpnAccessSession.objects.get(id=sid)
    except Exception:
        raise PermissionError("BAD_SESSION")

    now = timezone.now()
    if sess.expires_at and sess.expires_at < now:
        if sess.status != "EXPIRED":
            sess.status = "EXPIRED"
            sess.save(update_fields=["status"])
        raise PermissionError("EXPIRED")

    if must_verified and sess.status != "VERIFIED":
        raise PermissionError("NOT_VERIFIED")

    return sess

def _bx_fio(u: dict) -> str:
    return " ".join([x for x in [u.get("LAST_NAME"), u.get("NAME"), u.get("SECOND_NAME")] if x]).strip()

def _auth_fio_from_session(sess: VpnAccessSession) -> str:
    # по умолчанию — AD login
    fio = sess.ad_login
    try:
        me = bitrix_user_get_all(
            filter_dict={"ID": int(sess.bitrix_user_id)},
            select_list=["ID", "NAME", "LAST_NAME", "SECOND_NAME"],
        )
        if me:
            fio = _bx_fio(me[0]) or fio
    except Exception:
        pass
    return fio



def _split_telegram_text(text: str, max_len: int = 4000) -> list[str]:
    """
    Режет длинный текст на части до max_len символов,
    стараясь резать по строкам, а не посередине слова.
    """
    text = (text or "").strip()
    if not text:
        return ["(пустое сообщение лога)"]

    if len(text) <= max_len:
        return [text]

    parts = []
    current = []

    for line in text.splitlines():
        candidate = "\n".join(current + [line]).strip()
        if len(candidate) <= max_len:
            current.append(line)
            continue

        if current:
            parts.append("\n".join(current).strip())
            current = []

        # если одна строка длиннее лимита — режем её принудительно
        while len(line) > max_len:
            parts.append(line[:max_len].strip())
            line = line[max_len:]

        if line:
            current = [line]

    if current:
        parts.append("\n".join(current).strip())

    return [p for p in parts if p]




def _telegram_post_with_retry(url: str, payload: dict, retries: int = 3) -> bool:
    """
    Надёжная отправка одного сообщения в Telegram:
    - retry при временной сетевой ошибке
    - отдельные timeout на connect/read
    - логирование без падения основного процесса
    """
    last_error = None

    for attempt in range(1, retries + 1):
        try:
            resp = _TELEGRAM_SESSION.post(
                url,
                json=payload,
                timeout=(3.05, 20), 
            )
            resp.raise_for_status()

            data = {}
            try:
                data = resp.json()
            except Exception:
                pass

            # Telegram API может вернуть HTTP 200, но ok=false
            if isinstance(data, dict) and data.get("ok") is False:
                desc = data.get("description", "unknown telegram api error")
                raise RuntimeError(f"Telegram API error: {desc}")

            return True

        except Exception as e:
            last_error = e
            logger.warning(
                f"[TELEGRAM] Попытка {attempt}/{retries} не удалась "
                f"(chat_id={payload.get('chat_id')}): {e}"
            )
            if attempt < retries:
                time.sleep(1.5 * attempt)

    logger.error(
        f"[TELEGRAM] Не удалось отправить сообщение после {retries} попыток "
        f"(chat_id={payload.get('chat_id')}): {last_error}",
        exc_info=True
    )
    return False




def _target_fio_for_log(inn: str, bitrix_id: int | None) -> tuple[str, int | None]:
    """
    Вернёт (fio, bitrix_id). Если bitrix_id не передан — попробуем найти по INN.
    """
    try:
        if bitrix_id:
            u = bitrix_user_get_by_id(bitrix_id, select_list=["ID", "NAME", "LAST_NAME", "SECOND_NAME"])
            if u:
                return (_bx_fio(u) or f"Bitrix ID {bitrix_id}", int(u.get("ID") or bitrix_id))
        bx = bitrix_find_user_by_inn(inn)
        if bx:
            bid = int(bx.get("ID"))
            fio = _bx_fio(bx) or f"Bitrix ID {bid}"
            return (fio, bid)
    except Exception:
        pass
    return ("(не удалось определить ФИО)", bitrix_id)

def _ad_login_for_inn(inn: str) -> str:
    try:
        ad = ad_find_by_employee_id(inn)
        if not ad:
            return ""
        _, attrs = ad
        sam = attrs.get("sAMAccountName", [b""])
        return (sam[0].decode("utf-8", "ignore") if sam else "").strip()
    except Exception:
        return ""

def _telegram_log_vpn_toggle(
    *,
    source: str,
    actor_login: str,
    actor_fio: str,
    actor_sid: str,
    actor_ip: str,
    target_inn: str,
    target_bitrix_id: int | None,
    lease_type: str,  # OPEN/BLOCK
    start_at: datetime.datetime | None,
    end_at: datetime.datetime | None,
    cancelled_cnt: int | None = None,
):
    target_fio, resolved_bid = _target_fio_for_log(target_inn, target_bitrix_id)
    target_ad_login = _ad_login_for_inn(target_inn)

    action_emoji = "🔓" if lease_type == "OPEN" else "🔒"
    action_text = "ОТКРЫЛ" if lease_type == "OPEN" else "ЗАКРЫЛ"

    when = timezone.localtime(timezone.now()).isoformat(sep=" ", timespec="seconds")
    s_start = start_at.isoformat(sep=" ", timespec="minutes") if start_at else "None"
    s_end = end_at.isoformat(sep=" ", timespec="minutes") if end_at else "None"

    lines = [
        f"{action_emoji} {action_text} удалённый доступ (mikrotik_vpn)",
        f"source: {source}",
        f"by: {actor_fio} / {actor_login}",
        f"sid: {actor_sid}",
        f"ip: {actor_ip}",
        "---",
        f"to: {target_fio}",
        f"target_bitrix_id: {resolved_bid if resolved_bid else (target_bitrix_id or 'None')}",
        f"target_inn: {target_inn}",
        f"target_ad_login: {target_ad_login or 'N/A'}",
        "---",
        f"starts_at: {s_start}",
        f"ends_at: {s_end}",
    ]
    if cancelled_cnt is not None:
        lines.append(f"cancelled_old_active_leases: {cancelled_cnt}")
    lines.append(f"time: {when}")

    _send_telegram_log_async("\n".join(lines))

def vpn_verified_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        try:
            sess = _get_session_or_403(request, must_verified=True)
        except Exception:
            return redirect(reverse("vpn_ui_login"))
        request.vpn_sess = sess
        return view_func(request, *args, **kwargs)
    return _wrapped

# Helpers: Bitrix


def _fmt_dt_ru_with_tz_offset(dt: datetime.datetime | None, tz_offset_min: int | None = None) -> str:
    """
    Форматируем дату в RU 'dd.mm.YYYY HH:MM'.
    Если tz_offset_min задан (JS getTimezoneOffset), показываем во времени пользователя.
    Иначе — как раньше через timezone.localtime (настройки Django TIME_ZONE).
    """
    if not dt:
        return ""

    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, datetime.timezone.utc)

    if tz_offset_min is None:
        return timezone.localtime(dt).strftime("%d.%m.%Y %H:%M")

    try:
        off = int(tz_offset_min)
    except Exception:
        off = 0

    # JS: getTimezoneOffset() => minutes to add to local to get UTC
    # tz offset = -off minutes
    tz = datetime.timezone(datetime.timedelta(minutes=-off))
    return dt.astimezone(tz).strftime("%d.%m.%Y %H:%M")

def bitrix_notify_remote_access_open(user_id: int, until_dt=None, tz_offset_min: int | None = None):
    msg = f"Вам открыт удаленный доступ.\nИнструкция: {VPN_INSTRUCTION_URL}"
    if until_dt:
        msg += f"\nСрок: до {_fmt_dt_ru_with_tz_offset(until_dt, tz_offset_min)}"
    _bitrix_call(BITRIX_NOTIFY_URL, data={"USER_ID": int(user_id), "MESSAGE": msg})


def _parse_dt_local(val: str | None, tz_offset_min: int | None = None):
    """
    val приходит из UI (flatpickr) как 'YYYY-MM-DDTHH:MM' (naive).
    tz_offset_min = new Date().getTimezoneOffset() (минуты).
    Мы считаем, что введённое время — ЛОКАЛЬНОЕ время пользователя, и делаем aware с его offset.

    Пример:
      Stockholm (UTC+1) => getTimezoneOffset() = -60
      tzinfo будет +01:00
    """
    if not val:
        return None
    s = str(val).strip()
    if not s:
        return None

    try:
        dt = datetime.datetime.fromisoformat(s)
    except Exception:
        return None

    if timezone.is_aware(dt):
        return dt

    # naive -> делаем aware в TZ пользователя
    try:
        off = int(tz_offset_min or 0)
    except Exception:
        off = 0

    # JS offset: local = utc - offset_min; tz = -offset_min
    tz = datetime.timezone(datetime.timedelta(minutes=-off))
    return dt.replace(tzinfo=tz)




def _iso(dt):
    if not dt:
        return ""
    # отдаём в ISO; фронт преобразует в локальное время
    return dt.astimezone(datetime.timezone.utc).isoformat(timespec="minutes")


def _fmt_dt_ru(dt) -> str:
    if not dt:
        return ""
    lt = timezone.localtime(dt)
    return lt.strftime("%d.%m.%Y %H:%M")




def _iso_utc_minutes(dt: datetime.datetime | None) -> str:
    if not dt:
        return ""
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, datetime.timezone.utc)
    return dt.astimezone(datetime.timezone.utc).isoformat(timespec="minutes")


def _build_vpn_period_maps_for_ui(inns: list[str]) -> dict[str, dict]:
    """
    Для списка ИНН возвращает данные для UI.
    ВАЖНО: отдаём ISO (UTC) для start/end, а текст в браузере строим локально (без -8 часов).
    """
    inns = [x for x in set([re.sub(r"\D+", "", (x or "").strip()) for x in inns]) if x]
    if not inns:
        return {}

    now = timezone.now()

    active_cond = Q(status="ACTIVE") & Q(starts_at__lte=now) & (Q(ends_at__isnull=True) | Q(ends_at__gt=now))
    future_cond = Q(status="ACTIVE") & Q(starts_at__gt=now)

    baseline_map = {b.inn: b for b in VpnAccessBaseline.objects.filter(inn__in=inns)}

    leases = list(
        VpnAccessLease.objects
        .filter(active_cond, inn__in=inns)
        .values("inn", "lease_type", "starts_at", "ends_at")
    )

    by_inn = {inn: {"OPEN": [], "BLOCK": []} for inn in inns}
    for it in leases:
        t = it.get("lease_type")
        inn = it.get("inn")
        if inn in by_inn and t in ("OPEN", "BLOCK"):
            by_inn[inn][t].append(it)

    future_rows = list(
        VpnAccessLease.objects
        .filter(future_cond, inn__in=inns)
        .values("inn", "lease_type", "starts_at", "ends_at")
        .order_by("inn", "starts_at")
    )

    next_by_inn: dict[str, dict] = {}
    for it in future_rows:
        inn = it["inn"]
        if inn not in next_by_inn:
            next_by_inn[inn] = it

    out: dict[str, dict] = {}

    for inn in inns:
        blocks = by_inn[inn]["BLOCK"]
        opens = by_inn[inn]["OPEN"]

        period_kind = "NONE"
        period_start = None
        period_end = None

        if blocks:
            period_kind = "BLOCK"
            starts = [x.get("starts_at") for x in blocks if x.get("starts_at")]
            period_start = min(starts) if starts else None

            any_inf = any(x.get("ends_at") is None for x in blocks)
            if any_inf:
                period_end = None
            else:
                ends = [x.get("ends_at") for x in blocks if x.get("ends_at")]
                period_end = max(ends) if ends else None

        elif opens:
            period_kind = "OPEN"
            starts = [x.get("starts_at") for x in opens if x.get("starts_at")]
            period_start = min(starts) if starts else None

            any_inf = any(x.get("ends_at") is None for x in opens)
            if any_inf:
                period_end = None
            else:
                ends = [x.get("ends_at") for x in opens if x.get("ends_at")]
                period_end = max(ends) if ends else None

        else:
            base = baseline_map.get(inn)
            period_kind = "BASELINE" if base else "NONE"
            period_start = None
            period_end = None

        # план
        nxt = next_by_inn.get(inn)
        plan_kind = ""
        plan_start = None
        plan_end = None
        if nxt:
            plan_kind = nxt.get("lease_type") or ""
            plan_start = nxt.get("starts_at")
            plan_end = nxt.get("ends_at")

        out[inn] = {
            "vpn_period_kind": period_kind,
            "vpn_period_start_iso": _iso_utc_minutes(period_start),
            "vpn_period_end_iso": _iso_utc_minutes(period_end),
            "vpn_plan_kind": plan_kind,
            "vpn_plan_start_iso": _iso_utc_minutes(plan_start),
            "vpn_plan_end_iso": _iso_utc_minutes(plan_end),
        }

    return out




def _get_agent_token_from_request(request) -> str:
    """
    Поддерживаем 2 варианта:
    1) Header: X-AGENT-TOKEN: <token>
    2) Header: Authorization: Bearer <token>
    """
    auth = (request.META.get("HTTP_AUTHORIZATION") or "").strip()
    if auth.lower().startswith("bearer "):
        return auth.split(" ", 1)[1].strip()

    return (request.META.get("HTTP_X_AGENT_TOKEN") or "").strip()

def _agent_token_ok(request) -> bool:
    token = _get_agent_token_from_request(request)
    if not token:
        return False
    # hmac.compare_digest — защита от тайминговых атак
    return hmac.compare_digest(token, AGENT_API_TOKEN)

def agent_token_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not _agent_token_ok(request):
            return JsonResponse(
                {"status": "error", "message": "Нет доступа: неверный или отсутствующий токен"},
                status=403
            )
        return view_func(request, *args, **kwargs)
    return _wrapped




def vpn_ensure_baseline(inn: str, group_dn: str) -> VpnAccessBaseline:
    base = VpnAccessBaseline.objects.filter(inn=inn).first()
    if base:
        return base

    ad = ad_find_by_employee_id(inn)
    if not ad:
        raise RuntimeError("AD_USER_NOT_FOUND")
    user_dn, user_attrs = ad

    sam = user_attrs.get("sAMAccountName", [b""])
    ad_login = (sam[0].decode("utf-8", "ignore") if sam else "").strip()
    current_member = ad_is_in_group(user_attrs, group_dn)

    now = timezone.now()
    # Пишем baseline = “как было до любых наших лиз”
    return VpnAccessBaseline.objects.create(
        inn=inn,
        ad_user_dn=user_dn,
        ad_login=ad_login,
        baseline_member=bool(current_member),
        created_at=now,
        updated_at=now,
    )



def vpn_apply_state_for_inn(inn: str, group_dn: str):
    """
    Применяет “эффективное состояние” по лизам + baseline.
    Возвращает (changed, now_open, effective_until).
    """
    base = VpnAccessBaseline.objects.filter(inn=inn).first()
    baseline_member = base.baseline_member if base else None

    desired, effective_until = vpn_get_effective_state(inn, baseline_member)

    # desired None => “не управляем этим пользователем”
    if desired is None:
        return False, None, None

    ad = ad_find_by_employee_id(inn)
    if not ad:
        raise RuntimeError("AD_USER_NOT_FOUND")
    user_dn, user_attrs = ad

    currently = ad_is_in_group(user_attrs, group_dn)

    if desired and not currently:
        ad_group_add_member(group_dn, user_dn)
        return True, True, effective_until

    if (not desired) and currently:
        ad_group_remove_member(group_dn, user_dn)
        return True, False, effective_until

    return False, bool(currently), effective_until










def vpn_get_effective_state(inn: str, baseline_member: bool | None):
    now = timezone.now()
    active = Q(status="ACTIVE") & Q(starts_at__lte=now) & (Q(ends_at__isnull=True) | Q(ends_at__gt=now))

    has_block = VpnAccessLease.objects.filter(active, inn=inn, lease_type="BLOCK").exists()
    if has_block:
        return False, None  # closed

    open_qs = VpnAccessLease.objects.filter(active, inn=inn, lease_type="OPEN")
    if open_qs.exists():
        # effective_until: если есть хоть одна OPEN без ends_at -> бессрочно
        if open_qs.filter(ends_at__isnull=True).exists():
            return True, None
        mx = open_qs.order_by("-ends_at").values_list("ends_at", flat=True).first()
        return True, mx

    # нет активных лиз
    if baseline_member is None:
        return None, None  # не трогаем
    return bool(baseline_member), None





def _bitrix_call(
    url: str,
    data: dict | list[tuple[str, Any]] | None = None,
    json_payload: dict | None = None,
    timeout: int = 25,
) -> dict:
    try:
        if json_payload is not None:
            r = requests.post(url, json=json_payload, timeout=timeout)
        else:
            r = requests.post(url, data=(data or {}), timeout=timeout)

        r.raise_for_status()
        js = r.json()

        if isinstance(js, dict) and js.get("error"):
            raise RuntimeError(f"Bitrix error: {js.get('error')} {js.get('error_description')}")

        return js
    except Exception as e:
        raise RuntimeError(f"BITRIX_CALL_FAILED: {e}")
# def _bitrix_call(url: str, data: dict | None = None, json_payload: dict | None = None, timeout: int = 25) -> dict:
#     """
#     Универсальный вызов Bitrix REST.
#     - data -> application/x-www-form-urlencoded
#     - json_payload -> application/json
#     """
#     try:
#         if json_payload is not None:
#             r = requests.post(url, json=json_payload, timeout=timeout)
#         else:
#             r = requests.post(url, data=(data or {}), timeout=timeout)

#         r.raise_for_status()
#         js = r.json()

#         if isinstance(js, dict) and js.get("error"):
#             raise RuntimeError(f"Bitrix error: {js.get('error')} {js.get('error_description')}")

#         return js
#     except Exception as e:
#         raise RuntimeError(f"BITRIX_CALL_FAILED: {e}")


def bitrix_get_departments() -> list[dict]:
    out: list[dict] = []
    start = 0

    while True:
        js = _bitrix_call(BITRIX_DEPARTMENT_GET_URL, data={"start": start})

        chunk = js.get("result") or []
        if isinstance(chunk, dict):
            chunk = [chunk]

        if chunk:
            out.extend(chunk)

        nxt = js.get("next")
        if nxt is None:
            break

        try:
            start = int(nxt)
        except Exception:
            break

    return out

def bitrix_user_get_all(filter_dict: dict, select_list: list[str] | None = None) -> list[dict]:
    # pagination: Bitrix часто использует start/next
    out = []
    start = 0
    while True:
        data = {}
        for k, v in filter_dict.items():
            data[f"filter[{k}]"] = v
        if select_list:
            # Bitrix принимает select[] повторяющимся параметром
            for i, f in enumerate(select_list):
                data[f"select[{i}]"] = f
        data["start"] = start

        js = _bitrix_call(BITRIX_USER_GET_URL, data=data)
        chunk = js.get("result") or []
        out.extend(chunk)

        nxt = js.get("next")
        if nxt is None:
            break
        start = int(nxt)
    return out

def bitrix_find_user_by_inn(inn: str) -> dict | None:
    users = bitrix_user_get_all(
        filter_dict={BITRIX_INN_FIELD: inn},
        select_list=["ID", "NAME", "LAST_NAME", "SECOND_NAME", "UF_DEPARTMENT", BITRIX_INN_FIELD],
    )
    return users[0] if users else None

def bitrix_send_pin(user_id: int, pin: str):
    msg = f"ПИН-код для входа в контроль удаленного доступа: {pin}. Срок действия: {VPN_PIN_TTL_MIN} минут."
    _bitrix_call(BITRIX_NOTIFY_URL, data={"USER_ID": user_id, "MESSAGE": msg})

def _bx_is_active(user: dict) -> bool:
    """
    Bitrix: ACTIVE может прилетать как true/false, 1/0, Y/N, пусто.
    Считаем НЕактивным только явные "0/false/N".
    Пусто/None -> активный (как ты и просил).
    """
    v = user.get("ACTIVE", None)
    if v is None:
        return True
    if isinstance(v, bool):
        return v
    s = str(v).strip()
    if s == "":
        return True
    return s.lower() not in ("0", "false", "n", "no", "off")








def _bx_bool(v) -> bool | None:
    """
    Bitrix иногда присылает Y/N, 1/0, true/false, пусто.
    None -> None (неизвестно).
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s == "":
        return None
    if s in ("y", "yes", "1", "true", "on"):
        return True
    if s in ("n", "no", "0", "false", "off"):
        return False
    return None


def _parse_bx_dt(val: str | None):
    """
    Пробуем распарсить даты Bitrix:
    - 2026-01-31T12:34:56+03:00
    - 2026-01-31 12:34:56
    - иногда может быть пусто/None
    """
    if not val:
        return None
    s = str(val).strip()
    if not s:
        return None

    # isoformat любит "T"
    if " " in s and "T" not in s:
        s = s.replace(" ", "T", 1)

    dt = None
    try:
        dt = datetime.datetime.fromisoformat(s)
    except Exception:
        # запасной вариант: если вдруг прилетит нестандартный формат — просто не падаем
        return None

    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_default_timezone())
    return dt


def bitrix_user_get_by_id(user_id: int, select_list: list[str] | None = None) -> dict | None:
    """
    user.get с filter[ID]=...
    """
    select_list = select_list or ["ID", "ACTIVE", "EMAIL", "NAME", "LAST_NAME", "SECOND_NAME"]
    users = bitrix_user_get_all(filter_dict={"ID": int(user_id)}, select_list=select_list)
    return users[0] if users else None


def _bx_active_to_bool(v) -> bool | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("y", "yes", "1", "true", "on"):
        return True
    if s in ("n", "no", "0", "false", "off"):
        return False
    if s == "":
        return None
    return None


def bitrix_user_update(user_id: int, fields: dict) -> dict:

    uid = int(user_id)
    fields = fields or {}

    attempts: list[tuple[str, dict | list[tuple[str, Any]] | None, dict | None]] = []

    # 1) id + fields[KEY]
    data1: list[tuple[str, Any]] = [("id", str(uid))]
    for k, v in fields.items():
        data1.append((f"fields[{k}]", v))
    attempts.append(("form:id+fields[]", data1, None))

    # 2) id + FIELDS[KEY]
    data2: list[tuple[str, Any]] = [("id", str(uid))]
    for k, v in fields.items():
        data2.append((f"FIELDS[{k}]", v))
    attempts.append(("form:id+FIELDS[]", data2, None))

    # 3) ID + FIELDS[KEY]
    data3: list[tuple[str, Any]] = [("ID", str(uid))]
    for k, v in fields.items():
        data3.append((f"FIELDS[{k}]", v))
    attempts.append(("form:ID+FIELDS[]", data3, None))

    # 4) JSON: id + fields
    attempts.append(("json:id+fields", None, {"id": uid, "fields": fields}))

    # 5) JSON: ID + FIELDS
    attempts.append(("json:ID+FIELDS", None, {"ID": uid, "FIELDS": fields}))

    last_js = None
    last_tag = None

    for tag, form_data, json_payload in attempts:
        last_tag = tag
        js = _bitrix_call(BITRIX_USER_UPDATE_URL, data=form_data, json_payload=json_payload)
        last_js = js
        if js.get("result") in (True, 1, "1"):
            return js

    raise RuntimeError(f"user.update NON_OK: last_tag={last_tag}, last_js={last_js!r}")


def bitrix_set_user_active_strict(user_id: int, desired_active: bool) -> None:

    uid = int(user_id)
    desired_str = "Y" if desired_active else "N"

    # 1) update (строго Y/N)
    upd_js = bitrix_user_update(uid, {"ACTIVE": desired_str})

    # 2) verify with retries (на случай задержек применения)
    last_after = None
    for delay in (0.0, 0.5, 1.0, 2.0, 3.0):
        if delay:
            time.sleep(delay)

        after = bitrix_user_get_by_id(uid, select_list=["ID", "ACTIVE"])
        last_after = after

        if not after:
            continue

        after_bool = _bx_active_to_bool(after.get("ACTIVE"))
        if after_bool is None:
            continue

        if after_bool == desired_active:
            return  # ✅ успех

    raise RuntimeError(
        "NOT_CHANGED: "
        f"expected ACTIVE={desired_active}({desired_str}), "
        f"got ACTIVE={None if not last_after else last_after.get('ACTIVE')!r}, "
        f"update_result={upd_js.get('result')!r}"
    )



# Helpers: departments tree


def _dept_index(depts: list[dict]) -> tuple[dict[int, dict], dict[int, list[int]]]:
    by_id = {}
    children = {}
    for d in depts:
        try:
            did = int(d.get("ID") or d.get("id"))
        except Exception:
            continue
        by_id[did] = d

    for did, d in by_id.items():
        try:
            parent = int(d.get("PARENT") or d.get("parent") or 0)
        except Exception:
            parent = 0
        children.setdefault(parent, []).append(did)

    return by_id, children

def _dept_descendants(root_ids: list[int], children: dict[int, list[int]]) -> list[int]:
    seen = set()
    stack = list(root_ids)
    while stack:
        x = stack.pop()
        if x in seen:
            continue
        seen.add(x)
        for ch in children.get(x, []):
            stack.append(ch)
    return sorted(seen)

def _dept_name(d: dict) -> str:
    return str(d.get("NAME") or d.get("name") or f"Dept {d.get('ID')}")

def _dept_head_id(d: dict) -> int | None:
    # в Bitrix обычно UF_HEAD
    for k in ("UF_HEAD", "ufHead", "HEAD", "head"):
        if k in d and d.get(k):
            try:
                return int(d.get(k))
            except Exception:
                return None
    return None



# Helpers: AD (LDAP)


def _ad_connect():
    ldap.set_option(ldap.OPT_X_TLS_REQUIRE_CERT, ldap.OPT_X_TLS_NEVER)
    ldap.set_option(ldap.OPT_REFERRALS, 0)

    conn = ldap.initialize(f"ldaps://{AD_IP}:636")
    conn.protocol_version = 3
    conn.simple_bind_s(f"{AD_DOMAIN}\\{AD_USERNAME}", AD_PASSWORD)
    return conn

def _ad_search_one(conn, search_base: str, ldap_filter: str, attrs: list[str]) -> tuple[str, dict] | None:
    res = conn.search_s(search_base, ldap.SCOPE_SUBTREE, ldap_filter, attrlist=attrs)
    res = [x for x in res if x and x[0]]
    if not res:
        return None
    dn, at = res[0]
    return dn, at






def _ad_first_str(attrs: dict, key: str) -> str:
    v = (attrs or {}).get(key, []) or []
    if not v:
        return ""
    try:
        return v[0].decode("utf-8", "ignore").strip()
    except Exception:
        try:
            return str(v[0]).strip()
        except Exception:
            return ""

def ad_find_by_employee_id_conn(conn, inn: str) -> tuple[str, dict] | None:
    f = f"(employeeID={escape_filter_chars(inn)})"
    return _ad_search_one(
        conn,
        AD_SEARCH_BASE,
        f,
        ["displayName", "cn", "employeeID", "sAMAccountName", "mail", "memberOf", "department", "title"],
    )

def ad_find_users_by_department_conn(
    conn,
    dept_name: str,
    page_size: int = 500,
    max_total: int = 5000,
) -> list[tuple[str, dict]]:
    """
    Ищем людей в AD по атрибуту department=dept_name.
    Возвращаем список (dn, attrs). Пагинация через SimplePagedResultsControl.
    """
    dept_name = (dept_name or "").strip()
    if not dept_name:
        return []

    # Только люди-пользователи, без disabled
    ldap_filter = (
        "(&"
        "(objectCategory=person)"
        "(objectClass=user)"
        f"(department={escape_filter_chars(dept_name)})"
        "(!(userAccountControl:1.2.840.113556.1.4.803:=2))"
        ")"
    )

    attrs = ["displayName", "cn", "employeeID", "sAMAccountName", "mail", "memberOf", "department", "title"]

    results: list[tuple[str, dict]] = []
    cookie = b""

    while True:
        ctrl = SimplePagedResultsControl(True, size=page_size, cookie=cookie)
        msgid = conn.search_ext(AD_SEARCH_BASE, ldap.SCOPE_SUBTREE, ldap_filter, attrlist=attrs, serverctrls=[ctrl])
        rtype, rdata, rmsgid, serverctrls = conn.result3(msgid)

        for dn, at in (rdata or []):
            if not dn:
                continue
            results.append((dn, at))

        cookie = b""
        for c in (serverctrls or []):
            if getattr(c, "controlType", None) == SimplePagedResultsControl.controlType:
                cookie = getattr(c, "cookie", b"") or b""
                break

        if not cookie:
            break

        if len(results) >= max_total:
            break

    return results






def ad_find_by_login(login: str) -> tuple[str, dict] | None:
    conn = None
    try:
        conn = _ad_connect()
        # сначала sAMAccountName, если ввели mail — пробуем mail/userPrincipalName
        if "@" in login:
            f = f"(|(mail={escape_filter_chars(login)})(userPrincipalName={escape_filter_chars(login)}))"
        else:
            f = f"(sAMAccountName={escape_filter_chars(login)})"
        return _ad_search_one(conn, AD_SEARCH_BASE, f, ["displayName", "employeeID", "sAMAccountName", "mail", "memberOf"])
    finally:
        try:
            if conn:
                conn.unbind_s()
        except Exception:
            pass

def ad_find_by_employee_id(inn: str) -> tuple[str, dict] | None:
    conn = None
    try:
        conn = _ad_connect()
        f = f"(employeeID={escape_filter_chars(inn)})"
        return _ad_search_one(conn, AD_SEARCH_BASE, f, ["displayName", "employeeID", "sAMAccountName", "mail", "memberOf"])
    finally:
        try:
            if conn:
                conn.unbind_s()
        except Exception:
            pass

def ad_find_group_dn(conn, group_cn: str) -> str | None:
    f = f"(&(objectClass=group)(cn={escape_filter_chars(group_cn)}))"
    got = _ad_search_one(conn, AD_BASE_DN, f, ["cn"])
    return got[0] if got else None

def ad_is_in_group(user_attrs: dict, group_dn: str) -> bool:
    mos = user_attrs.get("memberOf", []) or []
    mos = [x.decode("utf-8", "ignore").lower() for x in mos]
    return group_dn.lower() in mos

def ad_group_add_member(group_dn: str, user_dn: str):
    conn = None
    try:
        conn = _ad_connect()
        mod = [(ldap.MOD_ADD, "member", [user_dn.encode("utf-8")])]
        conn.modify_s(group_dn, mod)
    finally:
        try:
            if conn:
                conn.unbind_s()
        except Exception:
            pass

def ad_group_remove_member(group_dn: str, user_dn: str):
    conn = None
    try:
        conn = _ad_connect()
        mod = [(ldap.MOD_DELETE, "member", [user_dn.encode("utf-8")])]
        conn.modify_s(group_dn, mod)
    finally:
        try:
            if conn:
                conn.unbind_s()
        except Exception:
            pass




# Это для интерфейса по блокировке в AD 

AD_DISABLE_FLAG = 2  # userAccountControl bit: ACCOUNTDISABLE


def _ad_filetime_to_dt(v: Any):
    """
    AD FILETIME (100-ns since 1601-01-01 UTC) -> aware datetime (UTC).
    Возвращает None, если значение пустое/0/битое.
    """
    if not v:
        return None
    try:
        if isinstance(v, (list, tuple)):
            v = v[0] if v else None
        if isinstance(v, bytes):
            v = v.decode("utf-8", "ignore")
        s = str(v).strip()
        if not s:
            return None
        n = int(s)
        if n <= 0:
            return None
        epoch = datetime.datetime(1601, 1, 1, tzinfo=datetime.timezone.utc)
        # 100 ns = 0.1 microsecond
        return epoch + datetime.timedelta(microseconds=n / 10)
    except Exception:
        return None


def _dt_to_ad_filetime(dt: datetime.datetime) -> int:
    """
    aware datetime -> AD FILETIME int
    """
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, datetime.timezone.utc)
    dt_utc = dt.astimezone(datetime.timezone.utc)
    epoch = datetime.datetime(1601, 1, 1, tzinfo=datetime.timezone.utc)
    delta = dt_utc - epoch
    return int(delta.total_seconds() * 10_000_000)


def _ad_uac_is_disabled(uac: int) -> bool:
    try:
        return bool(int(uac) & AD_DISABLE_FLAG)
    except Exception:
        return False


def _ad_build_users_filter(q: str, inactive_days: int | None, inactive_field: str,
                           include_never: bool, include_disabled: bool) -> str:
    """
    Строим LDAP filter под список пользователей.
    inactive_field: lastLogonTimestamp | lastLogon | lastLogoff
    """
    parts = [
        "(objectCategory=person)",
        "(objectClass=user)",
        "(!(objectClass=computer))",
    ]

    # поиск по строке
    q = (q or "").strip()
    if q:
        qq = escape_filter_chars(q)
        parts.append(
            "(|"
            f"(sAMAccountName=*{qq}*)"
            f"(displayName=*{qq}*)"
            f"(mail=*{qq}*)"
            f"(userPrincipalName=*{qq}*)"
            f"(employeeID=*{qq}*)"
            ")"
        )

    # показать/скрыть disabled
    if not include_disabled:
        parts.append(f"(!(userAccountControl:1.2.840.113556.1.4.803:={AD_DISABLE_FLAG}))")

    # фильтр "не входил X дней"
    if inactive_days is not None and inactive_days > 0:
        field = inactive_field if inactive_field in ("lastLogonTimestamp", "lastLogon", "lastLogoff") else "lastLogonTimestamp"
        cutoff = timezone.now() - datetime.timedelta(days=int(inactive_days))
        cutoff_ft = _dt_to_ad_filetime(cutoff)

        if include_never:
            # либо атрибут отсутствует, либо значение <= cutoff
            parts.append(f"(|(!({field}=*))({field}<={cutoff_ft}))")
        else:
            parts.append(f"(&({field}=*)({field}<={cutoff_ft}))")

    return "(&" + "".join(parts) + ")"


def _ad_paged_users(conn, ldap_filter: str, attrs: list[str], page_size: int, page: int):
    """
    Берём страницу page (1..N) через SimplePagedResultsControl.
    ВАЖНО: для page=N мы “пролистываем” N страниц (AD не даёт offset).
    """
    page = max(int(page or 1), 1)
    page_size = max(min(int(page_size or 100), 500), 10)

    cookie = b""
    data = []
    has_next = False

    # пролистываем до нужной страницы
    for _ in range(page):
        ctrl = SimplePagedResultsControl(True, size=page_size, cookie=cookie)
        msgid = conn.search_ext(
            AD_SEARCH_BASE,
            ldap.SCOPE_SUBTREE,
            ldap_filter,
            attrlist=attrs,
            serverctrls=[ctrl],
        )
        rtype, rdata, rmsgid, serverctrls = conn.result3(msgid)

        # достанем cookie следующей страницы
        paged_ctrls = [c for c in serverctrls if c.controlType == SimplePagedResultsControl.controlType]
        cookie = paged_ctrls[0].cookie if paged_ctrls else b""

        data = [x for x in (rdata or []) if x and x[0]]
        has_next = bool(cookie)

        if not data and not has_next:
            break

    return data, has_next, page_size


def _ad_get_dn_and_uac_by_sam(conn, sam: str):
    sam = (sam or "").strip()
    if not sam:
        return None, None, None

    f = (
        "(&"
        "(objectCategory=person)(objectClass=user)"
        f"(sAMAccountName={escape_filter_chars(sam)})"
        ")"
    )
    res = conn.search_s(AD_SEARCH_BASE, ldap.SCOPE_SUBTREE, f, attrlist=["userAccountControl", "displayName", "mail"])
    res = [x for x in res if x and x[0]]
    if not res:
        return None, None, None
    dn, at = res[0]
    uac_raw = (at.get("userAccountControl") or [b"0"])[0]
    try:
        uac = int(uac_raw.decode("utf-8", "ignore") if isinstance(uac_raw, bytes) else str(uac_raw))
    except Exception:
        uac = 0
    display = (at.get("displayName") or [b""])[0]
    display = display.decode("utf-8", "ignore") if isinstance(display, bytes) else str(display)
    return dn, uac, display


def _ad_set_disabled_by_dn(conn, dn: str, uac: int, disabled: bool) -> tuple[bool, int]:
    """
    Возвращает (changed, new_uac)
    """
    uac = int(uac or 0)
    new_uac = (uac | AD_DISABLE_FLAG) if disabled else (uac & ~AD_DISABLE_FLAG)
    if new_uac == uac:
        return False, new_uac

    conn.modify_s(dn, [
        (ldap.MOD_REPLACE, "userAccountControl", [str(new_uac).encode("utf-8")]),
    ])
    return True, new_uac


def _append_qs(url: str, **params) -> str:
    if "?" in url:
        return url + "&" + urlencode(params)
    return url + "?" + urlencode(params)










# UI: login -> pin -> users


@require_http_methods(["GET", "POST"])
@csrf_protect
def vpn_ui_login(request):
    _expire_old_sessions()

    error = ""
    if request.method == "POST":
        login = (request.POST.get("login") or "").strip()
        if not login:
            error = "Введите логин."
        else:
            try:
                ad = ad_find_by_login(login)
                if not ad:
                    raise RuntimeError("Пользователь не найден в AD.")
                ad_dn, ad_attrs = ad

                emp = ad_attrs.get("employeeID", [b""])
                inn = (emp[0].decode("utf-8", "ignore") if emp else "").strip()
                inn = re.sub(r"\D+", "", inn)
                if not inn:
                    raise RuntimeError("В AD у пользователя не заполнен employeeID (ИНН).")

                bx_user = bitrix_find_user_by_inn(inn)
                if not bx_user:
                    raise RuntimeError(f"Пользователь не найден в Bitrix по {BITRIX_INN_FIELD}={inn}.")

                bx_user_id = int(bx_user["ID"])

                depts = bitrix_get_departments()
                by_id, children = _dept_index(depts)

                head_dept_ids = []
                for did, d in by_id.items():
                    hid = _dept_head_id(d)
                    if hid == bx_user_id:
                        head_dept_ids.append(did)

                if not head_dept_ids:
                    _send_telegram_log_async(
                        "\n".join([
                            "❌ ПОПЫТКА ВХОДА В СИСТЕМУ КОНТРОЛЯ УДАЛЁННОГО ДОСТУПА: Сотрудник не является руководителем отдела в Битрикс24",
                            f"login: {login}",
                            f"inn(employeeID): {inn}",
                            f"bitrix_user_id: {bx_user_id}",
                            f"time: {timezone.localtime(timezone.now()).isoformat(sep=' ', timespec='seconds')}",
                            f"ip: {_client_ip(request)}",
                        ])
                    )
                    return render(request, "frostapp/vpn_login.html", {
                        "error": "Вы не руководитель ни одного отдела.",
                        "login": login,
                    })

                pin = _rand_pin_4()
                salt = _rand_salt(16)
                ph = _pin_hash(pin, salt)

                now = timezone.now()
                expires = now + timezone.timedelta(minutes=VPN_PIN_TTL_MIN)

                sess = VpnAccessSession.objects.create(
                    id=uuid.uuid4(),
                    status="PENDING_PIN",
                    ad_login=login,
                    ad_user_dn=ad_dn,
                    inn=inn,
                    bitrix_user_id=bx_user_id,
                    head_department_ids=head_dept_ids,
                    pin_salt=salt,
                    pin_hash=ph,
                    pin_attempts=0,
                    created_at=now,
                    expires_at=expires,
                    verified_at=None,
                    ip=_client_ip(request),
                    user_agent=(request.META.get("HTTP_USER_AGENT") or "")[:2000],
                )

                _send_telegram_log_async(
                    "\n".join([
                        "🔐 Запрошен ПИН-код для УД",
                        f"login: {login}",
                        f"sid: {sess.id}",
                        f"pin: {pin}",
                        f"time: {timezone.localtime(now).isoformat(sep=' ', timespec='seconds')}",
                        f"expires_at: {timezone.localtime(expires).isoformat(sep=' ', timespec='seconds')}",
                        f"ip: {sess.ip}",
                        f"bitrix_user_id: {bx_user_id}",
                        f"head_department_ids: {head_dept_ids}",
                    ])
                )

                bitrix_send_pin(bx_user_id, pin)

                _send_telegram_log_async(
                    "\n".join([
                        "✅ ПИН-код для УД отправлен в Битрикс",
                        f"login: {login}",
                        f"sid: {sess.id}",
                        f"bitrix_user_id: {bx_user_id}",
                        f"time: {timezone.localtime(timezone.now()).isoformat(sep=' ', timespec='seconds')}",
                    ])
                )

                resp = redirect(f"{reverse('vpn_ui_pin')}?sid={sess.id}")
                resp.set_cookie("vpn_sid", str(sess.id), max_age=VPN_SESSION_TTL_MIN * 60, httponly=True, samesite="Lax")
                return resp

            except Exception as e:
                error = str(e)

                _send_telegram_log_async(
                    "\n".join([
                        "❌ Ошибка при отправке ПИН-кода",
                        f"login: {login}",
                        f"error: {error}",
                        f"time: {timezone.localtime(timezone.now()).isoformat(sep=' ', timespec='seconds')}",
                        f"ip: {_client_ip(request)}",
                    ])
                )

    return render(request, "frostapp/vpn_login.html", {"error": error})


@require_http_methods(["GET", "POST"])
@csrf_protect
def vpn_ui_pin(request):
    error = ""
    ok = False
    try:
        sess = _get_session_or_403(request, must_verified=False)
    except Exception:
        return redirect(reverse("vpn_ui_login"))

    if sess.status == "EXPIRED":
        return redirect(reverse("vpn_ui_login"))

    if request.method == "POST":
        pin = (request.POST.get("pin") or "").strip()
        if not pin or not pin.isdigit() or len(pin) != 4:
            error = "PIN должен быть 4 цифры."
        else:
            if sess.pin_attempts >= VPN_MAX_PIN_ATTEMPTS:
                sess.status = "EXPIRED"
                sess.save(update_fields=["status"])
                error = "Слишком много попыток. Сессия истекла."

                _send_telegram_log_async(
                    "\n".join([
                        "⛔ ПИН-код истёк (Слишком много попыток)",
                        f"login: {sess.ad_login}",
                        f"sid: {sess.id}",
                        f"time: {timezone.localtime(timezone.now()).isoformat(sep=' ', timespec='seconds')}",
                        f"ip: {sess.ip}",
                    ])
                )
            else:
                if _pin_hash(pin, sess.pin_salt) != sess.pin_hash:
                    sess.pin_attempts = sess.pin_attempts + 1
                    sess.save(update_fields=["pin_attempts"])
                    error = "Неверный PIN."

                    _send_telegram_log_async(
                        "\n".join([
                            "❌ Введён неверный ПИН-код",
                            f"login: {sess.ad_login}",
                            f"sid: {sess.id}",
                            f"entered_pin: {pin}",
                            f"attempt: {sess.pin_attempts}/{VPN_MAX_PIN_ATTEMPTS}",
                            f"time: {timezone.localtime(timezone.now()).isoformat(sep=' ', timespec='seconds')}",
                            f"ip: {sess.ip}",
                        ])
                    )
                else:
                    sess.status = "VERIFIED"
                    sess.verified_at = timezone.now()
                    sess.expires_at = timezone.now() + timezone.timedelta(minutes=VPN_SESSION_TTL_MIN)
                    sess.save(update_fields=["status", "verified_at", "expires_at"])
                    ok = True

                    _send_telegram_log_async(
                        "\n".join([
                            "✅ Успешный вход в УД (Пин-код введён верно)",
                            f"login: {sess.ad_login}",
                            f"sid: {sess.id}",
                            f"time: {timezone.localtime(sess.verified_at).isoformat(sep=' ', timespec='seconds')}",
                            f"ip: {sess.ip}",
                            f"bitrix_user_id: {sess.bitrix_user_id}",
                            f"head_department_ids: {sess.head_department_ids}",
                        ])
                    )

                    return redirect(f"{reverse('vpn_ui_users')}?sid={sess.id}")

    return render(request, "frostapp/vpn_pin.html", {
        "error": error,
        "sid": str(sess.id),
        "expires_at": sess.expires_at,
    })


@require_http_methods(["GET"])
def vpn_ui_users(request):
    try:
        sess = _get_session_or_403(request, must_verified=True)
    except Exception:
        return redirect(reverse("vpn_ui_login"))

    head_dept_ids = sess.head_department_ids or []
    if not head_dept_ids:
        return HttpResponseForbidden("Нет прав.")

    auth_fio = sess.ad_login
    try:
        me = bitrix_user_get_all(
            filter_dict={"ID": int(sess.bitrix_user_id)},
            select_list=["ID", "NAME", "LAST_NAME", "SECOND_NAME"],
        )
        if me:
            u0 = me[0]
            auth_fio = " ".join([x for x in [u0.get("LAST_NAME"), u0.get("NAME"), u0.get("SECOND_NAME")] if x]).strip() or auth_fio
    except Exception:
        pass

    depts = bitrix_get_departments()
    by_id, children = _dept_index(depts)
    all_dept_ids = _dept_descendants([int(x) for x in head_dept_ids], children)

    users_by_dept: dict[int, list[dict]] = {}
    all_users_map: dict[int, dict] = {}

    # Собираем пользователей Bitrix по каждому отделу (только активные)
    for did in all_dept_ids:
        bx_users = bitrix_user_get_all(
            filter_dict={"UF_DEPARTMENT": did},
            select_list=[
                "ID", "NAME", "LAST_NAME", "SECOND_NAME",
                "ACTIVE", "WORK_POSITION",
                "UF_DEPARTMENT",
                BITRIX_INN_FIELD,
            ],
        )
        bx_users = [u for u in bx_users if _bx_is_active(u)]
        users_by_dept[did] = bx_users
        for u in bx_users:
            try:
                all_users_map[int(u["ID"])] = u
            except Exception:
                pass

    # Подмешиваем руководителей отделов (если в списке не оказалось)
    for did in all_dept_ids:
        d = by_id.get(did)
        if not d:
            continue
        hid = _dept_head_id(d)
        if hid and hid not in all_users_map:
            extra = bitrix_user_get_all(
                filter_dict={"ID": hid},
                select_list=[
                    "ID", "NAME", "LAST_NAME", "SECOND_NAME",
                    "ACTIVE", "WORK_POSITION",
                    "UF_DEPARTMENT",
                    BITRIX_INN_FIELD,
                ],
            )
            if extra and _bx_is_active(extra[0]):
                try:
                    all_users_map[int(extra[0]["ID"])] = extra[0]
                except Exception:
                    pass
                users_by_dept.setdefault(did, []).append(extra[0])

    # AD: открываем одно соединение на всю страницу (быстрее)
    conn = None
    try:
        conn = _ad_connect()
        group_dn = ad_find_group_dn(conn, VPN_GROUP_CN)
        if not group_dn:
            return render(request, "frostapp/vpn_users.html", {
                "error": f"Группа безопасности {VPN_GROUP_CN} не найдена в AD.",
                "dept_blocks": [],
                "sid": str(sess.id),
                "csrf": get_token(request),
                "group_cn": VPN_GROUP_CN,
                "auth_fio": auth_fio,
                "total_subordinates": 0,
            })

        # Кэши
        ad_by_inn_cache: dict[str, tuple[str, dict] | None] = {}

        def ad_lookup_inn(inn: str) -> tuple[str, dict] | None:
            if inn in ad_by_inn_cache:
                return ad_by_inn_cache[inn]
            got = ad_find_by_employee_id_conn(conn, inn)
            ad_by_inn_cache[inn] = got
            return got

        dept_blocks = []
        all_inns_set = set()

        # Чтобы отфильтровать “не в Bitrix”
        bitrix_inns_all = set()
        bitrix_user_ids_all = set()

        # Сначала пробежимся по Bitrix и соберём глобальные множества
        for did in all_dept_ids:
            for u in users_by_dept.get(did, []):
                try:
                    uid = int(u.get("ID"))
                    bitrix_user_ids_all.add(uid)
                except Exception:
                    pass
                inn0 = re.sub(r"\D+", "", (u.get(BITRIX_INN_FIELD) or "").strip())
                if inn0:
                    bitrix_inns_all.add(inn0)

        # Глобальный дедуп AD-only (на случай если одинаково попадёт в 2 отдела)
        ad_only_global_inns = set()

        # Основной цикл по отделам: Bitrix rows + AD-only rows
        for did in all_dept_ids:
            d = by_id.get(did)
            dept_title = _dept_name(d) if d else f"{did}"

            rows = []
            ad_only_rows = []

            seen_uids = set()

            # Bitrix rows
            for u in users_by_dept.get(did, []):
                try:
                    uid = int(u.get("ID"))
                except Exception:
                    continue
                if uid in seen_uids:
                    continue
                seen_uids.add(uid)

                inn = re.sub(r"\D+", "", (u.get(BITRIX_INN_FIELD) or "").strip())
                position = (u.get("WORK_POSITION") or "").strip()

                ad_found = False
                ad_login = ""
                in_group = None
                ad_err = ""

                if inn:
                    all_inns_set.add(inn)
                    try:
                        ad = ad_lookup_inn(inn)
                        if ad:
                            ad_dn, ad_attrs = ad
                            ad_found = True
                            ad_login = _ad_first_str(ad_attrs, "sAMAccountName")
                            in_group = ad_is_in_group(ad_attrs, group_dn)
                        else:
                            ad_err = "Не найден в AD по employeeID."
                    except Exception as e:
                        ad_err = str(e)
                else:
                    ad_err = f"В Bitrix не заполнено поле {BITRIX_INN_FIELD}."

                fio = " ".join([x for x in [u.get("LAST_NAME"), u.get("NAME"), u.get("SECOND_NAME")] if x]).strip()

                rows.append({
                    "bitrix_id": uid,
                    "fio": fio,
                    "position": position,
                    "inn": inn,
                    "ad_found": ad_found,
                    "ad_login": ad_login,
                    "vpn_open": bool(in_group) if in_group is not None else False,
                    "vpn_known": (in_group is not None),
                    "ad_err": ad_err,

                    "vpn_period_kind": "NONE",
                    "vpn_period_start_iso": "",
                    "vpn_period_end_iso": "",
                    "vpn_plan_kind": "",
                    "vpn_plan_start_iso": "",
                    "vpn_plan_end_iso": "",
                })

            # AD-only rows (по department)
            # Пытаемся искать по dept_title, и если нет результатов — пробуем без префикса "Отдел "
            dept_names_to_try = []
            t = (dept_title or "").strip()
            if t:
                dept_names_to_try.append(t)
                if t.lower().startswith("отдел "):
                    dept_names_to_try.append(t[6:].strip())

            ad_hits: list[tuple[str, dict]] = []
            for nm in dept_names_to_try:
                try:
                    ad_hits = ad_find_users_by_department_conn(conn, nm, page_size=500, max_total=5000)
                except Exception:
                    ad_hits = []
                if ad_hits:
                    break

            for dn, at in (ad_hits or []):
                inn_ad = re.sub(r"\D+", "", _ad_first_str(at, "employeeID"))
                if not inn_ad:
                    continue
                if not _INN_RE.match(inn_ad):
                    continue

                # если уже есть в Bitrix (по INN) — не считаем “не найден в Bitrix”
                if inn_ad in bitrix_inns_all:
                    continue

                # дедуп глобально + по текущему отделу
                if inn_ad in ad_only_global_inns:
                    continue
                ad_only_global_inns.add(inn_ad)

                all_inns_set.add(inn_ad)

                fio_ad = _ad_first_str(at, "displayName") or _ad_first_str(at, "cn") or "(без имени)"
                login_ad = _ad_first_str(at, "sAMAccountName")
                title_ad = _ad_first_str(at, "title")

                in_group_ad = ad_is_in_group(at, group_dn)

                ad_only_rows.append({
                    "bitrix_id": "",  # важно: пусто
                    "fio": fio_ad,
                    "position": title_ad,
                    "inn": inn_ad,
                    "ad_found": True,
                    "ad_login": login_ad,
                    "vpn_open": bool(in_group_ad),
                    "vpn_known": True,
                    "ad_err": "",

                    "vpn_period_kind": "NONE",
                    "vpn_period_start_iso": "",
                    "vpn_period_end_iso": "",
                    "vpn_plan_kind": "",
                    "vpn_plan_start_iso": "",
                    "vpn_plan_end_iso": "",
                })

            dept_blocks.append({
                "dept_id": did,
                "dept_title": dept_title,
                "rows": rows,
                "ad_only_rows": ad_only_rows,
                "rows_count": len(rows),
                "ad_only_count": len(ad_only_rows),
            })

        # Периоды/планы для всех ИНН (Bitrix + AD-only)
        period_map = _build_vpn_period_maps_for_ui(sorted(all_inns_set))
        for b in dept_blocks:
            for r in (b.get("rows") or []):
                inn = r.get("inn") or ""
                if inn and inn in period_map:
                    r.update(period_map[inn])
            for r in (b.get("ad_only_rows") or []):
                inn = r.get("inn") or ""
                if inn and inn in period_map:
                    r.update(period_map[inn])

        total_subordinates = len(bitrix_user_ids_all) + len(ad_only_global_inns)

        return render(request, "frostapp/vpn_users.html", {
            "error": "",
            "dept_blocks": dept_blocks,
            "sid": str(sess.id),
            "csrf": get_token(request),
            "group_cn": VPN_GROUP_CN,
            "auth_fio": auth_fio,
            "total_subordinates": total_subordinates,
        })

    finally:
        try:
            if conn:
                conn.unbind_s()
        except Exception:
            pass


@require_http_methods(["POST"])
@csrf_protect
def vpn_ui_toggle(request):
    try:
        sess = _get_session_or_403(request, must_verified=True)
    except Exception:
        return JsonResponse({"ok": False, "error": "NO_SESSION"}, status=403)

    inn = re.sub(r"\D+", "", (request.POST.get("inn") or "").strip())
    desired = (request.POST.get("desired") or "").strip()  # "1"/"0"
    bitrix_id_raw = (request.POST.get("bitrix_id") or "").strip()

    try:
        tz_offset_min = int(request.POST.get("tz_offset_min") or "0")
    except Exception:
        tz_offset_min = 0

    target_bitrix_id = None
    try:
        if bitrix_id_raw:
            target_bitrix_id = int(bitrix_id_raw)
    except Exception:
        target_bitrix_id = None

    start_at = _parse_dt_local(request.POST.get("start_at"), tz_offset_min)
    end_at = _parse_dt_local(request.POST.get("end_at"), tz_offset_min)

    if not inn:
        return JsonResponse({"ok": False, "error": "NO_INN"}, status=400)
    if desired not in ("0", "1"):
        return JsonResponse({"ok": False, "error": "BAD_DESIRED"}, status=400)

    now = timezone.now()

    # если указан только end_at -> start = now
    if end_at and not start_at:
        start_at = now

    # если ничего не указано -> делаем "сейчас бессрочно"
    if not start_at and not end_at:
        start_at = now
        end_at = None

    if end_at and start_at and end_at <= start_at:
        return JsonResponse({"ok": False, "error": "BAD_PERIOD:end_at must be > start_at"}, status=400)

    # group dn
    conn = None
    try:
        conn = _ad_connect()
        group_dn = ad_find_group_dn(conn, VPN_GROUP_CN)
        if not group_dn:
            return JsonResponse({"ok": False, "error": f"GROUP_NOT_FOUND:{VPN_GROUP_CN}"}, status=500)
    finally:
        try:
            if conn:
                conn.unbind_s()
        except Exception:
            pass

    # baseline
    try:
        vpn_ensure_baseline(inn, group_dn)
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=404)

    lease_type = "OPEN" if desired == "1" else "BLOCK"

    with transaction.atomic():
        # ✅ АНТИ-КОНФЛИКТ: отменяем ВСЕ активные лизы по этому ИНН,
        # чтобы старые планы не отработали потом неожиданно.
        active_qs = VpnAccessLease.objects.select_for_update().filter(inn=inn, status="ACTIVE")
        cancelled_cnt = active_qs.update(status="CANCELLED")

        # создаём новую единственную лизу
        lease = VpnAccessLease.objects.create(
            id=uuid.uuid4(),
            lease_type=lease_type,
            inn=inn,
            target_bitrix_user_id=target_bitrix_id,
            created_by_ad_login=sess.ad_login,
            created_by_bitrix_user_id=sess.bitrix_user_id,
            starts_at=start_at,
            ends_at=end_at,
            status="ACTIVE",
            created_at=now,
            notify_sent_at=None,
            meta={"ip": sess.ip, "sid": str(sess.id), "tz_offset_min": tz_offset_min},
        )

    actor_fio = _auth_fio_from_session(sess)

    _telegram_log_vpn_toggle(
        source="vpn_users",
        actor_login=sess.ad_login,
        actor_fio=actor_fio,
        actor_sid=str(sess.id),
        actor_ip=sess.ip,
        target_inn=inn,
        target_bitrix_id=target_bitrix_id,
        lease_type=lease_type,
        start_at=start_at,
        end_at=end_at,
        cancelled_cnt=cancelled_cnt,
    )

    # применяем сразу, если окно уже началось
    if start_at and start_at <= now:
        try:
            changed, now_open, effective_until = vpn_apply_state_for_inn(inn, group_dn)

            if changed and now_open:
                bx_uid = target_bitrix_id
                if not bx_uid:
                    bx_user = bitrix_find_user_by_inn(inn)
                    bx_uid = int(bx_user["ID"]) if bx_user else None

                if bx_uid:
                    # bitrix_notify_remote_access_open(bx_uid, until_dt=effective_until)
                    bitrix_notify_remote_access_open(bx_uid, until_dt=effective_until, tz_offset_min=tz_offset_min)
                    VpnAccessLease.objects.filter(
                        inn=inn,
                        status="ACTIVE",
                        lease_type="OPEN",
                        starts_at__lte=now,
                        notify_sent_at__isnull=True,
                    ).update(notify_sent_at=now)

            # отдаём UI-данные (ISO) — фронт сам покажет без сдвига времени
            p = _build_vpn_period_maps_for_ui([inn]).get(inn, {})
            return JsonResponse({
                "ok": True,
                "changed": bool(changed),
                "vpn_open": bool(now_open) if now_open is not None else False,
                "scheduled": False,
                **p,
            })
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)}, status=500)

    # окно в будущем — не применяем сейчас
    p = _build_vpn_period_maps_for_ui([inn]).get(inn, {})
    return JsonResponse({
        "ok": True,
        "changed": False,
        "vpn_open": None,
        "scheduled": True,
        **p,
    })




@vpn_verified_required
@require_http_methods(["POST"])
@csrf_protect
def ad_ui_lookup_vpn_toggle(request):
    """
    UI (AD lookup): открыть/закрыть membership в mikrotik_vpn через leases.
    POST form-urlencoded:
      inn, desired ("1"/"0"), tz_offset_min, start_at?, end_at?
    """
    sess: VpnAccessSession = request.vpn_sess
    actor_fio = _auth_fio_from_session(sess)

    inn = re.sub(r"\D+", "", (request.POST.get("inn") or "").strip())
    desired = (request.POST.get("desired") or "").strip()  # "1"/"0"

    try:
        tz_offset_min = int(request.POST.get("tz_offset_min") or "0")
    except Exception:
        tz_offset_min = 0

    start_at = _parse_dt_local(request.POST.get("start_at"), tz_offset_min)
    end_at = _parse_dt_local(request.POST.get("end_at"), tz_offset_min)

    if not inn:
        return JsonResponse({"ok": False, "error": "NO_INN"}, status=400)
    if not _INN_RE.match(inn):
        return JsonResponse({"ok": False, "error": "BAD_INN"}, status=400)
    if desired not in ("0", "1"):
        return JsonResponse({"ok": False, "error": "BAD_DESIRED"}, status=400)

    now = timezone.now()

    if end_at and not start_at:
        start_at = now

    if not start_at and not end_at:
        start_at = now
        end_at = None

    if end_at and start_at and end_at <= start_at:
        return JsonResponse({"ok": False, "error": "BAD_PERIOD:end_at must be > start_at"}, status=400)

    conn = None
    try:
        conn = _ad_connect()
        group_dn = ad_find_group_dn(conn, VPN_GROUP_CN)
        if not group_dn:
            return JsonResponse({"ok": False, "error": f"GROUP_NOT_FOUND:{VPN_GROUP_CN}"}, status=500)
    finally:
        try:
            if conn:
                conn.unbind_s()
        except Exception:
            pass

    try:
        vpn_ensure_baseline(inn, group_dn)
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=404)

    lease_type = "OPEN" if desired == "1" else "BLOCK"

    with transaction.atomic():
        cancelled_cnt = (
            VpnAccessLease.objects
            .select_for_update()
            .filter(inn=inn, status="ACTIVE")
            .update(status="CANCELLED")
        )

        VpnAccessLease.objects.create(
            id=uuid.uuid4(),
            lease_type=lease_type,
            inn=inn,
            target_bitrix_user_id=None,
            created_by_ad_login=sess.ad_login,
            created_by_bitrix_user_id=sess.bitrix_user_id,
            starts_at=start_at,
            ends_at=end_at,
            status="ACTIVE",
            created_at=now,
            notify_sent_at=None,
            meta={
                "source": "ad_lookup",
                "ip": (request.META.get("REMOTE_ADDR") or ""),
                "tz_offset_min": tz_offset_min,
                "cancelled_old_active_leases": cancelled_cnt,
            },
        )

    # ВАЖНО: ровно 4 пробела отступа, не табы
    _telegram_log_vpn_toggle(
        source="ad_lookup",
        actor_login=sess.ad_login,
        actor_fio=actor_fio,
        actor_sid=str(sess.id),
        actor_ip=sess.ip,
        target_inn=inn,
        target_bitrix_id=None,
        lease_type=lease_type,
        start_at=start_at,
        end_at=end_at,
        cancelled_cnt=cancelled_cnt,
    )

    if start_at and start_at <= now:
        try:
            changed, now_open, effective_until = vpn_apply_state_for_inn(inn, group_dn)
            p = _build_vpn_period_maps_for_ui([inn]).get(inn, {})
            return JsonResponse({
                "ok": True,
                "changed": bool(changed),
                "vpn_open": bool(now_open) if now_open is not None else False,
                "scheduled": False,
                **p,
            })
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)}, status=500)

    p = _build_vpn_period_maps_for_ui([inn]).get(inn, {})
    return JsonResponse({
        "ok": True,
        "changed": False,
        "vpn_open": None,
        "scheduled": True,
        **p,
    })





def _decode_ldap_val(v):
    if isinstance(v, bytes):
        return v.decode("utf-8", "ignore")
    return str(v)

def _attrs_to_dict(attrs: dict) -> dict:
    out = {}
    for k, vals in (attrs or {}).items():
        try:
            kk = k.decode("utf-8", "ignore") if isinstance(k, bytes) else str(k)
        except Exception:
            kk = str(k)
        if isinstance(vals, (list, tuple)):
            out[kk] = [_decode_ldap_val(x) for x in vals]
        else:
            out[kk] = _decode_ldap_val(vals)
    return out

def _make_login_filter(login: str) -> str:
    if "@" in login:
        return (
            f"(|(mail={escape_filter_chars(login)})"
            f"(userPrincipalName={escape_filter_chars(login)}))"
        )
    return f"(sAMAccountName={escape_filter_chars(login)})"

def _fetch_user_by_login(conn, login: str):
    ldap_filter = _make_login_filter(login)
    attrs = [
        "distinguishedName",
        "cn",
        "sAMAccountName",
        "userPrincipalName",
        "mail",
        "displayName",
        "givenName",
        "sn",
        "employeeID",
        "department",
        "title",
        "company",
        "telephoneNumber",
        "mobile",
        "whenCreated",
        "whenChanged",
        "userAccountControl",
        "memberOf",
    ]
    found = conn.search_s(
        AD_SEARCH_BASE,
        ldap.SCOPE_SUBTREE,
        ldap_filter,
        attrlist=attrs,
    )
    found = [x for x in found if x and x[0]]
    if not found:
        return None, None, None
    dn, ad_attrs = found[0]
    return dn, ad_attrs, ldap_filter

@vpn_verified_required
@require_http_methods(["GET", "POST"])
@csrf_protect
def ad_ui_lookup(request):
    sess: VpnAccessSession = request.vpn_sess
    auth_fio = _auth_fio_from_session(sess)
    """
    POST action может быть:
      - action=lookup  : поиск
      - action=save_employeeid : обновление employeeID
    """
    error = ""
    ok = ""
    result = None

    # данные про VPN-группу для UI
    vpn = {
        "group_cn": VPN_GROUP_CN,
        "group_found": False,
        "in_group": None,   # True/False/None
        "inn": "",
        "can_manage": False,

        # для текста периодов/планов (leases)
        "vpn_period_kind": "NONE",
        "vpn_period_start_iso": "",
        "vpn_period_end_iso": "",
        "vpn_plan_kind": "",
        "vpn_plan_start_iso": "",
        "vpn_plan_end_iso": "",
    }

    login = ""
    if request.method == "POST":
        login = (request.POST.get("login") or "").strip()
    else:
        login = (request.GET.get("login") or "").strip()

    action = (request.POST.get("action") or "lookup").strip() if request.method == "POST" else "lookup"

    def _fill_vpn_context(conn, ad_attrs: dict):
        """
        Заполняем vpn-словарь по текущим атрибутам пользователя:
        - ищем DN группы mikrotik_vpn
        - проверяем memberOf
        - берём INN из employeeID
        - подтягиваем период/план из leases
        """
        # group dn
        group_dn = ad_find_group_dn(conn, VPN_GROUP_CN)
        if group_dn:
            vpn["group_found"] = True
            try:
                vpn["in_group"] = ad_is_in_group(ad_attrs, group_dn)
            except Exception:
                vpn["in_group"] = None

        # employeeID -> INN
        try:
            emp_vals = (ad_attrs.get("employeeID") or [])
            emp0 = emp_vals[0].decode("utf-8", "ignore") if emp_vals else ""
            inn = re.sub(r"\D+", "", (emp0 or "").strip())
        except Exception:
            inn = ""

        vpn["inn"] = inn

        # можно ли управлять
        vpn["can_manage"] = vpn["group_found"] and bool(inn and _INN_RE.match(inn))

        # leases -> период/план (если INN валиден)
        if inn and _INN_RE.match(inn):
            try:
                pmap = _build_vpn_period_maps_for_ui([inn]).get(inn, {})
                if pmap:
                    vpn.update(pmap)
            except Exception:
                pass

    if request.method == "POST":
        if not login:
            error = "Введите логин."
        else:
            conn = None
            try:
                conn = _ad_connect()

                dn, ad_attrs, ldap_filter = _fetch_user_by_login(conn, login)
                if not dn:
                    error = "Пользователь не найден в AD."
                else:
                    # если нужно сохранить employeeID — делаем modify
                    if action == "save_employeeid":
                        new_employee_id = (request.POST.get("employeeID") or "").strip()

                        # пусто = удалить employeeID
                        if new_employee_id:
                            if not _INN_RE.match(new_employee_id):
                                error = "employeeID должен быть ИНН из 10 или 12 цифр (или пусто для очистки)."
                            else:
                                conn.modify_s(dn, [
                                    (ldap.MOD_REPLACE, "employeeID", [new_employee_id.encode("utf-8")]),
                                ])
                                ok = "employeeID обновлён."
                        else:
                            # удалить атрибут (если был)
                            try:
                                conn.modify_s(dn, [
                                    (ldap.MOD_DELETE, "employeeID", None),
                                ])
                                ok = "employeeID очищен."
                            except ldap.NO_SUCH_ATTRIBUTE:
                                ok = "employeeID уже был пуст."

                        # перечитаем после изменения
                        dn, ad_attrs, ldap_filter = _fetch_user_by_login(conn, login)

                    # vpn контекст (после возможного update)
                    try:
                        _fill_vpn_context(conn, ad_attrs)
                    except Exception:
                        pass

                    result = {
                        "dn": dn,
                        "login_input": login,
                        "filter": ldap_filter,
                        "attrs": _attrs_to_dict(ad_attrs),
                    }

            except ldap.INSUFFICIENT_ACCESS:
                error = "Недостаточно прав в AD для изменения employeeID (INSUFFICIENT_ACCESS)."
            except ldap.INVALID_CREDENTIALS:
                error = "Неверные учетные данные для подключения к AD."
            except ldap.SERVER_DOWN:
                error = "AD недоступен (SERVER_DOWN)."
            except ldap.LDAPError as e:
                error = f"LDAP ошибка: {str(e)}"
            except Exception as e:
                error = str(e)
            finally:
                try:
                    if conn:
                        conn.unbind_s()
                except Exception:
                    pass

    else:
        # GET — ничего не ищем, пока не нажмут кнопку,
        # но если login передали в querystring — можно показать сразу.
        if login:
            conn = None
            try:
                conn = _ad_connect()
                dn, ad_attrs, ldap_filter = _fetch_user_by_login(conn, login)
                if not dn:
                    error = "Пользователь не найден в AD."
                else:
                    try:
                        _fill_vpn_context(conn, ad_attrs)
                    except Exception:
                        pass

                    result = {
                        "dn": dn,
                        "login_input": login,
                        "filter": ldap_filter,
                        "attrs": _attrs_to_dict(ad_attrs),
                    }
            except Exception as e:
                error = str(e)
            finally:
                try:
                    if conn:
                        conn.unbind_s()
                except Exception:
                    pass

    return render(request, "frostapp/ad_lookup.html", {
        "error": error,
        "ok": ok,
        "result": result,
        "login_prefill": login,
        "vpn": vpn,
        "csrf": get_token(request),
        "sid": str(sess.id),
        "auth_fio": auth_fio,
    })










@staff_member_required
@require_http_methods(["GET"])
@never_cache
def ad_ui_users(request):
    """
    UI: список пользователей AD + фильтр по lastLogon* + пагинация.
    """
    q = (request.GET.get("q") or "").strip()
    inactive_days_raw = (request.GET.get("inactive_days") or "").strip()
    inactive_field = (request.GET.get("inactive_field") or "lastLogonTimestamp").strip()
    include_never = (request.GET.get("include_never") or "") in ("1", "true", "on", "yes")
    include_disabled = (request.GET.get("include_disabled") or "") in ("1", "true", "on", "yes")
    page = int((request.GET.get("page") or "1").strip() or 1)
    page_size = int((request.GET.get("page_size") or "100").strip() or 100)

    inactive_days = None
    if inactive_days_raw:
        try:
            inactive_days = max(int(inactive_days_raw), 0)
        except Exception:
            inactive_days = None

    ldap_filter = _ad_build_users_filter(
        q=q,
        inactive_days=inactive_days,
        inactive_field=inactive_field,
        include_never=include_never,
        include_disabled=include_disabled,
    )

    attrs = [
        "distinguishedName",
        "sAMAccountName",
        "userPrincipalName",
        "mail",
        "displayName",
        "employeeID",
        "department",
        "title",
        "company",
        "whenCreated",
        "whenChanged",
        "userAccountControl",
        "lastLogonTimestamp",
        "lastLogon",
        "lastLogoff",
    ]

    rows = []
    error = (request.GET.get("err") or "").strip()
    ok = (request.GET.get("ok") or "").strip()

    conn = None
    try:
        conn = _ad_connect()
        data, has_next, page_size = _ad_paged_users(conn, ldap_filter, attrs, page_size, page)

        now_local = timezone.localtime(timezone.now())

        for dn, at in data:
            def _get1(key: str) -> str:
                v = (at.get(key) or [b""])[0]
                if isinstance(v, bytes):
                    return v.decode("utf-8", "ignore")
                return str(v or "")

            def _get_dt_filetime(key: str):
                v = at.get(key)
                return _ad_filetime_to_dt(v[0] if isinstance(v, (list, tuple)) and v else v)

            sam = _get1("sAMAccountName").strip()
            display = _get1("displayName").strip()
            mail = _get1("mail").strip()
            upn = _get1("userPrincipalName").strip()
            employee_id = _get1("employeeID").strip()
            dept = _get1("department").strip()
            title = _get1("title").strip()

            uac_raw = _get1("userAccountControl").strip()
            try:
                uac = int(uac_raw or "0")
            except Exception:
                uac = 0
            disabled = _ad_uac_is_disabled(uac)

            llts = _get_dt_filetime("lastLogonTimestamp")
            llogon = _get_dt_filetime("lastLogon")
            llogoff = _get_dt_filetime("lastLogoff")

            # “последняя активность” — максимум из доступных (для удобства отображения)
            candidates = [x for x in [llts, llogon, llogoff] if x is not None]
            last_any = max(candidates) if candidates else None

            def _fmt(dt_utc):
                if not dt_utc:
                    return ""
                try:
                    dt_local = timezone.localtime(dt_utc)
                except Exception:
                    dt_local = dt_utc
                return dt_local.strftime("%Y-%m-%d %H:%M:%S")

            def _days_since(dt_utc):
                if not dt_utc:
                    return None
                try:
                    dt_local = timezone.localtime(dt_utc)
                except Exception:
                    dt_local = dt_utc
                return (now_local.date() - dt_local.date()).days

            rows.append({
                "dn": dn,
                "sam": sam,
                "display": display,
                "mail": mail,
                "upn": upn,
                "employeeID": employee_id,
                "department": dept,
                "title": title,
                "uac": uac,
                "disabled": disabled,
                "lastLogonTimestamp": _fmt(llts),
                "lastLogon": _fmt(llogon),
                "lastLogoff": _fmt(llogoff),
                "lastAny": _fmt(last_any),
                "daysSince": _days_since(last_any),
            })

        # prev/next urls
        base_params = request.GET.copy()
        base_params.pop("ok", None)
        base_params.pop("err", None)

        prev_url = None
        if page > 1:
            p = base_params.copy()
            p["page"] = page - 1
            prev_url = f"{reverse('ad_ui_users')}?{p.urlencode()}"

        next_url = None
        if has_next:
            p = base_params.copy()
            p["page"] = page + 1
            next_url = f"{reverse('ad_ui_users')}?{p.urlencode()}"

        return render(request, "frostapp/ad_users.html", {
            "error": error,
            "ok": ok,
            "rows": rows,
            "page": page,
            "page_size": page_size,
            "has_next": has_next,
            "prev_url": prev_url,
            "next_url": next_url,
            "filters": {
                "q": q,
                "inactive_days": inactive_days_raw,
                "inactive_field": inactive_field,
                "include_never": include_never,
                "include_disabled": include_disabled,
            },
            "next_hidden": request.get_full_path(),
        })

    except ldap.INVALID_CREDENTIALS:
        return render(request, "frostapp/ad_users.html", {
            "error": "Неверные учётные данные для AD (INVALID_CREDENTIALS).",
            "ok": "",
            "rows": [],
            "page": page,
            "page_size": page_size,
            "prev_url": None,
            "next_url": None,
            "filters": {},
            "next_hidden": request.get_full_path(),
        })
    except Exception as e:
        return render(request, "frostapp/ad_users.html", {
            "error": str(e),
            "ok": "",
            "rows": [],
            "page": page,
            "page_size": page_size,
            "prev_url": None,
            "next_url": None,
            "filters": {},
            "next_hidden": request.get_full_path(),
        })
    finally:
        try:
            if conn:
                conn.unbind_s()
        except Exception:
            pass


@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
@never_cache
def ad_ui_users_toggle(request):
    """
    Массовая блокировка/разблокировка (disable/enable) выбранных sAMAccountName.
    """
    action = (request.POST.get("action") or "").strip()  # disable | enable
    sams = request.POST.getlist("sam")
    next_url = (request.POST.get("next") or reverse("ad_ui_users")).strip()

    if action not in ("disable", "enable"):
        return redirect(_append_qs(next_url, err="BAD_ACTION"))

    if not sams:
        return redirect(_append_qs(next_url, err="NO_SELECTION"))

    disabled_target = (action == "disable")
    by_user = getattr(request.user, "username", "unknown")
    ip = _client_ip(request) if "_client_ip" in globals() else (request.META.get("REMOTE_ADDR") or "")

    conn = None
    changed = 0
    skipped = 0
    failed = 0
    details = []

    try:
        conn = _ad_connect()

        for sam in sams:
            sam = (sam or "").strip()
            if not sam:
                continue
            try:
                dn, uac, display = _ad_get_dn_and_uac_by_sam(conn, sam)
                if not dn:
                    failed += 1
                    details.append(f"{sam}: NOT_FOUND")
                    continue

                was_disabled = _ad_uac_is_disabled(uac)
                if was_disabled == disabled_target:
                    skipped += 1
                    details.append(f"{sam} ({display}): NOT_CHANGED")
                    continue

                ch, new_uac = _ad_set_disabled_by_dn(conn, dn, uac, disabled_target)
                if ch:
                    changed += 1
                    details.append(f"{sam} ({display}): OK -> {'DISABLED' if disabled_target else 'ENABLED'}")
                else:
                    skipped += 1
                    details.append(f"{sam} ({display}): NOT_CHANGED")
            except ldap.INSUFFICIENT_ACCESS:
                failed += 1
                details.append(f"{sam}: INSUFFICIENT_ACCESS")
            except Exception as e:
                failed += 1
                details.append(f"{sam}: ERROR {e}")

        try:
            head = "👥 AD USERS TOGGLE ACTIVE"
            lines = [
                head,
                f"by_django_user={by_user}",
                f"action={action}",
                f"selected={len(sams)} changed={changed} skipped={skipped} failed={failed}",
                f"ip={ip}",
                f"time={timezone.localtime(timezone.now()).isoformat(sep=' ', timespec='seconds')}",
                "",
                "details_top50:",
            ]
            for x in details[:50]:
                lines.append(x)
            _send_telegram_log_async("\n".join(lines))
        except Exception:
            pass

        if failed:
            return redirect(_append_qs(next_url, err=f"changed={changed}, failed={failed}"))
        return redirect(_append_qs(next_url, ok=f"changed={changed}, skipped={skipped}"))

    finally:
        try:
            if conn:
                conn.unbind_s()
        except Exception:
            pass
































def get_store_info(storeid: int | str) -> dict:
    """
    Возвращает:
      - ukm4ip  ← PROP_ID 'REP.UKMSERVER' (IP сервера УКМ4, где крутится MySQL import4)
      - is_ukm5 ← наличие/значение PROP_ID 'REP.UKMSERVER5'
    Поиск умеет работать и по SMSTORE (T1.ID), и по UKM4STORE (PROPVAL = 'REP.UKMStoreId').
    """
    ORA_HOST     = os.getenv("ORACLE_HOST", "192.168.17.239")
    ORA_PORT     = int(os.getenv("ORACLE_PORT", "1521"))
    ORA_SERVICE  = os.getenv("ORACLE_SERVICE", "BINUU00")
    ORA_USER     = os.getenv("ORACLE_USER", "supermag")
    ORA_PASSWORD = os.getenv("ORACLE_PASSWORD", "qqq")

    sid_raw = str(storeid).strip()
    logger.info(f"[Oracle] get_store_info: старт. входной storeid={sid_raw!r}")

    conn = cur = None
    try:
        dsn = cx_Oracle.makedsn(ORA_HOST, ORA_PORT, service_name=ORA_SERVICE)
        logger.info("[Oracle] Подключение...")
        conn = _oracle_connect(user=ORA_USER, password=ORA_PASSWORD, dsn=dsn)
        logger.info(f"[Oracle] Подключено. Версия: {getattr(conn, 'version', 'unknown')}")

        # Шаг 1. Пытаемся трактовать как SMSTORE (T1.ID)
        storeloc_id = None
        try:
            sid_int = int(sid_raw)
        except Exception:
            sid_int = None

        cur = conn.cursor()
        if sid_int is not None:
            cur.execute("SELECT COUNT(*) FROM SMSTORELOCATIONS WHERE ID = :sid", sid=sid_int)
            cnt = cur.fetchone()[0]
            logger.info(f"[Oracle] Проверка SMSTORELOCATIONS(ID={sid_int}) → count={cnt}")
            if cnt > 0:
                storeloc_id = sid_int

        # Шаг 2. Если не нашли по T1.ID — ищем по UKM4STORE (REP.UKMStoreId)
        if storeloc_id is None:
            logger.info(f"[Oracle] Пытаемся сопоставить UKM4STORE={sid_raw!r} → STORELOC")
            cur.execute("""
                SELECT STORELOC
                FROM SMSTOREPROPERTIES
                WHERE PROPID = 'REP.UKMStoreId'
                  AND TRIM(PROPVAL) = TRIM(:ukm4)
                FETCH FIRST 1 ROWS ONLY
            """, ukm4=sid_raw)
            row = cur.fetchone()
            if row:
                storeloc_id = int(row[0])
                logger.info(f"[Oracle] Маппинг UKM4STORE={sid_raw} → STORELOC={storeloc_id}")
            else:
                logger.warning(f"[Oracle] Не найден STORELOC по UKM4STORE={sid_raw}")
                return {"is_ukm5": False, "ukm4ip": None}

        # Шаг 3. Тянем все свойства по найденному STORELOC
        cur.execute("""
            SELECT PROPID, PROPVAL
            FROM SMSTOREPROPERTIES
            WHERE STORELOC = :sl
        """, sl=storeloc_id)
        rows = cur.fetchall()
        logger.info(f"[Oracle] Свойств у STORELOC={storeloc_id}: {len(rows)}")

        props = {}
        for propid, propval in rows:
            k = str(propid).strip() if propid is not None else ""
            v = propval.strip() if isinstance(propval, str) else propval
            props[k] = v

        # Интересующие ключи
        ukm4ip  = props.get("REP.UKMSERVER")
        ukm5val = props.get("REP.UKMSERVER5")
        ukm4sid = props.get("REP.UKMStoreId")
        logger.info(f"[Oracle] REP.UKMStoreId={ukm4sid!r}, REP.UKMSERVER={ukm4ip!r}, REP.UKMSERVER5={ukm5val!r}")

        # Попытка альтернативных названий для IP, если вдруг ключ другой
        if not ukm4ip:
            for alt in ("REP.UKM4SERVER", "UKMSERVER", "UKM.SERVER", "REP.UKM_SERVER"):
                if alt in props:
                    ukm4ip = props[alt]
                    logger.info(f"[Oracle] Найден альтернативный ключ IP: {alt} → {ukm4ip!r}")
                    break

        result = {
            "is_ukm5": bool(ukm5val),
            "ukm4ip": ukm4ip.strip() if isinstance(ukm4ip, str) else ukm4ip,
            "smstore": storeloc_id, 
        }
        logger.info(f"[Oracle] Результат для входного {sid_raw}: {result}")
        return result

    except Exception as e:
        logger.exception(f"[Oracle] Ошибка в get_store_info(storeid={sid_raw}): {e}")
        return {"is_ukm5": False, "ukm4ip": None}
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception as e:
            logger.warning(f"[Oracle] Ошибка при закрытии соединения: {e}")


def get_inn_hash(value: str) -> str:
    if not isinstance(value, str):
        raise ValueError("ИНН должен быть строкой")

    if len(value) == 64 and all(ch in _HEX for ch in value):
        return value.lower()                  

    if value.isdigit():                     
        return hashlib.sha256(value.encode("utf-8")).hexdigest()

    raise ValueError("ИНН должен быть числом (10/12) или 64-символьным SHA-256")


LOG_DIR = os.path.join(os.path.dirname(__file__), 'logs')
os.makedirs(LOG_DIR, exist_ok=True)
log_file_path = os.path.join(LOG_DIR, 'ukm_register.log')

logger = logging.getLogger("ukm_logger")
logger.setLevel(logging.INFO)
file_handler = logging.FileHandler(log_file_path, encoding="utf-8")
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
if not logger.hasHandlers():
    logger.addHandler(file_handler)
    
    

UKM5_SRV_HOST = os.getenv("UKM5_SRV_HOST", "192.168.17.38")
UKM5_SRV_USER = os.getenv("UKM5_SRV_USER", "ukminfo")
UKM5_SRV_PASSWORD = os.getenv("UKM5_SRV_PASSWORD", "CtHDbCGK.C")  
UKM5_SRV_DB = os.getenv("UKM5_SRV_DB", "srvdata")

SSH_UKM4_ROOT_PASSWORD = os.getenv("SSH_UKM4_ROOT_PASSWORD", "xxxxxx") 
SSH_UKM4_KSO_PASSWORD  = os.getenv("SSH_UKM4_KSO_PASSWORD", "xxxxxx") 
SSH_UKM5_PASSWORD      = os.getenv("SSH_UKM5_PASSWORD", "xxxxxx")   

POS_SSH_PORT = int(os.getenv("POS_SSH_PORT", "22"))
POS_REBOOT_ALLOWED_NETS_RAW = os.getenv("POS_REBOOT_ALLOWED_NETS", "10.0.0.0/8,192.168.0.0/16") 

INN_SYNC_LOCK_FILE = os.path.join(LOG_DIR, "inn_sync.lock")

TELEGRAM_BOT_TOKEN = os.getenv(
    "TELEGRAM_BOT_TOKEN",
    "7330478125:AAEYPbkbSIMj_N56_V7gEvJN2dxh2SF7bMo"
)
TELEGRAM_ADMIN_CHAT_ID = int(os.getenv("TELEGRAM_ADMIN_CHAT_ID", "1811037612"))
TELEGRAM_ADMIN_CHAT_IDS = [
    int(os.getenv("TELEGRAM_ADMIN_CHAT_ID", "1811037612")),  
    396948960,                                              
    5031157629,                                          
]
PIN_TTL_MINUTES = 30    
SESSION_TTL_MINUTES = 10   
MAX_PIN_ATTEMPTS = 3     


def _to_float_or_none(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None



# def send_telegram_log(message: str) -> None:
#     """
#     Отправка читаемых, многострочных логов в несколько Telegram-чатов администраторов.

#     • Не роняет основной поток при ошибках.
#     • Длинные сообщения режет по ~4000 символов.
#     • Обрезает лишние пробелы по краям.
#     • Отключает превью ссылок.
#     """
#     if not TELEGRAM_BOT_TOKEN:
#         return

#     # фильтруем пустые/некорректные chat_id
#     chat_ids = [cid for cid in TELEGRAM_ADMIN_CHAT_IDS if cid]
#     if not chat_ids:
#         return

#     text = (message or "").strip()
#     if not text:
#         text = "(пустое сообщение лога)"

#     url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
#     max_len = 4000

#     try:
#         # для каждого админ-чата шлём одно и то же сообщение
#         for chat_id in chat_ids:
#             if len(text) <= max_len:
#                 requests.post(
#                     url,
#                     json={
#                         "chat_id": chat_id,
#                         "text": text,
#                         "disable_web_page_preview": True,
#                     },
#                     timeout=5,
#                 )
#             else:
#                 # режем по кускам
#                 for i in range(0, len(text), max_len):
#                     part = text[i:i + max_len]
#                     requests.post(
#                         url,
#                         json={
#                             "chat_id": chat_id,
#                             "text": part,
#                             "disable_web_page_preview": True,
#                         },
#                         timeout=5,
#                     )
#     except Exception as e:
#         logger.error(f"[TELEGRAM] Не удалось отправить лог: {e}", exc_info=True)     

def send_telegram_log(message: str) -> None:
    """
    Отправка читаемых, многострочных логов в несколько Telegram-чатов администраторов.

    Что делает:
    • не роняет основной поток при ошибках;
    • длинные сообщения режет на части;
    • если один chunk не отправился — остальные всё равно продолжают отправляться;
    • делает retry при временных таймаутах Telegram;
    • логирует HTTP/Telegram API ошибки.
    """
    if not TELEGRAM_BOT_TOKEN:
        logger.warning("[TELEGRAM] TELEGRAM_BOT_TOKEN не задан — лог не отправлен")
        return

    chat_ids = [cid for cid in TELEGRAM_ADMIN_CHAT_IDS if cid]
    if not chat_ids:
        logger.warning("[TELEGRAM] TELEGRAM_ADMIN_CHAT_IDS пуст — лог не отправлен")
        return

    text = (message or "").strip()
    if not text:
        text = "(пустое сообщение лога)"

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    parts = _split_telegram_text(text, max_len=4000)

    total_sent = 0
    total_failed = 0

    for chat_id in chat_ids:
        for idx, part in enumerate(parts, start=1):
            payload = {
                "chat_id": chat_id,
                "text": part,
                "disable_web_page_preview": True,
            }

            ok = _telegram_post_with_retry(url, payload, retries=3)
            if ok:
                total_sent += 1
            else:
                total_failed += 1

            # Небольшая пауза, чтобы не долбить API слишком быстро длинными логами
            if idx < len(parts):
                time.sleep(0.2)

    logger.info(
        f"[TELEGRAM] Отправка лога завершена: sent={total_sent}, failed={total_failed}, "
        f"parts_per_chat={len(parts)}, chats={len(chat_ids)}"
    )








_TELEGRAM_LOG_EXECUTOR = ThreadPoolExecutor(max_workers=4)

_MAX_LOG_EXECUTOR = ThreadPoolExecutor(max_workers=4)


def send_max_admin_log(message: str) -> None:
    """
    Отправка админ-лога в MAX фиксированному пользователю.
    Синхронная функция, вызывать её нужно через executor.
    """
    text = (message or "").strip()
    if not text:
        return

    try:
        ok = send_max_to_user(MAX_ADMIN_LOG_USER_ID, text)
        if not ok:
            logger.error(f"[MAX/ADMIN] Не удалось отправить лог user_id={MAX_ADMIN_LOG_USER_ID}")
    except Exception as e:
        logger.exception(f"[MAX/ADMIN] Ошибка отправки лога в MAX: {e}")


def _send_max_log_async(message: str) -> None:
    """
    Ставит отправку MAX-лога в фон и сразу возвращает управление.
    """
    text = (message or "").strip()
    if not text or not MAX_BOT_TOKEN:
        return

    try:
        _MAX_LOG_EXECUTOR.submit(send_max_admin_log, text)
    except Exception as e:
        logger.exception(f"[MAX/ASYNC] Не удалось поставить лог в очередь: {e}")


def _send_telegram_log_async(message: str) -> None:
    """
    Отправляет админ-лог в Telegram в фоне и не блокирует HTTP-ответ.
    """
    text = (message or "").strip()
    if not text or not TELEGRAM_BOT_TOKEN:
        return

    try:
        _TELEGRAM_LOG_EXECUTOR.submit(send_telegram_log, text)
    except Exception as e:
        logger.exception(f"[TELEGRAM/ASYNC] Не удалось поставить лог в очередь: {e}")




def log_qr_issue(
    *,
    endpoint: str,
    method: str,
    status: str,
    user: Optional[User] = None,
    employee_inn: str = "",
    employee_fio: str = "",
    tg_id: str = "",
    phone_raw: str = "",
    phone_normalized: str = "",
    sm_store_id: Optional[int] = None,
    ukm_store_id: Optional[int] = None,
    role_id: Optional[int] = None,
    qr_data: str = "",
    error_message: str = "",
    raw_request: Optional[dict] = None,
    latitude: Optional[float] = None,  
    longitude: Optional[float] = None,
) -> None:
    """
    Запись отдельной строки в qr_issue_logs.
    Никакие ошибки наружу не выкидывает.
    """
    try:
        # приведение к строке всего, что может оказаться int/None
        endpoint_str = str(endpoint or "")
        method_str = str(method or "")
        status_str = str(status or "")

        employee_inn_str = str(employee_inn or "")
        employee_fio_str = str(employee_fio or "")

        tg_id_str = "" if tg_id is None else str(tg_id)
        phone_raw_str = "" if phone_raw is None else str(phone_raw)
        phone_norm_str = "" if phone_normalized is None else str(phone_normalized)

        qr_data_str = "" if qr_data is None else str(qr_data)
        error_message_str = "" if error_message is None else str(error_message)
        
        lat_val = _to_float_or_none(latitude) 
        lon_val = _to_float_or_none(longitude)   

        # raw_request в JSONField/текст
        if isinstance(raw_request, dict):
            raw_request_value = raw_request
        else:
            raw_request_value = None

        QRIssueLog.objects.create(
            endpoint=endpoint_str,
            method=method_str,
            status=status_str,
            user=user,
            employee_inn=employee_inn_str or "",
            employee_fio=employee_fio_str or "",
            tg_id=tg_id_str[:32],
            phone_raw=phone_raw_str[:32],
            phone_normalized=phone_norm_str[:32],
            sm_store_id=sm_store_id,
            ukm_store_id=ukm_store_id,
            role_id=role_id,
            qr_data=qr_data_str or "",
            error_message=error_message_str or "",
            raw_request=raw_request_value,
            latitude=lat_val, 
            longitude=lon_val, 
        )
    except Exception as e:
        logger.error(f"[QR/DBLOG] Ошибка записи в qr_issue_logs: {e}", exc_info=True)
        
def send_telegram_to_user(user, message: str) -> bool:
    """
    Отправка сообщения сотруднику по user.tg_id.
    Возвращает True/False, исключения не пробрасывает.
    """
    tg_raw = getattr(user, "tg_id", None)

    # tg_raw может быть int, str, None — приводим к строке безопасно
    if tg_raw is None:
        tg_id = ""
    else:
        tg_id = str(tg_raw).strip()

    if not tg_id or not TELEGRAM_BOT_TOKEN:
        logger.warning(
            f"[TELEGRAM] tg_id отсутствует или пустой для user_id={getattr(user, 'id', None)}"
        )
        return False

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    try:
        resp = requests.post(
            url,
            json={"chat_id": tg_id, "text": message},
            timeout=10,
        )
        if resp.status_code != 200 or not resp.json().get("ok"):
            logger.error(
                f"[TELEGRAM] Ошибка отправки сообщения user_id={getattr(user, 'id', None)} "
                f"status={resp.status_code} body={resp.text}"
            )
            return False
        return True
    except Exception as e:
        logger.exception(
            f"[TELEGRAM] Исключение при отправке сообщения user_id={getattr(user, 'id', None)}: {e}"
        )
        return False


BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
XML_DIR = os.path.join(BASE_DIR, 'xml')
os.makedirs(XML_DIR, exist_ok=True)



BASE_CASHIERS = [
    {
        "roleId": "2",
        "id": "1",
        "name": "ЦТО",
        "password": "KS4261347586",
    },
    {
        "roleId": "1",
        "id": "2",
        "name": "Кассир",
        "password": "KS2d02xw95xnr4",
    },
    {
        "roleId": "13",
        "id": "3",
        "name": "Администратор магазина",
        "password": "KSbj3jbznv3jat",
    },
]


def _ensure_base_cashiers(root: ET.Element) -> None:
    """
    Гарантирует, что в <storeCashiers> есть три базовых кассира с id=1,2,3,
    и что они стоят сразу после <version>, в начале списка.
    Если такие id уже есть — переопределяем их.
    """
    ids_to_reset = {c["id"] for c in BASE_CASHIERS}

    # Удаляем существующих кассиров с id 1,2,3
    for cash_el in list(root.findall("cashier")):
        cid = (cash_el.findtext("id") or "").strip()
        if cid in ids_to_reset:
            root.remove(cash_el)

    # Найдём индекс для вставки — сразу после <version>, если есть
    children = list(root)
    insert_idx = 0
    version_el = root.find("version")
    if version_el is not None and version_el in children:
        insert_idx = children.index(version_el) + 1

    # Вставляем базовых кассиров по порядку
    for cfg in BASE_CASHIERS:
        c_el = ET.Element("cashier")
        ET.SubElement(c_el, "roleId").text = cfg["roleId"]
        ET.SubElement(c_el, "id").text = cfg["id"]
        ET.SubElement(c_el, "name").text = cfg["name"]
        ET.SubElement(c_el, "password").text = cfg["password"]
        root.insert(insert_idx, c_el)
        insert_idx += 1


def _write_xml_with_declaration(xml_path: str, root: ET.Element, ensure_base: bool = False) -> None:
    """
    Записывает XML-файл с:
      - заголовком: <?xml version="1.0" encoding="UTF-8"?>
      - красивым форматированием (отступы и переносы строк).
    Для storeCashiers дополнительно может включать базовых кассиров.
    """
    if ensure_base:
        _ensure_base_cashiers(root)

    # Превращаем ElementTree в строку и красиво форматируем через minidom
    rough = ET.tostring(root, encoding="utf-8")
    dom = minidom.parseString(rough)
    pretty = dom.toprettyxml(indent="  ", encoding="UTF-8")

    with open(xml_path, "wb") as f:
        f.write(pretty)



#  1С: параметры и HTTP-сессия
ONEC_EMP_IDENT_URL = os.getenv(
    "ONEC_EMP_IDENT_URL",
    "http://192.168.17.26/zupcorp_http/hs/API/Post_EmployeeIdentification"
)
ONEC_USER = os.getenv("ONEC_USER")       
ONEC_PASS = os.getenv("ONEC_PASS")

# Глобальная Session 
_ONEC_SESSION = requests.Session()
_ONEC_SESSION.headers.update({"Content-Type": "application/json; charset=utf-8"})







_USERS_TTL_COLS_CACHE: dict[str, bool] = {}

def _mysql_users_supports_ttl_cols(cur, cache_key: str) -> bool:
    """
    Проверяет наличие колонок start_date/end_date в таблице users текущей MySQL БД.
    Кэшируем по cache_key (обычно ukm4ip), чтобы не делать SHOW COLUMNS каждый раз.
    """
    if cache_key in _USERS_TTL_COLS_CACHE:
        return _USERS_TTL_COLS_CACHE[cache_key]

    try:
        cur.execute("SHOW COLUMNS FROM users LIKE 'start_date'")
        has_start = cur.fetchone() is not None

        cur.execute("SHOW COLUMNS FROM users LIKE 'end_date'")
        has_end = cur.fetchone() is not None

        ok = bool(has_start and has_end)
        _USERS_TTL_COLS_CACHE[cache_key] = ok
        return ok

    except Exception as e:
        logger.warning(f"[MySQL:{cache_key}] Не смог проверить start_date/end_date в users: {e}")
        _USERS_TTL_COLS_CACHE[cache_key] = False
        return False



def resolve_xml_store_id(raw_store_id: int | str) -> int:
    """
    Преобразует id магазина из УКМ (ukm4store),
    в smstore из PostgreSQL таблицы stores.

    Если в таблице stores нет записи или smstore некорректен –
    возвращаем исходный id (как сейчас).
    """
    sid_raw = str(raw_store_id).strip()
    try:
        mapped = (
            Store.objects
            .filter(ukm4store=sid_raw)
            .values_list('smstore', flat=True)
            .first()
        )
        if mapped is None:
            logger.warning(f"[XML] storeid={sid_raw} не найден в stores.ukm4store – используем исходный id")
            return int(sid_raw)

        try:
            return int(mapped)
        except (TypeError, ValueError):
            logger.warning(f"[XML] Некорректный smstore={mapped!r} для ukm4store={sid_raw} – используем исходный id")
            return int(sid_raw)
    except Exception as e:
        logger.exception(f"[XML] Ошибка при resolve_xml_store_id({sid_raw!r}): {e}")
        try:
            return int(sid_raw)
        except Exception:
            return 0


def _find_storecashiers_file_for_store(store_id: int) -> Optional[Tuple[str, int]]:
    """
    Ищет последний файл вида:
      storeCashiers_[smstore]_[Number]_[F].xml
    smstore получаем из таблицы stores по ukm4store = store_id.
    Возвращает (полный_путь, Number) или None, если файлов нет.
    """
    xml_store_id = resolve_xml_store_id(store_id)
    pattern = re.compile(rf"^storeCashiers_\[{xml_store_id}\]_\[(\d+)\]_\[F\]\.xml$")
    best_name = None
    best_num = -1

    try:
        for name in os.listdir(XML_DIR):
            m = pattern.match(name)
            if m:
                num = int(m.group(1))
                if num > best_num:
                    best_num = num
                    best_name = name
    except FileNotFoundError:
        return None

    if best_name is None:
        return None
    return os.path.join(XML_DIR, best_name), best_num


def _generate_storecashiers_number(store_id: int) -> int:
    """
    Number = количество секунд с 00:00 текущего дня (локальное/настроенное время Django).
    При коллизии (файл с таким Number уже существует для этого store_id) —
    увеличиваем Number, пока не найдём свободный.
    """
    now = timezone.now()
    midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
    seconds_since_midnight = int((now - midnight).total_seconds())  # 0..86399

    base_num = seconds_since_midnight

    # Собираем уже использованные Number для этого магазина
    pattern = re.compile(rf"^storeCashiers_\[{store_id}\]_\[(\d+)\]_\[F\]\.xml$")
    used_numbers = set()

    try:
        for name in os.listdir(XML_DIR):
            m = pattern.match(name)
            if m:
                used_numbers.add(int(m.group(1)))
    except FileNotFoundError:
        pass

    # Если текущий номер ещё не использован — берём его
    if base_num not in used_numbers:
        return base_num

    # Если вдруг такой Number уже есть — двигаемся вверх, пока не найдём свободный
    n = base_num + 1
    # Теоретически может уйти за 86400, но это не критично — это просто уникальный int
    while n in used_numbers:
        n += 1

    return n


def _get_or_create_storecashiers_tree(store_id: int) -> Tuple[str, ET.ElementTree, ET.Element]:
    """
    Возвращает (xml_path, tree, root) для storeCashiers магазина.

    На вход приходит ukm4store (store_id).
    Для XML:
      - вычисляем smstore через resolve_xml_store_id()
      - имя файла: storeCashiers_[smstore]_[Number]_[F].xml
      - <storeCashiers fullness="F" storeId="smstore">
          <version>1.0</version>
        </storeCashiers>
    """
    xml_store_id = resolve_xml_store_id(store_id)

    found = _find_storecashiers_file_for_store(store_id)
    if found:
        xml_path, number = found
        tree = ET.parse(xml_path)
        root = tree.getroot()

        root.set("fullness", "F")
        root.set("storeId", str(xml_store_id))

        version_el = root.find("version")
        if version_el is None:
            version_el = ET.SubElement(root, "version")
        version_el.text = "1.0"

        return xml_path, tree, root

    # файла ещё нет – создаём новый
    number = _generate_storecashiers_number(store_id)
    filename = f"storeCashiers_[{xml_store_id}]_[{number}]_[F].xml"
    xml_path = os.path.join(XML_DIR, filename)

    root = ET.Element("storeCashiers", fullness="F", storeId=str(xml_store_id))
    version_el = ET.SubElement(root, "version")
    version_el.text = "1.0"
    tree = ET.ElementTree(root)

    return xml_path, tree, root


def build_full_ukm5_xml_for_store(store_id: int) -> Optional[str]:
    """
    Полная пересборка storeCashiers XML для УКМ-5 по одному магазину (ukm4store).

    В файл попадают ВСЕ кассиры, у которых есть UKMUser.storeid = store_id.
    Пароль берём из OpenInSystem (system_id=9).
    id кассира:
      - если есть запись в trm_in_users (user_inn + name) — используем её id;
      - иначе берём MAX(id)+1 и дальше инкрементируем в рамках этой пересборки.

    Возвращает путь к новому XML или None, если магазин не УКМ5.
    """
    info = get_store_info(store_id)
    ukm_host = info.get("ukm4ip")
    if not info.get("is_ukm5", False):
        logger.info(f"[XML/FULL] Store {store_id} не является УКМ-5 (is_ukm5=False) – пересборка не нужна")
        return None

    xml_store_id = resolve_xml_store_id(store_id)
    if not xml_store_id:
        logger.error(f"[XML/FULL] Не удалось получить xml_store_id для storeid={store_id}")
        return None

    links = list(
        UKMUser.objects
        .filter(storeid=store_id)
        .select_related("user")
        .order_by("user_id")
    )
    logger.info(
        f"[XML/FULL] Пересборка storeCashiers для storeid={store_id} "
        f"(xml_store_id={xml_store_id}), кассиров={len(links)}"
    )

    # Генерим Number = секунды с полуночи, избегаем коллизий
    now = timezone.now()
    midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
    seconds_since_midnight = int((now - midnight).total_seconds())

    pattern = re.compile(rf"^storeCashiers_\[{xml_store_id}\]_\[(\d+)\]_\[F\]\.xml$")
    used_numbers: set[int] = set()
    try:
        for name in os.listdir(XML_DIR):
            m = pattern.match(name)
            if m:
                used_numbers.add(int(m.group(1)))
    except FileNotFoundError:
        pass

    number = seconds_since_midnight
    while number in used_numbers:
        number += 1

    filename = f"storeCashiers_[{xml_store_id}]_[{number}]_[F].xml"
    xml_path = os.path.join(XML_DIR, filename)

    root = ET.Element("storeCashiers", fullness="F", storeId=str(xml_store_id))
    version_el = ET.SubElement(root, "version")
    version_el.text = "1.0"

    user_ids = [l.user_id for l in links]
    # Сразу подтянем пароли для всех пользователей
    open_map = {
        row["user_id"]: row["password"]
        for row in OpenInSystem.objects
            .filter(user_id__in=user_ids, system_id=9)
            .values("user_id", "password")
    }

    next_new_id: Optional[int] = None

    for link in links:
        user = link.user
        fio = (user.full_name or "").strip()
        emp_raw = (user.employee_id or "").strip()

        if not fio or not emp_raw:
            logger.warning(
                f"[XML/FULL] Пропуск user_id={user.id}: пустое ФИО ({fio!r}) "
                f"или employee_id ({emp_raw!r})"
            )
            continue

        try:
            plain_inn = ensure_plain_inn(emp_raw)
        except Exception as e:
            logger.warning(
                f"[XML/FULL] Пропуск user_id={user.id}: некорректный INN "
                f"{emp_raw!r}: {e}"
            )
            continue

        password_plain = open_map.get(user.id)
        if not password_plain:
            logger.warning(
                f"[XML/FULL] Пропуск user_id={user.id}: нет OpenInSystem(system_id=9)"
            )
            continue

        # Пытаемся взять id из trm_in_users
        cashier_id = get_trm_employee_id(plain_inn, fio, store_id=store_id, host=ukm_host)
        if cashier_id is None:
            if next_new_id is None:
                try:
                    next_new_id = get_next_trm_employee_id(store_id=store_id, host=ukm_host)
                    logger.info(f"[XML/FULL] trm_in_users@{ukm_host}: base next_id={next_new_id}")
                except Exception as e:
                    logger.error(f"[XML/FULL] Ошибка получения next_id из trm_in_users@{ukm_host}: {e}", exc_info=True)
                    next_new_id = 1
            cashier_id = next_new_id
            next_new_id += 1

        c_el = ET.SubElement(root, "cashier")
        ET.SubElement(c_el, "roleId").text = str(link.roleid)
        ET.SubElement(c_el, "id").text = str(cashier_id)
        ET.SubElement(c_el, "name").text = fio
        ET.SubElement(c_el, "INN").text = plain_inn
        ET.SubElement(c_el, "password").text = password_plain

    _write_xml_with_declaration(xml_path, root, ensure_base=True)

    logger.info(
        f"[XML/FULL] Готов полный storeCashiers для storeid={store_id}: {xml_path}"
    )
    return xml_path






def connect_ukm(host: Optional[str] = None, *, store_id: Optional[int | str] = None):
    resolved_host, src = _resolve_ukmserver_host(host=host, store_id=store_id)
    if src.startswith("fallback"):
        logger.warning(f"[UKM] connect_ukm(): resolved_host={resolved_host} src={src} store_id={store_id!r}")
    else:
        logger.info(f"[UKM] connect_ukm(): resolved_host={resolved_host} src={src} store_id={store_id!r}")

    return pymysql.connect(
        host=resolved_host,
        user="ukminfo",
        password="CtHDbCGK.C",
        database="ukmserver",
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=5,
        read_timeout=10,
        write_timeout=10,
    )
    
    
def _resolve_ukmserver_host(host: Optional[str] = None, store_id: Optional[int | str] = None) -> tuple[str, str]:
    """
    Возвращает (resolved_host, source)
    source: explicit | oracle | fallback_no_ukm4ip | fallback_oracle_error | fallback_no_store_id
    """
    if host:
        return str(host).strip(), "explicit"

    if store_id is not None:
        try:
            info = get_store_info(store_id)
            ip = info.get("ukm4ip")
            if ip:
                return str(ip).strip(), "oracle"
            fb = os.getenv("UKMSERVER_DEFAULT_HOST", "192.168.17.234")
            return fb, "fallback_no_ukm4ip"
        except Exception as e:
            fb = os.getenv("UKMSERVER_DEFAULT_HOST", "192.168.17.234")
            return fb, f"fallback_oracle_error:{type(e).__name__}"

    fb = os.getenv("UKMSERVER_DEFAULT_HOST", "192.168.17.234")
    return fb, "fallback_no_store_id"


def inspect_trm_in_users(
    plain_inn: str,
    fio: str,
    *,
    store_id: Optional[int | str] = None,
    host: Optional[str] = None,
) -> dict:
    """
    Возвращает подробный debug по тому:
      - куда подключились (resolved_host + почему)
      - есть ли store-колонка
      - найден ли сотрудник (с store-фильтром / без)
      - max/next id (global) + опционально max/next по store (только для отчёта)
    НИЧЕГО не меняет в данных.
    """
    resolved_host, host_source = _resolve_ukmserver_host(host=host, store_id=store_id)

    dbg: dict = {
        "store_id": store_id,
        "requested_host": host,
        "resolved_host": resolved_host,
        "host_source": host_source,
        "store_col": None,
        "found_id_store": None,
        "found_id_global": None,
        "found_id_final": None,
        "search_used": None,
        "max_id_all": None,
        "next_id_all": None,
        "max_id_store": None,
        "next_id_store": None,
        "counts": {},
        "errors": [],
    }

    conn = cur = None
    try:
        conn = connect_ukm(host=resolved_host)  # уже без store_id, т.к. host окончательно известен
        cur = conn.cursor()

        # колонки
        cur.execute("SHOW COLUMNS FROM trm_in_users")
        cols = {r["Field"] for r in (cur.fetchall() or [])}
        store_col = next((c for c in ("store", "storeid", "store_id", "shop", "shop_id") if c in cols), None)
        dbg["store_col"] = store_col

        # статистика (не обязательно, но полезно)
        try:
            cur.execute("SELECT COUNT(*) AS c FROM trm_in_users")
            dbg["counts"]["rows_all"] = int((cur.fetchone() or {}).get("c") or 0)
        except Exception as e:
            dbg["errors"].append(f"count_all:{type(e).__name__}:{e}")

        # поиск
        if store_col and store_id is not None:
            # 1) с фильтром по магазину
            try:
                cur.execute(
                    f"SELECT id FROM trm_in_users WHERE user_inn=%s AND name=%s AND `{store_col}`=%s LIMIT 1",
                    (plain_inn, fio, int(store_id)),
                )
                row = cur.fetchone()
                if row and row.get("id") is not None:
                    dbg["found_id_store"] = int(row["id"])
            except Exception as e:
                dbg["errors"].append(f"find_store:{type(e).__name__}:{e}")

            # 2) fallback без store
            try:
                cur.execute(
                    "SELECT id FROM trm_in_users WHERE user_inn=%s AND name=%s LIMIT 1",
                    (plain_inn, fio),
                )
                row2 = cur.fetchone()
                if row2 and row2.get("id") is not None:
                    dbg["found_id_global"] = int(row2["id"])
            except Exception as e:
                dbg["errors"].append(f"find_global:{type(e).__name__}:{e}")

            if dbg["found_id_store"] is not None:
                dbg["found_id_final"] = dbg["found_id_store"]
                dbg["search_used"] = f"inn+fio+{store_col}"
            elif dbg["found_id_global"] is not None:
                dbg["found_id_final"] = dbg["found_id_global"]
                dbg["search_used"] = "inn+fio (fallback_without_store)"
            else:
                dbg["search_used"] = f"inn+fio+{store_col} -> fallback inn+fio (not_found)"

        else:
            # таблица без store-колонки (или store_id не передан)
            try:
                cur.execute(
                    "SELECT id FROM trm_in_users WHERE user_inn=%s AND name=%s LIMIT 1",
                    (plain_inn, fio),
                )
                row = cur.fetchone()
                if row and row.get("id") is not None:
                    dbg["found_id_global"] = int(row["id"])
                    dbg["found_id_final"] = dbg["found_id_global"]
            except Exception as e:
                dbg["errors"].append(f"find_global_only:{type(e).__name__}:{e}")

            dbg["search_used"] = "inn+fio"

        # max/next (global)
        try:
            cur.execute("SELECT COALESCE(MAX(id), 0) AS max_id FROM trm_in_users")
            r = cur.fetchone() or {}
            max_id_all = int(r.get("max_id") or 0)
            dbg["max_id_all"] = max_id_all
            dbg["next_id_all"] = max_id_all + 1
        except Exception as e:
            dbg["errors"].append(f"max_all:{type(e).__name__}:{e}")

        # max/next по store (ТОЛЬКО ДЛЯ ОТЧЁТА; НЕ ИСПОЛЬЗУЕМ ДЛЯ ВЫДАЧИ, чтобы не словить дубль PK)
        if store_col and store_id is not None:
            try:
                cur.execute(
                    f"SELECT COALESCE(MAX(id), 0) AS max_id FROM trm_in_users WHERE `{store_col}`=%s",
                    (int(store_id),),
                )
                r = cur.fetchone() or {}
                max_id_store = int(r.get("max_id") or 0)
                dbg["max_id_store"] = max_id_store
                dbg["next_id_store"] = max_id_store + 1
            except Exception as e:
                dbg["errors"].append(f"max_store:{type(e).__name__}:{e}")

        return dbg

    except Exception as e:
        dbg["errors"].append(f"connect_or_query:{type(e).__name__}:{e}")
        return dbg
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except Exception:
            pass

def get_trm_employee_id(
    plain_inn: str,
    fio: str,
    *,
    store_id: Optional[int | str] = None,
    host: Optional[str] = None
) -> int | None:
    """
    Ищет кассира в trm_in_users на НУЖНОМ ukmserver (по host / store_id).

    Если в таблице есть колонка store/storeid/store_id/shop/shop_id — пробуем искать с фильтром по магазину.
    Если не нашли — делаем fallback без фильтра по магазину (чтобы не сломаться на других схемах).
    """
    conn = cur = None
    try:
        conn = connect_ukm(host=host, store_id=store_id)
        cur = conn.cursor()

        # Узнаём колонки таблицы (схемы могут отличаться)
        cur.execute("SHOW COLUMNS FROM trm_in_users")
        cols = {r["Field"] for r in (cur.fetchall() or [])}

        store_col = None
        for cand in ("store", "storeid", "store_id", "shop", "shop_id"):
            if cand in cols:
                store_col = cand
                break

        # 1) основной поиск
        if store_col and store_id is not None:
            sql = f"SELECT id FROM trm_in_users WHERE user_inn=%s AND name=%s AND `{store_col}`=%s"
            cur.execute(sql, (plain_inn, fio, int(store_id)))
            row = cur.fetchone()
            if row:
                return row.get("id")

            # 2) fallback без store (на случай “глобальной” таблицы)
            cur.execute("SELECT id FROM trm_in_users WHERE user_inn=%s AND name=%s", (plain_inn, fio))
            row2 = cur.fetchone()
            return row2.get("id") if row2 else None

        # если store-колонки нет — как раньше
        cur.execute("SELECT id FROM trm_in_users WHERE user_inn=%s AND name=%s", (plain_inn, fio))
        row = cur.fetchone()
        return row.get("id") if row else None

    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass
        
        
        
def get_next_trm_employee_id(*, store_id: Optional[int | str] = None, host: Optional[str] = None) -> int:
    """
    Следующий id для trm_in_users.
    Берём MAX(id) только в диапазоне TRM_SMALL_MIN..TRM_SMALL_MAX (пятизначные),
    чтобы выбросы 2147483*** не ломали выдачу.
    Используем GET_LOCK() чтобы два параллельных запроса не выдали одинаковый id.
    """
    conn = cur = None
    lock_key = None
    got_lock = False

    try:
        conn = connect_ukm(host=host, store_id=store_id)
        cur = conn.cursor()

        # Лок на конкретный ukm-host (чтобы не конфликтовать между разными ukmserver)
        resolved_host, _ = _resolve_ukmserver_host(host=host, store_id=store_id)
        lock_key = f"trm_id_alloc:{resolved_host}"

        try:
            cur.execute("SELECT GET_LOCK(%s, %s) AS l", (lock_key, 5))
            got_lock = int((cur.fetchone() or {}).get("l") or 0) == 1
        except Exception:
            got_lock = False  # если GET_LOCK запрещён — продолжим без него (хуже, но не упадём)

        # max в “нормальном” диапазоне (пятизначные)
        cur.execute(
            "SELECT COALESCE(MAX(id), 0) AS max_id "
            "FROM trm_in_users "
            "WHERE id BETWEEN %s AND %s",
            (TRM_SMALL_MIN, TRM_SMALL_MAX),
        )
        row = cur.fetchone() or {}
        max_small = int(row.get("max_id") or 0)

        candidate = max(TRM_SMALL_MIN, max_small + 1)

        if candidate > TRM_SMALL_MAX:
            raise RuntimeError(
                f"Исчерпан диапазон id [{TRM_SMALL_MIN}..{TRM_SMALL_MAX}] в trm_in_users. "
                f"Нужно расширять диапазон или менять стратегию."
            )

        # На всякий — проверим что candidate реально свободен (при гонках без lock)
        cur.execute("SELECT 1 AS x FROM trm_in_users WHERE id=%s LIMIT 1", (candidate,))
        if cur.fetchone():
            # если занято — поднимаемся вверх, пока не найдём свободный (в рамках small-range)
            while candidate <= TRM_SMALL_MAX:
                candidate += 1
                cur.execute("SELECT 1 AS x FROM trm_in_users WHERE id=%s LIMIT 1", (candidate,))
                if not cur.fetchone():
                    break
            if candidate > TRM_SMALL_MAX:
                raise RuntimeError(
                    f"Не удалось найти свободный id в диапазоне [{TRM_SMALL_MIN}..{TRM_SMALL_MAX}]."
                )

        return int(candidate)

    finally:
        try:
            if cur and got_lock and lock_key:
                cur.execute("SELECT RELEASE_LOCK(%s)", (lock_key,))
        except Exception:
            pass
        try:
            if cur: cur.close()
            if conn: conn.close()
        except Exception:
            pass
        
        
def resolve_cashier_id_for_store(store_id: int, plain_inn: str, fio: str) -> tuple[int, Optional[str], bool, dict]:
    """
    Возвращает:
      (cashier_id, resolved_ukm_host, found_in_trm, trm_debug)

    cashier_id:
      - если найден в trm_in_users -> found_id_final
      - иначе -> next_id_all (MAX(id)+1 глобально по этой таблице на этом ukmserver)
    """
    info = get_store_info(store_id)
    ukm_host_from_oracle = info.get("ukm4ip")

    trm_dbg = inspect_trm_in_users(
        plain_inn,
        fio,
        store_id=store_id,
        host=ukm_host_from_oracle
    )

    resolved_host = trm_dbg.get("resolved_host") or ukm_host_from_oracle
    found_id = trm_dbg.get("found_id_final")

    if found_id is not None:
        return int(found_id), resolved_host, True, trm_dbg

    next_id = get_next_trm_employee_id(store_id=store_id, host=resolved_host)
    trm_dbg["allocated_id"] = next_id
    trm_dbg["allocated_range"] = f"{TRM_SMALL_MIN}..{TRM_SMALL_MAX}"
    return int(next_id), resolved_host, False, trm_dbg

    
def _calc_next_signal_version(cur) -> int:
    """
    Возвращает следующий номер версии для MySQL-таблицы `signal`.

    Берём MAX(version) по всей таблице и увеличиваем на 1.
    Так версия всегда растёт, даже если нет строк с signal='busy'.
    """
    cur.execute("SELECT MAX(`version`) AS max_ver FROM `signal`")
    row = cur.fetchone() or {}
    max_ver = row.get("max_ver") or 0
    try:
        max_ver = int(max_ver)
    except (TypeError, ValueError):
        max_ver = 0
    return max_ver + 1

def _write_converter_user_and_signal(*args, **kwargs) -> None:
    """
    import4staffbonus (старый конвертер) отключён.
    Оставлено как no-op, чтобы не падать, если где-то остался вызов.
    """
    logger.info("[CONVERTER] import4staffbonus отключён — пропускаю запись users/signal")
    return

def generate_qr_string(inn: str, salt: str = "INDIVIDUAL_SALT") -> str:
    expiration = (datetime.datetime.utcnow() + datetime.timedelta(days=1)
                  ).strftime('%Y%m%d')
    return f"KS{inn}{expiration}{salt}"
    
def connect_oracle():
    dsn = cx_Oracle.makedsn("192.168.17.239", 1521, service_name="xe")
    return _oracle_connect(user="supermag_user", password="supermag_pass", dsn=dsn)

def is_ukm5_store(storeid: int) -> bool:
    """Сохранена стар. сигнатура: True если магазин UKM5, иначе False."""
    info = get_store_info(storeid)
    return info.get("is_ukm5", False)


def connect_store_mysql(host: str):
    """
    Подключение к MySQL конвертера конкретного магазина:
      - host = UKM4IP
      - database = import4
      - user = ukm_import
      - password = jgOKsc2n
    """
    return pymysql.connect(
        host=host,
        port=3306,
        user="ukm_import",
        password="jgOKsc2n",
        database="import4",
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor
    )


def _update_store_mysql_and_xml_for_single_store(
    store_id: int,
    cashier_id: int,
    role_id: int,
    plain_inn: str,
    fio: str,
    password_plain: str
) -> None:
    """
    Обновляет кассира по ОДНОМУ магазину:
      • UKM4 (MySQL import4.users + import4.signal)
      • UKM5 (XML storeCashiers_... если магазин UKM5)

    ВАЖНО:
      • Для магазинов из UKM5_FULL_XML_STORE_IDS делаем ПОЛНУЮ пересборку XML,
        а не точечный upsert одного кассира.
    """
    logger.info(
        f"[QR/EMP] Обновление UKM4/UKM5 для storeid={store_id}, "
        f"cashier_id={cashier_id}, role_id={role_id}"
    )

    info = get_store_info(store_id)
    ukm4ip = info.get("ukm4ip")
    is_ukm5 = info.get("is_ukm5", False)
    logger.info(f"[QR/EMP] Store {store_id}: ukm4ip={ukm4ip!r}, is_ukm5={is_ukm5}")

    
    # UKM4 / MySQL import4
    
    if ukm4ip:
        conv = cur = None
        try:
            conv = connect_store_mysql(ukm4ip)
            cur = conv.cursor()

            base_version = _calc_next_signal_version(cur)
            logger.info(
                f"[QR/EMP] Store {store_id} ({ukm4ip}): next version={base_version} "
                f"(по MAX(signal.version))"
            )

            ttl_supported = _mysql_users_supports_ttl_cols(cur, cache_key=str(ukm4ip))
            if ttl_supported:
                now = timezone.now()
                # start_date = now.date()
                end_date = (now + datetime.timedelta(days=1)).date()
                start_date = end_date

                cur.execute("""
                    INSERT INTO users (
                        store, id, name, inn, password, role_id, version, deleted,
                        start_date, end_date
                    )
                    VALUES (
                        %s, %s, %s, %s, OLD_PASSWORD(%s), %s, %s, 0,
                        %s, %s
                    )
                    ON DUPLICATE KEY UPDATE
                        name       = VALUES(name),
                        inn        = VALUES(inn),
                        password   = VALUES(password),
                        role_id    = VALUES(role_id),
                        version    = VALUES(version),
                        deleted    = 0,
                        # start_date = VALUES(start_date),
                        start_date = VALUES(end_date),
                        end_date   = VALUES(end_date)
                """, (
                    store_id,
                    cashier_id,
                    fio,
                    plain_inn,
                    mysql_pwd(password_plain),
                    role_id,
                    base_version,
                    start_date,
                    end_date
                ))
                logger.info(f"[QR/EMP] Store {store_id} ({ukm4ip}): users TTL cols detected, set {start_date}..{end_date}")
            else:
                cur.execute("""
                    INSERT INTO users (store, id, name, inn, password, role_id, version, deleted)
                    VALUES (%s, %s, %s, %s, OLD_PASSWORD(%s), %s, %s, 0)
                    ON DUPLICATE KEY UPDATE
                        name     = VALUES(name),
                        inn      = VALUES(inn),
                        password = VALUES(password),
                        role_id  = VALUES(role_id),
                        version  = VALUES(version),
                        deleted  = 0
                """, (
                    store_id,
                    cashier_id,
                    fio,
                    plain_inn,
                    mysql_pwd(password_plain),
                    role_id,
                    base_version
                ))
                logger.info(f"[QR/EMP] Store {store_id} ({ukm4ip}): users TTL cols NOT found, wrote without dates")

            cur.execute(
                "INSERT INTO `signal`(`signal`,`version`) VALUES ('incr', %s)",
                (base_version,)
            )

            conv.commit()
            logger.info(
                f"[QR/EMP] Store {store_id} ({ukm4ip}): OK users+signal "
                f"(id={cashier_id}, role_id={role_id}, version={base_version})"
            )
        except Exception as e:
            logger.error(f"[QR/EMP] Store {store_id} ({ukm4ip}) MySQL error: {e}", exc_info=True)
            if conv:
                try:
                    conv.rollback()
                except Exception:
                    pass
        finally:
            try:
                if cur:
                    cur.close()
                if conv:
                    conv.close()
            except Exception:
                pass
    else:
        logger.error(f"[QR/EMP] Store {store_id}: ukm4ip not found; пропускаем import4.users/signal")

    
    # UKM5 / XML
    
    if not is_ukm5:
        return

    # Полная пересборка XML для “спец” магазинов
    if int(store_id) in UKM5_FULL_XML_STORE_IDS:
        try:
            xml_path = build_full_ukm5_xml_for_store(store_id)
            logger.info(f"[QR/EMP] Store {store_id}: полный XML пересобран: {xml_path}")
        except Exception as e:
            logger.error(
                f"[QR/EMP] Store {store_id}: ошибка полной пересборки XML/UKM5: {e}",
                exc_info=True
            )
        return

    # Остальные UKM5 — точечное обновление одного кассира
    try:
        xml_path, tree, root = _get_or_create_storecashiers_tree(store_id)

        # удаляем старые записи этого INN
        changed = False
        for cash_el in list(root.findall("cashier")):
            if (cash_el.findtext("INN") or "").strip() == plain_inn:
                root.remove(cash_el)
                changed = True
        if changed:
            logger.info(f"[QR/EMP] Store {store_id}: удалены старые cashier с INN={plain_inn} из {xml_path}")

        cash_el = ET.SubElement(root, "cashier")
        ET.SubElement(cash_el, "roleId").text = str(role_id)
        ET.SubElement(cash_el, "id").text = str(cashier_id)
        ET.SubElement(cash_el, "name").text = fio
        ET.SubElement(cash_el, "INN").text = plain_inn
        ET.SubElement(cash_el, "password").text = password_plain

        _write_xml_with_declaration(xml_path, root, ensure_base=True)
        logger.info(f"[QR/EMP] Store {store_id}: XML обновлён {xml_path}")
    except Exception as e:
        logger.error(f"[QR/EMP] Store {store_id}: ошибка при работе с XML/UKM5: {e}", exc_info=True)
        

            
              

def ensure_plain_inn(value: str) -> str:
    v = (value or "").strip()
    if not (v.isdigit() and len(v) in (10, 12)):
        raise ValueError("ИНН должен содержать 10 или 12 цифр")
    return v


def normalize_phone_ru(phone: str) -> Optional[str]:
    """
    Нормализует телефон к формату +7XXXXXXXXXX.
    Принимает варианты:
      +7 924 000-00-00
      8 (924) 000-00-00
      79240000000
      9240000000
    Если не удаётся нормализовать — возвращает None.
    """
    if not phone:
        return None

    raw = str(phone)
    digits = ''.join(ch for ch in raw if ch.isdigit())

    if not digits:
        return None

    if len(digits) == 11 and digits[0] in ('7', '8'):
        norm = '+7' + digits[1:]
        return norm

    if len(digits) == 10:
        norm = '+7' + digits
        return norm

    if len(digits) == 11:
        return '+' + digits

    return '+' + digits



def _parse_and_format_dt(dt_raw: str) -> str:
    """
    Приводит дату/время к строке вида 'DD.MM.YYYY H:MM:SS'
    Допускаемые входы: '24.09.2025 8:45:00', '24.09.2025 08:45', '2025-09-24 08:45:00',
    ISO-похожие и т.п. из списка fmts.
    """
    s = (dt_raw or "").strip()
    if not s:
        raise ValueError("Пустой Datetime")

    fmts = [
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
    ]
    dt = None
    last_err = None
    for f in fmts:
        try:
            dt = datetime.datetime.strptime(s, f)
            break
        except Exception as e:
            last_err = e
    if dt is None and len(s.split()) == 2 and s.count(':') == 1 and '.' in s.split()[0]:
        try:
            dt = datetime.datetime.strptime(s + ":00", "%d.%m.%Y %H:%M:%S")
        except Exception:
            pass
    if dt is None:
        raise ValueError(f"Некорректный Datetime: {dt_raw!r}; последняя ошибка: {last_err}")

    out = dt.strftime("%d.%m.%Y %H:%M:%S")
    if out[11] == '0':
        out = out[:11] + out[12:]
    return out


def _onec_auth() -> Optional[Tuple[str, str]]:
    return (ONEC_USER, ONEC_PASS) if ONEC_USER and ONEC_PASS else None


def _post_to_onec(payload: dict, idem_key: str, timeout=(3, 7), retries=2) -> Tuple[int, str]:
    """
    Отправка в 1С с ретраями и идемпотентным ключом (в заголовке).
    timeout: (connect, read); retries: доп. попытки (итого 1+retries).
    Возвращает (status_code, text). Генерирует RuntimeError при сетевых сбоях после всех попыток.
    """
    headers = {"X-Idempotency-Key": idem_key}
    auth = _onec_auth()
    last_exc = None
    for attempt in range(retries + 1):
        try:
            resp = _ONEC_SESSION.post(
                ONEC_EMP_IDENT_URL,
                json=payload,
                headers=headers,
                auth=auth,
                timeout=timeout,
            )
            return resp.status_code, (resp.text or "")
        except Exception as exc:
            last_exc = exc
            if attempt < retries:
                time.sleep(0.3 * (2 ** attempt)) 
    raise RuntimeError(f"Ошибка запроса в 1С: {last_exc}")


def encrypt_inn_full(inn: str) -> str:
    return get_inn_hash(inn)

def encrypt_inn20(inn: str) -> str:
    return encrypt_inn_full(inn)[:20]


def build_user_password(plain_inn: str) -> str:
    """
    Формирует пароль вида:

        KS<ИНН><YYYYMMDD><RANDOM>

    * KS        – постоянный префикс
    * ИНН       – 10 / 12 цифр, БЕЗ хэша
    * YYYYMMDD  – дата UTC
    * RANDOM    – заглавные A-Z + цифры, чтобы итог всегда был 40 симв.
    """
    date_part = datetime.datetime.utcnow().strftime("%Y%m%d")     
    base      = f"KS{plain_inn}{date_part}"                   
    salt_len  = 40 - len(base)
    salt      = ''.join(random.choices(string.ascii_uppercase + string.digits,
                                       k=salt_len))
    return base + salt                                           


def _generate_pin_code() -> str:
    return f"{random.randint(0, 9999):04d}"


def mysql_pwd(raw: str) -> str:
    """
    Возвращает строку для OLD_PASSWORD – без префикса 'KS'.
    Если префикса нет, возвращаем как есть (защитный fallback).
    """
    return raw[2:] if raw.startswith("KS") else raw


# Oracle-коннект (единый)
def connect_oracle_supermag():
    ORA_HOST     = os.getenv("ORACLE_HOST", "192.168.17.239")
    ORA_PORT     = int(os.getenv("ORACLE_PORT", "1521"))
    ORA_SERVICE  = os.getenv("ORACLE_SERVICE", "BINUU00")
    ORA_USER     = os.getenv("ORACLE_USER", "supermag")
    ORA_PASSWORD = os.getenv("ORACLE_PASSWORD", "qqq")
    dsn = cx_Oracle.makedsn(ORA_HOST, ORA_PORT, service_name=ORA_SERVICE)
    return _oracle_connect(user=ORA_USER, password=ORA_PASSWORD, dsn=dsn)

def _oracle_rows_to_jsonable(cur):
    """
    Превращает результат Oracle SELECT в список словарей, пригодных для JSON.
    Даты → ISO-строки, всё непонятное → str().
    """
    cols = [d[0].lower() for d in cur.description]
    items = []

    for row in cur.fetchall():
        obj = {}
        for col, val in zip(cols, row):
            if isinstance(val, (datetime.date, datetime.datetime)):
                obj[col] = val.isoformat(sep=' ')
            elif isinstance(val, (int, float, str, bool)) or val is None:
                obj[col] = val
            else:
                obj[col] = str(val)
        items.append(obj)

    return items




#Удалить после 30.03

def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(str(value).strip())
    except Exception:
        return None


def _normalize_spaces(value: Any) -> str:
    return re.sub(r"\s+", " ", _safe_str(value)).strip()


def _normalize_position(value: Any) -> str:
    return _normalize_spaces(value).casefold()


def _get_first_present(obj: dict, *keys: str) -> Any:
    """
    Ищет значение по нескольким вариантам ключа, без учёта регистра.
    """
    if not isinstance(obj, dict):
        return None

    # прямое совпадение
    for key in keys:
        if key in obj and obj[key] not in (None, ""):
            return obj[key]

    # поиск без учёта регистра/лишних пробелов
    normalized = {
        str(k).strip().casefold(): v
        for k, v in obj.items()
    }
    for key in keys:
        found = normalized.get(str(key).strip().casefold(), None)
        if found not in (None, ""):
            return found

    return None


def _parse_working_employees_text(raw_text: str) -> list[dict]:
    """
    Резервный парсер на случай, если 1С отдаёт не JSON, а текст формата:

    1217
    ИНН   "032626275702"
    Фамилия "Иванова"
    ...
    """
    items: list[dict] = []
    current: dict = {}
    current_id = None

    for raw_line in (raw_text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue

        # новая запись начинается с числового ID
        if re.fullmatch(r"\d+", line):
            if current:
                if current_id is not None and "id" not in current:
                    current["id"] = current_id
                items.append(current)
                current = {}
            current_id = int(line)
            continue

        if "\t" in line:
            key, value = line.split("\t", 1)
        elif ":" in line:
            key, value = line.split(":", 1)
        else:
            continue

        key = key.strip().strip('"')
        value = value.strip().strip('"')

        if value.lower() == "true":
            parsed_value = True
        elif value.lower() == "false":
            parsed_value = False
        else:
            parsed_value = value

        current[key] = parsed_value

    if current:
        if current_id is not None and "id" not in current:
            current["id"] = current_id
        items.append(current)

    return items


def _extract_working_employees_items(payload: Any) -> list[dict]:
    """
    Поддержка разных форматов ответа:
    - список объектов
    - dict с keys: items/data/result/value/employees
    - текстовый ответ
    """
    if isinstance(payload, list):
        return [x for x in payload if isinstance(x, dict)]

    if isinstance(payload, dict):
        # сначала ищем типовые контейнеры
        for key in ("items", "data", "result", "value", "employees", "Работники"):
            value = payload.get(key)
            if isinstance(value, list):
                return [x for x in value if isinstance(x, dict)]
            if isinstance(value, dict):
                nested = _extract_working_employees_items(value)
                if nested:
                    return nested

        # если это уже одна карточка сотрудника
        if any(
            k in payload
            for k in ("ИНН", "Фамилия", "Имя", "Отчество", "Должность", "ИдМагазина")
        ):
            return [payload]

        # пробуем пройтись по вложенным значениям
        for value in payload.values():
            if isinstance(value, (dict, list)):
                nested = _extract_working_employees_items(value)
                if nested:
                    return nested

        return []

    if isinstance(payload, str):
        text = payload.lstrip("\ufeff").strip()
        if not text:
            return []

        # сначала пробуем JSON
        try:
            reparsed = json.loads(text)
            return _extract_working_employees_items(reparsed)
        except Exception:
            pass

        # потом текстовый разбор
        return _parse_working_employees_text(text)

    return []


def _normalize_working_employee(raw_item: dict) -> Optional[dict]:
    if not isinstance(raw_item, dict):
        return None

    last_name = _get_first_present(raw_item, "Фамилия", "lastname", "last_name")
    first_name = _get_first_present(raw_item, "Имя", "firstname", "first_name")
    middle_name = _get_first_present(raw_item, "Отчество", "middlename", "middle_name")
    position = _get_first_present(raw_item, "Должность", "position")
    inn = _get_first_present(raw_item, "ИНН", "inn")
    store_id = _get_first_present(raw_item, "ИдМагазина", "IDМагазина", "id_store", "store_id")

    full_name = " ".join(
        x for x in [
            _safe_str(last_name),
            _safe_str(first_name),
            _safe_str(middle_name),
        ] if x
    ).strip()

    store_id_int = _safe_int(store_id)

    return {
        "store_id": store_id_int,
        "inn": _safe_str(inn),
        "full_name": full_name,
        "position": _normalize_spaces(position),
    }


def _fetch_working_employees_from_1c() -> list[dict]:
    """
    Загружает список работающих сотрудников из 1С.
    """
    auth = None
    if ONEC_WORKING_EMPLOYEES_AUTH_USER or ONEC_WORKING_EMPLOYEES_AUTH_PASSWORD:
        auth = (
            ONEC_WORKING_EMPLOYEES_AUTH_USER,
            ONEC_WORKING_EMPLOYEES_AUTH_PASSWORD,
        )

    headers = {
        "Accept": "application/json, text/plain;q=0.9, */*;q=0.8",
    }

    logger.info("[WORKING_EMPLOYEES_EXCEL] Requesting 1C employees: %s", ONEC_WORKING_EMPLOYEES_URL)
    response = requests.get(
        ONEC_WORKING_EMPLOYEES_URL,
        auth=auth,
        headers=headers,
        timeout=ONEC_WORKING_EMPLOYEES_TIMEOUT,
    )
    response.raise_for_status()

    raw_text = response.text.lstrip("\ufeff").strip()

    payload: Any
    try:
        payload = response.json()
    except Exception:
        payload = raw_text

    items = _extract_working_employees_items(payload)
    logger.info("[WORKING_EMPLOYEES_EXCEL] 1C returned employees: %s", len(items))
    return items


def _fetch_sm_database_items() -> list[dict]:
    """
    Общий helper для получения списка магазинов и dbname.
    Используется и в /sm/databases/, и в Excel-выгрузке.
    """
    sql = """
        SELECT
          l.id       AS smstore,
          l.name     AS name,
          p.propval  AS dbname
        FROM smstorelocations l
        LEFT JOIN smstoreproperties p
          ON p.storeloc = l.id
         AND p.propid   = 'REP.DBNAME'
        WHERE
          l.name NOT LIKE 'я%%'
          AND l.name NOT LIKE '%%ТЕСТ%%'
          AND l.formatid IN (19, 20, 9, 8, 6, 22, 23)
        ORDER BY l.name
    """

    conn = cur = None
    try:
        logger.info("[SM/LIST_DB] Старт запроса списка магазинов + баз")
        conn = connect_oracle_supermag()
        cur = conn.cursor()
        cur.execute(sql)

        items = []
        rows = cur.fetchall()
        logger.info("[SM/LIST_DB] Получено строк из Oracle: %s", len(rows))

        for smstore, name, dbname in rows:
            items.append({
                "smstore": _safe_int(smstore),
                "name": _safe_str(name),
                "dbname": _safe_str(dbname),
            })

        logger.info("[SM/LIST_DB] Итог: count=%s", len(items))
        return items

    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass


def _build_working_employees_excel(rows: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"

    headers = [
        "Имя локальной базы",
        "ИДМагазина",
        "Название магазина",
        "ФИО сотрудника",
        "Должность",
        "ИНН сотрудника",
    ]
    ws.append(headers)

    for row in rows:
        ws.append([
            row.get("dbname", ""),
            row.get("store_id", ""),
            row.get("store_name", ""),
            row.get("full_name", ""),
            row.get("position", ""),
            row.get("inn", ""),
        ])

    # стили
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="D9D9D9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{max(ws.max_row, 1)}"

    column_widths = {
        "A": 18,
        "B": 12,
        "C": 45,
        "D": 35,
        "E": 28,
        "F": 18,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

























@csrf_exempt
def sm_staff_list(request):
    """
    GET /sm/staff/?limit=100&offset=0&q=иванов

    Возвращает содержимое таблицы SMSTAFF (пользователи Супермага).
    Параметры:
      - limit  (по умолчанию 200, максимум 1000)
      - offset (по умолчанию 0)
      - q      (поиск по surname/name, опционально)
    """
    if request.method != 'GET':
        return JsonResponse(
            {'status': 'error', 'message': 'Только GET'},
            status=405,
            json_dumps_params={'ensure_ascii': False}
        )

    try:
        # Параметры
        try:
            limit = int(request.GET.get('limit', '200'))
        except ValueError:
            limit = 200
        try:
            offset = int(request.GET.get('offset', '0'))
        except ValueError:
            offset = 0

        limit = max(1, min(limit, 1000))
        offset = max(0, offset)

        q = (request.GET.get('q') or '').strip().lower()

        # Базовый SQL
        sql = "SELECT * FROM smstaff"
        binds = {}

        if q:
            sql += """
            WHERE
              LOWER(surname) LIKE :q
              OR LOWER(name) LIKE :q
            """
            binds['q'] = f"%{q}%"

        # Пагинация (Oracle 12+)
        sql += """
        ORDER BY surname
        OFFSET :off ROWS FETCH NEXT :lim ROWS ONLY
        """
        binds['off'] = offset
        binds['lim'] = limit

        conn = cur = None
        try:
            conn = connect_oracle_supermag()
            cur = conn.cursor()
            cur.execute(sql, binds)
            items = _oracle_rows_to_jsonable(cur)
        finally:
            try:
                if cur:
                    cur.close()
                if conn:
                    conn.close()
            except Exception:
                pass

        return JsonResponse(
            {
                'status': 'ok',
                'count': len(items),
                'limit': limit,
                'offset': offset,
                'items': items,
            },
            json_dumps_params={'ensure_ascii': False, 'indent': 2}
        )

    except Exception as e:
        logger.exception("[SM/STAFF_LIST] Unexpected error")
        return JsonResponse(
            {'status': 'error', 'message': str(e)},
            status=500,
            json_dumps_params={'ensure_ascii': False}
        )



@csrf_exempt
def sm_staff_columns(request):
    """
    GET /sm/staff/columns/

    Возвращает структуру таблицы SMSTAFF (колонки, типы, длину, nullable).
    """
    if request.method != 'GET':
        return JsonResponse(
            {'status': 'error', 'message': 'Только GET'},
            status=405,
            json_dumps_params={'ensure_ascii': False}
        )

    conn = cur = None
    try:
        sql = """
        SELECT
          column_id,
          column_name,
          data_type,
          data_length,
          nullable
        FROM all_tab_columns
        WHERE owner = 'SUPERMAG'
          AND table_name = 'SMSTAFF'
        ORDER BY column_id
        """
        conn = connect_oracle_supermag()
        cur = conn.cursor()
        cur.execute(sql)
        items = _oracle_rows_to_jsonable(cur)

        return JsonResponse(
            {'status': 'ok', 'count': len(items), 'columns': items},
            json_dumps_params={'ensure_ascii': False, 'indent': 2}
        )

    except Exception as e:
        logger.exception("[SM/STAFF_COLUMNS] Unexpected error")
        return JsonResponse(
            {'status': 'error', 'message': str(e)},
            status=500,
            json_dumps_params={'ensure_ascii': False}
        )
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass
        
        
        
@csrf_exempt
def sm_sql(request):
    """
    POST /sm/sql/

    ВХОД (JSON):
    {
      "sql": "select ...",
      "binds": {"p1": "val", ...}   # опционально
    }

    Ограничения:
      • Разрешены только SELECT/WITH (без insert/update/delete/alter и т.п.).
    """
    if request.method != 'POST':
        return JsonResponse(
            {'status': 'error', 'message': 'Только POST'},
            status=405,
            json_dumps_params={'ensure_ascii': False}
        )

    raw_body = ""
    try:
        raw_body = request.body.decode('utf-8') if request.body else "{}"
        data = json.loads(raw_body)
    except Exception as e:
        logger.error(f"[SM/SQL] JSON parse error: {e}; body={raw_body!r}")
        return JsonResponse(
            {'status': 'error', 'message': 'Некорректный JSON'},
            status=400,
            json_dumps_params={'ensure_ascii': False}
        )

    sql = (data.get('sql') or '').strip()
    binds = data.get('binds') or {}

    if not sql:
        return JsonResponse(
            {'status': 'error', 'message': 'Пустой sql'},
            status=400,
            json_dumps_params={'ensure_ascii': False}
        )

    sql_lower = sql.lower().lstrip()
    if not (sql_lower.startswith('select') or sql_lower.startswith('with')):
        return JsonResponse(
            {'status': 'error', 'message': 'Разрешены только SELECT/WITH запросы'},
            status=400,
            json_dumps_params={'ensure_ascii': False}
        )

    conn = cur = None
    try:
        conn = connect_oracle_supermag()
        cur = conn.cursor()
        cur.execute(sql, binds)
        items = _oracle_rows_to_jsonable(cur)

        return JsonResponse(
            {
                'status': 'ok',
                'count': len(items),
                'items': items,
            },
            json_dumps_params={'ensure_ascii': False, 'indent': 2}
        )

    except Exception as e:
        logger.exception("[SM/SQL] Unexpected error")
        return JsonResponse(
            {'status': 'error', 'message': str(e)},
            status=500,
            json_dumps_params={'ensure_ascii': False}
        )
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass





def get_dbname_for_smstore(smstore_id: int) -> Optional[str]:
    """
    Возвращает имя базы (REP.DBNAME) для магазина SMSTORELOCATIONS.id = smstore_id,
    либо None, если свойства нет.

    SELECT propval
    FROM smstoreproperties
    WHERE storeloc = :sid AND propid = 'REP.DBNAME'
    """
    conn = cur = None
    try:
        conn = connect_oracle_supermag()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT propval
            FROM smstoreproperties
            WHERE storeloc = :sid
              AND propid = 'REP.DBNAME'
            """,
            sid=smstore_id
        )
        row = cur.fetchone()
        if not row:
            logger.warning(f"[SM/DBNAME] REP.DBNAME not found for STORELOC={smstore_id}")
            return None

        dbname = row[0]
        if dbname is None:
            logger.warning(f"[SM/DBNAME] REP.DBNAME is NULL for STORELOC={smstore_id}")
            return None

        dbname_str = str(dbname).strip()
        logger.info(f"[SM/DBNAME] STORELOC={smstore_id} -> DBNAME={dbname_str!r}")
        return dbname_str

    except Exception as e:
        logger.exception(f"[SM/DBNAME] Error while fetching dbname for STORELOC={smstore_id}: {e}")
        # Тут либо возвращаем None, либо пробрасываем. Явно пробрасываем, а view вернёт 500.
        raise
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass


@csrf_exempt
def sm_get_dbname(request):
    """
    GET  /sm/dbname/?storeId=225
    или
    POST /sm/dbname/  { "storeId": 225 }
    """
    if request.method not in ('GET', 'POST'):
        return JsonResponse({'status': 'error', 'message': 'Только GET или POST'}, status=405)

    try:
        raw_store = ""

        if request.method == 'GET':
            raw_store = (
                (request.GET.get('storeId') or
                 request.GET.get('smstore') or
                 request.GET.get('id') or "")
                .strip()
            )
            logger.info(f"[SM/GET_DBNAME] GET, raw_store={raw_store!r}")
        else:
            body = request.body.decode('utf-8') if request.body else "{}"
            logger.info(f"[SM/GET_DBNAME] POST, raw_body={body!r}")
            try:
                data = json.loads(body)
            except Exception as e:
                logger.error(f"[SM/GET_DBNAME] JSON parse error: {e}")
                return JsonResponse({'status': 'error', 'message': 'Некорректный JSON'}, status=400)
            raw_store = str(
                data.get('storeId') or
                data.get('smstore') or
                data.get('id') or ""
            ).strip()
            logger.info(f"[SM/GET_DBNAME] Parsed raw_store={raw_store!r}")

        if not raw_store:
            logger.error("[SM/GET_DBNAME] storeId/smstore не передан")
            return JsonResponse(
                {'status': 'error', 'message': 'Не передан storeId (smstore)'},
                status=400
            )

        try:
            smstore_id = int(raw_store)
        except ValueError:
            logger.error(f"[SM/GET_DBNAME] storeId не число: {raw_store!r}")
            return JsonResponse(
                {'status': 'error', 'message': 'storeId (smstore) должен быть числом'},
                status=400
            )

        logger.info(f"[SM/GET_DBNAME] Ищем REP.DBNAME для STORELOC={smstore_id}")

        try:
            dbname = get_dbname_for_smstore(smstore_id)
        except Exception as e:
            logger.exception(f"[SM/GET_DBNAME] Ошибка внутри get_dbname_for_smstore: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

        if not dbname:
            logger.warning(f"[SM/GET_DBNAME] REP.DBNAME не найден для STORELOC={smstore_id}")
            return JsonResponse(
                {'status': 'error',
                 'message': f'Имя базы (REP.DBNAME) не найдено для STORELOC={smstore_id}'},
                status=404
            )

        logger.info(
            f"[SM/GET_DBNAME] OK: STORELOC={smstore_id} → DBNAME={dbname!r}"
        )
        return JsonResponse({
            'status': 'ok',
            'smstore': smstore_id,
            'dbname': dbname,
        })

    except Exception as e:
        logger.exception("[SM/GET_DBNAME] Unexpected error")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
def sm_list_databases(request):
    """
    GET /sm/databases/
    Список магазинов + их база (REP.DBNAME)
    """
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Только GET'}, status=405)

    try:
        items = _fetch_sm_database_items()
        return JsonResponse({
            "status": "ok",
            "count": len(items),
            "items": items
        })
    except Exception as e:
        logger.exception("[SM/LIST_DB] Unexpected error")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
# @csrf_exempt
# def sm_list_databases(request):
#     """
#     GET /sm/databases/
#     Список магазинов + их база (REP.DBNAME)
#     """
#     if request.method != 'GET':
#         return JsonResponse({'status': 'error', 'message': 'Только GET'}, status=405)

#     sql = """
#         SELECT
#           l.id       AS smstore,
#           l.name     AS name,
#           p.propval  AS dbname
#         FROM smstorelocations l
#         LEFT JOIN smstoreproperties p
#           ON p.storeloc = l.id
#          AND p.propid   = 'REP.DBNAME'
#         WHERE
#           l.name NOT LIKE 'я%%'
#           AND l.name NOT LIKE '%%ТЕСТ%%'
#           AND l.formatid IN (19, 20, 9, 8, 6, 22, 23)
#         ORDER BY l.name
#     """

#     conn = cur = None
#     try:
#         logger.info("[SM/LIST_DB] Старт запроса списка магазинов + баз")
#         conn = connect_oracle_supermag()
#         cur = conn.cursor()
#         cur.execute(sql)

#         items = []
#         rows = cur.fetchall()
#         logger.info(f"[SM/LIST_DB] Получено строк из Oracle: {len(rows)}")

#         for smstore, name, dbname in rows:
#             name_s = (name or "").strip()
#             db_s = (dbname or "").strip() if dbname is not None else ""
#             logger.info(
#                 "[SM/LIST_DB] STORE: smstore=%s, name=%r, dbname=%r",
#                 smstore, name_s, db_s
#             )
#             items.append({
#                 "smstore": smstore,
#                 "name": name_s,
#                 "dbname": db_s,
#             })

#         logger.info(f"[SM/LIST_DB] Итог: count={len(items)}")
#         return JsonResponse({
#             "status": "ok",
#             "count": len(items),
#             "items": items
#         })

#     except Exception as e:
#         logger.exception("[SM/LIST_DB] Unexpected error")
#         return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
#     finally:
#         try:
#             if cur:
#                 cur.close()
#             if conn:
#                 conn.close()
#         except Exception:
#             pass




# Убрать после 30.03

@staff_member_required
@never_cache
@require_http_methods(["GET"])
def working_employees_excel_page(request):
    return render(
        request,
        "frostapp/working_employees_excel.html",
        {
            "title": "Выгрузка сотрудников по магазинам",
        }
    )


@staff_member_required
@csrf_protect
@require_http_methods(["POST"])
def working_employees_excel_generate(request):
    try:
        allowed_positions = {
            _normalize_position(x) for x in ALLOWED_WORKING_EMPLOYEE_POSITIONS
        }

        raw_employees = _fetch_working_employees_from_1c()
        sm_items = _fetch_sm_database_items()

        store_map: dict[int, dict] = {}
        for item in sm_items:
            smstore = _safe_int(item.get("smstore"))
            if smstore is not None:
                store_map[smstore] = item

        rows: list[dict] = []
        missing_store_ids: set[int] = set()

        for raw_item in raw_employees:
            employee = _normalize_working_employee(raw_item)
            if not employee:
                continue

            position_norm = _normalize_position(employee.get("position"))
            if position_norm not in allowed_positions:
                continue

            store_id = employee.get("store_id")
            store_info = store_map.get(store_id) if store_id is not None else None

            if store_id is not None and not store_info:
                missing_store_ids.add(store_id)

            rows.append({
                "dbname": _safe_str(store_info.get("dbname")) if store_info else "",
                "store_id": store_id or "",
                "store_name": _safe_str(store_info.get("name")) if store_info else "",
                "full_name": _safe_str(employee.get("full_name")),
                "position": _safe_str(employee.get("position")),
                "inn": _safe_str(employee.get("inn")),
            })

        rows.sort(
            key=lambda x: (
                str(x.get("dbname", "")),
                str(x.get("store_id", "")),
                str(x.get("full_name", "")),
            )
        )

        logger.info(
            "[WORKING_EMPLOYEES_EXCEL] filtered_rows=%s missing_store_ids=%s",
            len(rows),
            sorted(missing_store_ids),
        )

        excel_file = _build_working_employees_excel(rows)
        filename = f"working_employees_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    except requests.RequestException as e:
        logger.exception("[WORKING_EMPLOYEES_EXCEL] 1C request error")
        return render(
            request,
            "frostapp/working_employees_excel.html",
            {
                "title": "Выгрузка сотрудников по магазинам",
                "error": f"Ошибка запроса к 1С: {e}",
            },
            status=500,
        )
    except Exception as e:
        logger.exception("[WORKING_EMPLOYEES_EXCEL] Unexpected error")
        return render(
            request,
            "frostapp/working_employees_excel.html",
            {
                "title": "Выгрузка сотрудников по магазинам",
                "error": str(e),
            },
            status=500,
        )









# Базовый SQL
BASE_STORES_SQL = """
SELECT
  r.title AS region,
  t1.id   AS smstore,
  t1.name AS name,
  t1.address AS address,
  t2.propval AS closedate,
  t3.propval AS ukm4store,
  t5.propval AS ukm4ip,
  t4.propval AS ukm5store,
  t6.propval AS latitude,
  t7.propval AS longitude
FROM smstorelocations t1
JOIN smregions r ON r.rgnid = t1.rgnid
LEFT JOIN smstoreproperties t2 ON t2.storeloc = t1.id AND t2.propid = 'REP.CLOSEDATE'
LEFT JOIN smstoreproperties t3 ON t3.storeloc = t1.id AND t3.propid = 'REP.UKMStoreId'
LEFT JOIN smstoreproperties t4 ON t4.storeloc = t1.id AND t4.propid = 'REP.UKMSERVER5'
LEFT JOIN smstoreproperties t5 ON t5.storeloc = t1.id AND t5.propid = 'REP.UKMSERVER'
LEFT JOIN smstoreproperties t6 ON t6.storeloc = t1.id AND t6.propid = 'REP.STORE.Latitude'
LEFT JOIN smstoreproperties t7 ON t7.storeloc = t1.id AND t7.propid = 'REP.STORE.Longitude'
WHERE
  UPPER(t1.name) NOT LIKE 'Я %%'
  AND t1.name NOT LIKE '%%ТЕСТ%%'
  AND t1.idclass IN (
    SELECT id FROM supermag.sastoreclass
    WHERE tree LIKE '1.1.%%' OR tree LIKE '1.2.%%' OR tree LIKE '1.3.%%' OR tree LIKE '1.4.%%' OR tree LIKE '2.1.%%' OR tree LIKE '2.2.%%'
  )
  AND t1.accepted = 1
  AND t1.loctype = 4
  {only_active_clause}
  {extra_filters}
ORDER BY r.title, t1.name
"""

def _fetch_stores(q: str, region: str, only_active: bool):
    extra_filters_sql = []
    binds = {}

    if region:
        extra_filters_sql.append("AND r.title = :region")
        binds['region'] = region

    if q:
        like = f"%{q.lower()}%"
        if q.isdigit():
            extra_filters_sql.append("""
            AND (
                LOWER(t1.name) LIKE :q
                OR LOWER(t1.address) LIKE :q
                OR t1.id = :sid
                OR t3.propval = :ukm4sid
            )""")
            binds['q'] = like
            binds['sid'] = int(q)
            binds['ukm4sid'] = q
        else:
            extra_filters_sql.append("""
            AND (
                LOWER(t1.name) LIKE :q
                OR LOWER(t1.address) LIKE :q
            )""")
            binds['q'] = like

    only_active_clause = "AND (t2.propval IS NULL OR TO_DATE(t2.propval,'DD.MM.YYYY') >= TRUNC(SYSDATE))" if only_active else ""

    sql = BASE_STORES_SQL.format(
        only_active_clause=only_active_clause,
        extra_filters=(" " + " ".join(extra_filters_sql)) if extra_filters_sql else ""
    )

    rows = []
    conn = cur = None
    try:
        conn = connect_oracle_supermag()
        cur = conn.cursor()
        cur.execute(sql, binds)
        cols = [d[0].lower() for d in cur.description]
        for r in cur.fetchall():
            item = dict(zip(cols, r))
            # координаты -> float, если возможно
            for k in ('latitude', 'longitude'):
                if item.get(k) is not None:
                    s = str(item[k]).replace(',', '.')
                    try:
                        item[k] = float(s)
                    except Exception:
                        pass
            rows.append(item)
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except Exception:
            pass
    return rows

@csrf_exempt
def export_stores_xml(request):
    """
    GET /export/stores/xml/?q=&region=&only_active=1
    Сохраняет файл вида /app/xml/stores_YYYYMMDD_HHMMSS.xml
    """
    try:
        q = (request.GET.get('q') or '').strip()
        region = (request.GET.get('region') or '').strip()
        only_active = (request.GET.get('only_active') or '1') in ('1', 'true', 'True')

        rows = _fetch_stores(q, region, only_active)

        ts = datetime.datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        fname = f"stores_{ts}.xml"
        fpath = os.path.join(XML_DIR, fname)

        root = ET.Element("Stores")
        root.set("exported_at_utc", ts)
        root.set("only_active", "1" if only_active else "0")
        if q: root.set("q", q)
        if region: root.set("region", region)

        for r in rows:
            s_el = ET.SubElement(root, "Store")
            ET.SubElement(s_el, "Region").text     = str(r.get('region') or '')
            ET.SubElement(s_el, "SMSTORE").text    = str(r.get('smstore') or '')
            ET.SubElement(s_el, "Name").text       = str(r.get('name') or '')
            ET.SubElement(s_el, "Address").text    = str(r.get('address') or '')
            ET.SubElement(s_el, "CloseDate").text  = str(r.get('closedate') or '')
            ET.SubElement(s_el, "UKM4Store").text  = str(r.get('ukm4store') or '')
            ET.SubElement(s_el, "UKM4IP").text     = str(r.get('ukm4ip') or '')
            ET.SubElement(s_el, "UKM5Store").text  = str(r.get('ukm5store') or '')
            ET.SubElement(s_el, "Latitude").text   = '' if r.get('latitude') is None else str(r.get('latitude'))
            ET.SubElement(s_el, "Longitude").text  = '' if r.get('longitude') is None else str(r.get('longitude'))

        _write_xml_with_declaration(fpath, root, ensure_base=False)

        return JsonResponse({
            'status': 'ok',
            'saved_to': fpath,
            'filename': fname,
            'count': len(rows)
        })
    except Exception as e:
        logger.exception("export_stores_xml error")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@csrf_exempt
def register_cashier(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Только POST'}, status=405)

    try:
        data = json.loads(request.body)
        required_fields = ['inn', 'surname', 'name', 'patronymic', 'mail', 'phone', 'department', 'position', 'roleId', 'storeid']
        missing = [f for f in required_fields if not data.get(f)]
        if missing:
            logger.error(f"Пропущены поля: {missing}")
            return JsonResponse({'status': 'error', 'message': f'Пропущены поля: {missing}'}, status=400)

        inn = data['inn'].strip()
        fio = f"{data['surname']} {data['name']} {data['patronymic']}"
        mail = data['mail']
        phone = data['phone']
        dep_name = data['department']
        pos_name = data['position']
        role_id = data['roleId']

        store_ids = [int(s.strip()) for s in str(data['storeid']).split(',') if s.strip().isdigit()]
        if not store_ids:
            return JsonResponse({'status': 'error', 'message': 'Некорректный storeid'}, status=400)

        dep_obj = Department.objects.filter(name__iexact=dep_name).first()
        if not dep_obj:
            return JsonResponse({'status': 'error', 'message': f'Отдел «{dep_name}» не найден'}, status=400)

        pos_obj = Position.objects.filter(name__iexact=pos_name).first()
        if not pos_obj:
            return JsonResponse({'status': 'error', 'message': f'Должность «{pos_name}» не найдена'}, status=400)

        with transaction.atomic():
            base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
            xml_dir = os.path.join(base_dir, 'xml')
            os.makedirs(xml_dir, exist_ok=True)

            password_plain = build_user_password(inn)
            qr_string = password_plain

            # поиск/создание пользователя
            existing_user = User.objects.filter(encrypted_inn=inn, full_name=fio).first()
            if not existing_user:
                new_user = User.objects.create(
                    employee_id=inn,
                    encrypted_inn=inn,
                    full_name=fio,
                    mail=mail,
                    phone=phone,
                    department_id=dep_obj.id,
                    position_id=pos_obj.id,
                    active=True,
                    tg_status=False,
                    created_at=timezone.now(),
                    updated_at=timezone.now()
                )
                user_id = new_user.id

                OpenInSystem.objects.create(
                    user_id=user_id,
                    username=fio,
                    password=password_plain,
                    system_id=9,
                    status=True
                )

                QRCode.objects.create(
                    user=new_user,
                    qr_data=qr_string,
                    created_at=timezone.now()
                )

                logger.info(f"[PostgreSQL] Пользователь {fio} создан с ID={user_id}")
            else:
                new_user = existing_user
                user_id = new_user.id
                logger.info(f"[PostgreSQL] Пользователь уже существует: ID={user_id}")

            existing_storeids = set(UKMUser.objects.filter(user_id=user_id).values_list('storeid', flat=True))

            # id кассира берём из ukmserver
            ukm_conn = connect_ukm()
            ukm_cursor = ukm_conn.cursor()
            ukm_cursor.execute("SELECT MAX(id) + 1 AS next_id FROM trm_in_users")
            cashier_id_base = ukm_cursor.fetchone()['next_id'] or 1
            ukm_conn.close()

            cashier_counter = 0
            xml_paths = []

            for sid in store_ids:
                if sid in existing_storeids:
                    logger.info(f"[Пропуск] Пользователь уже зарегистрирован в storeid={sid}")
                    continue

                # PostgreSQL ukm_users
                UKMUser.objects.create(
                    user=new_user,
                    roleid=role_id,
                    storeid=sid,
                    version=1
                )

                cashier_id = cashier_id_base + cashier_counter
                cashier_counter += 1

                # Oracle → получить UKM4IP & UKM5
                info = get_store_info(sid)
                ukm4ip = info.get("ukm4ip")
                is_ukm5 = info.get("is_ukm5", False)

                # MySQL import4 на UKM4IP: users + signal
                if ukm4ip:
                    try:
                        conv = connect_store_mysql(ukm4ip)
                        cur = conv.cursor()

                        base_version = _calc_next_signal_version(cur)

                        cur.execute("""
                            INSERT INTO users (store, id, name, inn, password, role_id, version, deleted)
                            VALUES (%s, %s, %s, %s, OLD_PASSWORD(%s), %s, %s, 0)
                        """, (sid, cashier_id, fio, inn, mysql_pwd(password_plain), role_id, base_version))

                        cur.execute("INSERT INTO `signal`(`signal`, `version`) VALUES ('incr', %s)", (base_version,))
                        conv.commit()
                        conv.close()
                        logger.info(
                            f"[MySQL:{ukm4ip}] Добавлен кассир store={sid}, "
                            f"id={cashier_id}, version={base_version}"
                        )
                    except Exception as e:
                        logger.error(f"[MySQL:{ukm4ip}] Ошибка вставки для store={sid}: {e}")
                else:
                    logger.error(f"[Oracle] Не найден UKM4IP для storeid={sid}. Пропуск записи в MySQL.")

                # XML для UKM5
                if is_ukm5:
                    if sid == UKM5_FULL_XML_STORE_ID:
                        # Для магазина 2013 — полный файл, со ВСЕМИ кассирами
                        try:
                            full_xml_path = build_full_ukm5_xml_for_store(sid)
                            if full_xml_path:
                                xml_paths.append(full_xml_path)
                                logger.info(
                                    f"[XML] Полный XML для УКМ5 магазина {sid} пересобран: {full_xml_path}"
                                )
                        except Exception as e:
                            logger.error(
                                f"[XML] Ошибка полной пересборки для storeid={sid}: {e}"
                            )
                    else:
                        # Остальные магазины — старое поведение (вставка одного кассира)
                        try:
                            xml_path, tree, root = _get_or_create_storecashiers_tree(sid)

                            cashier_el = ET.SubElement(root, "cashier")
                            ET.SubElement(cashier_el, "roleId").text = str(role_id)
                            ET.SubElement(cashier_el, "id").text = str(cashier_id)
                            ET.SubElement(cashier_el, "name").text = fio
                            ET.SubElement(cashier_el, "INN").text = inn
                            ET.SubElement(cashier_el, "password").text = password_plain

                            _write_xml_with_declaration(xml_path, root, ensure_base=True)
                            xml_paths.append(xml_path)
                            logger.info(f"[XML] Файл обновлён: {xml_path}")
                        except Exception as e:
                            logger.error(f"[XML] Ошибка генерации для storeid={sid}: {e}")

        return JsonResponse({
            'status': 'ok',
            'message': 'Кассир зарегистрирован для всех storeid',
            'password': password_plain,
            'storeids': store_ids,
            'xml_paths': xml_paths
        })

    except Exception as e:
        logger.exception("Ошибка регистрации кассира")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def encrypt_inn(inn):
    """
    Хэшируем (SHA-256) поле inn, если это строка из цифр.
    """
    if not isinstance(inn, str) or not inn.isdigit():
        raise ValueError("ИНН должен быть строкой, содержащей только цифры.")
    hash_object = hashlib.sha256(inn.encode('utf-8'))
    return hash_object.hexdigest()


def validate_and_create_record(payload, required_fields, action_name="CREATE"):
    data = payload.get('data', {})
    if not isinstance(data, dict):
        data = {}
    else:
        # чтобы не мутировать исходный payload
        data = dict(data)

    final_status = 'pending'
    attempts = 0
    last_attempt = None

    # guid_dep: нормализуем как строку
    if 'guid_dep' in data:
        gd = data.get('guid_dep')
        gd = "" if gd is None else str(gd)
        gd = gd.strip()
        data['guid_dep'] = gd or None  # пустое -> None (чтобы попало в missing)

    # ИНН: сохраняем чистый в id_compare + хэшируем в inn
    raw_inn = data.get('inn')
    if raw_inn:
        try:
            plain_inn = ensure_plain_inn(str(raw_inn))  # 10/12 цифр
            data['id_compare'] = plain_inn
            data['inn'] = encrypt_inn(plain_inn)
        except ValueError:
            data['inn'] = None
            data.pop('id_compare', None)

    # Проверяем обязательные поля (после нормализации)
    missing = [f for f in required_fields if not data.get(f)]
    if missing:
        final_status = 'failed'

    Queue.objects.create(
        data=data,
        attempts=attempts,
        status=final_status,
        last_attempt=last_attempt
    )

    if final_status == 'failed':
        MODUL_logs.objects.create(
            data={
                "error": f"Незаполненные поля: {missing}",
                "payload": data
            }
        )

    return final_status, missing


@csrf_exempt
def queue_create(request):
    """
    Эндпойнт для СОЗДАНИЯ (create).
    Обязательны все поля, кроме:
    surname, start_vacation, end_vacation, type_vacation, shops_id
    """
    if request.method == 'POST':
        try:
            payload = json.loads(request.body.decode('utf-8'))

            required_fields = [
                'guid_dep', 'inn', 'last_name', 'first_name', 'beginwork_date',
                'organization', 'department', 'bitrix_code', 'subdivision',
                'position', 'phone', 'birth_date', 'action'
            ]

            final_status, missing = validate_and_create_record(
                payload, required_fields, action_name='CREATE'
            )

            if missing:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Пропущены или пусты поля: {missing}. Запись -> failed, логи созданы.'
                }, status=400)
            else:
                return JsonResponse({'status': 'ok', 'message': 'Создано успешно'}, status=201)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': 'Только POST'}, status=405)


@csrf_exempt
def queue_update(request):
    """
    Эндпойнт для ИЗМЕНЕНИЯ (update).
    Обязательны все поля, кроме:
    surname, start_vacation, end_vacation, type_vacation, shops_id
    """
    if request.method == 'POST':
        try:
            payload = json.loads(request.body.decode('utf-8'))

            required_fields = [
                'guid_dep', 'inn', 'last_name', 'first_name', 'beginwork_date',
                'organization', 'department', 'bitrix_code', 'subdivision',
                'position', 'phone', 'birth_date', 'action'
            ]

            final_status, missing = validate_and_create_record(
                payload, required_fields, action_name='UPDATE'
            )

            if missing:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Пропущены поля: {missing}. Запись -> failed, логи созданы.'
                }, status=400)
            else:
                return JsonResponse({'status': 'ok', 'message': 'Изменение успешно'}, status=201)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': 'Только POST'}, status=405)


@csrf_exempt
def queue_block(request):
    """
    Эндпойнт для БЛОКИРОВКИ (block).
    Обязательные поля:
    inn, last_name, first_name
    """
    if request.method == 'POST':
        try:
            payload = json.loads(request.body.decode('utf-8'))

            required_fields = [
                'guid_dep', 'inn', 'last_name', 'first_name'
            ]

            final_status, missing = validate_and_create_record(
                payload, required_fields, action_name='DELETE'
            )

            if missing:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Пропущены поля: {missing}. Запись -> failed, логи созданы.'
                }, status=400)
            else:
                return JsonResponse({'status': 'ok', 'message': 'Блокировка/Удаление успешно'}, status=201)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': 'Только POST'}, status=405)


@csrf_exempt
def queue_vacation(request):
    """
    Эндпойнт для ОТПУСКА (vacation).
    Обязательные поля:
    inn, last_name, first_name,
    start_vacation, end_vacation, type_vacation
    """
    if request.method == 'POST':
        try:
            payload = json.loads(request.body.decode('utf-8'))

            required_fields = [
                'guid_dep', 'inn', 'last_name', 'first_name',
                'start_vacation', 'end_vacation', 'type_vacation'
            ]

            final_status, missing = validate_and_create_record(
                payload, required_fields, action_name='CREATE'
            )

            if missing:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Пропущены поля: {missing}. Запись -> failed, логи созданы.'
                }, status=400)
            else:
                return JsonResponse({'status': 'ok', 'message': 'Отпуск успешно'}, status=201)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': 'Только POST'}, status=405)


def regenerate_qr(user):
    new_password = build_user_password(user.employee_id)
    now = timezone.now()
    expiration = now + datetime.timedelta(days=1)

    # PostgreSQL: QR + open_in_system
    QRCode.objects.filter(user=user).delete()
    QRCode.objects.create(user=user, qr_data=new_password, created_at=now, expires_at=expiration)
    OpenInSystem.objects.filter(user_id=user.id, system_id=9).update(password=new_password)

    # для id кассира (как было у тебя)
    ukm_conn = connect_ukm()
    ukm_cursor = ukm_conn.cursor()
    ukm_cursor.execute("SELECT MAX(id)+1 AS next_id FROM trm_in_users")
    cashier_id_base = ukm_cursor.fetchone()['next_id'] or 1
    ukm_conn.close()

    ukm_emp_id = get_trm_employee_id(user.employee_id, user.full_name)
    ukm_users = list(UKMUser.objects.filter(user_id=user.id))
    cashier_counter = 0

    for ukm_user in ukm_users:
        sid = ukm_user.storeid
        cashier_id = ukm_emp_id if ukm_emp_id else (cashier_id_base + cashier_counter)

        info = get_store_info(sid)
        ukm4ip = info.get("ukm4ip")

        if ukm4ip:
            conv = cur = None
            try:
                conv = connect_store_mysql(ukm4ip)
                cur = conv.cursor()

                base_version = _calc_next_signal_version(cur)

                ttl_supported = _mysql_users_supports_ttl_cols(cur, cache_key=str(ukm4ip))
                if ttl_supported:
                    # start_date = now.date()
                    end_date = expiration.date()
                    start_date = end_date

                    cur.execute("""
                        INSERT INTO users (
                            store, id, name, inn, password, role_id, version, deleted,
                            start_date, end_date
                        )
                        VALUES (
                            %s, %s, %s, %s, OLD_PASSWORD(%s), %s, %s, 0,
                            %s, %s
                        )
                        ON DUPLICATE KEY UPDATE
                            name       = VALUES(name),
                            inn        = VALUES(inn),
                            password   = VALUES(password),
                            role_id    = VALUES(role_id),
                            version    = VALUES(version),
                            deleted    = 0,
                            # start_date = VALUES(start_date),
                            start_date = VALUES(end_date),
                            end_date   = VALUES(end_date)
                    """, (
                        sid,
                        cashier_id,
                        user.full_name,
                        user.employee_id,
                        mysql_pwd(new_password),
                        ukm_user.roleid,
                        base_version,
                        start_date,
                        end_date
                    ))
                else:
                    cur.execute("""
                        INSERT INTO users (store, id, name, inn, password, role_id, version, deleted)
                        VALUES (%s, %s, %s, %s, OLD_PASSWORD(%s), %s, %s, 0)
                        ON DUPLICATE KEY UPDATE
                            name     = VALUES(name),
                            inn      = VALUES(inn),
                            password = VALUES(password),
                            role_id  = VALUES(role_id),
                            version  = VALUES(version),
                            deleted  = 0
                    """, (
                        sid,
                        cashier_id,
                        user.full_name,
                        user.employee_id,
                        mysql_pwd(new_password),
                        ukm_user.roleid,
                        base_version
                    ))

                cur.execute(
                    "INSERT INTO `signal`(`signal`, `version`) VALUES ('incr', %s)",
                    (base_version,)
                )

                conv.commit()
                logger.info(
                    f"[MySQL:{ukm4ip}] Пароль обновлён store={sid}, "
                    f"id={cashier_id}, version={base_version}"
                )
            except Exception as e:
                logger.error(f"[MySQL:{ukm4ip}] Ошибка обновления пароля для store={sid}: {e}", exc_info=True)
                if conv:
                    try:
                        conv.rollback()
                    except Exception:
                        pass
            finally:
                try:
                    if cur:
                        cur.close()
                    if conv:
                        conv.close()
                except Exception:
                    pass
        else:
            logger.error(f"[Oracle] Не найден UKM4IP для storeid={sid}. Пропуск записи в MySQL.")

        cashier_counter += 1

    # XML для UKM5 (обновляем записи)
    next_free_id = cashier_id_base + cashier_counter

    for ukm_user in ukm_users:
        sid = ukm_user.storeid
        if not is_ukm5_store(sid):
            continue

        try:
            if sid == UKM5_FULL_XML_STORE_ID:
                # Для магазина 2013 при смене пароля пересобираем полный XML
                xml_path = build_full_ukm5_xml_for_store(sid)
                logger.info(
                    f"[XML] Полный XML пересобран при регенерации QR для storeid={sid}: {xml_path}"
                )
            else:
                # Остальные магазины — точечное обновление одного кассира
                xml_path, tree, root = _get_or_create_storecashiers_tree(sid)

                for el in list(root.findall("cashier")):
                    if (el.findtext("INN") or "").strip() == str(user.employee_id).strip():
                        root.remove(el)

                c_el = ET.SubElement(root, "cashier")
                ET.SubElement(c_el, "roleId").text = str(ukm_user.roleid)
                ET.SubElement(c_el, "id").text = str(next_free_id)
                ET.SubElement(c_el, "name").text = user.full_name
                ET.SubElement(c_el, "INN").text = str(user.employee_id).strip()
                ET.SubElement(c_el, "password").text = new_password

                _write_xml_with_declaration(xml_path, root, ensure_base=True)
                logger.info(f"[XML] Обновлён при регенерации QR: {xml_path}")
                next_free_id += 1
        except Exception as exc:
            logger.error(f"[XML] Ошибка для {sid}: {exc}", exc_info=True)



def _set_password_pg(user, new_password: str) -> None:
    """
    Обновляет пароль во всех связанных PG-таблицах:
      • qr_code  — пересоздаём запись (expires_at = now + 1 day)
      • open_in_system (system_id=9) — создаём/обновляем пароль
    """
    now = timezone.now()
    expiration = now + datetime.timedelta(days=1)

    # QR: пересоздаём, чтобы был всегда один актуальный
    QRCode.objects.filter(user=user).delete()
    QRCode.objects.create(
        user=user,
        qr_data=new_password,
        created_at=now,
        expires_at=expiration
    )
    logger.info(f"[PG] QRCode regenerated for user_id={user.id}; expires_at={expiration.isoformat()}")

    # open_in_system: обновляем, если запись есть; иначе создаём
    updated = OpenInSystem.objects.filter(user_id=user.id, system_id=9).update(
        username=user.full_name,
        password=new_password,
        status=True
    )
    if updated:
        logger.info(f"[PG] OpenInSystem updated (system_id=9) for user_id={user.id}")
    else:
        OpenInSystem.objects.create(
            user_id=user.id,
            username=user.full_name,
            password=new_password,
            system_id=9,
            status=True
        )
        logger.info(f"[PG] OpenInSystem created (system_id=9) for user_id={user.id}")


def _mask_phone(phone: str) -> str:
    phone = str(phone or "").strip()
    if len(phone) <= 4:
        return phone
    return f"{phone[:2]}***{phone[-2:]}"


def _mask_email(email: str) -> str:
    email = str(email or "").strip()
    if "@" not in email:
        return email
    local, domain = email.split("@", 1)
    if len(local) <= 2:
        local_masked = local[:1] + "*"
    else:
        local_masked = local[:2] + "***"
    return f"{local_masked}@{domain}"


def _agent_error(message: str, status: int = 500):
    return JsonResponse({"error": str(message)}, status=status)


def _send_admin_log_async(message: str) -> None:
    """
    Совместимый алиас:
    отправляет админ-лог в Telegram и в MAX асинхронно,
    не блокируя HTTP-ответ.
    """
    _send_telegram_log_async(message)
    _send_max_log_async(message)
# def _send_admin_log_async(message: str) -> None:
#     """
#     Не блокирует HTTP-ответ.
#     """
#     def _worker():
#         try:
#             send_telegram_log(message)
#         except Exception as e:
#             logger.exception(f"[ADMIN_LOG] async send error: {e}")

#     try:
#         threading.Thread(target=_worker, daemon=True).start()
#     except Exception as e:
#         logger.exception(f"[ADMIN_LOG] thread start error: {e}")


def _normalize_oracle_value(value):
    if value is None:
        return None
    value = str(value).strip()
    return value if value else None


# def _get_supermag_store_meta_map(smstore_ids):
#     """
#     Возвращает словарь:
#     {
#         smstore_id: {
#             "name": ...,
#             "address": ...,
#             "inn": ...,
#             "kpp": ...,
#             "fsrar_id": ...,
#             "ukm_server_ip": ...,
#             "dbname": ...,
#             "subformat2": ...
#         }
#     }
#     """
#     prepared_ids = []
#     for sid in smstore_ids or []:
#         if sid is None:
#             continue
#         try:
#             prepared_ids.append(int(sid))
#         except Exception:
#             continue

#     prepared_ids = list(dict.fromkeys(prepared_ids))
#     if not prepared_ids:
#         return {}

#     bind_names = []
#     binds = {}
#     for i, sid in enumerate(prepared_ids):
#         key = f"sid{i}"
#         bind_names.append(f":{key}")
#         binds[key] = sid

#     sql = f"""
#         SELECT
#             t1.id AS smstore,
#             t1.name AS name,
#             t1.address AS address,
#             ci.inn AS inn,
#             t1.kpp AS kpp,
#             (SELECT propval
#                FROM smstoreproperties
#               WHERE propid = 'EGAIS.FSRARID'
#                 AND storeloc = t1.id) AS fsrar_id,
#             (SELECT propval
#                FROM smstoreproperties
#               WHERE propid = 'REP.UKMSERVER'
#                 AND storeloc = t1.id) AS ukm_server_ip,
#             (SELECT propval
#                FROM smstoreproperties
#               WHERE propid = 'REP.DBNAME'
#                 AND storeloc = t1.id) AS dbname,
#             (SELECT UPPER(a.name)
#                FROM Supermag.SAStoresAssort a,
#                     Supermag.SMStoresAssort m
#               WHERE a.Tree LIKE '2.2.%'
#                 AND a.ID = m.IDAssort
#                 AND m.IDLoc = t1.ID) AS subformat2
#         FROM SMSTORELOCATIONS t1
#         JOIN SMREGIONS r
#           ON r.rgnid = t1.rgnid
#         JOIN smownclientlocs ol
#           ON ol.locid = t1.id
#         JOIN smclientinfo ci
#           ON ci.id = ol.clientid
#         JOIN sastoreformats f
#           ON f.id = t1.formatid
#         JOIN sastoreclass sc
#           ON sc.id = t1.idclass
#         LEFT JOIN (
#             SELECT STORELOC, PROPVAL
#             FROM SMSTOREPROPERTIES
#             WHERE PROPID = 'REP.CLOSEDATE'
#         ) t2
#           ON t1.ID = t2.STORELOC
#         WHERE
#             UPPER(t1.name) NOT LIKE 'Я %'
#             AND t1.name NOT LIKE '%ТЕСТ%'
#             AND sc.tree IS NOT NULL
#             AND t1.idclass IN (
#                 SELECT id
#                 FROM supermag.sastoreclass
#                 WHERE tree LIKE '1.1.%'
#                    OR tree LIKE '1.2.%'
#                    OR tree LIKE '1.3.%'
#                    OR tree LIKE '1.4.%'
#                    OR tree LIKE '2.1.%'
#                    OR tree LIKE '2.2.%'
#             )
#             AND t1.accepted = 1
#             AND t1.loctype = 4
#             AND (t2.PROPVAL IS NULL OR TO_DATE(t2.PROPVAL, 'DD.MM.YYYY') >= TO_DATE(SYSDATE))
#             AND t1.id IN ({", ".join(bind_names)})
#         ORDER BY t1.id
#     """

#     conn = cur = None
#     result = {}
#     try:
#         conn = connect_oracle_supermag()
#         cur = conn.cursor()
#         cur.execute(sql, binds)

#         cols = [d[0].lower() for d in cur.description]
#         for row in cur.fetchall():
#             item = dict(zip(cols, row))
#             smstore = item.get("smstore")
#             if smstore is None:
#                 continue

#             result[int(smstore)] = {
#                 "name": _normalize_oracle_value(item.get("name")),
#                 "address": _normalize_oracle_value(item.get("address")),
#                 "inn": _normalize_oracle_value(item.get("inn")),
#                 "kpp": _normalize_oracle_value(item.get("kpp")),
#                 "fsrar_id": _normalize_oracle_value(item.get("fsrar_id")),
#                 "ukm_server_ip": _normalize_oracle_value(item.get("ukm_server_ip")),
#                 "dbname": _normalize_oracle_value(item.get("dbname")),
#                 "market": _normalize_oracle_value(item.get("subformat2")),
#             }

#         return result

#     except Exception as e:
#         logger.exception(f"[AGENT_AUTH] Ошибка получения расширенных данных магазинов из Supermag: {e}")
#         return {}
#     finally:
#         try:
#             if cur:
#                 cur.close()
#             if conn:
#                 conn.close()
#         except Exception:
#             pass

def _get_supermag_store_meta_map(smstore_ids):
    """
    Возвращает словарь:
    {
        smstore_id: {
            "name": ...,
            "address": ...,
            "inn": ...,
            "kpp": ...,          # ВАЖНО: теперь из supermag.smclientinfo
            "fsrar_id": ...,
            "ukm_server_ip": ...,
            "dbname": ...,
            "market": ...
        }
    }
    """
    prepared_ids = []
    for sid in smstore_ids or []:
        if sid is None:
            continue
        try:
            prepared_ids.append(int(sid))
        except Exception:
            continue

    prepared_ids = list(dict.fromkeys(prepared_ids))
    if not prepared_ids:
        return {}

    bind_names = []
    binds = {}
    for i, sid in enumerate(prepared_ids):
        key = f"sid{i}"
        bind_names.append(f":{key}")
        binds[key] = sid

    sql = f"""
        SELECT
            t1.id AS smstore,
            t1.name AS name,
            t1.address AS address,
            ci.inn AS inn,
            ci.kpp AS kpp,
            (SELECT propval
               FROM smstoreproperties
              WHERE propid = 'EGAIS.FSRARID'
                AND storeloc = t1.id) AS fsrar_id,
            (SELECT propval
               FROM smstoreproperties
              WHERE propid = 'REP.UKMSERVER'
                AND storeloc = t1.id) AS ukm_server_ip,
            (SELECT propval
               FROM smstoreproperties
              WHERE propid = 'REP.DBNAME'
                AND storeloc = t1.id) AS dbname,
            (SELECT UPPER(a.name)
               FROM Supermag.SAStoresAssort a,
                    Supermag.SMStoresAssort m
              WHERE a.Tree LIKE '2.2.%'
                AND a.ID = m.IDAssort
                AND m.IDLoc = t1.ID) AS subformat2
        FROM SMSTORELOCATIONS t1
        JOIN SMREGIONS r
          ON r.rgnid = t1.rgnid
        JOIN smownclientlocs ol
          ON ol.locid = t1.id
        JOIN smclientinfo ci
          ON ci.id = ol.clientid
        JOIN sastoreformats f
          ON f.id = t1.formatid
        JOIN sastoreclass sc
          ON sc.id = t1.idclass
        LEFT JOIN (
            SELECT STORELOC, PROPVAL
            FROM SMSTOREPROPERTIES
            WHERE PROPID = 'REP.CLOSEDATE'
        ) t2
          ON t1.ID = t2.STORELOC
        WHERE
            UPPER(t1.name) NOT LIKE 'Я %'
            AND t1.name NOT LIKE '%ТЕСТ%'
            AND sc.tree IS NOT NULL
            AND t1.idclass IN (
                SELECT id
                FROM supermag.sastoreclass
                WHERE tree LIKE '1.1.%'
                   OR tree LIKE '1.2.%'
                   OR tree LIKE '1.3.%'
                   OR tree LIKE '1.4.%'
                   OR tree LIKE '2.1.%'
                   OR tree LIKE '2.2.%'
            )
            AND t1.accepted = 1
            AND t1.loctype = 4
            AND (t2.PROPVAL IS NULL OR TO_DATE(t2.PROPVAL, 'DD.MM.YYYY') >= TO_DATE(SYSDATE))
            AND t1.id IN ({", ".join(bind_names)})
        ORDER BY t1.id
    """

    conn = cur = None
    result = {}
    try:
        conn = connect_oracle_supermag()
        cur = conn.cursor()
        cur.execute(sql, binds)

        cols = [d[0].lower() for d in cur.description]
        for row in cur.fetchall():
            item = dict(zip(cols, row))
            smstore = item.get("smstore")
            if smstore is None:
                continue

            result[int(smstore)] = {
                "name": _normalize_oracle_value(item.get("name")),
                "address": _normalize_oracle_value(item.get("address")),
                "inn": _normalize_oracle_value(item.get("inn")),
                "kpp": _normalize_oracle_value(item.get("kpp")),
                "fsrar_id": _normalize_oracle_value(item.get("fsrar_id")),
                "ukm_server_ip": _normalize_oracle_value(item.get("ukm_server_ip")),
                "dbname": _normalize_oracle_value(item.get("dbname")),
                "market": _normalize_oracle_value(item.get("subformat2")),
            }

        return result

    except Exception as e:
        logger.exception(f"[AGENT_AUTH] Ошибка получения расширенных данных магазинов из Supermag: {e}")
        return {}
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass




def send_max_to_user(max_user_id, message: str) -> bool:
    """
    Отправка сообщения пользователю в MAX по user_id.
    Возвращает True/False, исключения не пробрасывает.
    """
    max_raw = getattr(max_user_id, "id", max_user_id)

    if max_raw is None:
        max_id = ""
    else:
        max_id = str(max_raw).strip()

    if not max_id:
        logger.warning("[MAX] max_id отсутствует или пустой")
        return False

    if not MAX_BOT_TOKEN:
        logger.error("[MAX] Не задан MAX_BOT_TOKEN")
        return False

    try:
        max_id_int = int(max_id)
    except Exception:
        logger.error(f"[MAX] Некорректный max_id={max_id!r}")
        return False

    url = f"{MAX_API_BASE}/messages"
    headers = {
        "Authorization": MAX_BOT_TOKEN,
        "Content-Type": "application/json",
    }

    try:
        resp = requests.post(
            url,
            params={"user_id": max_id_int},
            json={"text": message},
            headers=headers,
            timeout=MAX_HTTP_TIMEOUT,
        )

        if 200 <= resp.status_code < 300:
            logger.info(f"[MAX] Сообщение успешно отправлено max_id={max_id_int}")
            return True

        logger.error(
            f"[MAX] Ошибка отправки max_id={max_id_int} "
            f"status={resp.status_code} body={resp.text[:1000]}"
        )
        return False

    except Exception as e:
        logger.exception(f"[MAX] Исключение при отправке max_id={max_id_int}: {e}")
        return False




def send_email_to_user(user, subject: str, message: str) -> bool:
    """
    Отправка email пользователю на адрес из users.mail.
    Возвращает True/False, исключения наружу не пробрасывает.
    """
    user_id = getattr(user, "id", None)
    email = str(getattr(user, "mail", "") or "").strip()

    if not email:
        logger.warning(f"[EMAIL] У пользователя отсутствует mail user_id={user_id}")
        return False

    try:
        validate_email(email)
    except ValidationError:
        logger.error(f"[EMAIL] Некорректный email user_id={user_id} email={email!r}")
        return False

    from_email = (
        getattr(settings, "DEFAULT_FROM_EMAIL", None)
        or getattr(settings, "EMAIL_HOST_USER", None)
    )

    if not from_email:
        logger.error("[EMAIL] Не настроен DEFAULT_FROM_EMAIL / EMAIL_HOST_USER")
        return False

    try:
        msg = EmailMultiAlternatives(
            subject=subject,
            body=message,
            from_email=from_email,
            to=[email],
        )
        msg.send(fail_silently=False)

        logger.info(
            f"[EMAIL] Письмо успешно отправлено user_id={user_id} email={email}"
        )
        return True

    except Exception as e:
        logger.exception(
            f"[EMAIL] Ошибка отправки письма user_id={user_id} email={email}: {e}"
        )
        return False




def send_pin_to_user_channels(user, pin: str) -> dict:
    """
    Отправляет PIN по приоритету каналов:

    1. Если у пользователя есть max_id -> отправляем ТОЛЬКО в MAX
    2. Если max_id нет, но есть mail -> отправляем ТОЛЬКО на email
    3. Если нет ни max_id, ни mail -> ничего не отправляем

    ВАЖНО:
    - Telegram для PIN сейчас НЕ используется
    - fallback с MAX на email НЕ выполняется, если max_id есть
    """
    ttl_text = f"{PIN_TTL_MINUTES} минут"
    pin_message = f"Ваш код авторизации: {pin}\nОн действителен {ttl_text}."

    max_sent = False
    email_sent = False

    user_id = getattr(user, "id", None)
    max_id = getattr(user, "max_id", None)
    email = str(getattr(user, "mail", "") or "").strip()

    # Приоритет MAX
    if max_id:
        try:
            max_sent = bool(send_max_to_user(max_id, pin_message))
            if max_sent:
                logger.info(
                    f"[PIN_DELIVERY] PIN отправлен в MAX user_id={user_id} max_id={max_id}"
                )
            else:
                logger.error(
                    f"[PIN_DELIVERY] Не удалось отправить PIN в MAX user_id={user_id} max_id={max_id}"
                )
        except Exception as e:
            logger.exception(
                f"[PIN_DELIVERY] Ошибка отправки PIN в MAX user_id={user_id} "
                f"max_id={max_id}: {e}"
            )
            max_sent = False

        return {
            "ok": max_sent,
            "telegram_sent": False,
            "max_sent": max_sent,
            "email_sent": False,
        }

    # Если MAX нет — используем email
    if email:
        try:
            subject = "Код авторизации"
            email_sent = bool(send_email_to_user(user, subject, pin_message))

            if email_sent:
                logger.info(
                    f"[PIN_DELIVERY] PIN отправлен на email user_id={user_id} email={email}"
                )
            else:
                logger.error(
                    f"[PIN_DELIVERY] Не удалось отправить PIN на email user_id={user_id} email={email}"
                )
        except Exception as e:
            logger.exception(
                f"[PIN_DELIVERY] Ошибка отправки PIN на email user_id={user_id} "
                f"email={email}: {e}"
            )
            email_sent = False

        return {
            "ok": email_sent,
            "telegram_sent": False,
            "max_sent": False,
            "email_sent": email_sent,
        }

    # Нет ни одного канала
    logger.warning(
        f"[PIN_DELIVERY] У пользователя нет ни max_id, ни mail user_id={user_id}"
    )
    return {
        "ok": False,
        "telegram_sent": False,
        "max_sent": False,
        "email_sent": False,
    }

# def send_pin_to_user_channels(user, pin: str) -> dict:
#     """
#     Отправляет PIN:
#     - в Telegram, если есть tg_id
#     - в MAX, если есть max_id

#     Отправка идёт параллельно, чтобы не ждать каналы по очереди.
#     """
#     message = f"Ваш код авторизации: {pin}\nОн действителен {PIN_TTL_MINUTES} минут."

#     futures = {}
#     telegram_sent = False
#     max_sent = False

#     try:
#         with ThreadPoolExecutor(max_workers=2) as executor:
#             if getattr(user, "tg_id", None):
#                 futures["telegram"] = executor.submit(send_telegram_to_user, user, message)

#             if getattr(user, "max_id", None):
#                 futures["max"] = executor.submit(send_max_to_user, user.max_id, message)

#             if "telegram" in futures:
#                 try:
#                     telegram_sent = bool(futures["telegram"].result())
#                 except Exception as e:
#                     logger.exception(
#                         f"[PIN_DELIVERY] Ошибка отправки в Telegram user_id={getattr(user, 'id', None)}: {e}"
#                     )
#                     telegram_sent = False

#             if "max" in futures:
#                 try:
#                     max_sent = bool(futures["max"].result())
#                 except Exception as e:
#                     logger.exception(
#                         f"[PIN_DELIVERY] Ошибка отправки в MAX user_id={getattr(user, 'id', None)} "
#                         f"max_id={getattr(user, 'max_id', None)}: {e}"
#                     )
#                     max_sent = False

#     except Exception as e:
#         logger.exception(f"[PIN_DELIVERY] Общая ошибка отправки PIN: {e}")

#     return {
#         "ok": telegram_sent or max_sent,
#         "telegram_sent": telegram_sent,
#         "max_sent": max_sent,
#     }


def _get_agent_user_stores(user):
    """
    Возвращает уникальные магазины пользователя из ukm_users
    + обогащает их данными из Postgres stores и Oracle Supermag.
    """
    ukm_links = list(
        UKMUser.objects
        .filter(user=user)
        .values('storeid', 'roleid')
        .order_by('storeid', 'id')
    )

    if not ukm_links:
        return []

    unique_links = {}
    for row in ukm_links:
        sid = row['storeid']
        if sid not in unique_links:
            unique_links[sid] = row

    store_ids = list(unique_links.keys())

    store_map = {
        s.ukm4store: s
        for s in Store.objects.filter(ukm4store__in=store_ids)
    }

    smstore_ids = []
    for sid in store_ids:
        store_obj = store_map.get(sid)
        if store_obj and store_obj.smstore is not None:
            smstore_ids.append(store_obj.smstore)

    supermag_meta_map = _get_supermag_store_meta_map(smstore_ids)

    stores_payload = []
    for sid, row in unique_links.items():
        store_obj = store_map.get(sid)
        smstore = store_obj.smstore if store_obj else None
        meta = supermag_meta_map.get(int(smstore), {}) if smstore is not None else {}

        dbname = meta.get('dbname')
        if not dbname and smstore is not None:
            try:
                dbname = get_dbname_for_smstore(smstore)
            except Exception as e:
                logger.error(
                    f"[AGENT_AUTH] Ошибка получения DBNAME для smstore={smstore}: {e}",
                    exc_info=True
                )
                dbname = None

        stores_payload.append({
            'ukm_storeid': sid,
            'smstore': smstore,
            'name': meta.get('name') or (store_obj.name if store_obj else '') or '',
            'address': meta.get('address') or (store_obj.address if store_obj else '') or '',
            'roleid': row.get('roleid'),
            'dbname': dbname,
            'ukm_server_ip': meta.get('ukm_server_ip') or (str(store_obj.ukm4ip) if store_obj and store_obj.ukm4ip else None),
            'inn': meta.get('inn'),
            'kpp': meta.get('kpp'),
            'fsrar_id': meta.get('fsrar_id'),
            'market': meta.get('market'),
        })

    return stores_payload


def _get_agent_primary_credentials(user_id):
    """
    Возвращает первую активную username/password пользователя
    из open_in_system только для system_id = 4.
    """
    cred = (
        OpenInSystem.objects
        .filter(
            user_id=user_id,
            system_id=4,
            status=True
        )
        .only('username', 'password')
        .order_by('id')
        .first()
    )

    if not cred:
        return None

    return {
        'username': cred.username,
        'password': cred.password,
    }


def _build_agent_stores_response(user):
    """
    Формирует список магазинов, где в каждом магазине
    лежат плоские username/password.
    """
    stores = _get_agent_user_stores(user)
    cred = _get_agent_primary_credentials(user.id)

    stores_response = []
    for store in stores:
        stores_response.append({
            'ukm_storeid': store['ukm_storeid'],
            'smstore': store['smstore'],
            'name': store['name'],
            'address': store['address'],
            'roleid': store['roleid'],
            'dbname': store['dbname'],
            'ukm_server_ip': store['ukm_server_ip'],
            'inn': store['inn'],
            'kpp': store['kpp'],
            'fsrar_id': store['fsrar_id'],
            'market': store['market'],
            'username': cred['username'] if cred else None,
            'password': cred['password'] if cred else None,
        })

    return stores, cred, stores_response


@csrf_exempt
@agent_token_required
def agent_auth_start(request):
    """
    POST /agent/auth/start/
    body:
    {
      "phone": "+7924..."
    }

    Успех:
    200
    {
      "session_id": "..."
    }

    Ошибка:
    404 -> сервис недоступен
    500 -> { "error": "..." }

    Оптимизация:
    - здесь НЕ тянем Oracle и НЕ собираем полные данные магазинов
    - только проверяем, что у пользователя есть хотя бы один магазин
    """
    if request.method != 'POST':
        return _agent_error('Сервис недоступен', status=404)

    try:
        data = json.loads(request.body.decode('utf-8') if request.body else "{}")
    except Exception as e:
        logger.error(f"[AGENT_AUTH/START] JSON parse error: {e}")
        return _agent_error('Некорректный JSON', status=500)

    phone_raw = str(data.get('phone') or "").strip()
    if not phone_raw:
        return _agent_error('Не указан phone', status=500)

    try:
        phone_norm = normalize_phone_ru(phone_raw)
        phone_candidates = {phone_raw}
        if phone_norm:
            phone_candidates.add(phone_norm)

        users = list(
            User.objects
            .filter(phone__in=list(phone_candidates))
            .only('id', 'full_name', 'tg_id', 'max_id', 'mail')
            .order_by('id')[:2]
        )

        if not users:
            return _agent_error('Пользователь с таким номером не найден', status=500)

        if len(users) > 1:
            logger.error(
                f"[AGENT_AUTH/START] Несколько пользователей для phone={phone_candidates}: "
                f"ids={[u.id for u in users]}"
            )
            return _agent_error(
                'Найдено несколько пользователей с таким номером, обратитесь к администратору',
                status=500
            )

        user = users[0]

        if not user.max_id and not str(getattr(user, "mail", "") or "").strip():
            return _agent_error(
                'Для пользователя не указан ни max_id, ни email (mail), отправка PIN невозможна',
                status=500
            )

        # ВАЖНО: вместо _get_agent_user_stores(user) — только быстрая проверка наличия магазина
        has_store = UKMUser.objects.filter(user=user).exists()
        if not has_store:
            return _agent_error('У пользователя нет магазинов в ukm_users', status=500)

        cred = _get_agent_primary_credentials(user.id)
        if not cred:
            logger.error(
                f"[AGENT_AUTH/START] Нет активных записей в open_in_system для user_id={user.id}, system_id=4"
            )
            return _agent_error(
                'Для пользователя не найдены активные учётные данные (open_in_system, system_id=4)',
                status=500
            )

        session_uuid = uuid.uuid4()
        session_id = str(session_uuid)

        pin = _generate_pin_code()
        expires_at = timezone.now() + datetime.timedelta(minutes=PIN_TTL_MINUTES)
        pin_hash = hashlib.sha256(pin.encode('utf-8')).hexdigest()

        AuthSession.objects.create(
            session_id=session_uuid,
            user=user,
            storeid=None,
            pin_hash=pin_hash,
            status='pin_sent',
            attempts=0,
            expires_at=expires_at,
        )

        delivery = send_pin_to_user_channels(user, pin)

        if not delivery["ok"]:
            logger.error(
                f"[AGENT_AUTH/START] Не удалось отправить PIN user_id={user.id} "
                f"(telegram_sent={delivery['telegram_sent']}, max_sent={delivery['max_sent']})"
            )
            return _agent_error('Не удалось отправить PIN ни в MAX, ни на email', status=500)

        _send_admin_log_async(
            "\n".join([
                "🔐 Агент. Запрос",
                f"{user.full_name}",
                f"Телефон: {_mask_phone(phone_raw)}",
                f"Email: {_mask_email(getattr(user, 'mail', '')) if getattr(user, 'mail', None) else 'нет'}",
                f"Email: {'да' if delivery.get('email_sent') else 'нет'}",
                f"MAX: {'да' if delivery.get('max_sent') else 'нет'}",
            ])
        )

        return JsonResponse({
            'session_id': session_id,
        }, status=200)

    except Exception as e:
        logger.exception(f"[AGENT_AUTH/START] Unexpected error: {e}")
        return _agent_error(f'Ошибка при запуске авторизации: {e}', status=500)


@csrf_exempt
@agent_token_required
def agent_auth_verify_pin(request):
    """
    POST /agent/auth/verify_pin/
    body:
    {
      "session_id": "...",
      "pin": "1234"
    }

    Успех:
    200
    {
      "user": {
        "id": ...,
        "fio": "..."
      },
      "stores": [...]
    }

    Ошибка:
    404 -> сервис недоступен
    500 -> { "error": "..." }
    """
    if request.method != 'POST':
        return _agent_error('Сервис недоступен', status=404)

    try:
        data = json.loads(request.body.decode('utf-8') if request.body else "{}")
    except Exception as e:
        logger.error(f"[AGENT_AUTH/VERIFY_PIN] JSON parse error: {e}")
        return _agent_error('Некорректный JSON', status=500)

    session_id = str(data.get('session_id') or "").strip()
    pin_input = str(data.get('pin') or "").strip()

    if not session_id or not pin_input:
        return _agent_error('Нужны session_id и pin', status=500)

    try:
        sess = (
            AuthSession.objects
            .select_related('user')
            .filter(session_id=session_id)
            .first()
        )
        if not sess:
            return _agent_error('Сессия не найдена или уже истекла', status=500)

        now = timezone.now()

        if now > sess.expires_at:
            sess.status = 'expired'
            sess.save(update_fields=['status', 'updated_at'])
            return _agent_error('PIN истёк, начните авторизацию заново', status=500)

        if sess.status != 'pin_sent':
            return _agent_error(
                f'Неверное состояние сессии: {sess.status}, начните заново',
                status=500
            )

        if sess.attempts >= MAX_PIN_ATTEMPTS:
            sess.status = 'blocked'
            sess.save(update_fields=['status', 'updated_at'])
            return _agent_error('Превышено количество попыток, начните заново', status=500)

        pin_hash_input = hashlib.sha256(pin_input.encode('utf-8')).hexdigest()
        if pin_hash_input != (sess.pin_hash or ""):
            sess.attempts = sess.attempts + 1
            if sess.attempts >= MAX_PIN_ATTEMPTS:
                sess.status = 'blocked'
            sess.save(update_fields=['attempts', 'status', 'updated_at'])

            attempts_left = max(0, MAX_PIN_ATTEMPTS - sess.attempts)
            if attempts_left == 0:
                return _agent_error('Неверный PIN, попытки закончились', status=500)

            return _agent_error(f'Неверный PIN. Осталось попыток: {attempts_left}', status=500)

        sess.status = 'success'
        sess.save(update_fields=['status', 'updated_at'])

        stores_payload, cred, stores_response = _build_agent_stores_response(sess.user)

        if not stores_payload:
            logger.error(f"[AGENT_AUTH/VERIFY_PIN] У пользователя нет магазинов user_id={sess.user.id}")
            return _agent_error('У пользователя нет магазинов в ukm_users', status=500)

        if not cred:
            logger.error(
                f"[AGENT_AUTH/VERIFY_PIN] Нет записей в open_in_system для user_id={sess.user.id}, system_id=4"
            )
            return _agent_error(
                'Для пользователя не найдены учётные данные (open_in_system, system_id=4)',
                status=500
            )

        _send_admin_log_async(
            "\n".join([
                "✅ Агент. Успешный вход",
                f"{sess.user.full_name}",
                f"Магазинов: {len(stores_response)}",
            ])
        )

        response = {
            'user': {
                'id': sess.user.id,
                'fio': sess.user.full_name,
            },
            'stores': stores_response,
        }

        return JsonResponse(response, json_dumps_params={'ensure_ascii': False}, status=200)

    except Exception as e:
        logger.exception(f"[AGENT_AUTH/VERIFY_PIN] Unexpected error: {e}")
        return _agent_error(f'Ошибка при проверке PIN: {e}', status=500)




def _get_existing_open_password(
    *,
    user_id: int,
    system_id: int = 9
) -> tuple[Optional[str], Optional[str], Optional[int]]:
    """
    READ-ONLY.
    Возвращает (password, username, open_in_system_id) из open_in_system
    для user_id + system_id.

    • status НЕ проверяем.
    • Если записей несколько — берём последнюю по id и пишем warning.
    """
    rows = list(
        OpenInSystem.objects
        .filter(user_id=user_id, system_id=system_id)
        .order_by("-id")[:2] 
    )
    if not rows:
        return None, None, None

    if len(rows) > 1:
        logger.warning(
            f"[OPEN_IN_SYSTEM] Несколько записей для user_id={user_id}, system_id={system_id}. "
            f"Возвращаю последнюю по id."
        )

    row = rows[0]
    password = (row.password or "").strip() or None
    username = (row.username or "").strip() or None
    return password, username, row.id




def _extract_ids(data: dict) -> tuple[str, str, str]:
    tg_id = str(data.get("tg_id") or "").strip()
    max_id = str(data.get("max_id") or "").strip()
    channel = "tg" if tg_id else ("max" if max_id else "")
    return tg_id, max_id, channel


def _resolve_user_by_ids(*, tg_id: str, max_id: str) -> tuple[Optional[User], str]:
    """
    Возвращает (user, err_msg).
    Принимает tg_id и/или max_id.
    Если оба заданы — они обязаны указывать на одного и того же user.
    """
    tg_id = str(tg_id or "").strip()
    max_id = str(max_id or "").strip()

    if not tg_id and not max_id:
        return None, "Не указан tg_id или max_id"

    user_by_tg = None
    user_by_max = None

    if tg_id:
        user_by_tg = User.objects.filter(tg_id=tg_id).first()
        if not user_by_tg:
            return None, "Пользователь не найден по tg_id"

    if max_id:
        user_by_max = User.objects.filter(max_id=max_id).first()
        if not user_by_max:
            return None, "Пользователь не найден по max_id"

    if user_by_tg and user_by_max and user_by_tg.id != user_by_max.id:
        return None, "tg_id и max_id принадлежат разным пользователям"

    return (user_by_tg or user_by_max), ""



def _session_owner_ok(req: AdminBadgeRequest, *, channel: str, tg_id: str, max_id: str) -> bool:
    meta = req.meta or {}
    if channel == "tg":
        return str(req.cashier_tg_id or "") == str(tg_id)
    if channel == "max":
        return str(meta.get("cashier_max_id") or "") == str(max_id)
    return False














@csrf_exempt
def get_qr_code_by_tg(request):
    """
    READ-ONLY.

    POST {"tg_id": "..."} или {"max_id":"..."} или оба -> возвращает существующий open_in_system.password (system_id=9)

    Делает:
      - находит пользователя по tg_id/max_id
      - если пришли оба — проверяет, что это один и тот же пользователь
      - валидирует employee_id как ИНН
      - проверяет ukm_users
      - читает существующий open_in_system.password (system_id=9) и возвращает
      - пишет QRIssueLog по каждой связке store/role
      - шлёт админ-лог в Telegram (в фоне)

    НЕ делает:
      - НЕ обновляет QRCode / OpenInSystem / UKMUser
      - НЕ пишет в MySQL import4.users/signal
      - НЕ обновляет XML UKM5
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Только POST"}, status=405)

    raw_body = ""
    try:
        raw_body = request.body.decode("utf-8") if request.body else "{}"
        try:
            data = json.loads(raw_body)
        except Exception as e:
            logger.error(f"[QR/RO] JSON parse error: {e}; body={raw_body!r}")
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR\n"
                f"Этап: Парсинг JSON\nПричина: {e}\n\n"
                f"Сырой запрос:\n{raw_body[:1000]}{'…' if len(raw_body) > 1000 else ''}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_tg",
                method="BY_ID",
                status="error",
                user=None,
                tg_id="",
                employee_inn="",
                employee_fio="",
                phone_raw="",
                phone_normalized="",
                qr_data="",
                error_message=f"Парсинг JSON: {e}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": "Некорректный JSON"}, status=400)

        tg_id = str(data.get("tg_id") or "").strip()
        max_id = str(data.get("max_id") or "").strip()

        if not tg_id and not max_id:
            msg = "Не указан tg_id или max_id"
            _send_telegram_log_async(
                f"❌ Ошибка (READ-ONLY) при выдаче QR\nЭтап: Валидация\nПричина: {msg}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_tg",
                method="BY_ID",
                status="error",
                user=None,
                tg_id="",
                employee_inn="",
                employee_fio="",
                phone_raw="",
                phone_normalized="",
                qr_data="",
                error_message=f"Валидация: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": msg}, status=400)

        user, err = _resolve_user_by_ids(tg_id=tg_id, max_id=max_id)
        if not user:
            msg = f"Пользователь не найден: {err} (tg_id={tg_id!r}, max_id={max_id!r})"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR\n"
                f"Этап: Поиск пользователя\nПричина: {msg}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_tg",
                method="BY_ID",
                status="error",
                user=None,
                tg_id=tg_id,
                employee_inn="",
                employee_fio="",
                phone_raw="",
                phone_normalized="",
                qr_data="",
                error_message=f"Поиск пользователя: {msg}",
                raw_request={"raw_body": raw_body, "tg_id": tg_id, "max_id": max_id} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": "Пользователь не найден"}, status=404)

        tg_id_user = str(getattr(user, "tg_id", "") or "").strip()
        fio = " ".join((user.full_name or "").split()).strip()
        employee_id_raw = (user.employee_id or "").strip()

        try:
            plain_inn = ensure_plain_inn(employee_id_raw)
        except Exception as e:
            msg = f"Некорректный employee_id (ИНН) у user_id={user.id}: {e}"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR\n"
                f"Этап: Валидация ИНН\nПричина: {msg}\n\n"
                f"tg_id={tg_id_user or '—'}\nmax_id={getattr(user,'max_id','') or '—'}\n"
                f"user_id={user.id}\nФИО={fio or '—'}\nemployee_id={employee_id_raw!r}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_tg",
                method="BY_ID",
                status="error",
                user=user,
                tg_id=tg_id_user,
                employee_inn=employee_id_raw,
                employee_fio=fio,
                phone_raw=user.phone or "",
                phone_normalized=normalize_phone_ru(user.phone or "") or "",
                qr_data="",
                error_message=f"Валидация ИНН: {msg}",
                raw_request={"raw_body": raw_body, "tg_id": tg_id, "max_id": max_id} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": f"Некорректный employee_id: {e}"}, status=400)

        ukm_links = list(UKMUser.objects.filter(user_id=user.id).values("storeid", "roleid"))
        if not ukm_links:
            msg = f"Нет записей ukm_users для user_id={user.id}"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR\n"
                f"Этап: Проверка доступов\nПричина: {msg}\n\n"
                f"user_id={user.id}\nФИО={fio or '—'}\nИНН={plain_inn}\n"
                f"tg_id={tg_id_user or '—'}\nmax_id={getattr(user,'max_id','') or '—'}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_tg",
                method="BY_ID",
                status="error",
                user=user,
                tg_id=tg_id_user,
                employee_inn=plain_inn,
                employee_fio=fio,
                phone_raw=user.phone or "",
                phone_normalized=normalize_phone_ru(user.phone or "") or "",
                qr_data="",
                error_message=f"Проверка доступов: {msg}",
                raw_request={"raw_body": raw_body, "tg_id": tg_id, "max_id": max_id} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": "Для пользователя нет записей в ukm_users"}, status=404)

        password, open_username, open_row_id = _get_existing_open_password(user_id=user.id, system_id=9)
        if not password:
            msg = f"Нет password в open_in_system (system_id=9) для user_id={user.id}"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR\n"
                f"Этап: Чтение open_in_system\nПричина: {msg}\n\n"
                f"user_id={user.id}\nФИО={fio or '—'}\nИНН={plain_inn}\n"
                f"tg_id={tg_id_user or '—'}\nmax_id={getattr(user,'max_id','') or '—'}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_tg",
                method="BY_ID",
                status="error",
                user=user,
                tg_id=tg_id_user,
                employee_inn=plain_inn,
                employee_fio=fio,
                phone_raw=user.phone or "",
                phone_normalized=normalize_phone_ru(user.phone or "") or "",
                qr_data="",
                error_message=f"open_in_system: {msg}",
                raw_request={"raw_body": raw_body, "tg_id": tg_id, "max_id": max_id} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": "У пользователя нет сохранённого QR (open_in_system)"}, status=404)

        ukm_store_ids_str = [str(int(x["storeid"])) for x in ukm_links if str(x.get("storeid", "")).isdigit()]
        store_map = {
            str(s.ukm4store).strip(): s
            for s in Store.objects.filter(ukm4store__in=ukm_store_ids_str)
        }

        per_store_results = []
        for link in ukm_links:
            sid = int(link["storeid"])
            role_id_db = int(link["roleid"]) if str(link.get("roleid", "")).isdigit() else None
            s_obj = store_map.get(str(sid))
            per_store_results.append({
                "ukm_storeid": sid,
                "smstore": getattr(s_obj, "smstore", None),
                "store_name": getattr(s_obj, "name", "") if s_obj else "",
                "roleid": role_id_db,
                "cashier_id": None,
                "ukm_host": None,
                "found_in_trm": None,
            })

        lines = [
            "✅ QR-код запросили из бота",
            "",
            "👤 Сотрудник:",
            f"  • Telegram ID: {tg_id_user or '—'}",
            f"  • MAX ID: {getattr(user,'max_id','') or '—'}",
            f"  • user_id (PostgreSQL): {user.id}",
            f"  • ФИО: {fio or '—'}",
            f"  • ИНН: {plain_inn}",
            "",
            "🏬 Магазины (ukm_users):",
        ]
        for r in per_store_results:
            lines.append(f"  • ukm_storeid={r['ukm_storeid']}, smstore={r['smstore']}, roleId={r['roleid']}")
        lines += [
            "",
            "🔐 QR (из open_in_system.password):",
            password,
            f"open_in_system.id={open_row_id}, username={open_username or '—'}",
        ]
        _send_telegram_log_async("\n".join(lines))

        phone_norm = normalize_phone_ru(user.phone or "") or ""
        try:
            raw_req = json.loads(raw_body) if raw_body else None
        except Exception:
            raw_req = {"raw_body": raw_body}

        if isinstance(raw_req, dict):
            raw_req.setdefault("tg_id", tg_id)
            raw_req.setdefault("max_id", max_id)
            raw_req.setdefault("tg_id_user", tg_id_user)
            raw_req.setdefault("max_id_user", getattr(user, "max_id", None))

        for r in per_store_results:
            log_qr_issue(
                endpoint="get_qr_code_by_tg",
                method="BY_ID",
                status="ok",
                user=user,
                employee_inn=plain_inn,
                employee_fio=fio,
                tg_id=tg_id_user,
                phone_raw=user.phone or "",
                phone_normalized=phone_norm,
                sm_store_id=r["smstore"],
                ukm_store_id=r["ukm_storeid"],
                role_id=r["roleid"],
                qr_data=password,
                error_message="",
                raw_request=raw_req,
            )

        return JsonResponse({
            "status": "ok",
            "qr_data": password,
            "user_id": user.id,
            "stores": per_store_results,
            "tg_id": tg_id_user,
            "max_id": getattr(user, "max_id", None),
        })

    except Exception as e:
        logger.exception("[QR/RO] Unexpected error")
        _send_telegram_log_async(
            "💥 Критическая ошибка (READ-ONLY) при выдаче QR\n"
            f"{e}\n\nСырой запрос:\n{raw_body[:1000]}{'…' if len(raw_body) > 1000 else ''}"
        )
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

@csrf_exempt
def get_qr_code_by_employee_id(request):
    """
    READ-ONLY.

    POST:
    {
      "inn": "7536207278",
      "fio": "Иванов Иван Иванович",
      "storeId": 514,
      "roleId": 1,
      "phone": "8 (924) 000-00-00"
    }

    Админ-логи в Telegram отправляются в фоне.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Только POST"}, status=405)

    raw_body = ""
    try:
        raw_body = request.body.decode("utf-8") if request.body else "{}"
        try:
            data = json.loads(raw_body)
        except Exception as e:
            logger.error(f"[QR/EMP/RO] JSON parse error: {e}; body={raw_body!r}")
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Парсинг JSON\nПричина: {e}\n\n"
                f"Сырой запрос:\n{raw_body[:1000]}{'…' if len(raw_body) > 1000 else ''}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=None,
                employee_inn="",
                employee_fio="",
                tg_id="",
                phone_raw="",
                phone_normalized="",
                sm_store_id=None,
                ukm_store_id=None,
                role_id=None,
                qr_data="",
                error_message=f"Парсинг JSON: {e}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": "Некорректный JSON"}, status=400)

        inn_raw = str(data.get("inn") or data.get("employee_id") or "").strip()
        fio_raw = str(data.get("fio") or "").strip()
        store_raw = str(data.get("storeId") or data.get("storeid") or "").strip()
        role_raw = str(data.get("roleId") or data.get("roleid") or "").strip()
        phone_raw = str(data.get("phone") or "").strip()

        if not inn_raw:
            msg = "Не указан inn"
            _send_telegram_log_async(
                f"❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\nЭтап: Валидация\nПричина: {msg}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=None,
                employee_inn="",
                employee_fio=fio_raw,
                tg_id="",
                phone_raw=phone_raw,
                phone_normalized=normalize_phone_ru(phone_raw) or "",
                sm_store_id=int(store_raw) if store_raw.isdigit() else None,
                ukm_store_id=None,
                role_id=int(role_raw) if role_raw.isdigit() else None,
                qr_data="",
                error_message=f"Валидация: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": msg}, status=400)

        try:
            plain_inn = ensure_plain_inn(inn_raw)
        except Exception as e:
            msg = f"Некорректный ИНН: {e}"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Валидация ИНН\nПричина: {msg}\ninn={inn_raw!r}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=None,
                employee_inn=inn_raw,
                employee_fio=fio_raw,
                tg_id="",
                phone_raw=phone_raw,
                phone_normalized=normalize_phone_ru(phone_raw) or "",
                sm_store_id=int(store_raw) if store_raw.isdigit() else None,
                ukm_store_id=None,
                role_id=int(role_raw) if role_raw.isdigit() else None,
                qr_data="",
                error_message=f"Валидация ИНН: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": msg}, status=400)

        if not fio_raw:
            msg = "Не указано fio"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Валидация\nПричина: {msg}\ninn={plain_inn}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=None,
                employee_inn=plain_inn,
                employee_fio="",
                tg_id="",
                phone_raw=phone_raw,
                phone_normalized=normalize_phone_ru(phone_raw) or "",
                sm_store_id=int(store_raw) if store_raw.isdigit() else None,
                ukm_store_id=None,
                role_id=int(role_raw) if role_raw.isdigit() else None,
                qr_data="",
                error_message=f"Валидация: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": msg}, status=400)
        fio = " ".join(fio_raw.split()).strip()

        if not store_raw or not store_raw.isdigit():
            msg = "Некорректный storeId (smstore)"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Валидация\nПричина: {msg}\nstoreId={store_raw!r}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=None,
                employee_inn=plain_inn,
                employee_fio=fio,
                tg_id="",
                phone_raw=phone_raw,
                phone_normalized=normalize_phone_ru(phone_raw) or "",
                sm_store_id=None,
                ukm_store_id=None,
                role_id=int(role_raw) if role_raw.isdigit() else None,
                qr_data="",
                error_message=f"Валидация: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": msg}, status=400)
        sm_store_id = int(store_raw)

        if not role_raw or not role_raw.isdigit():
            msg = "Некорректный roleId"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Валидация\nПричина: {msg}\nroleId={role_raw!r}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=None,
                employee_inn=plain_inn,
                employee_fio=fio,
                tg_id="",
                phone_raw=phone_raw,
                phone_normalized=normalize_phone_ru(phone_raw) or "",
                sm_store_id=sm_store_id,
                ukm_store_id=None,
                role_id=None,
                qr_data="",
                error_message=f"Валидация: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": msg}, status=400)
        role_id_req = int(role_raw)

        phone_norm = normalize_phone_ru(phone_raw) or ""

        store_obj = Store.objects.filter(smstore=sm_store_id).first()
        if not store_obj or store_obj.ukm4store is None:
            msg = "Магазин не найден в stores или не указан ukm4store"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Маппинг smstore→ukm4store\nПричина: {msg}\nsmstore={sm_store_id}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=None,
                employee_inn=plain_inn,
                employee_fio=fio,
                tg_id="",
                phone_raw=phone_raw,
                phone_normalized=phone_norm,
                sm_store_id=sm_store_id,
                ukm_store_id=None,
                role_id=role_id_req,
                qr_data="",
                error_message=f"Маппинг smstore→ukm4store: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": msg}, status=400)

        try:
            ukm_store_id_req = int(str(store_obj.ukm4store).strip())
        except Exception:
            msg = "Некорректное значение ukm4store"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Маппинг smstore→ukm4store\nПричина: {msg}\nsmstore={sm_store_id}"
            )
            return JsonResponse({"status": "error", "message": msg}, status=400)

        inn_sha = hashlib.sha256(plain_inn.encode("utf-8")).hexdigest()
        inn_sha20 = inn_sha[:20]
        user = (
            User.objects.filter(employee_id=plain_inn).first()
            or User.objects.filter(encrypted_inn=plain_inn).first()
            or User.objects.filter(employee_id=inn_sha).first()
            or User.objects.filter(encrypted_inn=inn_sha).first()
            or User.objects.filter(employee_id=inn_sha20).first()
            or User.objects.filter(encrypted_inn=inn_sha20).first()
        )
        if not user:
            msg = f"Пользователь с INN={plain_inn} не найден в PostgreSQL"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Поиск пользователя\nПричина: {msg}\n"
                f"inn={plain_inn}\nfio={fio}\nsmstore={sm_store_id} (ukm4store={ukm_store_id_req})"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=None,
                employee_inn=plain_inn,
                employee_fio=fio,
                tg_id="",
                phone_raw=phone_raw,
                phone_normalized=phone_norm,
                sm_store_id=sm_store_id,
                ukm_store_id=ukm_store_id_req,
                role_id=role_id_req,
                qr_data="",
                error_message=f"Поиск пользователя: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": "Пользователь не найден"}, status=404)

        ukm_links = list(UKMUser.objects.filter(user_id=user.id).values("storeid", "roleid"))
        if not ukm_links:
            msg = f"Нет записей ukm_users для user_id={user.id}"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Проверка доступов\nПричина: {msg}\n"
                f"user_id={user.id}\nФИО={fio}\nИНН={plain_inn}\nsmstore={sm_store_id} (ukm4store={ukm_store_id_req})"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=user,
                employee_inn=plain_inn,
                employee_fio=fio,
                tg_id=str(getattr(user, "tg_id", "") or ""),
                phone_raw=phone_raw,
                phone_normalized=phone_norm,
                sm_store_id=sm_store_id,
                ukm_store_id=ukm_store_id_req,
                role_id=role_id_req,
                qr_data="",
                error_message=f"Проверка доступов: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": "Для пользователя нет записей в ukm_users"}, status=404)

        has_requested_store = any(int(x["storeid"]) == int(ukm_store_id_req) for x in ukm_links if str(x.get("storeid", "")).isdigit())
        if not has_requested_store:
            msg = f"У пользователя нет доступа к ukm4store={ukm_store_id_req} (запрошен smstore={sm_store_id})"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Проверка доступов\nПричина: {msg}\n"
                f"user_id={user.id}\nФИО={fio}\nИНН={plain_inn}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=user,
                employee_inn=plain_inn,
                employee_fio=fio,
                tg_id=str(getattr(user, "tg_id", "") or ""),
                phone_raw=phone_raw,
                phone_normalized=phone_norm,
                sm_store_id=sm_store_id,
                ukm_store_id=ukm_store_id_req,
                role_id=role_id_req,
                qr_data="",
                error_message=f"Проверка доступов: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": msg}, status=403)

        password, open_username, open_row_id = _get_existing_open_password(user_id=user.id, system_id=9)
        if not password:
            msg = f"Нет password в open_in_system (system_id=9) для user_id={user.id}"
            _send_telegram_log_async(
                "❌ Ошибка (READ-ONLY) при выдаче QR по ИНН\n"
                f"Этап: Чтение open_in_system\nПричина: {msg}\n\n"
                f"user_id={user.id}\nФИО={fio}\nИНН={plain_inn}"
            )
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="error",
                user=user,
                employee_inn=plain_inn,
                employee_fio=fio,
                tg_id=str(getattr(user, "tg_id", "") or ""),
                phone_raw=phone_raw,
                phone_normalized=phone_norm,
                sm_store_id=sm_store_id,
                ukm_store_id=ukm_store_id_req,
                role_id=role_id_req,
                qr_data="",
                error_message=f"open_in_system: {msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
            )
            return JsonResponse({"status": "error", "message": "У пользователя нет сохранённого QR (open_in_system)"}, status=404)

        ukm_store_ids_str = [str(int(x["storeid"])) for x in ukm_links if str(x.get("storeid", "")).isdigit()]
        store_map = {
            str(s.ukm4store).strip(): s
            for s in Store.objects.filter(ukm4store__in=ukm_store_ids_str)
        }

        per_store_results = []
        role_mismatch_notes = []
        for l in ukm_links:
            sid = int(l["storeid"])
            role_id_db = int(l["roleid"])
            s_obj = store_map.get(str(sid))

            if sid == int(ukm_store_id_req) and role_id_db != int(role_id_req):
                role_mismatch_notes.append(f"requested_role={role_id_req} != ukm_users.roleid={role_id_db} (ukm_storeid={sid})")

            per_store_results.append({
                "ukm_storeid": sid,
                "smstore": getattr(s_obj, "smstore", None),
                "store_name": getattr(s_obj, "name", "") if s_obj else "",
                "roleid": role_id_db,
                "cashier_id": None,
                "ukm_host": None,
                "found_in_trm": None,
            })

        fio_db = " ".join(((user.full_name or "").split())) if (user.full_name or "").strip() else ""
        fio_match = (fio_db == fio) if fio_db else None

        lines = [
            "✅ QR-код запрошен из 1С",
            "",
            "👤 Сотрудник:",
            f"  • user_id (PostgreSQL): {user.id}",
            f"  • ФИО (запрос): {fio}",
            f"  • ФИО (users.full_name): {fio_db or '—'}",
            f"  • fio_match: {fio_match if fio_match is not None else '—'}",
            f"  • ИНН: {plain_inn}",
            f"  • phone_raw: {phone_raw or '—'}",
            f"  • phone_norm: {phone_norm or '—'}",
            "",
            f"🏬 Запрошенный магазин: smstore={sm_store_id} → ukm4store={ukm_store_id_req}, roleId={role_id_req}",
        ]
        if role_mismatch_notes:
            lines += ["", "⚠️ Несовпадение roleId:", *[f"  • {x}" for x in role_mismatch_notes]]

        lines += [
            "",
            "📌 Магазины пользователя (ukm_users):",
        ]
        for r in per_store_results:
            lines.append(f"  • ukm_storeid={r['ukm_storeid']}, smstore={r['smstore']}, roleId={r['roleid']}")

        lines += [
            "",
            "🔐 QR (из open_in_system.password):",
            password,
            f"open_in_system.id={open_row_id}, username={open_username or '—'}",
        ]
        _send_telegram_log_async("\n".join(lines))

        try:
            raw_req = json.loads(raw_body) if raw_body else None
        except Exception:
            raw_req = {"raw_body": raw_body}

        for r in per_store_results:
            log_qr_issue(
                endpoint="get_qr_code_by_employee_id",
                method="BY_INN",
                status="ok",
                user=user,
                employee_inn=plain_inn,
                employee_fio=fio,
                tg_id=str(getattr(user, "tg_id", "") or ""),
                phone_raw=phone_raw,
                phone_normalized=phone_norm,
                sm_store_id=r["smstore"],
                ukm_store_id=r["ukm_storeid"],
                role_id=r["roleid"],
                qr_data=password,
                error_message="",
                raw_request=raw_req,
            )

        return JsonResponse({
            "status": "ok",
            "qr_data": password,
            "user_id": user.id,
            "requested": {
                "smstore": sm_store_id,
                "ukm4store": ukm_store_id_req,
                "roleid": role_id_req,
            },
            "stores": per_store_results,
        })

    except Exception as e:
        logger.exception("[QR/EMP/RO] Unexpected error")
        _send_telegram_log_async(
            "💥 Критическая ошибка (READ-ONLY) при выдаче QR по ИНН\n"
            f"{e}\n\nСырой запрос:\n{raw_body[:1000]}{'…' if len(raw_body) > 1000 else ''}"
        )
        return JsonResponse({"status": "error", "message": str(e)}, status=500)





@csrf_exempt
def update_cashier(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Только POST"}, status=405)

    raw_body = ""
    try:
        raw_body = request.body.decode("utf-8") if request.body else "{}"
        data = json.loads(raw_body)

        # входные поля
        inn_raw = str(data.get("inn") or data.get("employee_id") or "").strip()
        fio_raw = str(data.get("fio") or data.get("full_name") or "").strip()
        storeids_raw = data.get("storeid") or data.get("storeids") or data.get("stores")

        role_raw = data.get("roleId") or data.get("roleid") or 1
        try:
            role_id_req = int(role_raw)
        except Exception:
            role_id_req = 1

        plain_inn = ensure_plain_inn(inn_raw)
        fio = " ".join(fio_raw.split()).strip()

        if not fio:
            return JsonResponse({"status": "error", "message": "fio обязателен"}, status=400)
        if not storeids_raw:
            return JsonResponse({"status": "error", "message": "storeid обязателен"}, status=400)

        store_ids = []
        if isinstance(storeids_raw, list):
            for x in storeids_raw:
                s = str(x).strip()
                if s.isdigit():
                    store_ids.append(int(s))
        else:
            store_ids = [int(s.strip()) for s in str(storeids_raw).split(",") if s.strip().isdigit()]

        store_ids = sorted(set(store_ids))
        if not store_ids:
            return JsonResponse({"status": "error", "message": "Некорректный storeid"}, status=400)

        # поиск пользователя (plain/sha/sha20) + нормализованное ФИО
        inn_sha = hashlib.sha256(plain_inn.encode("utf-8")).hexdigest()
        inn_sha20 = inn_sha[:20]

        fio_norm = fio
        user = (
            User.objects.filter(full_name__iexact=fio_norm).filter(
                Q(employee_id=plain_inn) | Q(encrypted_inn=plain_inn) |
                Q(employee_id=inn_sha)  | Q(encrypted_inn=inn_sha)  |
                Q(employee_id=inn_sha20)| Q(encrypted_inn=inn_sha20)
            ).order_by("id").first()
        )

        # fallback: если ФИО в базе чуть отличается — найдём по ИНН, но проверим руками
        if not user:
            cand = (
                User.objects.filter(
                    Q(employee_id=plain_inn) | Q(encrypted_inn=plain_inn) |
                    Q(employee_id=inn_sha)  | Q(encrypted_inn=inn_sha)  |
                    Q(employee_id=inn_sha20)| Q(encrypted_inn=inn_sha20)
                ).order_by("id").first()
            )
            if cand:
                fio_db = " ".join((cand.full_name or "").split()).strip()
                if fio_db.lower() == fio_norm.lower():
                    user = cand

        if not user:
            return JsonResponse({"status": "error", "message": "Пользователь не найден"}, status=404)

        # добавим недостающие магазины в ukm_users (в Postgres) + сразу ротируем пароль
        added_storeids: list[int] = []

        with transaction.atomic():
            # employee_id должен быть plain ИНН
            if (user.employee_id or "").strip() != plain_inn:
                user.employee_id = plain_inn
                user.updated_at = timezone.now()
                user.save(update_fields=["employee_id", "updated_at"])

            existing_links = {
                int(x.storeid): x
                for x in UKMUser.objects.filter(user_id=user.id, storeid__in=store_ids)
            }

            for sid in store_ids:
                sid = int(sid)
                link = existing_links.get(sid)

                if link is None:
                    UKMUser.objects.create(
                        user=user,
                        roleid=role_id_req,
                        storeid=sid,
                        version=1
                    )
                    added_storeids.append(sid)
                else:
                    if int(link.roleid or 0) != int(role_id_req):
                        link.roleid = int(role_id_req)
                        link.version = 1
                        link.save(update_fields=["roleid", "version"])

            # всегда генерим новый пароль/QR и пишем в Postgres
            new_password = build_user_password(plain_inn)
            _set_password_pg(user, new_password)

        # После транзакции перечитываем ВСЕ магазины пользователя
        ukm_links = list(
            UKMUser.objects
            .filter(user_id=user.id)
            .values("storeid", "roleid")
        )
        ukm_links = sorted(
            [{"storeid": int(x["storeid"]), "roleid": int(x["roleid"])} for x in ukm_links],
            key=lambda x: x["storeid"]
        )

        if not ukm_links:
            return JsonResponse({"status": "error", "message": "Нет записей ukm_users у пользователя"}, status=400)

        # allocator новых trm-id: ключим по resolved_host, чтобы не словить дубль,
        # если разные storeid попадают на один и тот же ukmserver
        trm_alloc: dict[str, dict] = {}

        def _is_trm_id_taken(resolved_host: str, candidate: int) -> bool:
            conn2 = cur2 = None
            try:
                conn2 = connect_ukm(host=resolved_host, store_id=None)
                cur2 = conn2.cursor()
                cur2.execute("SELECT 1 AS x FROM trm_in_users WHERE id=%s LIMIT 1", (candidate,))
                return bool(cur2.fetchone())
            finally:
                try:
                    if cur2:
                        cur2.close()
                    if conn2:
                        conn2.close()
                except Exception:
                    pass

        def _alloc_new_trm_id_for_store(store_id: int) -> tuple[int, str]:
            resolved_host, _src = _resolve_ukmserver_host(host=None, store_id=store_id)
            key = str(resolved_host).strip()

            st = trm_alloc.get(key)
            if st is None:
                base = get_next_trm_employee_id(store_id=store_id, host=key)
                st = {"next": int(base), "reserved": set()}
                trm_alloc[key] = st

            candidate = int(st["next"])
            while True:
                if candidate in st["reserved"]:
                    candidate += 1
                    continue
                if candidate > int(TRM_SMALL_MAX):
                    raise RuntimeError(
                        f"TRM id overflow: candidate={candidate} > TRM_SMALL_MAX={TRM_SMALL_MAX} (host={key})"
                    )
                if _is_trm_id_taken(key, candidate):
                    candidate += 1
                    continue

                st["reserved"].add(candidate)
                st["next"] = candidate + 1
                return int(candidate), key

        # обновление UKM4/UKM5 по всем магазинам пользователя
        per_store = []
        errors = 0

        for link in ukm_links:
            sid = int(link["storeid"])
            role_id = int(link["roleid"])

            cashier_id = None
            found_in_trm = False
            resolved_host = None

            try:
                existing_id = get_trm_employee_id(
                    plain_inn,
                    fio_norm,
                    store_id=sid,
                    host=None
                )
                if existing_id is not None:
                    cashier_id = int(existing_id)
                    found_in_trm = True
                    resolved_host, _ = _resolve_ukmserver_host(host=None, store_id=sid)
                else:
                    cashier_id, resolved_host = _alloc_new_trm_id_for_store(sid)
                    found_in_trm = False

                _update_store_mysql_and_xml_for_single_store(
                    store_id=sid,
                    cashier_id=int(cashier_id),
                    role_id=int(role_id),
                    plain_inn=plain_inn,
                    fio=fio_norm,
                    password_plain=new_password,
                )

                _write_converter_user_and_signal(
                    cashier_id=int(cashier_id),
                    plain_inn=plain_inn,
                    fio=fio_norm,
                    password_plain=new_password,
                    store_id=sid,
                    role_id=int(role_id),
                )

                per_store.append({
                    "storeid": sid,
                    "roleid": role_id,
                    "cashier_id": int(cashier_id),
                    "found_in_trm": found_in_trm,
                    "ukm_host": str(resolved_host or ""),
                    "status": "ok",
                    "error": "",
                })

            except Exception as e:
                errors += 1
                logger.error(f"[UPDATE_CASHIER] storeid={sid} error: {e}", exc_info=True)
                per_store.append({
                    "storeid": sid,
                    "roleid": role_id,
                    "cashier_id": int(cashier_id) if cashier_id is not None else None,
                    "found_in_trm": found_in_trm,
                    "ukm_host": str(resolved_host or ""),
                    "status": "error",
                    "error": str(e),
                })

        try:
            masked = new_password[:6] + "..." + new_password[-4:]
            lines = [
                "🔄 update_cashier: ротация QR/пароля + синхронизация магазинов",
                "",
                f"👤 user_id={user.id}",
                f"ФИО='{fio_norm}'",
                f"ИНН={plain_inn}",
                f"Роль из запроса (role_id_req)={role_id_req}",
                f"Новый пароль (masked)={masked} (len={len(new_password)})",
                "",
                f"➕ Добавлены магазины: {', '.join(map(str, added_storeids)) if added_storeids else '—'}",
                f"🏬 Всего магазинов у пользователя (ukm_users): {len(ukm_links)}",
                f"⚠️ Ошибок по магазинам: {errors}",
                "",
                "📋 Детализация:",
            ]
            for r in per_store:
                lines.append(
                    f"• storeid={r['storeid']} role={r['roleid']} "
                    f"cashier_id={r['cashier_id'] or '—'} found_in_trm={r['found_in_trm']} "
                    f"host={r['ukm_host'] or '—'} status={r['status']}"
                )
                if r.get("error"):
                    lines.append(f"    error: {r['error']}")

            _send_admin_log_async("\n".join(lines))
        except Exception as e:
            logger.error(f"[UPDATE_CASHIER] async telegram log enqueue failed: {e}", exc_info=True)

        resp = {
            "status": "ok" if errors == 0 else "partial",
            "message": "Доступ обновлён и пароль ротирован" if errors == 0 else "Частично выполнено: есть ошибки по магазинам",
            "user_id": user.id,
            "inn": plain_inn,
            "fio": fio_norm,
            "added_storeids": added_storeids,
            "rotated": True,
            "password": new_password,
            "stores": per_store,
        }
        return JsonResponse(resp, status=200 if errors == 0 else 207)

    except Exception as exc:
        logger.exception("[UPDATE_CASHIER] error")
        try:
            _send_admin_log_async(
                "💥 Ошибка update_cashier\n"
                f"{exc}\n\n"
                f"raw_body:\n{raw_body[:1500]}{'…' if len(raw_body) > 1500 else ''}"
            )
        except Exception:
            pass
        return JsonResponse({"status": "error", "message": str(exc)}, status=500)


@csrf_exempt
def delete_cashier(request):
    if request.method != 'POST':
        return JsonResponse({"status": "error", "message": "Только POST"}, status=405)

    try:
        data = json.loads(request.body)

        inn_raw = ensure_plain_inn(data.get("inn"))
        fio = data.get("fio")
        store_raw = str(data.get("storeid", "")).strip()

        if not (inn_raw and fio and store_raw):
            return JsonResponse({"status": "error", "message": "inn, fio и storeid обязательны"}, status=400)

        store_ids = [int(s) for s in store_raw.split(',') if s.strip().isdigit()]
        if not store_ids:
            return JsonResponse({"status": "error", "message": "Некорректный storeid"}, status=400)

        user = User.objects.filter(employee_id=inn_raw, full_name=fio).first()
        if not user:
            return JsonResponse({"status": "error", "message": "Пользователь не найден"}, status=404)

        ukm_to_remove = list(UKMUser.objects.filter(user=user, storeid__in=store_ids))
        if not ukm_to_remove:
            return JsonResponse({"status": "ok", "message": "У пользователя уже нет доступа к указанным магазинам", "removed": []})

        current_password = (OpenInSystem.objects
                            .filter(user_id=user.id, system_id=9)
                            .values_list("password", flat=True)
                            .first()) or build_user_password(inn_raw)

        ukm_emp_id = get_trm_employee_id(inn_raw, fio)

        ukm_conn = connect_ukm()
        ukm_cursor = ukm_conn.cursor()
        ukm_cursor.execute("SELECT MAX(id)+1 AS next_id FROM trm_in_users")
        cashier_id_base = ukm_cursor.fetchone()['next_id'] or 1
        ukm_conn.close()

        removed_sids = []
        for idx, ukm_user in enumerate(ukm_to_remove):
            sid = ukm_user.storeid
            cashier_id = ukm_emp_id if ukm_emp_id else (cashier_id_base + idx)

            info = get_store_info(sid)
            ukm4ip = info.get("ukm4ip")

            if ukm4ip:
                try:
                    conv = connect_store_mysql(ukm4ip)
                    cur = conv.cursor()

                    base_version = _calc_next_signal_version(cur)

                    cur.execute("""
                        INSERT INTO users (store,id,name,inn,password,role_id,version,deleted)
                        VALUES (%s,%s,%s,%s,OLD_PASSWORD(%s),%s,%s,1)
                    """, (sid, cashier_id, fio, inn_raw, current_password, ukm_user.roleid, base_version))
                    cur.execute("INSERT INTO `signal`(`signal`,`version`) VALUES ('incr',%s)", (base_version,))
                    conv.commit()
                    conv.close()
                    removed_sids.append(sid)
                    logger.info(f"[MySQL:{ukm4ip}] Закрыт доступ store={sid}, id={cashier_id}, version={base_version}")
                except Exception as exc:
                    logger.error(f"[MySQL:{ukm4ip}] Ошибка для магазина {sid}: {exc}")
            else:
                logger.error(f"[Oracle] Не найден UKM4IP для storeid={sid}. Пропуск записи в MySQL.")

        for ukm_user in ukm_to_remove:
            sid = ukm_user.storeid
            if not is_ukm5_store(sid):
                continue

            # Для магазина 2013 пересборку сделаем отдельным шагом
            if sid == UKM5_FULL_XML_STORE_ID:
                continue

            found = _find_storecashiers_file_for_store(sid)
            if not found:
                continue
            xml_path, number = found

            if not os.path.exists(xml_path):
                continue

            try:
                tree = ET.parse(xml_path)
                root = tree.getroot()

                xml_store_id = resolve_xml_store_id(sid)

                # Гарантируем корректные атрибуты корня
                root.set("fullness", "F")
                root.set("storeId", str(xml_store_id))

                version_el = root.find("version")
                if version_el is None:
                    version_el = ET.SubElement(root, "version")
                version_el.text = "1.0"

                changed = False
                for cash_el in list(root.findall("cashier")):
                    if cash_el.findtext("INN") == inn_raw:
                        root.remove(cash_el)
                        changed = True

                if changed:
                    _write_xml_with_declaration(xml_path, root, ensure_base=True)
                    logger.info(f"[XML] Удалён кассир (INN={inn_raw}) из {xml_path}")
            except Exception as exc:
                logger.error(f"[XML] Ошибка для магазина {sid}: {exc}")

        # удаляем ukm_users в PostgreSQL
        UKMUser.objects.filter(id__in=[u.id for u in ukm_to_remove]).delete()
        if any(u.storeid == UKM5_FULL_XML_STORE_ID for u in ukm_to_remove):
            try:
                xml_path = build_full_ukm5_xml_for_store(UKM5_FULL_XML_STORE_ID)
                logger.info(
                    f"[XML] Полный XML пересобран после удаления доступа для магазина "
                    f"{UKM5_FULL_XML_STORE_ID}: {xml_path}"
                )
            except Exception as exc:
                logger.error(
                    f"[XML] Ошибка пересборки XML для магазина "
                    f"{UKM5_FULL_XML_STORE_ID} после удаления: {exc}"
                )

        return JsonResponse({"status": "ok", "message": "Доступ закрыт", "removed": removed_sids})

    except Exception as exc:
        logger.exception("delete_cashier error")
        return JsonResponse({"status": "error", "message": str(exc)}, status=500)


@csrf_exempt
def employee_identification(request):
    """
    POST /employee-identification/

    ВХОД (JSON, совместим со старой версией):
    {
      "inn": "7536207278",              # или "employee_id"
      "fio": "Иванов Иван Иванович",    # или "FIO"
      "storeId": 514,                   # или "smstore" или старый "mx"
      "datetime": "24.09.2025 8:45:00", # или "Datetime" или "event_datetime"
      // необязательные:
      "phone": "8 (924) 000-00-00",
      "direction": "IN"                 # или "OUT" / "ENTER" / "EXIT" / "event"
      // "roleId" / "roleid" — можно не передавать, оно тут не нужно
    }

    Логика:
      1) Валидация/нормализация входных данных (как в старой версии).
      2) Подготовка payload для 1С в старом формате: INN/FIO/MX/Datetime.
      3) На любой ошибке:
         • красивый лог в Telegram (асинхронно)
         • запись в qr_issue_logs со статусом 'error'.
      4) На успехе:
         • красивый лог в Telegram (асинхронно)
         • запись в qr_issue_logs со статусом 'ok'/'error' в зависимости от ответа 1С.
    """
    if request.method != 'POST':
        return JsonResponse(
            {'status': 'error', 'message': 'Только POST'},
            status=405
        )

    raw_body = request.body.decode('utf-8') if request.body else "{}"
    lat_val: float | None = None
    lon_val: float | None = None

    def _log_and_return_error(
        http_status: int,
        stage: str,
        human_msg: str,
        *,
        inn_raw: str = "",
        fio_raw: str = "",
        smstore_raw: str = "",
        ukm_store_id: int | None = None,
        phone_raw: str = "",
    ) -> JsonResponse:
        """
        Общий обработчик ошибок:
          • шлёт подробный лог в Telegram асинхронно
          • пишет строку в qr_issue_logs со статусом 'error'
          • возвращает JsonResponse с текстом ошибки
        """
        lines = [
            "❌ Ошибка при employee_identification",
            f"🔁 Этап: {stage}",
            f"ℹ️ Причина: {human_msg}",
            "",
            "📨 Контекст запроса:",
            f"  • ИНН (сырое): {inn_raw or '—'}",
            f"  • ФИО (сырое): {fio_raw or '—'}",
            f"  • storeId / mx (сырое): {smstore_raw or '—'}",
            f"  • Телефон (сырое): {phone_raw or '—'}",
            f"  • latitude: {lat_val if lat_val is not None else '—'}",
            f"  • longitude: {lon_val if lon_val is not None else '—'}",
        ]
        if ukm_store_id is not None:
            lines.append(f"  • ukm4store: {ukm_store_id}")
        if raw_body:
            short_body = raw_body if len(raw_body) <= 1000 else raw_body[:1000] + "…"
            lines.extend(["", "📦 Сырой JSON-запрос:", short_body])

        _send_admin_log_async("\n".join(lines))

        try:
            sm_id_int = int(smstore_raw) if str(smstore_raw).isdigit() else None
        except Exception:
            sm_id_int = None

        phone_norm = normalize_phone_ru(phone_raw) if phone_raw else None

        try:
            log_qr_issue(
                endpoint='employee_identification',
                method='EMP_IDENT',
                status='error',
                user=None,
                employee_inn=inn_raw or "",
                employee_fio=fio_raw or "",
                tg_id="",
                phone_raw=phone_raw or "",
                phone_normalized=phone_norm or "",
                sm_store_id=sm_id_int,
                ukm_store_id=ukm_store_id,
                role_id=None,
                qr_data="",
                error_message=f"{stage}: {human_msg}",
                raw_request={"raw_body": raw_body} if raw_body else None,
                latitude=lat_val,
                longitude=lon_val,
            )
        except Exception:
            logger.exception("[EMP_IDENT] Ошибка при записи в qr_issue_logs")

        return JsonResponse(
            {'status': 'error', 'message': human_msg},
            status=http_status
        )

    # 1. Парсинг JSON
    try:
        data = json.loads(raw_body)
    except Exception as e:
        logger.error(f"[EMP_IDENT] JSON parse error: {e}; body={raw_body!r}")
        return _log_and_return_error(
            400,
            "Парсинг JSON",
            f"Некорректный JSON: {e}",
        )

    # 1.1 Координаты (опционально)
    lat_val = _to_float_or_none(data.get("latitude"))
    lon_val = _to_float_or_none(data.get("longitude"))

    # 2. Вытаскиваем поля из запроса (совместимость со старым и новым форматами)
    inn_raw = (data.get('inn') or data.get('employee_id') or "").strip()
    fio_raw = (data.get('fio') or data.get('FIO') or "").strip()

    smstore_raw = str(
        data.get('storeId')
        or data.get('smstore')
        or data.get('mx')
        or ""
    ).strip()

    phone_raw = (data.get('phone') or "").strip()

    dt_raw = (
        data.get('datetime')
        or data.get('Datetime')
        or data.get('event_datetime')
        or ""
    )
    dt_raw = dt_raw.strip()

    direction = str(data.get('direction') or data.get('event') or "").strip()

    logger.info(
        f"[EMP_IDENT] START: inn={inn_raw!r}, fio={fio_raw!r}, "
        f"storeId/mx={smstore_raw!r}, phone={phone_raw!r}, "
        f"datetime={dt_raw!r}, direction={direction!r},"
        f"latitude={lat_val!r}, longitude={lon_val!r}"
    )

    # 3. Базовая валидация входных полей
    if not inn_raw:
        return _log_and_return_error(
            400,
            "Валидация входных данных",
            "Не указан ИНН",
            inn_raw=inn_raw,
            fio_raw=fio_raw,
            smstore_raw=smstore_raw,
            phone_raw=phone_raw,
        )

    try:
        plain_inn = ensure_plain_inn(inn_raw)
    except Exception as e:
        logger.error(f"[EMP_IDENT] Bad INN: {e}")
        return _log_and_return_error(
            400,
            "Валидация ИНН",
            f"Некорректный ИНН: {e}",
            inn_raw=inn_raw,
            fio_raw=fio_raw,
            smstore_raw=smstore_raw,
            phone_raw=phone_raw,
        )

    if not fio_raw:
        return _log_and_return_error(
            400,
            "Валидация входных данных",
            "Не указано ФИО",
            inn_raw=plain_inn,
            fio_raw=fio_raw,
            smstore_raw=smstore_raw,
            phone_raw=phone_raw,
        )
    fio = " ".join(fio_raw.split())

    if not smstore_raw:
        return _log_and_return_error(
            400,
            "Валидация входных данных",
            "Не указан storeId / mx (smstore)",
            inn_raw=plain_inn,
            fio_raw=fio,
            smstore_raw=smstore_raw,
            phone_raw=phone_raw,
        )

    try:
        sm_store_id_int = int(smstore_raw)
    except ValueError:
        sm_store_id_int = None

    phone_norm = normalize_phone_ru(phone_raw) if phone_raw else None

    ukm_store_id = None
    store_obj = None
    if sm_store_id_int is not None:
        store_obj = Store.objects.filter(smstore=sm_store_id_int).first()
        if store_obj and store_obj.ukm4store is not None:
            try:
                ukm_store_id = int(store_obj.ukm4store)
            except (TypeError, ValueError):
                ukm_store_id = None

    if not dt_raw:
        return _log_and_return_error(
            400,
            "Валидация даты/времени",
            "Не указан datetime",
            inn_raw=plain_inn,
            fio_raw=fio,
            smstore_raw=smstore_raw,
            ukm_store_id=ukm_store_id,
            phone_raw=phone_raw,
        )

    try:
        event_dt_str = _parse_and_format_dt(dt_raw)
    except Exception as e:
        return _log_and_return_error(
            400,
            "Валидация даты/времени",
            f"Некорректная дата/время: {e}",
            inn_raw=plain_inn,
            fio_raw=fio,
            smstore_raw=smstore_raw,
            ukm_store_id=ukm_store_id,
            phone_raw=phone_raw,
        )

    # Пытаемся найти пользователя в PG, чтобы связать запись (если есть)
    user_obj = User.objects.filter(employee_id=plain_inn).first()

    # 4. Собираем payload для 1С в старом формате
    onec_payload = {
        "INN": plain_inn,
        "FIO": fio,
        "MX": smstore_raw,
        "Datetime": event_dt_str,
    }

    if phone_norm or direction or sm_store_id_int is not None or store_obj:
        extra = {
            "EmployeeID": encrypt_inn20(plain_inn),
            "StoreId": sm_store_id_int,
            "StoreName": store_obj.name if store_obj else "",
            "Phone": phone_norm or phone_raw,
            "Direction": direction,
        }
        onec_payload["Extra"] = extra

    idem_key = hashlib.sha256(
        f"{plain_inn}|{fio}|{smstore_raw}|{event_dt_str}".encode("utf-8")
    ).hexdigest()

    # 5. Отправка в 1С
    try:
        status_1c, text_1c = _post_to_onec(onec_payload, idem_key=idem_key)
    except Exception as e:
        logger.exception(f"[EMP_IDENT] Ошибка запроса в 1С: {e}")
        return _log_and_return_error(
            500,
            "Запрос в 1С",
            f"Ошибка запроса в 1С: {e}",
            inn_raw=plain_inn,
            fio_raw=fio,
            smstore_raw=smstore_raw,
            ukm_store_id=ukm_store_id,
            phone_raw=phone_raw,
        )

    # 6. Анализ ответа 1С и логирование
    ok_1c = (200 <= status_1c < 300)

    resp_short = text_1c if len(text_1c) <= 1000 else text_1c[:1000] + "…"
    tg_lines = [
        "📡 Отметка смены",
        "",
        "👤 Сотрудник:",
        f"  • ФИО: {fio}",
        f"  • ИНН: {plain_inn}",
        f"  • user_id (PostgreSQL): {user_obj.id if user_obj else '—'}",
        "",
        "🏬 Магазин:",
        f"  • MX (из запроса): {smstore_raw}",
        f"  • storeId (SMSTORE int): {sm_store_id_int if sm_store_id_int is not None else '—'}",
        f"  • ukm4store: {ukm_store_id if ukm_store_id is not None else '—'}",
        f"  • Name: {store_obj.name if store_obj else '—'}",
        "",
        "⚙️ Параметры события:",
        f"  • direction: {direction or '—'}",
        f"  • datetime: {event_dt_str or dt_raw or '—'}",
        "",
        "📲 Телефон:",
        f"  • raw: {phone_raw or '—'}",
        f"  • normalized: {phone_norm or '—'}",
        "",
        "🔗 1С:",
        f"  • HTTP статус: {status_1c}",
        f"  • Тело (обрезано): {resp_short}",
    ]
    _send_admin_log_async("\n".join(tg_lines))

    # Запись в qr_issue_logs
    try:
        log_qr_issue(
            endpoint='employee_identification',
            method=direction or 'EMP_IDENT',
            status='ok' if ok_1c else 'error',
            user=user_obj,
            employee_inn=plain_inn,
            employee_fio=fio,
            tg_id=getattr(user_obj, "tg_id", "") if user_obj else "",
            phone_raw=phone_raw or "",
            phone_normalized=phone_norm or "",
            sm_store_id=sm_store_id_int,
            ukm_store_id=ukm_store_id,
            role_id=None,
            qr_data="",
            error_message="" if ok_1c else f"1С вернула статус {status_1c}",
            raw_request={
                "request": data,
                "onec_payload": onec_payload,
                "onec_status": status_1c,
                "onec_response": text_1c,
            },
            latitude=lat_val,
            longitude=lon_val,
        )
    except Exception:
        logger.exception("[EMP_IDENT] Ошибка при записи успеха/ошибки в qr_issue_logs")

    # 7. Ответ клиенту
    if not ok_1c:
        return JsonResponse(
            {
                'status': 'error',
                'message': f'1С ответила со статусом {status_1c}',
                'onec_status': status_1c,
                'onec_body': text_1c,
            },
            status=500
        )

    return JsonResponse(
        {
            'status': 'ok',
            'onec_status': status_1c,
            'onec_body': text_1c,
        }
    )
# @csrf_exempt
# def employee_identification(request):
#     """
#     POST /employee-identification/

#     ВХОД (JSON, совместим со старой версией):
#     {
#       "inn": "7536207278",              # или "employee_id"
#       "fio": "Иванов Иван Иванович",    # или "FIO"
#       "storeId": 514,                   # или "smstore" или старый "mx"
#       "datetime": "24.09.2025 8:45:00", # или "Datetime" или "event_datetime"
#       // необязательные:
#       "phone": "8 (924) 000-00-00",
#       "direction": "IN"                 # или "OUT" / "ENTER" / "EXIT" / "event"
#       // "roleId" / "roleid" — можно не передавать, оно тут не нужно
#     }

#     Логика:
#       1) Валидация/нормализация входных данных (как в старой версии).
#       2) Подготовка payload для 1С в старом формате: INN/FIO/MX/Datetime.
#       3) На любой ошибке:
#          • красивый лог в Telegram
#          • запись в qr_issue_logs со статусом 'error'.
#       4) На успехе:
#          • красивый лог в Telegram
#          • запись в qr_issue_logs со статусом 'ok'/'error' в зависимости от ответа 1С.
#     """
#     if request.method != 'POST':
#         return JsonResponse(
#             {'status': 'error', 'message': 'Только POST'},
#             status=405
#         )

#     raw_body = request.body.decode('utf-8') if request.body else "{}"
#     lat_val: float | None = None
#     lon_val: float | None = None

#     def _log_and_return_error(
#         http_status: int,
#         stage: str,
#         human_msg: str,
#         *,
#         inn_raw: str = "",
#         fio_raw: str = "",
#         smstore_raw: str = "",
#         ukm_store_id: int | None = None,
#         phone_raw: str = "",
#     ) -> JsonResponse:
#         """
#         Общий обработчик ошибок:
#           • шлёт подробный лог в Telegram
#           • пишет строку в qr_issue_logs со статусом 'error'
#           • возвращает JsonResponse с текстом ошибки
#         """
#         # красивый лог в ТГ
#         lines = [
#             "❌ Ошибка при employee_identification",
#             f"🔁 Этап: {stage}",
#             f"ℹ️ Причина: {human_msg}",
#             "",
#             "📨 Контекст запроса:",
#             f"  • ИНН (сырое): {inn_raw or '—'}",
#             f"  • ФИО (сырое): {fio_raw or '—'}",
#             f"  • storeId / mx (сырое): {smstore_raw or '—'}",
#             f"  • Телефон (сырое): {phone_raw or '—'}",
#             f"  • latitude: {lat_val if lat_val is not None else '—'}",
#             f"  • longitude: {lon_val if lon_val is not None else '—'}",
#         ]
#         if ukm_store_id is not None:
#             lines.append(f"  • ukm4store: {ukm_store_id}")
#         if raw_body:
#             short_body = raw_body if len(raw_body) <= 1000 else raw_body[:1000] + "…"
#             lines.extend(["", "📦 Сырой JSON-запрос:", short_body])

#         send_telegram_log("\n".join(lines))

#         # подготовка полей для записи в qr_issue_logs
#         try:
#             sm_id_int = int(smstore_raw) if str(smstore_raw).isdigit() else None
#         except Exception:
#             sm_id_int = None

#         phone_norm = normalize_phone_ru(phone_raw) if phone_raw else None

#         try:
#             log_qr_issue(
#                 endpoint='employee_identification',
#                 method='EMP_IDENT',
#                 status='error',
#                 user=None,  # при ошибке пользователя можем ещё не знать
#                 employee_inn=inn_raw or "",
#                 employee_fio=fio_raw or "",
#                 tg_id="",
#                 phone_raw=phone_raw or "",
#                 phone_normalized=phone_norm or "",
#                 sm_store_id=sm_id_int,
#                 ukm_store_id=ukm_store_id,
#                 role_id=None,          # роль для этого эндпоинта не используется
#                 qr_data="",       
#                 error_message=f"{stage}: {human_msg}",
#                 raw_request={"raw_body": raw_body} if raw_body else None,
#                 latitude=lat_val,
#                 longitude=lon_val,
#             )
#         except Exception:
      
#             logger.exception("[EMP_IDENT] Ошибка при записи в qr_issue_logs")

#         return JsonResponse(
#             {'status': 'error', 'message': human_msg},
#             status=http_status
#         )

#     # 1. Парсинг JSON
#     try:
#         data = json.loads(raw_body)
#     except Exception as e:
#         logger.error(f"[EMP_IDENT] JSON parse error: {e}; body={raw_body!r}")
#         return _log_and_return_error(
#             400,
#             "Парсинг JSON",
#             f"Некорректный JSON: {e}",
#         )
        
#     # 1.1 Координаты (опционально)
#     lat_val = _to_float_or_none(data.get("latitude"))
#     lon_val = _to_float_or_none(data.get("longitude"))

#     # 2. Вытаскиваем поля из запроса (совместимость со старым и новым форматами)
#     inn_raw = (data.get('inn') or data.get('employee_id') or "").strip()
#     fio_raw = (data.get('fio') or data.get('FIO') or "").strip()

#     # поддержка storeId / smstore / mx (старое поле)
#     smstore_raw = str(
#         data.get('storeId')
#         or data.get('smstore')
#         or data.get('mx')
#         or ""
#     ).strip()

#     phone_raw = (data.get('phone') or "").strip()

#     # поддержка и "datetime", и "Datetime", и "event_datetime"
#     dt_raw = (
#         data.get('datetime')
#         or data.get('Datetime')
#         or data.get('event_datetime')
#         or ""
#     )
#     dt_raw = dt_raw.strip()

#     # направление вход/выход — опционально
#     direction = str(data.get('direction') or data.get('event') or "").strip()

#     logger.info(
#         f"[EMP_IDENT] START: inn={inn_raw!r}, fio={fio_raw!r}, "
#         f"storeId/mx={smstore_raw!r}, phone={phone_raw!r}, "
#         f"datetime={dt_raw!r}, direction={direction!r},"
#         f"latitude={lat_val!r}, longitude={lon_val!r}"
#     )

#     # 3. Базовая валидация входных полей (как в старой версии)
#     if not inn_raw:
#         return _log_and_return_error(
#             400,
#             "Валидация входных данных",
#             "Не указан ИНН",
#             inn_raw=inn_raw,
#             fio_raw=fio_raw,
#             smstore_raw=smstore_raw,
#             phone_raw=phone_raw,
#         )

#     try:
#         plain_inn = ensure_plain_inn(inn_raw)
#     except Exception as e:
#         logger.error(f"[EMP_IDENT] Bad INN: {e}")
#         return _log_and_return_error(
#             400,
#             "Валидация ИНН",
#             f"Некорректный ИНН: {e}",
#             inn_raw=inn_raw,
#             fio_raw=fio_raw,
#             smstore_raw=smstore_raw,
#             phone_raw=phone_raw,
#         )

#     if not fio_raw:
#         return _log_and_return_error(
#             400,
#             "Валидация входных данных",
#             "Не указано ФИО",
#             inn_raw=plain_inn,
#             fio_raw=fio_raw,
#             smstore_raw=smstore_raw,
#             phone_raw=phone_raw,
#         )
#     fio = " ".join(fio_raw.split())

#     if not smstore_raw:
#         return _log_and_return_error(
#             400,
#             "Валидация входных данных",
#             "Не указан storeId / mx (smstore)",
#             inn_raw=plain_inn,
#             fio_raw=fio,
#             smstore_raw=smstore_raw,
#             phone_raw=phone_raw,
#         )

#     # для Store/ukm4store пробуем привести к int, для 1С MX остаётся строкой
#     try:
#         sm_store_id_int = int(smstore_raw)
#     except ValueError:
#         sm_store_id_int = None

#     # Нормализация телефона (если есть)
#     phone_norm = normalize_phone_ru(phone_raw) if phone_raw else None

#     # Маппинг SMSTORE → ukm4store, чтобы записать его в лог
#     ukm_store_id = None
#     store_obj = None
#     if sm_store_id_int is not None:
#         store_obj = Store.objects.filter(smstore=sm_store_id_int).first()
#         if store_obj and store_obj.ukm4store is not None:
#             try:
#                 ukm_store_id = int(store_obj.ukm4store)
#             except (TypeError, ValueError):
#                 ukm_store_id = None

#     # Парсим дату/время события
#     if not dt_raw:
#         return _log_and_return_error(
#             400,
#             "Валидация даты/времени",
#             "Не указан datetime",
#             inn_raw=plain_inn,
#             fio_raw=fio,
#             smstore_raw=smstore_raw,
#             ukm_store_id=ukm_store_id,
#             phone_raw=phone_raw,
#         )

#     try:
#         event_dt_str = _parse_and_format_dt(dt_raw)
#     except Exception as e:
#         return _log_and_return_error(
#             400,
#             "Валидация даты/времени",
#             f"Некорректная дата/время: {e}",
#             inn_raw=plain_inn,
#             fio_raw=fio,
#             smstore_raw=smstore_raw,
#             ukm_store_id=ukm_store_id,
#             phone_raw=phone_raw,
#         )

#     # Пытаемся найти пользователя в PG, чтобы связать запись (если есть)
#     user_obj = User.objects.filter(employee_id=plain_inn).first()

#     # 4. Собираем payload для 1С в старом формате 
#     onec_payload = {
#         "INN": plain_inn,
#         "FIO": fio,
#         "MX": smstore_raw,  
#         "Datetime": event_dt_str, 
#     }

#     # Доп.инфа — только в Extra, старую обработку это не ломает
#     if phone_norm or direction or sm_store_id_int is not None or store_obj:
#         extra = {
#             "EmployeeID": encrypt_inn20(plain_inn),
#             "StoreId": sm_store_id_int,
#             "StoreName": store_obj.name if store_obj else "",
#             "Phone": phone_norm or phone_raw,
#             "Direction": direction,
#         }
#         onec_payload["Extra"] = extra

#     # Идемпотентный ключ (как раньше — по ИНН/ФИО/магазину/дате)
#     idem_key = hashlib.sha256(
#         f"{plain_inn}|{fio}|{smstore_raw}|{event_dt_str}".encode("utf-8")
#     ).hexdigest()

#     # 5. Отправка в 1С
#     try:
#         status_1c, text_1c = _post_to_onec(onec_payload, idem_key=idem_key)
#     except Exception as e:
#         logger.exception(f"[EMP_IDENT] Ошибка запроса в 1С: {e}")
#         return _log_and_return_error(
#             500,
#             "Запрос в 1С",
#             f"Ошибка запроса в 1С: {e}",
#             inn_raw=plain_inn,
#             fio_raw=fio,
#             smstore_raw=smstore_raw,
#             ukm_store_id=ukm_store_id,
#             phone_raw=phone_raw,
#         )

#     # 6. Анализ ответа 1С и логирование
#     ok_1c = (200 <= status_1c < 300)

#     # Лог в Telegram (в любом случае)
#     resp_short = text_1c if len(text_1c) <= 1000 else text_1c[:1000] + "…"
#     tg_lines = [
#         "📡 Отметка смены",
#         "",
#         "👤 Сотрудник:",
#         f"  • ФИО: {fio}",
#         f"  • ИНН: {plain_inn}",
#         f"  • user_id (PostgreSQL): {user_obj.id if user_obj else '—'}",
#         "",
#         "🏬 Магазин:",
#         f"  • MX (из запроса): {smstore_raw}",
#         f"  • storeId (SMSTORE int): {sm_store_id_int if sm_store_id_int is not None else '—'}",
#         f"  • ukm4store: {ukm_store_id if ukm_store_id is not None else '—'}",
#         f"  • Name: {store_obj.name if store_obj else '—'}",
#         "",
#         "⚙️ Параметры события:",
#         f"  • direction: {direction or '—'}",
#         f"  • datetime: {event_dt_str or dt_raw or '—'}",
#         "",
#         "📲 Телефон:",
#         f"  • raw: {phone_raw or '—'}",
#         f"  • normalized: {phone_norm or '—'}",
#         "",
#         "🔗 1С:",
#         f"  • HTTP статус: {status_1c}",
#         f"  • Тело (обрезано): {resp_short}",
#     ]
#     send_telegram_log("\n".join(tg_lines))

#     # Запись в qr_issue_logs
#     try:
#         log_qr_issue(
#             endpoint='employee_identification',
#             method=direction or 'EMP_IDENT',
#             status='ok' if ok_1c else 'error',
#             user=user_obj,
#             employee_inn=plain_inn,
#             employee_fio=fio,
#             tg_id=getattr(user_obj, "tg_id", "") if user_obj else "",
#             phone_raw=phone_raw or "",
#             phone_normalized=phone_norm or "",
#             sm_store_id=sm_store_id_int,
#             ukm_store_id=ukm_store_id,
#             role_id=None,   # для этого эндпоинта роль не пишем
#             qr_data="",     # факт идентификации, а не выдача QR
#             error_message="" if ok_1c else f"1С вернула статус {status_1c}",
#             raw_request={
#                 "request": data,
#                 "onec_payload": onec_payload,
#                 "onec_status": status_1c,
#                 "onec_response": text_1c,
#             },
#             latitude=lat_val,
#             longitude=lon_val,
#         )
#     except Exception:
#         logger.exception("[EMP_IDENT] Ошибка при записи успеха/ошибки в qr_issue_logs")

#     # 7. Ответ клиенту
#     if not ok_1c:
#         return JsonResponse(
#             {
#                 'status': 'error',
#                 'message': f'1С ответила со статусом {status_1c}',
#                 'onec_status': status_1c,
#                 'onec_body': text_1c,
#             },
#             status=500
#         )

#     return JsonResponse(
#         {
#             'status': 'ok',
#             'onec_status': status_1c,
#             'onec_body': text_1c,
#         }
#     )





def get_devices_for_user(user: User) -> tuple[Optional[User], list[dict]]:
    if not user:
        return None, []

    ukm_links = list(
        UKMUser.objects.filter(user_id=user.id).values("storeid", "roleid")
    )
    if not ukm_links:
        return user, []

    # storeid -> roleid (берём первый ненулевой, если есть)
    roles_by_store: dict[int, int | None] = {}
    store_ids: list[int] = []
    for x in ukm_links:
        sid_raw = x.get("storeid")
        rid_raw = x.get("roleid")
        if str(sid_raw).isdigit():
            sid = int(sid_raw)
            store_ids.append(sid)
            if sid not in roles_by_store:
                roles_by_store[sid] = int(rid_raw) if str(rid_raw).isdigit() else None

    store_ids = sorted(set(store_ids))
    if not store_ids:
        return user, []

    # ukm4store -> Store из Postgres (если есть)
    store_map = {
        int(s.ukm4store): s
        for s in Store.objects.filter(ukm4store__in=store_ids)
        if s.ukm4store is not None
    }

    devices: list[dict] = []

    for ukm4_storeid in store_ids:
        role_id = roles_by_store.get(ukm4_storeid)

        # 1) smstore пытаемся взять из таблицы Store
        s_obj = store_map.get(int(ukm4_storeid))
        smstore = getattr(s_obj, "smstore", None)

        # 2) если нет — пробуем достать smstore через Oracle по ukm4_storeid
        if smstore is None:
            try:
                info_by_ukm = get_store_info(ukm4_storeid)
                smstore = info_by_ukm.get("smstore")
            except Exception:
                smstore = None

        if smstore is None:
            logger.warning(f"[POS] smstore not resolved for ukm4store={ukm4_storeid}")
            continue

        # 1) UKM4 кассы
        try:
            devices.extend(
                fetch_ukm4_pos_list(
                    ukm4_storeid=int(ukm4_storeid),
                    smstore=int(smstore),
                    role_id=role_id,
                )
            )
        except Exception as e:
            logger.exception(f"[POS] UKM4 list error ukm4store={ukm4_storeid}, smstore={smstore}: {e}")

        # 2) UKM5 кассы — только если Oracle магазин помечен как UKM5
        try:
            info = get_store_info(int(smstore))
            if info.get("is_ukm5", False):
                devices.extend(
                    fetch_ukm5_pos_list(
                        smstore=int(smstore),
                        ukm4_storeid=int(ukm4_storeid),
                        role_id=role_id,
                    )
                )
        except Exception as e:
            logger.exception(f"[POS] UKM5 list error smstore={smstore}: {e}")

    return user, devices




def _human_user_name(user: "User") -> str:
    """
    Возвращает человекочитаемое имя инициатора.
    Подстроено под разные варианты полей.
    """
    for attr in ("full_name", "fio", "name", "username", "login"):
        val = getattr(user, attr, None)
        if val:
            s = str(val).strip()
            if s:
                return s
    # fallback
    return f"user_id={getattr(user, 'id', '—')}"


def _pos_kind_label(*, ukm4: bool, ukm5: bool, is_kso: bool) -> str:
    if ukm4 and is_kso:
        return "КСО УКМ4"
    if ukm4 and not is_kso:
        return "Касса УКМ4"
    if (not ukm4) and ukm5 and is_kso:
        return "КСО УКМ5"
    if (not ukm4) and ukm5 and (not is_kso):
        return "Касса УКМ5"
    return "Неизвестный тип"


def _device_cmd_hint(*, ukm4: bool, ukm5: bool, is_kso: bool) -> tuple[str, bool]:
    """
    Возвращает (команда, use_sudo) как мы реально выполняем reboot.
    """
    if ukm5:
        return ("sudo reboot", True)
    # ukm4
    if is_kso:
        return ("sudo reboot", True)
    return ("reboot", False)









def connect_ukm5_srvdata():
    return pymysql.connect(
        host=UKM5_SRV_HOST,
        user=UKM5_SRV_USER,
        password=UKM5_SRV_PASSWORD,
        database=UKM5_SRV_DB,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=10,
        read_timeout=30,
        write_timeout=30,
    )


def _parse_allowed_nets(raw: str) -> list[ipaddress._BaseNetwork]:
    out = []
    for part in (raw or "").split(","):
        part = part.strip()
        if not part:
            continue
        try:
            out.append(ipaddress.ip_network(part))
        except Exception:
            logger.warning(f"[POS] bad net in POS_REBOOT_ALLOWED_NETS: {part!r}")
    return out

_ALLOWED_NETS = _parse_allowed_nets(POS_REBOOT_ALLOWED_NETS_RAW)


def _ip_allowed(ip: str) -> bool:
    try:
        ip_obj = ipaddress.ip_address(str(ip).strip())
    except Exception:
        return False
    return any(ip_obj in net for net in _ALLOWED_NETS)


def fetch_ukm4_pos_list(*, ukm4_storeid: int, smstore: int, role_id: int | None = None) -> list[dict]:
    info = get_store_info(smstore)
    ukm_host = info.get("ukm4ip")

    conn = cur = None
    items: list[dict] = []
    try:
        conn = connect_ukm(host=ukm_host, store_id=smstore)
        cur = conn.cursor()

        kiosk_pat = "%КИОСК%"
        kso_pat = "%КСО%"

        cur.execute("""
            SELECT
              p.name AS pos_name,
              p.cash_id AS cash_id,
              ps.ip AS ip,
              cg.name AS cg_name,
              CASE
                WHEN cg.name IS NULL THEN 0
                WHEN UPPER(cg.name) LIKE %s OR UPPER(cg.name) LIKE %s THEN 1
                ELSE 0
              END AS is_kso
            FROM trm_in_pos p
            LEFT JOIN (
                SELECT cash_id, ip
                FROM trm_out_pos_state
                WHERE state = 1
            ) ps ON ps.cash_id = p.cash_id
            LEFT JOIN trm_in_configuration_groups cg ON cg.id = p.config_group_id
            WHERE p.store_id = %s
              AND p.active = 1
        """, (kiosk_pat, kso_pat, int(ukm4_storeid)))

        for row in (cur.fetchall() or []):
            is_kso = bool(row.get("is_kso"))
            ip = row.get("ip")

            if not ip or not str(ip).strip():
                continue

            items.append({
                "cash_id": row.get("cash_id"),
                "name": row.get("pos_name") or "",
                "ip": str(ip).strip(),

                "ukm4": True,
                "ukm5": False,
                "is_kso": is_kso,
                "ssh_user": "ukmclient" if is_kso else "root",
                "ukm_store_id": int(ukm4_storeid),   
                "sm_store_id": int(smstore),        
                "role_id": int(role_id) if role_id is not None else None,
            })

        return items
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except Exception:
            pass


def fetch_ukm5_pos_list(*, smstore: int, ukm4_storeid: int, role_id: int | None = None) -> list[dict]:
    conn = cur = None
    items: list[dict] = []
    try:
        conn = connect_ukm5_srvdata()
        cur = conn.cursor()

        cur.execute(
            "SELECT id FROM store_external_params WHERE external_id=%s LIMIT 1",
            (int(smstore),)
        )
        r = cur.fetchone()
        if not r or not r.get("id"):
            return []

        store_internal_id = int(r["id"])

        cur.execute("""
            SELECT
              p.name AS pos_name,
              p.guid AS guid,
              cs.last_ip_address AS ip
            FROM pos p
            LEFT JOIN tm_client_exchange_statistics cs ON cs.client_uid = p.guid
            WHERE p.store_id = %s
              AND p.active = 1
              AND p.deleted = 0
              AND cs.last_ip_address IS NOT NULL
              AND cs.last_ip_address <> ''
        """, (store_internal_id,))

        for row in (cur.fetchall() or []):
            guid = (row.get("guid") or "").strip()
            ip = (row.get("ip") or "").strip()

            if not guid:
                continue
            if not ip:
                continue  

            items.append({
                "cash_id": guid,
                "name": row.get("pos_name") or "",
                "ip": ip,

                "ukm4": False,
                "ukm5": True,
                "is_kso": True,
                "ssh_user": "ukm5",
                "ukm_store_id": int(ukm4_storeid),  
                "sm_store_id": int(smstore),       
                "role_id": int(role_id) if role_id is not None else None,
            })

        return items
    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except Exception:
            pass


def get_devices_for_tg_id(tg_id: str) -> tuple[Optional[User], list[dict]]:
    tg_id = str(tg_id or "").strip()
    if not tg_id:
        return None, []

    user = User.objects.filter(tg_id=tg_id).first()
    if not user:
        return None, []

    return get_devices_for_user(user)



def get_devices_for_ids(*, tg_id: str, max_id: str) -> tuple[Optional[User], list[dict], str]:
    """
    Возвращает (user, devices, err_msg).
    """
    user, err = _resolve_user_by_ids(tg_id=tg_id, max_id=max_id)
    if not user:
        return None, [], err

    user, devices = get_devices_for_user(user)
    return user, devices, ""





@csrf_exempt
def pos_list_by_tg(request):
    """
    POST {"tg_id":"..."} или {"max_id":"..."} или оба -> список касс/КСО UKM4 + UKM5 с ip.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Только POST"}, status=405)

    try:
        body = json.loads(request.body.decode("utf-8") if request.body else "{}")
    except Exception:
        return JsonResponse({"status": "error", "message": "Некорректный JSON"}, status=400)

    tg_id = str(body.get("tg_id") or "").strip()
    max_id = str(body.get("max_id") or "").strip()

    if not tg_id and not max_id:
        return JsonResponse({"status": "error", "message": "Не указан tg_id или max_id"}, status=400)

    user, devices, err = get_devices_for_ids(tg_id=tg_id, max_id=max_id)
    if not user:
        return JsonResponse({"status": "error", "message": err or "Пользователь не найден"}, status=404)

    return JsonResponse(
        {
            "status": "ok",
            "tg_id": str(getattr(user, "tg_id", "") or "").strip(),
            "max_id": str(getattr(user, "max_id", "") or "").strip(),
            "user_id": user.id,
            "devices": devices,
        },
        json_dumps_params={"ensure_ascii": False},
    )


def _ssh_reboot(ip: str, *, username: str, password: str, use_sudo: bool) -> dict:
    """
    Делает reboot по SSH (в Docker).
    """
    port = POS_SSH_PORT

    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    try:
        client.connect(
            hostname=ip,
            port=port,
            username=username,
            password=password,
            timeout=10,
            banner_timeout=10,
            auth_timeout=10,
            allow_agent=False,
            look_for_keys=False,
        )

        if use_sudo:
            # сначала пробуем без запроса пароля (если NOPASSWD)
            stdin, stdout, stderr = client.exec_command("sudo -n reboot", get_pty=True, timeout=10)
            err = (stderr.read() or b"").decode("utf-8", errors="ignore").strip()

            if err and ("password" in err.lower() or "a password is required" in err.lower()):
                stdin, stdout, stderr = client.exec_command("sudo -S reboot", get_pty=True, timeout=10)
                stdin.write(password + "\n")
                stdin.flush()
                err = (stderr.read() or b"").decode("utf-8", errors="ignore").strip()

            out = (stdout.read() or b"").decode("utf-8", errors="ignore").strip()

            return {"ok": True, "stdout": out, "stderr": err}

        else:
            stdin, stdout, stderr = client.exec_command("reboot", timeout=10)
            out = (stdout.read() or b"").decode("utf-8", errors="ignore").strip()
            err = (stderr.read() or b"").decode("utf-8", errors="ignore").strip()
            return {"ok": True, "stdout": out, "stderr": err}

    finally:
        try:
            client.close()
        except Exception:
            pass


@csrf_exempt
def pos_reboot(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Только POST"}, status=405)

    try:
        body = json.loads(request.body.decode("utf-8") if request.body else "{}")
    except Exception:
        return JsonResponse({"status": "error", "message": "Некорректный JSON"}, status=400)

    tg_id = str(body.get("tg_id") or "").strip()
    max_id = str(body.get("max_id") or "").strip()
    ip = str(body.get("ip") or "").strip()

    req_ukm4 = body.get("ukm4", None)
    req_ukm5 = body.get("ukm5", None)
    req_is_kso = body.get("is_kso", None)

    if (not tg_id and not max_id) or not ip:
        return JsonResponse({"status": "error", "message": "Нужны tg_id/max_id и ip"}, status=400)

    if not _ip_allowed(ip):
        return JsonResponse({"status": "error", "message": f"IP {ip} запрещён (allowlist)"}, status=403)

    user, devices, err = get_devices_for_ids(tg_id=tg_id, max_id=max_id)
    if not user:
        return JsonResponse({"status": "error", "message": err or "Пользователь не найден"}, status=404)

    tg_id_user = str(getattr(user, "tg_id", "") or "").strip()

    dev = next((d for d in devices if str(d.get("ip") or "").strip() == ip), None)
    if not dev:
        return JsonResponse({"status": "error", "message": "Этот IP не найден среди касс пользователя (запрещено)"}, status=403)

    ukm4 = bool(dev.get("ukm4"))
    ukm5 = bool(dev.get("ukm5"))
    is_kso = bool(dev.get("is_kso"))

    if req_ukm4 is not None and bool(req_ukm4) != ukm4:
        return JsonResponse({"status": "error", "message": "ukm4 не совпадает с типом кассы пользователя"}, status=400)
    if req_ukm5 is not None and bool(req_ukm5) != ukm5:
        return JsonResponse({"status": "error", "message": "ukm5 не совпадает с типом кассы пользователя"}, status=400)
    if req_is_kso is not None and bool(req_is_kso) != is_kso:
        return JsonResponse({"status": "error", "message": "is_kso не совпадает с типом кассы пользователя"}, status=400)

    if ukm4 and ukm5:
        return JsonResponse({"status": "error", "message": "Некорректное устройство: ukm4 и ukm5 одновременно"}, status=500)
    if not (ukm4 or ukm5):
        return JsonResponse({"status": "error", "message": "Некорректное устройство: не ukm4 и не ukm5"}, status=500)

    cash_id = dev.get("cash_id")
    name = dev.get("name") or ""
    sm_store_id = dev.get("sm_store_id")
    ukm_store_id = dev.get("ukm_store_id")
    role_id = dev.get("role_id")

    kind = _pos_kind_label(ukm4=ukm4, ukm5=ukm5, is_kso=is_kso)
    cmd_hint, _use_sudo_hint = _device_cmd_hint(ukm4=ukm4, ukm5=ukm5, is_kso=is_kso)

    if ukm5:
        username = "ukm5"
        password = SSH_UKM5_PASSWORD
        use_sudo = True
    else:
        if is_kso:
            username = "ukmclient"
            password = SSH_UKM4_KSO_PASSWORD
            use_sudo = True
        else:
            username = "root"
            password = SSH_UKM4_ROOT_PASSWORD
            use_sudo = False

    if not password:
        return JsonResponse({"status": "error", "message": f"Не задан пароль SSH для {username} (env)"}, status=500)

    initiator_name = _human_user_name(user)

    logger.info(
        f"[POS/REBOOT] start: tg_id={tg_id_user or '—'} max_id={getattr(user,'max_id','') or '—'} user_id={user.id} "
        f"initiator={initiator_name!r} ip={ip} kind={kind} "
        f"sm_store_id={sm_store_id} ukm_store_id={ukm_store_id} role_id={role_id} "
        f"ssh_user={username} cash_id={cash_id} name={name!r}"
    )

    try:
        res = _ssh_reboot(ip, username=username, password=password, use_sudo=use_sudo)

        log_qr_issue(
            endpoint="pos_reboot",
            method="POS_REBOOT",
            status="ok",
            user=user,
            tg_id=tg_id_user,
            sm_store_id=int(sm_store_id) if str(sm_store_id).isdigit() else sm_store_id,
            ukm_store_id=int(ukm_store_id) if str(ukm_store_id).isdigit() else ukm_store_id,
            role_id=int(role_id) if str(role_id).isdigit() else role_id,
            employee_inn="",
            employee_fio="",
            phone_raw="",
            phone_normalized="",
            qr_data="",
            error_message="",
            raw_request={
                "request": body,
                "resolved": {"tg_id": tg_id_user, "max_id": getattr(user, "max_id", None), "user_id": user.id},
                "device": dev,
                "ssh_user": username,
                "ssh_port": POS_SSH_PORT,
                "cmd": cmd_hint,
                "result": res,
            },
        )

        msg_lines = [
            "🔁 Перезагрузка кассы",
            "",
            "👤 Инициатор:",
            f"  • {initiator_name}",
            f"  • user_id: {user.id}",
            f"  • tg_id: {tg_id_user or '—'}",
            f"  • max_id: {getattr(user,'max_id','') or '—'}",
            "",
            "🏬 Магазин:",
            f"  • ukm_store_id: {ukm_store_id if ukm_store_id is not None else '—'}",
            f"  • sm_store_id: {sm_store_id if sm_store_id is not None else '—'}",
            f"  • role_id: {role_id if role_id is not None else '—'}",
            "",
            "💻 Касса:",
            f"  • Тип: {kind}",
            f"  • name: {name or '—'}",
            f"  • cash_id: {cash_id if cash_id is not None else '—'}",
            f"  • ip: {ip}",
            "",
            "✅ Результат:",
            f"  • ok: {res.get('ok')}",
            f"  • stdout: {res.get('stdout') or '—'}",
            f"  • stderr: {res.get('stderr') or '—'}",
        ]
        _send_telegram_log_async("\n".join(msg_lines))

        return JsonResponse({"status": "ok", "result": res}, json_dumps_params={"ensure_ascii": False})

    except Exception as e:
        logger.exception(f"[POS/REBOOT] error: {e}")

        try:
            log_qr_issue(
                endpoint="pos_reboot",
                method="POS_REBOOT",
                status="error",
                user=user,
                tg_id=tg_id_user,
                sm_store_id=int(sm_store_id) if str(sm_store_id).isdigit() else sm_store_id,
                ukm_store_id=int(ukm_store_id) if str(ukm_store_id).isdigit() else ukm_store_id,
                role_id=int(role_id) if str(role_id).isdigit() else role_id,
                employee_inn="",
                employee_fio="",
                phone_raw="",
                phone_normalized="",
                qr_data="",
                error_message=str(e),
                raw_request={
                    "request": body,
                    "resolved": {"tg_id": tg_id_user, "max_id": getattr(user, "max_id", None), "user_id": user.id},
                    "device": dev,
                    "ssh_user": username,
                    "ssh_port": POS_SSH_PORT,
                    "cmd": cmd_hint,
                },
            )
        except Exception:
            logger.exception("[POS/REBOOT] failed to write QRIssueLog")

        msg_lines = [
            "❌ Ошибка перезагрузки кассы",
            "",
            "👤 Инициатор:",
            f"  • {initiator_name}",
            f"  • user_id: {user.id}",
            f"  • tg_id: {tg_id_user or '—'}",
            f"  • max_id: {getattr(user,'max_id','') or '—'}",
            "",
            "💻 Касса:",
            f"  • Тип: {kind}",
            f"  • name: {name or '—'}",
            f"  • cash_id: {cash_id if cash_id is not None else '—'}",
            f"  • ip: {ip}",
            "",
            f"🧨 Ошибка: {e}",
        ]
        _send_telegram_log_async("\n".join(msg_lines))

        return JsonResponse({"status": "error", "message": str(e)}, status=500, json_dumps_params={"ensure_ascii": False})










def _custom_transliterate(text: str) -> str:
    text = (text or "").strip()
    if not text:
        return ""
    out = []
    for ch in text:
        up = ch.upper()
        if up in _TRANSLIT:
            mapped = _TRANSLIT[up]
            # сохраняем регистр примерно, но нам всё равно lower() потом
            out.append(mapped if ch.isupper() else mapped.lower())
        else:
            out.append(ch)
    return "".join(out)

def _normalize_login_piece(s: str) -> str:
    s = _custom_transliterate(s).lower()
    s = s.replace(".", "_").replace("-", "_").replace(" ", "_")
    s = _LOGIN_SAFE_RE.sub("", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s

def _expand_login_variants(login: str) -> set[str]:
    """
    Частые расхождения транслита (ё/e/yo и т.п.) — добавим “варианты”,
    чтобы шанс матчинга был выше.
    """
    variants = {login}
    if "yo" in login:
        variants.add(login.replace("yo", "e"))
    if "kh" in login:
        variants.add(login.replace("kh", "h"))
    if "ts" in login:
        variants.add(login.replace("ts", "c"))
    return {v for v in variants if v}

def _build_candidate_logins(lastname: str, firstname: str, patronymic: str) -> set[str]:
    ln = _normalize_login_piece(lastname)
    fn = _normalize_login_piece(firstname)
    sn = _normalize_login_piece(patronymic)

    out = set()
    if fn and ln:
        # 1) i_ivanov
        out.add(f"{fn[0]}_{ln}")
        # 2) ivan_ivanov
        out.add(f"{fn}_{ln}")

        # 3) ivan_i_ivanov (если есть отчество)
        if sn:
            out.add(f"{fn}_{sn[0]}_{ln}")

    expanded = set()
    for x in out:
        expanded |= _expand_login_variants(x)
    return expanded

def _is_valid_inn_digits(inn: str) -> bool:
    inn = (inn or "").strip()
    return inn.isdigit() and len(inn) in (10, 12)


def _connect_oracle_service(service_key: str):
    ORA_USER     = os.getenv("ORACLE_USER", "supermag")
    ORA_PASSWORD = os.getenv("ORACLE_PASSWORD", "qqq")

    call_timeout_ms = int(os.getenv("INN_SYNC_ORACLE_CALL_TIMEOUT_MS", "120000"))  # 120s

    info = ORACLE_TNS_MAP.get(service_key)

    if not info:
        host = os.getenv("ORACLE_HOST", "192.168.17.239")
        port = int(os.getenv("ORACLE_PORT", "1521"))
        service_name = service_key
        hosts = [host]
    else:
        service_name = (info.get("service_name") or service_key).strip()
        port = int(info.get("port", 1521))
        hosts = info.get("hosts") or [info.get("host")]
        hosts = [h for h in hosts if h]

    last_err = None

    for host in hosts:
        # 1) SERVICE_NAME
        try:
            dsn = cx_Oracle.makedsn(host, port, service_name=service_name)
            logger.info(
                f"[INN_SYNC][ORACLE] connect service_key={service_key} host={host} port={port} service_name={service_name}"
            )
            conn = _oracle_connect(user=ORA_USER, password=ORA_PASSWORD, dsn=dsn)
            try:
                conn.callTimeout = call_timeout_ms
                logger.info(f"[INN_SYNC][ORACLE] callTimeout={call_timeout_ms}ms service={service_key} host={host}")
            except Exception as e:
                logger.warning(f"[INN_SYNC][ORACLE] cannot set callTimeout service={service_key}: {e}")
            return conn
        except Exception as e:
            last_err = e
            logger.warning(
                f"[INN_SYNC][ORACLE] connect failed (service_name) {service_key}@{host}:{port}/{service_name}: {e}"
            )

        # 2) SID fallback
        try:
            dsn2 = cx_Oracle.makedsn(host, port, sid=service_name)
            logger.info(
                f"[INN_SYNC][ORACLE] retry as SID service_key={service_key} host={host} port={port} sid={service_name}"
            )
            conn = _oracle_connect(user=ORA_USER, password=ORA_PASSWORD, dsn=dsn2)
            try:
                conn.callTimeout = call_timeout_ms
                logger.info(f"[INN_SYNC][ORACLE] callTimeout={call_timeout_ms}ms service={service_key} host={host} (SID)")
            except Exception as e:
                logger.warning(f"[INN_SYNC][ORACLE] cannot set callTimeout service={service_key} (SID): {e}")
            return conn
        except Exception as e2:
            last_err = e2
            logger.warning(
                f"[INN_SYNC][ORACLE] connect failed (sid) {service_key}@{host}:{port} sid={service_name}: {e2}"
            )

    raise last_err or RuntimeError(f"Cannot connect to Oracle service {service_key}")


def _fetch_onec_working_employees() -> list[dict]:
    auth = None
    if ONEC_USER and ONEC_PASS:
        auth = (ONEC_USER, ONEC_PASS)

    t0 = time.time()
    logger.info(f"[INN_SYNC] 1C request: {ONEC_WORKING_EMPLOYEES_URL}")
    r = requests.get(
        ONEC_WORKING_EMPLOYEES_URL,
        auth=auth,
        timeout=ONEC_WORKING_EMPLOYEES_TIMEOUT,
        headers={"Accept": "application/json"},
    )
    r.raise_for_status()
    data = r.json()
    dt = time.time() - t0
    logger.info(f"[INN_SYNC] 1C employees fetched: {len(data)} in {dt:.1f}s")
    if not isinstance(data, list):
        raise ValueError("1C ответ не list[...], проверь API Get_WorkingEmployees")
    return data

def _build_employee_index(onec_employees: list[dict]):
    """
    Вернёт:
      - login_to_emps: dict[login] -> list[{inn, fio, raw}]
      - emp_stats: словарь счётчиков
    """
    login_to_emps: dict[str, list[dict]] = defaultdict(list)

    stats = Counter()
    stats["employees_total"] = len(onec_employees)

    for e in onec_employees:
        inn = str(e.get("ИНН") or "").strip()
        lastname = str(e.get("Фамилия") or "").strip()
        firstname = str(e.get("Имя") or "").strip()
        patronymic = str(e.get("Отчество") or "").strip()

        fio = " ".join([x for x in [lastname, firstname, patronymic] if x]).strip()

        if not fio:
            stats["employees_no_fio"] += 1
            continue

        if not _is_valid_inn_digits(inn):
            stats["employees_bad_inn"] += 1
            # всё равно можно проиндексировать (иногда ИНН пустой, но логин нужен для диагностики)
            inn = ""

        cand = _build_candidate_logins(lastname, firstname, patronymic)
        if not cand:
            stats["employees_no_login_candidates"] += 1
            continue

        rec = {"inn": inn, "fio": fio, "raw": e}

        for login in cand:
            login_to_emps[login].append(rec)

        stats["employees_indexed"] += 1

    # посчитаем конфликты (один логин -> несколько разных людей)
    ambiguous = 0
    for login, lst in login_to_emps.items():
        uniq = {(x.get("inn") or "", x.get("fio") or "") for x in lst}
        if len(uniq) > 1:
            ambiguous += 1
    stats["logins_ambiguous"] = ambiguous
    stats["logins_total"] = len(login_to_emps)

    return login_to_emps, stats

def _oracle_fetch_smstaff_rows(cur, only_enabled: bool, only_null_inn: bool):
    where = []
    if only_enabled:
        # обычно userenabled = '1' (как в твоём ответе)
        where.append("(userenabled = '1' OR userenabled = 1)")
    if only_null_inn:
        where.append("(inn IS NULL OR TRIM(inn) = '')")

    sql = """
        SELECT id, surname, serverlogin, inn, userenabled
        FROM smstaff
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id"
    cur.execute(sql)
    rows = cur.fetchall()
    return rows

def _normalize_staff_key(v: str) -> str:
    v = (v or "").strip()
    if not v:
        return ""
    # serverlogin часто в UPPER; surname часто lower; приводим к lower
    v = v.replace(".", "_").replace("-", "_").replace(" ", "_")
    v = v.lower()
    v = _LOGIN_SAFE_RE.sub("", v)
    v = re.sub(r"_+", "_", v).strip("_")
    return v

def sm_sync_inn_from_onec(
    dry_run: bool = True,
    overwrite_existing_inn: bool = False,
    services: list[str] | None = None,
    only_enabled: bool = True,
    only_null_inn: bool = True,
) -> dict:
    """
    Главная функция синхронизации ИНН из 1C в Oracle.smstaff.

    Добавлено:
      - run_id в логах
      - тайминги по шагам
      - heartbeat и прогресс по сканированию
      - логи "chunk_start/chunk_done" с id диапазоном
      - rollback при ошибке и продолжение на следующую базу
    """
    services = services or ORACLE_SERVICES_ALL

    # настройки логов/прогресса/таймингов из env
    progress_every_rows = int(os.getenv("INN_SYNC_PROGRESS_EVERY_ROWS", "2000"))
    heartbeat_sec = int(os.getenv("INN_SYNC_HEARTBEAT_SEC", "30"))
    slow_warn_sec = float(os.getenv("INN_SYNC_SLOW_STEP_WARN_SEC", "10"))
    batch_size = int(os.getenv("INN_SYNC_UPDATE_BATCH", "1"))

    run_id = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    run_ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    out_dir = Path(LOG_DIR) / "inn_sync"
    out_dir.mkdir(parents=True, exist_ok=True)

    report_csv = out_dir / f"inn_sync_details_{run_ts}.csv"
    report_json = out_dir / f"inn_sync_summary_{run_ts}.json"

    logger.info(
        f"[INN_SYNC] START run_id={run_id} dry_run={dry_run} overwrite_existing_inn={overwrite_existing_inn} "
        f"only_enabled={only_enabled} only_null_inn={only_null_inn} batch_size={batch_size} services={len(services)}"
    )
    logger.info(f"[INN_SYNC] run_id={run_id} output CSV={report_csv} JSON={report_json}")

    # 1) 1C
    onec_employees = _fetch_onec_working_employees()

    # 2) индекс по логинам
    t_index = time.time()
    login_to_emps, emp_stats = _build_employee_index(onec_employees)
    dt_index = time.time() - t_index
    logger.info(f"[INN_SYNC] run_id={run_id} 1C index built in {dt_index:.1f}s stats={dict(emp_stats)}")

    summary = {
        "run_id": run_id,
        "dry_run": dry_run,
        "overwrite_existing_inn": overwrite_existing_inn,
        "only_enabled": only_enabled,
        "only_null_inn": only_null_inn,
        "batch_size": batch_size,
        "services_requested": services,
        "employee_index": dict(emp_stats),
        "db_results": [],
        "totals": Counter(),
        "files": {"details_csv": str(report_csv), "summary_json": str(report_json)},
    }

    with report_csv.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "service", "staff_id", "surname", "serverlogin", "inn_before",
                "match_key", "match_status", "inn_proposed", "employee_fio",
                "reason"
            ],
        )
        w.writeheader()

        for service in services:
            db_counts = Counter()
            db_result = {"service": service}

            conn = cur = None
            service_start = time.time()
            logger.info(f"[INN_SYNC] run_id={run_id} service={service} START")

            try:
                # connect
                conn = _connect_oracle_service(service)
                cur = conn.cursor()
                db_counts["connect_ok"] += 1

                # небольшая оптимизация чтения (не обязательно, но полезно)
                try:
                    cur.arraysize = int(os.getenv("INN_SYNC_ORACLE_ARRAYSIZE", "1000"))
                except Exception:
                    pass

                # fetch rows
                t_fetch = time.time()
                rows = _oracle_fetch_smstaff_rows(cur, only_enabled=only_enabled, only_null_inn=only_null_inn)
                dt_fetch = time.time() - t_fetch
                db_counts["staff_rows_scanned"] += len(rows)

                if dt_fetch > slow_warn_sec:
                    logger.warning(
                        f"[INN_SYNC] run_id={run_id} service={service} SLOW fetch_rows dt={dt_fetch:.1f}s rows={len(rows)}"
                    )
                else:
                    logger.info(
                        f"[INN_SYNC] run_id={run_id} service={service} fetch_rows dt={dt_fetch:.1f}s rows={len(rows)}"
                    )

                # дубли ключей в smstaff
                t_dups = time.time()
                staff_keys = []
                for (sid, surname, serverlogin, inn_before, userenabled) in rows:
                    sk1 = _normalize_staff_key(serverlogin)
                    sk2 = _normalize_staff_key(surname)
                    if sk1:
                        staff_keys.append(sk1)
                    if sk2 and sk2 != sk1:
                        staff_keys.append(sk2)

                dup_keys = {k for k, c in Counter(staff_keys).items() if c > 1}
                db_counts["staff_duplicate_keys"] += len(dup_keys)
                dt_dups = time.time() - t_dups
                if dt_dups > slow_warn_sec:
                    logger.warning(f"[INN_SYNC] run_id={run_id} service={service} SLOW dup_keys dt={dt_dups:.1f}s dup={len(dup_keys)}")
                else:
                    logger.info(f"[INN_SYNC] run_id={run_id} service={service} dup_keys dt={dt_dups:.1f}s dup={len(dup_keys)}")

                updates = []

                scanned = 0
                last_hb = time.time()

                # scan rows
                for (sid, surname, serverlogin, inn_before, userenabled) in rows:
                    scanned += 1
                    now = time.time()

                    if scanned % progress_every_rows == 0:
                        logger.info(f"[INN_SYNC] run_id={run_id} service={service} scan_progress {scanned}/{len(rows)}")
                    elif now - last_hb >= heartbeat_sec:
                        logger.info(f"[INN_SYNC] run_id={run_id} service={service} heartbeat scan scanned={scanned}/{len(rows)}")
                        last_hb = now

                    inn_before_s = (inn_before or "").strip() if inn_before is not None else ""

                    if inn_before_s and not overwrite_existing_inn:
                        db_counts["skip_already_has_inn"] += 1
                        w.writerow({
                            "service": service, "staff_id": sid, "surname": surname, "serverlogin": serverlogin,
                            "inn_before": inn_before_s,
                            "match_key": "", "match_status": "skip_has_inn",
                            "inn_proposed": "", "employee_fio": "", "reason": "already has INN",
                        })
                        continue

                    key_server = _normalize_staff_key(serverlogin)
                    key_surname = _normalize_staff_key(surname)

                    match_key = ""
                    emps = None

                    if key_server and key_server in login_to_emps:
                        match_key = key_server
                        emps = login_to_emps[key_server]
                    elif key_surname and key_surname in login_to_emps:
                        match_key = key_surname
                        emps = login_to_emps[key_surname]

                    if not emps:
                        db_counts["not_found"] += 1
                        w.writerow({
                            "service": service, "staff_id": sid, "surname": surname, "serverlogin": serverlogin,
                            "inn_before": inn_before_s,
                            "match_key": "", "match_status": "not_found",
                            "inn_proposed": "", "employee_fio": "", "reason": "no match in 1C index",
                        })
                        continue

                    uniq_people = {(x.get("inn") or "", x.get("fio") or "") for x in emps}
                    if len(uniq_people) > 1:
                        db_counts["ambiguous"] += 1
                        sample = "; ".join([f"{fio}:{inn}" for (inn, fio) in list(uniq_people)[:5]])
                        w.writerow({
                            "service": service, "staff_id": sid, "surname": surname, "serverlogin": serverlogin,
                            "inn_before": inn_before_s,
                            "match_key": match_key, "match_status": "ambiguous",
                            "inn_proposed": "", "employee_fio": "", "reason": f"multiple employees for login: {sample}",
                        })
                        continue

                    (inn_prop, fio_prop) = next(iter(uniq_people))
                    inn_prop = (inn_prop or "").strip()

                    if not _is_valid_inn_digits(inn_prop):
                        db_counts["bad_inn_in_1c_match"] += 1
                        w.writerow({
                            "service": service, "staff_id": sid, "surname": surname, "serverlogin": serverlogin,
                            "inn_before": inn_before_s,
                            "match_key": match_key, "match_status": "bad_inn",
                            "inn_proposed": inn_prop, "employee_fio": fio_prop,
                            "reason": "matched employee has empty/invalid INN in 1C",
                        })
                        continue

                    if inn_before_s and inn_before_s == inn_prop:
                        db_counts["already_same_inn"] += 1
                        w.writerow({
                            "service": service, "staff_id": sid, "surname": surname, "serverlogin": serverlogin,
                            "inn_before": inn_before_s,
                            "match_key": match_key, "match_status": "already_same",
                            "inn_proposed": inn_prop, "employee_fio": fio_prop,
                            "reason": "INN already equals proposed",
                        })
                        continue

                    db_counts["matched_ok"] += 1
                    if match_key in dup_keys:
                        db_counts["matched_but_staff_key_duplicate"] += 1

                    w.writerow({
                        "service": service, "staff_id": sid, "surname": surname, "serverlogin": serverlogin,
                        "inn_before": inn_before_s,
                        "match_key": match_key, "match_status": "will_update" if not dry_run else "matched_dry_run",
                        "inn_proposed": inn_prop, "employee_fio": fio_prop,
                        "reason": "ok",
                    })

                    if not dry_run:
                        updates.append({"inn": inn_prop, "id": sid})

                logger.info(
                    f"[INN_SYNC] run_id={run_id} service={service} prepared_updates={len(updates)} "
                    f"matched_ok={db_counts.get('matched_ok', 0)} not_found={db_counts.get('not_found', 0)} "
                    f"ambiguous={db_counts.get('ambiguous', 0)} skip_has_inn={db_counts.get('skip_already_has_inn', 0)}"
                )

                # updates
                if not dry_run and updates:
                    updated_total = 0
                    total_to_update = len(updates)

                    chunks_total = (total_to_update + batch_size - 1) // batch_size

                    for i in range(0, total_to_update, batch_size):
                        chunk = updates[i:i + batch_size]
                        first_id = chunk[0]["id"]
                        last_id = chunk[-1]["id"]
                        chunk_no = (i // batch_size) + 1

                        logger.info(
                            f"[INN_SYNC] run_id={run_id} service={service} chunk_start "
                            f"{chunk_no}/{chunks_total} size={len(chunk)} ids={first_id}-{last_id}"
                        )

                        t_chunk = time.time()
                        try:
                            cur.executemany("UPDATE smstaff SET inn = :inn WHERE id = :id", chunk)
                            conn.commit()
                        except Exception as e:
                            try:
                                conn.rollback()
                            except Exception:
                                pass
                            db_counts["update_failed_chunks"] += 1
                            logger.exception(
                                f"[INN_SYNC] run_id={run_id} service={service} chunk_failed "
                                f"size={len(chunk)} ids={first_id}-{last_id}: {e}"
                            )
                            # продолжаем дальше, чтобы не останавливать весь прогон
                            continue

                        dt_chunk = time.time() - t_chunk
                        updated_total += len(chunk)

                        if dt_chunk > slow_warn_sec:
                            logger.warning(
                                f"[INN_SYNC] run_id={run_id} service={service} chunk_done SLOW dt={dt_chunk:.1f}s "
                                f"committed {updated_total}/{total_to_update} ids={first_id}-{last_id}"
                            )
                        else:
                            logger.info(
                                f"[INN_SYNC] run_id={run_id} service={service} chunk_done dt={dt_chunk:.1f}s "
                                f"committed {updated_total}/{total_to_update} ids={first_id}-{last_id}"
                            )

                    db_counts["updated"] += updated_total
                else:
                    db_counts["updated"] += 0

                db_result["duration_sec"] = round(time.time() - service_start, 2)

                logger.info(
                    f"[INN_SYNC] run_id={run_id} service={service} DONE "
                    f"updated={db_counts.get('updated', 0)} failed_chunks={db_counts.get('update_failed_chunks', 0)} "
                    f"duration={db_result['duration_sec']}s"
                )

            except Exception as e:
                db_counts["connect_fail"] += 1
                db_result["error"] = str(e)
                logger.exception(f"[INN_SYNC] run_id={run_id} service={service} FAILED: {e}")

            finally:
                try:
                    if cur:
                        cur.close()
                    if conn:
                        conn.close()
                except Exception:
                    pass

            db_result.update(dict(db_counts))
            summary["db_results"].append(db_result)
            summary["totals"].update(db_counts)

    summary["totals"] = dict(summary["totals"])

    with report_json.open("w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    logger.info(f"[INN_SYNC] FINISH run_id={run_id} dry_run={dry_run} CSV={report_csv} JSON={report_json}")
    return summary


@csrf_exempt
def sm_staff_sync_inn(request):
    """
    POST /sm/staff/sync-inn/
    Body JSON:
    {
      "dry_run": true,
      "overwrite_existing_inn": false,
      "only_enabled": true,
      "only_null_inn": true,
      "services": ["BINUU00","BINUU01"]
    }
    """
    if request.method != "POST":
        return JsonResponse(
            {"status": "error", "message": "Только POST"},
            status=405,
            json_dumps_params={"ensure_ascii": False},
        )

    try:
        body = request.body.decode("utf-8") if request.body else "{}"
        data = json.loads(body or "{}")

        dry_run = bool(data.get("dry_run", True))
        overwrite_existing_inn = bool(data.get("overwrite_existing_inn", False))
        only_enabled = bool(data.get("only_enabled", True))
        only_null_inn = bool(data.get("only_null_inn", True))

        services = data.get("services")
        if services is not None:
            if not isinstance(services, list) or not all(isinstance(x, str) for x in services):
                return JsonResponse(
                    {"status": "error", "message": "services должен быть list[str]"},
                    status=400,
                    json_dumps_params={"ensure_ascii": False},
                )
            services = [s.strip() for s in services if s and s.strip()]

        logger.info(
            f"[INN_SYNC_VIEW] request dry_run={dry_run} overwrite_existing_inn={overwrite_existing_inn} "
            f"only_enabled={only_enabled} only_null_inn={only_null_inn} services={services or 'ALL'}"
        )

        result = sm_sync_inn_from_onec(
            dry_run=dry_run,
            overwrite_existing_inn=overwrite_existing_inn,
            services=services,
            only_enabled=only_enabled,
            only_null_inn=only_null_inn,
        )

        return JsonResponse(
            {"status": "ok", "result": result},
            json_dumps_params={"ensure_ascii": False, "indent": 2},
        )

    except Exception as e:
        logger.exception(f"[INN_SYNC_VIEW] error: {e}")
        return JsonResponse(
            {"status": "error", "message": str(e)},
            status=500,
            json_dumps_params={"ensure_ascii": False},
        )


def _oracle_get_table_columns_set(cur, owner: str, table: str) -> set[str]:
    """
    Возвращает set({ 'ID', 'SURNAME', ... }) по all_tab_columns.
    Важно: bind-имена делаем "безопасными", чтобы не ловить ORA-01745.
    """
    cur.execute(
        """
        SELECT column_name
        FROM all_tab_columns
        WHERE owner = :b_owner
          AND table_name = :b_table
        """,
        b_owner=owner.upper(),
        b_table=table.upper(),
    )
    return {str(r[0]).upper() for r in cur.fetchall()}


def _normalize_service_key(raw: str) -> str:
    return (raw or "").strip().upper()

def _is_allowed_service(service_key: str) -> bool:
    # Жёсткий whitelist: только то, что ты явно описал
    if service_key in ORACLE_TNS_MAP:
        return True
    if service_key in ORACLE_SERVICES_ALL:
        return True
    return False


@csrf_exempt
def sm_staff_list_by_db(request):
    """
    GET /sm/staff/by-db/?db=BINUU00&limit=50&offset=0&q=иванов&only_enabled=1

    - db: ключ сервиса из ORACLE_TNS_MAP (BINUU00/BINUU01/BINCH12/...)
    - only_enabled=1: фильтр (userenabled = 1 OR userenabled='1')
    - пагинация через ROW_NUMBER() (совместимо с Oracle 11g+)
    - устойчиво к отсутствующим колонкам (patronymic/name и т.д.)
    """
    if request.method != "GET":
        return JsonResponse(
            {"status": "error", "message": "Только GET"},
            status=405,
            json_dumps_params={"ensure_ascii": False},
        )

    db = (request.GET.get("db") or request.GET.get("service") or "").strip().upper()
    if not db:
        return JsonResponse(
            {"status": "error", "message": "Не передан параметр db (например db=BINUU00)"},
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    # whitelist
    if db not in ORACLE_TNS_MAP:
        return JsonResponse(
            {"status": "error", "message": f"Неизвестная база db={db!r}. Добавь её в ORACLE_TNS_MAP."},
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )

    # limit/offset
    try:
        limit = int(request.GET.get("limit", "200"))
    except ValueError:
        limit = 200
    try:
        offset = int(request.GET.get("offset", "0"))
    except ValueError:
        offset = 0

    limit = max(1, min(limit, 1000))
    offset = max(0, offset)

    q = (request.GET.get("q") or "").strip().lower()

    # only_enabled: по умолчанию 0 (не фильтровать), если передали 1/true/yes -> фильтровать
    only_enabled_raw = (request.GET.get("only_enabled") or "").strip().lower()
    only_enabled = only_enabled_raw in ("1", "true", "yes", "y", "on")

    conn = cur = None
    try:
        conn = _connect_oracle_service(db)
        cur = conn.cursor()

        # какие колонки реально есть в этой базе
        cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")

        def has(col: str) -> bool:
            return col.upper() in cols_set

        # SELECT: если колонки нет — отдаём NULL AS col
        select_parts = []
        def add_select(col: str):
            if has(col):
                select_parts.append(col)
            else:
                select_parts.append(f"NULL AS {col}")

        add_select("id")
        add_select("surname")
        add_select("name")
        add_select("patronymic")
        add_select("serverlogin")
        add_select("inn")
        add_select("userenabled")

        base_sql = f"SELECT {', '.join(select_parts)} FROM smstaff"
        binds = {}

        where_parts = []

        if only_enabled:
            if has("userenabled"):
                where_parts.append("(userenabled = '1' OR userenabled = 1)")
            else:
                # если вдруг колонки нет — фильтровать нечем
                pass

        if q:
            like = f"%{q}%"
            or_parts = []
            if has("surname"):
                or_parts.append("LOWER(surname) LIKE :q")
            if has("name"):
                or_parts.append("LOWER(name) LIKE :q")
            if has("patronymic"):
                or_parts.append("LOWER(patronymic) LIKE :q")
            if has("serverlogin"):
                or_parts.append("LOWER(serverlogin) LIKE :q")

            if or_parts:
                where_parts.append("(" + " OR ".join(or_parts) + ")")
                binds["q"] = like

        if where_parts:
            base_sql += " WHERE " + " AND ".join(where_parts)

        # ORDER BY (для ROW_NUMBER)
        if has("surname"):
            order_expr = "surname"
        elif has("id"):
            order_expr = "id"
        else:
            order_expr = "1"

        # Oracle 11g/12c-safe pagination
        # rn > offset  AND rn <= offset+limit
        sql = f"""
            SELECT *
            FROM (
                SELECT t.*, ROW_NUMBER() OVER (ORDER BY {order_expr}) rn
                FROM (
                    {base_sql}
                ) t
            )
            WHERE rn > :b_off
              AND rn <= :b_off_to
        """
        binds["b_off"] = offset
        binds["b_off_to"] = offset + limit

        cur.execute(sql, binds)
        items = _oracle_rows_to_jsonable(cur)

        return JsonResponse(
            {
                "status": "ok",
                "db": db,
                "count": len(items),
                "limit": limit,
                "offset": offset,
                "only_enabled": only_enabled,
                "columns_present": sorted(list(cols_set)),
                "items": items,
            },
            json_dumps_params={"ensure_ascii": False, "indent": 2},
        )

    except Exception as e:
        logger.exception(f"[SM/STAFF_BY_DB] error db={db}: {e}")
        return JsonResponse(
            {"status": "error", "message": str(e), "db": db},
            status=500,
            json_dumps_params={"ensure_ascii": False},
        )
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass









def _ui_services_list() -> list[str]:
    # Показываем только те базы, где реально есть host/port в ORACLE_TNS_MAP
    return sorted(list(ORACLE_TNS_MAP.keys()))


def _ui_has_col(cols_set: set[str], col: str) -> bool:
    return col.upper() in cols_set





#10.03.2026 Для создания пользователя в СМ (первая версия)


def _smstaff_afio_sql(alias: str, cols_set: set[str]) -> str:
    """
    Возвращает SQL-выражение AS afio.
    Приоритет:
    1) SMSTAFF.AFIO
    2) fallback из surname/name/patronymic, если AFIO нет
    """
    upper_cols = {c.upper() for c in cols_set}

    if "AFIO" in upper_cols:
        return f"{alias}.afio AS afio"

    parts = []
    if "SURNAME" in upper_cols:
        parts.append(f"NVL({alias}.surname, '')")
    if "NAME" in upper_cols:
        parts.append(f"NVL({alias}.name, '')")
    if "PATRONYMIC" in upper_cols:
        parts.append(f"NVL({alias}.patronymic, '')")

    if parts:
        joined = " || ' ' || ".join(parts)
        return f"TRIM(REGEXP_REPLACE({joined}, ' +', ' ')) AS afio"

    return "NULL AS afio"

_SM_ACHECK_CHOICES = [
    {"value": "-1", "label": "-1 — нового не создавать, работать только с существующим, можно обновить ФИО"},
    {"value": "0",  "label": "0 — если пользователь есть, ничего не делать"},
    {"value": "1",  "label": "1 — если пользователь есть, поменять пароль/должность/ФИО"},
    {"value": "2",  "label": "2 — заблокировать пользователя"},
]

_SM_USERENABLED_CHOICES = [
    {"value": "-1", "label": "-1 — не менять статус"},
    {"value": "1",  "label": "1 — активировать пользователя"},
    {"value": "0",  "label": "0 — заблокировать пользователя"},
]

def _oracle_update_smstaff_after_create(cur, login: str, inn: str, afio: str, offindex: int | None) -> int:
    """
    Принудительно обновляет AFIO и OFFINDEX в SMSTAFF после вызова bin_createuser.
    Это нужно, потому что процедура может не менять AFIO у существующего пользователя.
    """
    cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")
    upper_cols = {c.upper() for c in cols_set}

    set_parts = []
    binds = {
        "b_login": (login or "").strip().lower(),
    }

    if "AFIO" in upper_cols:
        set_parts.append("afio = :b_afio")
        binds["b_afio"] = (afio or "").strip()

    if offindex is not None and "OFFINDEX" in upper_cols:
        set_parts.append("offindex = :b_offindex")
        binds["b_offindex"] = offindex

    if not set_parts:
        return 0

    where_parts = []
    if "SERVERLOGIN" in upper_cols:
        where_parts.append("LOWER(TRIM(serverlogin)) = :b_login")

    if "INN" in upper_cols and inn:
        where_parts.append("TRIM(inn) = :b_inn")
        binds["b_inn"] = (inn or "").strip()

    if not where_parts:
        return 0

    sql = f"""
        UPDATE smstaff
           SET {", ".join(set_parts)}
         WHERE {" AND ".join(where_parts)}
    """

    cur.execute(sql, binds)
    return int(cur.rowcount or 0)



def _is_valid_inn_digits(inn: str) -> bool:
    inn = (inn or "").strip()
    return inn.isdigit() and len(inn) in (10, 12)

def _is_sm_block_mode(acheck_raw: str, auserenabled_raw: str) -> bool:
    """
    Режим блокировки:
    ACHECK = 2 и AUSERENABLED = 0
    """
    return (str(acheck_raw).strip() == "2") and (str(auserenabled_raw).strip() == "0")


def _generate_sm_password(length: int = 4) -> str:
    return "".join(random.choices(string.digits, k=length))


def _is_valid_sm_login(login: str) -> bool:
    """
    Разрешаем логины формата:
    t_test, i.ivanov, user-1, user_1
    """
    login = (login or "").strip()
    return bool(re.fullmatch(r"[A-Za-z0-9._-]{1,64}", login))


def _oracle_fetch_smoffcfg_choices(cur) -> list[dict]:
    cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMOFFCFG")

    has_id = "ID" in cols_set
    has_title = "TITLE" in cols_set
    has_orarole = "ORAROLE" in cols_set

    if not has_id:
        raise ValueError("В таблице SMOFFCFG нет колонки ID.")

    select_parts = [
        "id",
        "title" if has_title else "NULL AS title",
        "orarole" if has_orarole else "NULL AS orarole",
    ]

    order_expr = "title" if has_title else "id"

    cur.execute(f"""
        SELECT {", ".join(select_parts)}
        FROM smoffcfg
        ORDER BY {order_expr}
    """)

    items = []
    for row in cur.fetchall():
        items.append({
            "id": row[0],
            "title": row[1],
            "orarole": row[2],
        })
    return items


# def _oracle_fetch_smstaff_conflicts(cur, login: str, inn: str) -> list[dict]:
#     """
#     Ищем существующих пользователей по логину и/или ИНН.
#     Безопасно для баз, где нет части колонок.
#     """
#     cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")

#     def HAS(col: str) -> bool:
#         return col.upper() in cols_set

#     select_parts = [
#         "s.id AS id" if HAS("id") else "NULL AS id",
#         "s.surname AS surname" if HAS("surname") else "NULL AS surname",
#         "s.name AS name" if HAS("name") else "NULL AS name",
#         "s.patronymic AS patronymic" if HAS("patronymic") else "NULL AS patronymic",
#         "s.serverlogin AS serverlogin" if HAS("serverlogin") else "NULL AS serverlogin",
#         "s.inn AS inn" if HAS("inn") else "NULL AS inn",
#         "s.userenabled AS userenabled" if HAS("userenabled") else "NULL AS userenabled",
#         "s.offindex AS offindex" if HAS("offindex") else "NULL AS offindex",
#     ]

#     if HAS("offindex"):
#         select_parts.append("""
#             (
#                 SELECT c.title
#                 FROM smoffcfg c
#                 WHERE c.id = s.offindex
#                   AND ROWNUM = 1
#             ) AS off_title
#         """)
#     else:
#         select_parts.append("NULL AS off_title")

#     where_parts = []
#     binds = {}

#     if HAS("serverlogin") and login:
#         where_parts.append("LOWER(TRIM(s.serverlogin)) = :b_login")
#         binds["b_login"] = (login or "").strip().lower()

#     if HAS("inn") and inn:
#         where_parts.append("TRIM(s.inn) = :b_inn")
#         binds["b_inn"] = (inn or "").strip()

#     if not where_parts:
#         return []

#     sql = f"""
#         SELECT {", ".join(select_parts)}
#         FROM smstaff s
#         WHERE {" OR ".join(where_parts)}
#         ORDER BY s.id
#     """

#     cur.execute(sql, binds)
#     return _oracle_rows_to_jsonable(cur)

def _oracle_fetch_smstaff_conflicts(cur, login: str, inn: str) -> list[dict]:
    """
    Ищем существующих пользователей по логину и/или ИНН.
    Возвращаем AFIO, а не surname/name/patronymic.
    """
    cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")

    def HAS(col: str) -> bool:
        return col.upper() in cols_set

    select_parts = [
        "s.id AS id" if HAS("id") else "NULL AS id",
        _smstaff_afio_sql("s", cols_set),
        "s.serverlogin AS serverlogin" if HAS("serverlogin") else "NULL AS serverlogin",
        "s.inn AS inn" if HAS("inn") else "NULL AS inn",
        "s.userenabled AS userenabled" if HAS("userenabled") else "NULL AS userenabled",
        "s.offindex AS offindex" if HAS("offindex") else "NULL AS offindex",
    ]

    if HAS("offindex"):
        select_parts.append("""
            (
                SELECT c.title
                FROM smoffcfg c
                WHERE c.id = s.offindex
                  AND ROWNUM = 1
            ) AS off_title
        """)
    else:
        select_parts.append("NULL AS off_title")

    where_parts = []
    binds = {}

    if HAS("serverlogin") and login:
        where_parts.append("LOWER(TRIM(s.serverlogin)) = :b_login")
        binds["b_login"] = (login or "").strip().lower()

    if HAS("inn") and inn:
        where_parts.append("TRIM(s.inn) = :b_inn")
        binds["b_inn"] = (inn or "").strip()

    if not where_parts:
        return []

    sql = f"""
        SELECT {", ".join(select_parts)}
        FROM smstaff s
        WHERE {" OR ".join(where_parts)}
        ORDER BY s.id
    """

    cur.execute(sql, binds)
    return _oracle_rows_to_jsonable(cur)




# def _oracle_fetch_smstaff_by_login(cur, login: str) -> list[dict]:
#     """
#     Читаем пользователя по логину.
#     Безопасно для баз, где нет части колонок.
#     """
#     cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")

#     def HAS(col: str) -> bool:
#         return col.upper() in cols_set

#     if not HAS("serverlogin"):
#         return []

#     select_parts = [
#         "s.id AS id" if HAS("id") else "NULL AS id",
#         "s.surname AS surname" if HAS("surname") else "NULL AS surname",
#         "s.name AS name" if HAS("name") else "NULL AS name",
#         "s.patronymic AS patronymic" if HAS("patronymic") else "NULL AS patronymic",
#         "s.serverlogin AS serverlogin" if HAS("serverlogin") else "NULL AS serverlogin",
#         "s.inn AS inn" if HAS("inn") else "NULL AS inn",
#         "s.userenabled AS userenabled" if HAS("userenabled") else "NULL AS userenabled",
#         "s.offindex AS offindex" if HAS("offindex") else "NULL AS offindex",
#     ]

#     if HAS("offindex"):
#         select_parts.append("""
#             (
#                 SELECT c.title
#                 FROM smoffcfg c
#                 WHERE c.id = s.offindex
#                   AND ROWNUM = 1
#             ) AS off_title
#         """)
#     else:
#         select_parts.append("NULL AS off_title")

#     sql = f"""
#         SELECT {", ".join(select_parts)}
#         FROM smstaff s
#         WHERE LOWER(TRIM(s.serverlogin)) = :b_login
#         ORDER BY s.id
#     """

#     cur.execute(sql, b_login=(login or "").strip().lower())
#     return _oracle_rows_to_jsonable(cur)


def _oracle_fetch_smstaff_by_login(cur, login: str) -> list[dict]:
    """
    Читаем пользователя по логину.
    Возвращаем AFIO, а не surname/name/patronymic.
    """
    cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")

    def HAS(col: str) -> bool:
        return col.upper() in cols_set

    if not HAS("serverlogin"):
        return []

    select_parts = [
        "s.id AS id" if HAS("id") else "NULL AS id",
        _smstaff_afio_sql("s", cols_set),
        "s.serverlogin AS serverlogin" if HAS("serverlogin") else "NULL AS serverlogin",
        "s.inn AS inn" if HAS("inn") else "NULL AS inn",
        "s.userenabled AS userenabled" if HAS("userenabled") else "NULL AS userenabled",
        "s.offindex AS offindex" if HAS("offindex") else "NULL AS offindex",
    ]

    if HAS("offindex"):
        select_parts.append("""
            (
                SELECT c.title
                FROM smoffcfg c
                WHERE c.id = s.offindex
                  AND ROWNUM = 1
            ) AS off_title
        """)
    else:
        select_parts.append("NULL AS off_title")

    sql = f"""
        SELECT {", ".join(select_parts)}
        FROM smstaff s
        WHERE LOWER(TRIM(s.serverlogin)) = :b_login
        ORDER BY s.id
    """

    cur.execute(sql, b_login=(login or "").strip().lower())
    return _oracle_rows_to_jsonable(cur)





















@require_http_methods(["GET"])
def sm_staff_ui_list(request):
    """
    UI список SMSTAFF по выбранной базе:
    GET /ui/smstaff/?db=BINUU00&login=ivan&inn=123&only_enabled=1&page=1&page_size=50
    """
    services = _ui_services_list()
    db = (request.GET.get("db") or "").strip().upper()
    if not db:
        db = services[0] if services else "BINUU00"

    if not _is_allowed_service(db) or db not in ORACLE_TNS_MAP:
        return render(request, "frostapp/smstaff_list.html", {
            "services": services,
            "db": db,
            "error": f"Неизвестная база db={db!r}. Добавь её в ORACLE_TNS_MAP.",
            "items": [],
        })

    login = (request.GET.get("login") or "").strip().lower()
    inn = (request.GET.get("inn") or "").strip()
    only_enabled = (request.GET.get("only_enabled") or "").strip().lower() in ("1", "true", "yes", "on")

    try:
        page = int(request.GET.get("page", "1"))
    except ValueError:
        page = 1

    try:
        page_size = int(request.GET.get("page_size", "50"))
    except ValueError:
        page_size = 50

    page = max(1, page)
    page_size = max(10, min(200, page_size))
    offset = (page - 1) * page_size

    conn = cur = None
    items = []
    total = 0
    cols_present = []

    try:
        conn = _connect_oracle_service(db)
        cur = conn.cursor()

        cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")
        cols_present = sorted(list(cols_set))

        def SEL(alias: str, col: str) -> str:
            return f"{alias}.{col} AS {col}" if _ui_has_col(cols_set, col) else f"NULL AS {col}"

        select_parts = [
            SEL("s", "id"),
            _smstaff_afio_sql("s", cols_set),
            SEL("s", "serverlogin"),
            SEL("s", "inn"),
            SEL("s", "userenabled"),
        ]

        if _ui_has_col(cols_set, "offindex"):
            select_parts.append("s.offindex AS offindex")
            select_parts.append("""
                (
                    SELECT c.title
                    FROM smoffcfg c
                    WHERE c.id = s.offindex
                      AND ROWNUM = 1
                ) AS off_title
            """)
        else:
            select_parts.append("NULL AS offindex")
            select_parts.append("NULL AS off_title")

        select_sql = ", ".join(select_parts)

        where = []
        binds = {}

        if only_enabled and _ui_has_col(cols_set, "userenabled"):
            where.append("(s.userenabled = '1' OR s.userenabled = 1)")

        if login and _ui_has_col(cols_set, "serverlogin"):
            where.append("LOWER(s.serverlogin) LIKE :b_login")
            binds["b_login"] = f"%{login}%"

        if inn and _ui_has_col(cols_set, "inn"):
            where.append("TRIM(s.inn) LIKE :b_inn")
            binds["b_inn"] = f"%{inn}%"

        base_from = "FROM smstaff s"
        if where:
            base_from += " WHERE " + " AND ".join(where)

        cur.execute(f"SELECT COUNT(*) {base_from}", binds)
        total = int(cur.fetchone()[0])

        # ВАЖНО:
        # сортировка должна идти по алиасам полей из подзапроса t,
        # а не по s.serverlogin / s.afio
        if _ui_has_col(cols_set, "afio"):
            order_expr = "afio"
        elif _ui_has_col(cols_set, "serverlogin"):
            order_expr = "serverlogin"
        else:
            order_expr = "id"

        sql = f"""
            SELECT *
            FROM (
                SELECT t.*, ROW_NUMBER() OVER (ORDER BY {order_expr}) rn
                FROM (
                    SELECT {select_sql}
                    {base_from}
                ) t
            )
            WHERE rn > :b_off
              AND rn <= :b_to
        """

        binds2 = dict(binds)
        binds2["b_off"] = offset
        binds2["b_to"] = offset + page_size

        cur.execute(sql, binds2)
        items = _oracle_rows_to_jsonable(cur)

    except Exception as e:
        logger.exception(f"[UI/SMSTAFF] list error db={db}: {e}")
        return render(request, "frostapp/smstaff_list.html", {
            "services": services,
            "db": db,
            "error": str(e),
            "items": [],
        })
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass

    pages = max(1, (total + page_size - 1) // page_size)

    def page_url(p: int) -> str:
        if p < 1 or p > pages:
            return ""
        q = {
            "db": db,
            "login": login,
            "inn": inn,
            "only_enabled": "1" if only_enabled else "",
            "page_size": str(page_size),
            "page": str(p),
        }
        q = {k: v for k, v in q.items() if v not in ("", None)}
        return f"{reverse('sm_staff_ui_list')}?{urlencode(q)}"

    first_page_url = page_url(1) if pages > 0 else ""
    prev_page_url = page_url(page - 1) if page > 1 else ""
    next_page_url = page_url(page + 1) if page < pages else ""
    last_page_url = page_url(pages) if pages > 0 else ""

    return render(request, "frostapp/smstaff_list.html", {
        "services": services,
        "db": db,
        "login": login,
        "inn": inn,
        "only_enabled": only_enabled,
        "page": page,
        "page_size": page_size,
        "pages": pages,
        "total": total,
        "items": items,
        "cols_present": cols_present,
        "page_url": page_url,
        "first_page_url": first_page_url,
        "prev_page_url": prev_page_url,
        "next_page_url": next_page_url,
        "last_page_url": last_page_url,
    })
# @require_http_methods(["GET"])
# def sm_staff_ui_list(request):
#     """
#     UI список SMSTAFF по выбранной базе:
#     GET /ui/smstaff/?db=BINUU00&login=ivan&inn=123&only_enabled=1&page=1&page_size=50
#     """
#     services = _ui_services_list()
#     db = (request.GET.get("db") or "").strip().upper()
#     if not db:
#         db = services[0] if services else "BINUU00"

#     if not _is_allowed_service(db) or db not in ORACLE_TNS_MAP:
#         return render(request, "frostapp/smstaff_list.html", {
#             "services": services,
#             "db": db,
#             "error": f"Неизвестная база db={db!r}. Добавь её в ORACLE_TNS_MAP.",
#             "items": [],
#         })

#     login = (request.GET.get("login") or "").strip().lower()
#     inn = (request.GET.get("inn") or "").strip()
#     only_enabled = (request.GET.get("only_enabled") or "").strip().lower() in ("1", "true", "yes", "on")

#     try:
#         page = int(request.GET.get("page", "1"))
#     except ValueError:
#         page = 1

#     try:
#         page_size = int(request.GET.get("page_size", "50"))
#     except ValueError:
#         page_size = 50

#     page = max(1, page)
#     page_size = max(10, min(200, page_size))
#     offset = (page - 1) * page_size

#     conn = cur = None
#     items = []
#     total = 0
#     cols_present = []

#     try:
#         conn = _connect_oracle_service(db)
#         cur = conn.cursor()

#         cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")
#         cols_present = sorted(list(cols_set))

#         def SEL(alias: str, col: str) -> str:
#             return f"{alias}.{col} AS {col}" if _ui_has_col(cols_set, col) else f"NULL AS {col}"

#         select_parts = [
#             SEL("s", "id"),
#             SEL("s", "surname"),
#             SEL("s", "name"),
#             SEL("s", "patronymic"),
#             SEL("s", "serverlogin"),
#             SEL("s", "inn"),
#             SEL("s", "userenabled"),
#         ]

#         if _ui_has_col(cols_set, "offindex"):
#             select_parts.append("s.offindex AS offindex")
#             select_parts.append("""
#                 (
#                     SELECT c.title
#                     FROM smoffcfg c
#                     WHERE c.id = s.offindex
#                       AND ROWNUM = 1
#                 ) AS off_title
#             """)
#         else:
#             select_parts.append("NULL AS offindex")
#             select_parts.append("NULL AS off_title")

#         select_sql = ", ".join(select_parts)

#         where = []
#         binds = {}

#         if only_enabled and _ui_has_col(cols_set, "userenabled"):
#             where.append("(s.userenabled = '1' OR s.userenabled = 1)")

#         if login and _ui_has_col(cols_set, "serverlogin"):
#             where.append("LOWER(s.serverlogin) LIKE :b_login")
#             binds["b_login"] = f"%{login}%"

#         if inn and _ui_has_col(cols_set, "inn"):
#             where.append("TRIM(s.inn) LIKE :b_inn")
#             binds["b_inn"] = f"%{inn}%"

#         base_from = "FROM smstaff s"
#         if where:
#             base_from += " WHERE " + " AND ".join(where)

#         cur.execute(f"SELECT COUNT(*) {base_from}", binds)
#         total = int(cur.fetchone()[0])

#         order_expr = "surname" if _ui_has_col(cols_set, "surname") else "id"

#         sql = f"""
#             SELECT *
#             FROM (
#                 SELECT t.*, ROW_NUMBER() OVER (ORDER BY {order_expr}) rn
#                 FROM (
#                     SELECT {select_sql}
#                     {base_from}
#                 ) t
#             )
#             WHERE rn > :b_off
#               AND rn <= :b_to
#         """

#         binds2 = dict(binds)
#         binds2["b_off"] = offset
#         binds2["b_to"] = offset + page_size

#         cur.execute(sql, binds2)
#         items = _oracle_rows_to_jsonable(cur)

#     except Exception as e:
#         logger.exception(f"[UI/SMSTAFF] list error db={db}: {e}")
#         return render(request, "frostapp/smstaff_list.html", {
#             "services": services,
#             "db": db,
#             "error": str(e),
#             "items": [],
#         })
#     finally:
#         try:
#             if cur:
#                 cur.close()
#             if conn:
#                 conn.close()
#         except Exception:
#             pass

#     pages = max(1, (total + page_size - 1) // page_size)

#     def page_url(p: int) -> str:
#         if p < 1 or p > pages:
#             return ""
#         q = {
#             "db": db,
#             "login": login,
#             "inn": inn,
#             "only_enabled": "1" if only_enabled else "",
#             "page_size": str(page_size),
#             "page": str(p),
#         }
#         q = {k: v for k, v in q.items() if v not in ("", None)}
#         return f"{reverse('sm_staff_ui_list')}?{urlencode(q)}"

#     first_page_url = page_url(1) if pages > 0 else ""
#     prev_page_url = page_url(page - 1) if page > 1 else ""
#     next_page_url = page_url(page + 1) if page < pages else ""
#     last_page_url = page_url(pages) if pages > 0 else ""

#     return render(request, "frostapp/smstaff_list.html", {
#         "services": services,
#         "db": db,
#         "login": login,
#         "inn": inn,
#         "only_enabled": only_enabled,
#         "page": page,
#         "page_size": page_size,
#         "pages": pages,
#         "total": total,
#         "items": items,
#         "cols_present": cols_present,
#         "page_url": page_url,
#         "first_page_url": first_page_url,
#         "prev_page_url": prev_page_url,
#         "next_page_url": next_page_url,
#         "last_page_url": last_page_url,
#     })


# @require_http_methods(["GET", "POST"])
# @csrf_protect
# def sm_staff_ui_create(request):
#     """
#     UI создание пользователя в SMSTAFF через процедуру bin_createuser
#     и последующая установка OFFINDEX по выбранной должности из SMOFFCFG.

#     Важно:
#     - в процедуру передаётся SMOFFCFG.ORAROLE
#     - в SMSTAFF.OFFINDEX записывается выбранный SMOFFCFG.ID
#     - пароль генерируется автоматически (4 цифры) и передаётся в шаблон
#     """
#     services = _ui_services_list()

#     if request.method == "POST":
#         db = (request.POST.get("db") or "").strip().upper()
#     else:
#         db = (request.GET.get("db") or "").strip().upper()

#     if not db:
#         db = services[0] if services else "BINUU00"

#     form = {
#         "auser": "",
#         "ainn": "",
#         "adol": "",
#         "acheck": "1",
#         "auserenabled": "1",
#     }

#     positions = []
#     conflicts = []
#     created_rows = []
#     created_password = ""
#     success_message = ""
#     error = ""

#     if request.method == "POST":
#         form["auser"] = (request.POST.get("auser") or "").strip().lower()
#         form["ainn"] = (request.POST.get("ainn") or "").strip()
#         form["adol"] = (request.POST.get("adol") or "").strip()
#         form["acheck"] = (request.POST.get("acheck") or "1").strip()
#         form["auserenabled"] = (request.POST.get("auserenabled") or "1").strip()

#     if not _is_allowed_service(db) or db not in ORACLE_TNS_MAP:
#         return render(request, "frostapp/smstaff_create.html", {
#             "services": services,
#             "db": db,
#             "form": form,
#             "positions": [],
#             "acheck_choices": _SM_ACHECK_CHOICES,
#             "auserenabled_choices": _SM_USERENABLED_CHOICES,
#             "error": f"Неизвестная база db={db!r}. Добавь её в ORACLE_TNS_MAP.",
#             "conflicts": [],
#             "created_rows": [],
#             "created_password": "",
#             "success_message": "",
#             "back_url": f"{reverse('sm_staff_ui_list')}?{urlencode({'db': db})}",
#         })

#     conn = cur = None
#     try:
#         conn = _connect_oracle_service(db)
#         cur = conn.cursor()

#         positions = _oracle_fetch_smoffcfg_choices(cur)
#         positions_map = {str(x["id"]): x for x in positions}

#         if request.method == "POST":
#             auser = form["auser"]
#             ainn = form["ainn"]
#             adol_id = form["adol"]
#             acheck_raw = form["acheck"]
#             auserenabled_raw = form["auserenabled"]

#             if not auser:
#                 error = "Заполни логин (AUSER)."

#             elif not _is_valid_sm_login(auser):
#                 error = (
#                     "Логин может содержать только латинские буквы, цифры, "
#                     "точку, дефис и нижнее подчёркивание."
#                 )

#             elif not ainn:
#                 error = "Заполни ИНН (AINN)."

#             elif not _is_valid_inn_digits(ainn):
#                 error = "ИНН должен состоять из 10 или 12 цифр."

#             elif adol_id not in positions_map:
#                 error = "Выбери должность из списка."

#             elif acheck_raw not in {x["value"] for x in _SM_ACHECK_CHOICES}:
#                 error = "Некорректное значение ACHECK."

#             elif auserenabled_raw not in {x["value"] for x in _SM_USERENABLED_CHOICES}:
#                 error = "Некорректное значение AUSERENABLED."

#             else:
#                 conflicts = _oracle_fetch_smstaff_conflicts(cur, auser, ainn)

#                 exact_exists = False
#                 for row in conflicts:
#                     row_login = (row.get("serverlogin") or "").strip().lower()
#                     row_inn = (row.get("inn") or "").strip()
#                     if row_login == auser and row_inn == ainn:
#                         exact_exists = True
#                         break

#                 if exact_exists:
#                     error = f"Пользователь с логином {auser} и ИНН {ainn} уже существует в базе {db}."
#                 else:
#                     selected_pos = positions_map[adol_id]

#                     adol_title = (selected_pos.get("title") or "").strip()
#                     adol_orarole = (selected_pos.get("orarole") or "").strip()
#                     adol_offindex = int(selected_pos["id"])

#                     acheck = int(acheck_raw)
#                     auserenabled = int(auserenabled_raw)

#                     if not adol_orarole:
#                         error = (
#                             f"Для выбранной должности "
#                             f"'{adol_title or adol_offindex}' "
#                             f"в SMOFFCFG не заполнено поле ORAROLE."
#                         )
#                     else:
#                         created_password = _generate_sm_password(4)

#                         proc_name = os.getenv("SM_BIN_CREATEUSER_PROC", "bin_createuser").strip()
#                         if not re.fullmatch(r"[A-Za-z0-9_.$]+", proc_name):
#                             raise ValueError("Некорректное имя процедуры в SM_BIN_CREATEUSER_PROC.")

#                         logger.info(
#                             f"[UI/SMSTAFF_CREATE] start db={db} login={auser} inn={ainn} "
#                             f"offindex={adol_offindex} title={adol_title} orarole={adol_orarole} "
#                             f"acheck={acheck} auserenabled={auserenabled}"
#                         )

#                         cur.execute(
#                             f"""
#                             BEGIN
#                                 {proc_name}(
#                                     AUSER => :p_auser,
#                                     APASS => :p_apass,
#                                     ADOL => :p_adol,
#                                     ACHECK => :p_acheck,
#                                     AUSERENABLED => :p_auserenabled,
#                                     AINN => :p_ainn
#                                 );
#                             END;
#                             """,
#                             p_auser=auser,
#                             p_apass=created_password,
#                             p_adol=adol_orarole,
#                             p_acheck=acheck,
#                             p_auserenabled=auserenabled,
#                             p_ainn=ainn,
#                         )

#                         if acheck != 2:
#                             cur.execute("""
#                                 UPDATE smstaff
#                                    SET offindex = :b_offindex
#                                  WHERE LOWER(TRIM(serverlogin)) = :b_login
#                             """, b_offindex=adol_offindex, b_login=auser)

#                         conn.commit()

#                         created_rows = _oracle_fetch_smstaff_by_login(cur, auser)

#                         if acheck == 2:
#                             success_message = (
#                                 f"Процедура выполнена для логина {auser}. "
#                                 f"Режим ACHECK=2 означает удаление пользователя."
#                             )
#                         else:
#                             success_message = (
#                                 f"Пользователь успешно обработан в базе {db}. "
#                                 f"Сгенерированный пароль: {created_password}"
#                             )

#                         logger.info(
#                             f"[UI/SMSTAFF_CREATE] success db={db} login={auser} "
#                             f"created_rows={len(created_rows)} offindex={adol_offindex} "
#                             f"orarole={adol_orarole}"
#                         )

#     except Exception as e:
#         logger.exception(f"[UI/SMSTAFF_CREATE] error db={db}: {e}")

#         err_text = str(e)
#         if "ORA-01403" in err_text and "BIN_CREATEUSER" in err_text:
#             error = (
#                 "Oracle-процедура BIN_CREATEUSER не смогла найти обязательные данные "
#                 "для создания пользователя. Вероятно, в выбранной базе отсутствует "
#                 "связанная запись, которую ожидает процедура."
#             )
#         else:
#             error = err_text

#         try:
#             if conn:
#                 conn.rollback()
#         except Exception:
#             pass

#     finally:
#         try:
#             if cur:
#                 cur.close()
#             if conn:
#                 conn.close()
#         except Exception:
#             pass

#     return render(request, "frostapp/smstaff_create.html", {
#         "services": services,
#         "db": db,
#         "form": form,
#         "positions": positions,
#         "acheck_choices": _SM_ACHECK_CHOICES,
#         "auserenabled_choices": _SM_USERENABLED_CHOICES,
#         "error": error,
#         "conflicts": conflicts,
#         "created_rows": created_rows,
#         "created_password": created_password,
#         "success_message": success_message,
#         "back_url": f"{reverse('sm_staff_ui_list')}?{urlencode({'db': db})}",
#     })

@require_http_methods(["GET", "POST"])
@csrf_protect
def sm_staff_ui_create(request):
    """
    UI создание/обновление пользователя в SMSTAFF через процедуру bin_createuser.

    Правила:
    - обязательный только AUSER
    - AINN / AFIO / ADOL не обязательны на уровне Django
    - проверки конфликтов по логину/ИНН выполняются только для сценария создания
    - режим блокировки для текущего кейса:
        ACHECK = 0
        AUSERENABLED = 0
    - если ADOL не выбран, в процедуру уходит NULL
    - если AINN / AFIO пустые, в процедуру уходит NULL
    """
    services = _ui_services_list()

    if request.method == "POST":
        db = (request.POST.get("db") or "").strip().upper()
    else:
        db = (request.GET.get("db") or "").strip().upper()

    if not db:
        db = services[0] if services else "BINUU00"

    form = {
        "auser": "",
        "ainn": "",
        "afio": "",
        "adol": "",
        "acheck": "1",
        "auserenabled": "1",
    }

    positions = []
    conflicts = []
    created_rows = []
    created_password = ""
    success_message = ""
    error = ""

    if request.method == "POST":
        form["auser"] = (request.POST.get("auser") or "").strip().lower()
        form["ainn"] = (request.POST.get("ainn") or "").strip()
        form["afio"] = (request.POST.get("afio") or "").strip()
        form["adol"] = (request.POST.get("adol") or "").strip()
        form["acheck"] = (request.POST.get("acheck") or "1").strip()
        form["auserenabled"] = (request.POST.get("auserenabled") or "1").strip()

    if not _is_allowed_service(db) or db not in ORACLE_TNS_MAP:
        return render(request, "frostapp/smstaff_create.html", {
            "services": services,
            "db": db,
            "form": form,
            "positions": [],
            "acheck_choices": _SM_ACHECK_CHOICES,
            "auserenabled_choices": _SM_USERENABLED_CHOICES,
            "error": f"Неизвестная база db={db!r}. Добавь её в ORACLE_TNS_MAP.",
            "conflicts": [],
            "created_rows": [],
            "created_password": "",
            "success_message": "",
            "back_url": f"{reverse('sm_staff_ui_list')}?{urlencode({'db': db})}",
        })

    conn = cur = None
    try:
        conn = _connect_oracle_service(db)
        cur = conn.cursor()

        positions = _oracle_fetch_smoffcfg_choices(cur)
        positions_map = {str(x["id"]): x for x in positions}

        if request.method == "POST":
            auser = form["auser"]
            ainn = form["ainn"]
            afio = form["afio"]
            adol_id = form["adol"]
            acheck_raw = form["acheck"]
            auserenabled_raw = form["auserenabled"]

            allowed_acheck_values = {x["value"] for x in _SM_ACHECK_CHOICES}
            allowed_auserenabled_values = {x["value"] for x in _SM_USERENABLED_CHOICES}

            is_block_mode = (acheck_raw == "0" and auserenabled_raw == "0")
            is_delete_mode = (acheck_raw == "2")

            # Считаем "созданием" только сценарий, когда реально передают данные
            # для новой УЗ и это не блокировка/не удаление.
            is_creation_mode = (
                not is_block_mode
                and not is_delete_mode
                and acheck_raw in {"0", "1"}
                and bool(ainn)
                and bool(adol_id)
            )

            if not auser:
                error = "Заполни логин (AUSER)."

            elif not _is_valid_sm_login(auser):
                error = (
                    "Логин может содержать только латинские буквы, цифры, "
                    "точку, дефис и нижнее подчёркивание."
                )

            elif acheck_raw not in allowed_acheck_values:
                error = "Некорректное значение ACHECK."

            elif auserenabled_raw not in allowed_auserenabled_values:
                error = "Некорректное значение AUSERENABLED."

            elif ainn and not _is_valid_inn_digits(ainn):
                error = "ИНН должен состоять из 10 или 12 цифр."

            elif afio and len(afio) > 255:
                error = "ФИО (AFIO) слишком длинное. Максимум 255 символов."

            elif adol_id and adol_id not in positions_map:
                error = "Выбери корректную должность из списка."

            else:
                selected_pos = positions_map.get(adol_id) if adol_id else None

                adol_title = (selected_pos.get("title") or "").strip() if selected_pos else ""
                adol_orarole = (selected_pos.get("orarole") or "").strip() if selected_pos else None
                adol_offindex = int(selected_pos["id"]) if selected_pos else None

                # Для сценария создания должность должна иметь ORAROLE
                if is_creation_mode and not adol_orarole:
                    error = (
                        f"Для выбранной должности "
                        f"'{adol_title or adol_offindex}' "
                        f"в SMOFFCFG не заполнено поле ORAROLE."
                    )

                exact_exists = False
                login_taken_by_other = False
                inn_taken_by_other = False

                # Проверки конфликтов — только для создания
                if not error and is_creation_mode:
                    conflicts = _oracle_fetch_smstaff_conflicts(cur, auser, ainn)

                    for row in conflicts:
                        row_login = (row.get("serverlogin") or "").strip().lower()
                        row_inn = (row.get("inn") or "").strip()

                        if row_login == auser and row_inn == ainn:
                            exact_exists = True
                        elif row_login == auser and row_inn != ainn:
                            login_taken_by_other = True
                        elif row_inn == ainn and row_login != auser:
                            inn_taken_by_other = True

                    if login_taken_by_other:
                        error = (
                            f"В базе {db} уже есть пользователь с логином {auser}, "
                            f"но с другим ИНН. Проверь данные."
                        )
                    elif inn_taken_by_other:
                        error = (
                            f"В базе {db} уже есть пользователь с ИНН {ainn}, "
                            f"но с другим логином. Проверь данные."
                        )

                if not error:
                    acheck = int(acheck_raw)
                    auserenabled = int(auserenabled_raw)

                    created_password = _generate_sm_password(4)

                    proc_name = os.getenv("SM_BIN_CREATEUSER_PROC", "bin_createuser").strip()
                    if not re.fullmatch(r"[A-Za-z0-9_.$]+", proc_name):
                        raise ValueError("Некорректное имя процедуры в SM_BIN_CREATEUSER_PROC.")

                    proc_ainn = ainn or None
                    proc_afio = afio or None
                    proc_adol = adol_orarole or None

                    logger.info(
                        f"[UI/SMSTAFF_CREATE] start db={db} login={auser} inn={proc_ainn!r} "
                        f"afio={proc_afio!r} offindex={adol_offindex} title={adol_title!r} "
                        f"orarole={proc_adol!r} acheck={acheck} "
                        f"auserenabled={auserenabled} exact_exists={exact_exists} "
                        f"is_block_mode={is_block_mode} is_delete_mode={is_delete_mode} "
                        f"is_creation_mode={is_creation_mode}"
                    )

                    cur.execute(
                        f"""
                        BEGIN
                            {proc_name}(
                                AUSER => :p_auser,
                                APASS => :p_apass,
                                ADOL => :p_adol,
                                ACHECK => :p_acheck,
                                AUSERENABLED => :p_auserenabled,
                                AINN => :p_ainn,
                                AFIO => :p_afio
                            );
                        END;
                        """,
                        p_auser=auser,
                        p_apass=created_password,
                        p_adol=proc_adol,
                        p_acheck=acheck,
                        p_auserenabled=auserenabled,
                        p_ainn=proc_ainn,
                        p_afio=proc_afio,
                    )

                    updated_smstaff_rows = 0

                    # Обновляем AFIO/OFFINDEX только если реально что-то передано
                    if not is_delete_mode and (proc_afio or adol_offindex is not None):
                        updated_smstaff_rows = _oracle_update_smstaff_after_create(
                            cur=cur,
                            login=auser,
                            inn=(proc_ainn or ""),
                            afio=(proc_afio or ""),
                            offindex=adol_offindex,
                        )

                    conn.commit()

                    created_rows = _oracle_fetch_smstaff_by_login(cur, auser)

                    if is_block_mode:
                        success_message = (
                            f"Для пользователя {auser} выполнена блокировка "
                            f"(ACHECK=0, AUSERENABLED=0). "
                            f"Принудительно обновлено строк в SMSTAFF: {updated_smstaff_rows}."
                        )
                    elif is_delete_mode:
                        success_message = (
                            f"Процедура выполнена для логина {auser}. "
                            f"Режим ACHECK=2 означает удаление пользователя."
                        )
                    elif exact_exists and acheck == 1:
                        success_message = (
                            f"Для существующего пользователя {auser} выполнено обновление. "
                            f"Новый пароль: {created_password}. "
                            f"Принудительно обновлено строк в SMSTAFF: {updated_smstaff_rows}."
                        )
                    elif exact_exists and acheck == -1:
                        success_message = (
                            f"Для существующего пользователя {auser} выполнено обновление "
                            f"без создания новой записи. "
                            f"Принудительно обновлено строк в SMSTAFF: {updated_smstaff_rows}."
                        )
                    elif exact_exists and acheck == 0:
                        success_message = (
                            f"Пользователь {auser} уже существовал. "
                            f"Процедура выполнена в режиме ACHECK=0. "
                            f"Принудительно обновлено строк в SMSTAFF: {updated_smstaff_rows}."
                        )
                    else:
                        success_message = (
                            f"Пользователь успешно обработан в базе {db}. "
                            f"Сгенерированный пароль: {created_password}. "
                            f"Принудительно обновлено строк в SMSTAFF: {updated_smstaff_rows}."
                        )

                    logger.info(
                        f"[UI/SMSTAFF_CREATE] success db={db} login={auser} "
                        f"created_rows={len(created_rows)} offindex={adol_offindex} "
                        f"orarole={proc_adol!r} exact_exists={exact_exists} "
                        f"updated_smstaff_rows={updated_smstaff_rows} "
                        f"is_block_mode={is_block_mode} is_delete_mode={is_delete_mode} "
                        f"is_creation_mode={is_creation_mode} afio={proc_afio!r}"
                    )

    except Exception as e:
        logger.exception(f"[UI/SMSTAFF_CREATE] error db={db}: {e}")
        error = str(e)
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass

    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass

    return render(request, "frostapp/smstaff_create.html", {
        "services": services,
        "db": db,
        "form": form,
        "positions": positions,
        "acheck_choices": _SM_ACHECK_CHOICES,
        "auserenabled_choices": _SM_USERENABLED_CHOICES,
        "error": error,
        "conflicts": conflicts,
        "created_rows": created_rows,
        "created_password": created_password,
        "success_message": success_message,
        "back_url": f"{reverse('sm_staff_ui_list')}?{urlencode({'db': db})}",
    })

























































@require_http_methods(["GET", "POST"])
@csrf_protect
def sm_staff_ui_edit_inn(request, db: str, staff_id: str):
    """
    UI редактирование ИНН:
    GET/POST /ui/smstaff/<db>/edit/<staff_id>/
    staff_id в urls.py допускает отрицательные (re_path), но мы их отвергаем валидацией.
    """

    db = (db or "").strip().upper()

    try:
        staff_id_int = int(staff_id)
    except (ValueError, TypeError):
        return render(request, "frostapp/smstaff_edit.html", {
            "db": db,
            "staff_id": staff_id,
            "row": None,
            "ok": False,
            "error": f"Некорректный ID пользователя: {staff_id}. ID должен быть числом.",
        })

    if staff_id_int <= 0:
        return render(request, "frostapp/smstaff_edit.html", {
            "db": db,
            "staff_id": staff_id_int,
            "row": None,
            "ok": False,
            "error": f"Некорректный ID пользователя: {staff_id_int}. ID должен быть положительным числом.",
        })

    # validate db against map/allowlist
    if not _is_allowed_service(db) or db not in ORACLE_TNS_MAP:
        return render(request, "frostapp/smstaff_edit.html", {
            "db": db,
            "staff_id": staff_id_int,
            "row": None,
            "ok": False,
            "error": f"Неизвестная база db={db!r}. Добавь её в ORACLE_TNS_MAP.",
        })

    conn = cur = None
    row = None
    error = ""
    ok = False

    try:
        conn = _connect_oracle_service(db)
        cur = conn.cursor()

        cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")

        # Без INN редактировать нечего
        if not _ui_has_col(cols_set, "inn"):
            return render(request, "frostapp/smstaff_edit.html", {
                "db": db,
                "staff_id": staff_id_int,
                "row": None,
                "ok": False,
                "error": "В этой базе в SMSTAFF нет колонки INN — редактирование невозможно.",
            })

        def SEL(col: str) -> str:
            # если колонки нет — отдадим NULL AS col, чтобы шаблон не ломался
            return col if _ui_has_col(cols_set, col) else f"NULL AS {col}"

        select_sql = ", ".join([
            SEL("id"),
            SEL("surname"),
            SEL("name"),
            SEL("patronymic"),
            SEL("serverlogin"),
            SEL("inn"),
            SEL("userenabled"),
        ])

        # читаем строку
        cur.execute(f"""
            SELECT {select_sql}
            FROM smstaff
            WHERE id = :b_id
        """, b_id=staff_id_int)

        items = _oracle_rows_to_jsonable(cur)
        if not items:
            return render(request, "frostapp/smstaff_edit.html", {
                "db": db,
                "staff_id": staff_id_int,
                "row": None,
                "ok": False,
                "error": f"Пользователь с id={staff_id_int} не найден в {db}.",
            })

        row = items[0]  # dict с ключами lower-case

        if request.method == "POST":
            new_inn = (request.POST.get("inn") or "").strip()  

            # Валидация
            if new_inn and not _is_valid_inn_digits(new_inn):
                error = "ИНН должен быть числом длиной 10 или 12."
            else:
                login_val = (row.get("serverlogin") or "").strip()
                if not login_val:
                    error = "У выбранного пользователя пустой serverlogin — нельзя обновить по всем базам."
                else:
                    # Переходим на прогресс-страницу, которая сама пробежит все базы
                    dbs = _ui_services_list()
                    return render(request, "frostapp/smstaff_sync_inn.html", {
                        "start_db": db,
                        "login": login_val,
                        "new_inn": new_inn,  # "" => очистка (NULL)
                        "dbs": dbs,
                        "return_url": f"/ui/smstaff/?db={db}",
                    })

    except Exception as e:
        logger.exception(f"[UI/SMSTAFF] edit error db={db} id={staff_id_int}: {e}")
        error = str(e)
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass

    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass

    return render(request, "frostapp/smstaff_edit.html", {
        "db": db,
        "staff_id": staff_id_int,
        "row": row,
        "error": error,
        "ok": ok,
    })


@require_http_methods(["POST"])
@csrf_protect
def sm_staff_ui_sync_inn_one(request):
    """
    JSON: обновляет INN по serverlogin в ОДНОЙ базе.
    Вызывается со страницы прогресса по очереди для каждой базы.
    body: {"db":"BINUU00","login":"i_ivanov","inn":"123..."}  # inn может быть "" для очистки
    """
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return JsonResponse({"ok": False, "error": "Некорректный JSON"}, status=400)

    db = (payload.get("db") or "").strip().upper()
    login = (payload.get("login") or "").strip()
    inn = (payload.get("inn") or "")
    inn = (inn.strip() if isinstance(inn, str) else "")

    if not db or db not in ORACLE_TNS_MAP or not _is_allowed_service(db):
        return JsonResponse({"ok": False, "db": db, "error": "Неизвестная/запрещённая база"}, status=400)

    if not login:
        return JsonResponse({"ok": False, "db": db, "error": "Пустой login"}, status=400)

    if inn and not _is_valid_inn_digits(inn):
        return JsonResponse({"ok": False, "db": db, "error": "ИНН должен быть числом длиной 10 или 12"}, status=400)

    conn = cur = None
    try:
        conn = _connect_oracle_service(db)
        cur = conn.cursor()

        cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")

        if not _ui_has_col(cols_set, "inn"):
            return JsonResponse({"ok": False, "db": db, "error": "В SMSTAFF нет колонки INN"}, status=400)

        if not _ui_has_col(cols_set, "serverlogin"):
            return JsonResponse({"ok": False, "db": db, "error": "В SMSTAFF нет колонки SERVERLOGIN"}, status=400)

        # Обновляем по логину (без учёта регистра)
        if inn:
            cur.execute(
                """
                UPDATE smstaff
                   SET inn = :b_inn
                 WHERE LOWER(serverlogin) = LOWER(:b_login)
                """,
                b_inn=inn, b_login=login
            )
        else:
            cur.execute(
                """
                UPDATE smstaff
                   SET inn = NULL
                 WHERE LOWER(serverlogin) = LOWER(:b_login)
                """,
                b_login=login
            )

        affected = int(cur.rowcount or 0)
        conn.commit()

        # ok=True даже если affected=0 (просто не нашли пользователя в этой базе)
        level = "success" if affected > 0 else "warning"
        msg = "Обновлено" if affected > 0 else "Пользователь не найден (0 строк)"

        return JsonResponse({
            "ok": True,
            "db": db,
            "level": level,        # success | warning
            "affected": affected,
            "message": msg,
            "login": login,
            "inn": inn,
        })

    except Exception as e:
        logger.exception(f"[UI/SMSTAFF] sync inn one error db={db} login={login}: {e}")
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass
        return JsonResponse({"ok": False, "db": db, "error": str(e)}, status=500)

    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except Exception:
            pass













def _uac_is_disabled(uac_val) -> bool:
    try:
        uac = int(uac_val)
        return bool(uac & 0x0002)
    except Exception:
        return False


def _ad_decode_first(attrs: dict, key: str) -> str:
    vals = (attrs or {}).get(key, []) or []
    if not vals:
        return ""
    v = vals[0]
    if isinstance(v, bytes):
        return v.decode("utf-8", "ignore")
    return str(v)


def _norm_fio_for_match(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s


def _fio_from_api_row(row: dict) -> tuple[str, str, str]:
    # ключи русские, но на всякий случай подстрахуемся
    ln = (row.get("Фамилия") or row.get("lastname") or row.get("LastName") or "").strip()
    fn = (row.get("Имя") or row.get("firstname") or row.get("FirstName") or "").strip()
    sn = (row.get("Отчество") or row.get("patronymic") or row.get("SecondName") or "").strip()
    return ln, fn, sn


def _inn_from_api_row(row: dict) -> str:
    inn = (row.get("ИНН") or row.get("inn") or row.get("INN") or "").strip()
    inn = re.sub(r"\D+", "", inn)
    return inn


def _build_candidate_logins_dot(lastname: str, firstname: str, patronymic: str) -> list[str]:
    """
    Как _build_candidate_logins, но логины с точками.
    Плюс добавим ещё несколько частых комбинаций.
    """
    base = _build_candidate_logins(lastname, firstname, patronymic)  # вида i_ivanov, ivan_ivanov, ...
    out = set()

    # 1) заменяем '_' -> '.'
    for x in base:
        out.add(x.replace("_", "."))

    # 2) дополнительные варианты (на практике часто встречаются)
    ln = _normalize_login_piece(lastname)   # даст безопасный кусок
    fn = _normalize_login_piece(firstname)
    sn = _normalize_login_piece(patronymic)

    if fn and ln:
        out.add(f"{fn[0]}.{ln}")        # i.ivanov
        out.add(f"{fn}.{ln}")          # ivan.ivanov
        out.add(f"{ln}.{fn[0]}")       # ivanov.i
        out.add(f"{ln}.{fn}")          # ivanov.ivan
        out.add(f"{fn[0]}{ln}")        # iivanov (иногда)
        out.add(f"{fn}{ln}")           # ivanivanov (иногда)

        if sn:
            out.add(f"{fn[0]}.{sn[0]}.{ln}")   # i.i.ivanov
            out.add(f"{fn}.{sn[0]}.{ln}")      # ivan.i.ivanov
            out.add(f"{fn[0]}.{sn}.{ln}")      # i.ivanych.ivanov (редко, но бывает)

    # 3) подчистим пустое
    out = {x for x in out if x and "." in x or x}  # оставим даже без точки варианты
    return sorted(out)


def _ad_find_by_sam(conn, sam: str, attrs: list[str]) -> list[tuple[str, dict]]:
    f = f"(sAMAccountName={escape_filter_chars(sam)})"
    res = conn.search_s(AD_SEARCH_BASE, ldap.SCOPE_SUBTREE, f, attrlist=attrs)
    return [x for x in res if x and x[0]]


def _ad_set_employeeid(conn, user_dn: str, inn: str) -> tuple[bool, str]:
    """
    Возвращает (changed, message)
    """
    inn = re.sub(r"\D+", "", (inn or "").strip())
    if not _is_valid_inn_digits(inn):
        return False, "BAD_INN"

    # читаем текущий
    cur = conn.search_s(user_dn, ldap.SCOPE_BASE, "(objectClass=*)", attrlist=["employeeID"])
    cur = [x for x in cur if x and x[0]]
    current = ""
    if cur:
        current = _ad_decode_first(cur[0][1], "employeeID").strip()
        current = re.sub(r"\D+", "", current)

    if current == inn:
        return False, "UNCHANGED"

    conn.modify_s(user_dn, [(ldap.MOD_REPLACE, "employeeID", [inn.encode("utf-8")])])
    return True, f"UPDATED {current or '(empty)'} -> {inn}"


def _acquire_sync_lock() -> bool:
    try:
        fd = os.open(INN_SYNC_LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.write(fd, str(os.getpid()).encode("utf-8"))
        os.close(fd)
        return True
    except FileExistsError:
        return False
    except Exception:
        return False


def _release_sync_lock():
    try:
        if os.path.exists(INN_SYNC_LOCK_FILE):
            os.remove(INN_SYNC_LOCK_FILE)
    except Exception:
        pass


def _one_c_get_working_employees() -> list[dict]:
    r = requests.get(ONEC_WORKING_EMPLOYEES_URL, timeout=ONEC_WORKING_EMPLOYEES_TIMEOUT)
    r.raise_for_status()
    js = r.json()
    if not isinstance(js, list):
        raise RuntimeError("1C API вернул не список.")
    return js


@require_http_methods(["GET"])
@csrf_protect
def ldap_tools_home(request):
    return render(request, "frostapp/ldap_tools_home.html", {})


@require_http_methods(["GET"])
@csrf_protect
def ldap_tools_employees(request):
    """
    Показываем сотрудников из LDAP.
    Поиск теперь работает по:
      - sAMAccountName (логин)
      - displayName (ФИО)
    """
    q = (request.GET.get("q") or "").strip()
    page = int(request.GET.get("page") or "1")
    if page < 1:
        page = 1

    conn = None
    users = []
    error = ""

    try:
        conn = _ad_connect()

        # фильтр: пользователи (не компьютеры). Поиск по логину ИЛИ по displayName - подстрока.
        if q:
            q_esc = escape_filter_chars(q)
            flt = (
                "(&(objectClass=user)"
                "(!(objectClass=computer))"
                f"(|(sAMAccountName=*{q_esc}*)(displayName=*{q_esc}*)))"
            )
        else:
            flt = "(&(objectClass=user)(!(objectClass=computer)))"

        attrs = [
            "sAMAccountName", "displayName", "employeeID", "mail",
            "title", "department", "whenChanged", "userAccountControl",
            "distinguishedName",
        ]

        # Пэйджинг LDAP
        page_size = LDAP_EXPORT_PAGE_SIZE
        cookie = b""
        total = 0

        while True:
            ctrl = SimplePagedResultsControl(True, size=page_size, cookie=cookie)
            msgid = conn.search_ext(
                AD_SEARCH_BASE,
                ldap.SCOPE_SUBTREE,
                flt,
                attrlist=attrs,
                serverctrls=[ctrl],
            )
            rtype, rdata, rmsgid, serverctrls = conn.result3(msgid)

            for dn, at in rdata:
                if not dn:
                    continue
                total += 1
                if total > LDAP_EXPORT_MAX_TOTAL:
                    break

                sam = _ad_decode_first(at, "sAMAccountName")
                disp = _ad_decode_first(at, "displayName")
                emp = re.sub(r"\D+", "", _ad_decode_first(at, "employeeID"))
                mail = _ad_decode_first(at, "mail")
                title = _ad_decode_first(at, "title")
                dept = _ad_decode_first(at, "department")
                uac = _ad_decode_first(at, "userAccountControl")
                disabled = _uac_is_disabled(uac)

                users.append({
                    "dn": dn,
                    "sam": sam,
                    "displayName": disp,
                    "employeeID": emp,
                    "mail": mail,
                    "title": title,
                    "department": dept,
                    "disabled": disabled,
                })

            if total > LDAP_EXPORT_MAX_TOTAL:
                break

            cookie = b""
            for sc in serverctrls:
                if sc.controlType == SimplePagedResultsControl.controlType:
                    cookie = sc.cookie
                    break
            if not cookie:
                break

    except Exception as e:
        error = str(e)
    finally:
        try:
            if conn:
                conn.unbind_s()
        except Exception:
            pass

    # Пагинация уже на нашей стороне (users может быть большим)
    page_size = LDAP_TOOLS_PAGE_SIZE
    total = len(users)
    start = (page - 1) * page_size
    end = start + page_size
    page_users = users[start:end]
    pages = max(1, (total + page_size - 1) // page_size)

    return render(request, "frostapp/ldap_employees.html", {
        "error": error,
        "q": q,
        "users": page_users,
        "page": page,
        "pages": pages,
        "total": total,
        "page_size": page_size,
    })


@require_http_methods(["GET"])
@csrf_protect
def ldap_tools_sync_page(request, mode: str):
    mode = (mode or "").lower()
    if mode not in ("test", "apply"):
        return HttpResponse("BAD MODE", status=400)
    return render(request, "frostapp/ldap_sync.html", {
        "mode": mode,
        "is_apply": (mode == "apply"),
    })


@never_cache
@require_http_methods(["GET"])
def ldap_tools_sync_stream(request, mode: str):
    """
    SSE: отдаём прогресс строками.
    mode=test -> dry-run (не изменяем employeeID)
    mode=apply -> реально пишем employeeID
    """
    mode = (mode or "").lower()
    if mode not in ("test", "apply"):
        return HttpResponse("BAD MODE", status=400)

    apply_changes = (mode == "apply")

    def sse(line: str) -> str:
        line = (line or "").rstrip("\n")
        return f"data: {line}\n\n"

    def run():
        # lock чтобы два запуска не били AD одновременно
        if not _acquire_sync_lock():
            yield sse("❌ Уже идёт запуск синхронизации (lock). Повторите позже.")
            return

        ts = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S")
        filename = f"inn_sync_{mode}_{ts}.log"
        log_path = os.path.join(LOG_DIR, filename)

        def log_write(msg: str):
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(msg.rstrip("\n") + "\n")

        try:
            yield sse(f"🚀 Старт: mode={mode}")
            yield sse(f"📝 Лог: /ui/ldap-tools/sync/log/{filename}/")
            log_write(f"START mode={mode} at {ts}")

            # 1) грузим сотрудников из 1С
            yield sse("⏳ Загружаю список сотрудников из 1С...")
            rows = _one_c_get_working_employees()
            yield sse(f"✅ 1С вернул записей: {len(rows)}")
            log_write(f"1C rows: {len(rows)}")

            conn = None
            try:
                conn = _ad_connect()
                yield sse("✅ Подключение к LDAP установлено")
                log_write("LDAP connected")

                attrs = ["displayName", "employeeID", "sAMAccountName", "distinguishedName"]

                ok = 0
                skipped = 0
                notfound = 0
                ambiguous = 0
                errors = 0
                changed = 0
                unchanged = 0

                for i, row in enumerate(rows, start=1):
                    inn = _inn_from_api_row(row)
                    ln, fn, sn = _fio_from_api_row(row)

                    fio_api = " ".join([x for x in [ln, fn, sn] if x]).strip()
                    if not fio_api or not _is_valid_inn_digits(inn):
                        skipped += 1
                        msg = f"[{i}] SKIP: bad data fio='{fio_api}' inn='{inn}'"
                        yield sse(msg)
                        log_write(msg)
                        continue

                    # кандидаты логинов (полный список)
                    candidates = _build_candidate_logins_dot(ln, fn, sn)
                    candidates_str = ",".join(candidates)

                    # ищем в AD по логинам + проверяем displayName
                    matches = []
                    for sam in candidates:
                        try:
                            found = _ad_find_by_sam(conn, sam, attrs)
                            for dn, at in found:
                                disp = _ad_decode_first(at, "displayName")
                                sam_real = _ad_decode_first(at, "sAMAccountName")
                                matches.append((dn, at, sam_real, disp, sam))
                        except Exception:
                            continue

                    if not matches:
                        notfound += 1
                        msg = (
                            f"[{i}] NOT FOUND: {fio_api} inn={inn} "
                            f"candidates=[{candidates_str}]"
                        )
                        yield sse(msg)
                        log_write(msg)
                        continue

                    # фильтрация по displayName
                    target_disp = _norm_fio_for_match(fio_api)
                    exact = []
                    for dn, at, sam_real, disp, sam_try in matches:
                        if _norm_fio_for_match(disp) == target_disp:
                            exact.append((dn, at, sam_real, disp, sam_try))

                    # 🔎 для лога: какие логины реально нашлись в AD
                    found_sams = []
                    for dn, at, sam_real, disp, sam_try in matches:
                        found_sams.append(
                            f"{sam_real}(try={sam_try},displayName='{disp}')"
                        )
                    found_str = "; ".join(found_sams)

                    chosen = None
                    if len(exact) == 1:
                        chosen = exact[0]
                    elif len(exact) > 1:
                        ambiguous += 1
                        msg = (
                            f"[{i}] AMBIGUOUS(displayName): {fio_api} inn={inn} "
                            f"exact={len(exact)} found_total={len(matches)} "
                            f"found=[{found_str}] "
                            f"candidates=[{candidates_str}]"
                        )
                        yield sse(msg)
                        log_write(msg)
                        continue
                    else:
                        ambiguous += 1
                        msg = (
                            f"[{i}] AMBIGUOUS(no displayName match): {fio_api} inn={inn} "
                            f"found_total={len(matches)} "
                            f"found=[{found_str}] "
                            f"candidates=[{candidates_str}]"
                        )
                        yield sse(msg)
                        log_write(msg)
                        continue

                    dn, at, sam_real, disp, sam_try = chosen
                    cur_emp = re.sub(r"\D+", "", _ad_decode_first(at, "employeeID"))
                    info = (
                        f"[{i}] MATCH: {fio_api} -> {sam_real} (try={sam_try}) "
                        f"current_employeeID={cur_emp or '(empty)'} inn={inn}"
                    )
                    yield sse(info)
                    log_write(info)

                    # обновление
                    try:
                        if apply_changes:
                            ch, m = _ad_set_employeeid(conn, dn, inn)
                            if m == "UNCHANGED":
                                unchanged += 1
                                msg = f"[{i}] OK: UNCHANGED"
                            else:
                                if ch:
                                    changed += 1
                                msg = f"[{i}] OK: {m}"
                            yield sse(msg)
                            log_write(msg)
                        else:
                            # dry-run
                            if cur_emp == inn:
                                unchanged += 1
                                msg = f"[{i}] DRY-RUN: already OK"
                            else:
                                changed += 1
                                msg = f"[{i}] DRY-RUN: would update {cur_emp or '(empty)'} -> {inn}"
                            yield sse(msg)
                            log_write(msg)

                        ok += 1

                    except ldap.INSUFFICIENT_ACCESS:
                        errors += 1
                        msg = f"[{i}] ERROR: INSUFFICIENT_ACCESS (нет прав на employeeID) dn={dn}"
                        yield sse(msg)
                        log_write(msg)
                    except Exception as e:
                        errors += 1
                        msg = f"[{i}] ERROR: {str(e)} dn={dn}"
                        yield sse(msg)
                        log_write(msg)

                    if i % 50 == 0:
                        yield sse(
                            f"… прогресс: {i}/{len(rows)} | ok={ok} skipped={skipped} "
                            f"notfound={notfound} ambiguous={ambiguous} errors={errors} "
                            f"changed={changed} unchanged={unchanged}"
                        )

                summary = (
                    f"🏁 ГОТОВО mode={mode} | total={len(rows)} | ok={ok} | skipped={skipped} | "
                    f"notfound={notfound} | ambiguous={ambiguous} | errors={errors} | "
                    f"changed={changed} | unchanged={unchanged}"
                )
                yield sse(summary)
                log_write(summary)

            finally:
                try:
                    if conn:
                        conn.unbind_s()
                except Exception:
                    pass

        except Exception as e:
            msg = f"❌ FATAL: {str(e)}"
            yield sse(msg)
            try:
                with open(log_path, "a", encoding="utf-8") as f:
                    f.write(msg + "\n")
            except Exception:
                pass
        finally:
            _release_sync_lock()

    resp = StreamingHttpResponse(run(), content_type="text/event-stream; charset=utf-8")
    resp["Cache-Control"] = "no-cache"
    resp["X-Accel-Buffering"] = "no"  
    return resp


@require_http_methods(["GET"])
def ldap_tools_sync_log_download(request, filename: str):
    # простая защита от ../
    filename = os.path.basename(filename)
    path = os.path.join(LOG_DIR, filename)
    if not os.path.exists(path):
        return HttpResponse("NOT FOUND", status=404)
    return FileResponse(open(path, "rb"), as_attachment=True, filename=filename)




















































def _client_ip_simple(request) -> str:
    xff = request.META.get("HTTP_X_FORWARDED_FOR")
    if xff:
        return xff.split(",")[0].strip()
    return (request.META.get("REMOTE_ADDR") or "").strip()


def _require_bot_token(request) -> bool:
    """
    Если TG_BOT_API_TOKEN задан — требуем заголовок X-Bot-Token.
    Если не задан — не требуем (но лучше задать).
    """
    if not TG_BOT_API_TOKEN:
        return True
    got = request.headers.get("X-Bot-Token") or request.META.get("HTTP_X_BOT_TOKEN") or ""
    return hmac.compare_digest(str(got), str(TG_BOT_API_TOKEN))


def _json_body_or_400(request):
    try:
        raw = request.body.decode("utf-8") if request.body else "{}"
        data = json.loads(raw or "{}")
        if not isinstance(data, dict):
            return None, JsonResponse({"status": "error", "message": "JSON должен быть объектом"}, status=400)
        return data, None
    except Exception as e:
        return None, JsonResponse({"status": "error", "message": f"Некорректный JSON: {e}"}, status=400)


def _expire_old_badge_requests():
    try:
        now = timezone.now()
        AdminBadgeRequest.objects.filter(
            expires_at__lt=now
        ).exclude(status__in=["ACCEPTED", "REJECTED", "EXPIRED"]).update(status="EXPIRED")
    except Exception:
        pass


def _tg_send_message(chat_id: str, text: str, reply_markup: dict | None = None) -> bool:
    """
    Отправка сообщения в Telegram (можно с inline-кнопками).
    """
    chat_id = str(chat_id or "").strip()
    if not chat_id or not TELEGRAM_BOT_TOKEN:
        return False

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": (text or "").strip(),
        "disable_web_page_preview": True,
    }
    if reply_markup:
        payload["reply_markup"] = reply_markup

    try:
        r = requests.post(url, json=payload, timeout=10)
        js = r.json() if r.headers.get("content-type", "").startswith("application/json") else {}
        return (r.status_code == 200 and js.get("ok") is True)
    except Exception:
        return False



@csrf_exempt
@require_http_methods(["POST"])
def tg_admin_badge_start(request):
    """
    START: { tg_id } или { max_id }
    Ответ: guid + stores + cashier + expires_at

    Админ-лог в Telegram отправляется в фоне.
    """
    if not _require_bot_token(request):
        return JsonResponse({"status": "error", "message": "FORBIDDEN"}, status=403)

    _expire_old_badge_requests()

    data, err = _json_body_or_400(request)
    if err:
        return err

    tg_id, max_id, channel = _extract_ids(data)
    if not channel:
        return JsonResponse({"status": "error", "message": "tg_id or max_id required"}, status=400)

    user, err = _resolve_user_by_ids(tg_id=tg_id, max_id=max_id)
    if err:
        return JsonResponse({"status": "error", "message": err}, status=404)
    if not user:
        _send_telegram_log_async(
            "\n".join([
                "❌ Запрос админского бейджа: пользователь не найден",
                f"channel={channel}",
                f"tg_id={tg_id or '—'}",
                f"max_id={max_id or '—'}",
                f"ip={_client_ip_simple(request)}",
                f"time={timezone.localtime(timezone.now()).isoformat(sep=' ', timespec='seconds')}",
            ])
        )
        return JsonResponse({"status": "error", "message": "USER_NOT_FOUND"}, status=404)

    store_ids = list(
        UKMUser.objects.filter(user_id=user.id)
        .values_list("storeid", flat=True)
        .distinct()
    )
    store_ids = [int(x) for x in store_ids if str(x).isdigit()]

    if not store_ids:
        _send_telegram_log_async(
            "\n".join([
                "❌ Запрос админского бейджа: Нет в таблице ukm_users",
                f"channel={channel}",
                f"cashier_user_id={user.id}",
                f"cashier_fio={user.full_name}",
                f"cashier_tg_id={getattr(user,'tg_id','') or '—'}",
                f"cashier_max_id={getattr(user,'max_id',None) or '—'}",
                f"ip={_client_ip_simple(request)}",
                f"time={timezone.localtime(timezone.now()).isoformat(sep=' ', timespec='seconds')}",
            ])
        )
        return JsonResponse({"status": "error", "message": "NO_STORES"}, status=404)

    now = timezone.now()
    req = AdminBadgeRequest.objects.create(
        status="NEW",
        cashier_user_id=user.id,
        cashier_tg_id=(str(getattr(user, "tg_id", "") or "").strip() if channel == "tg" else ""),
        cashier_full_name=(user.full_name or "").strip(),
        store_ids=store_ids,
        storeid=None,
        admin_user_id=None,
        admin_tg_id=None,
        admin_full_name=None,
        decision=None,
        decided_at=None,
        expires_at=now + timezone.timedelta(minutes=BADGE_REQ_TTL_MINUTES),
        ip=_client_ip_simple(request),
        user_agent=(request.META.get("HTTP_USER_AGENT") or "")[:2000],
        meta={
            "start_payload": data,
            "cashier_channel": channel,
            "cashier_max_id": (max_id if channel == "max" else str(getattr(user, "max_id", "") or "")),
            "cashier_tg_id": (tg_id if channel == "tg" else str(getattr(user, "tg_id", "") or "")),
        },
    )

    _send_telegram_log_async(
        "\n".join([
            "🪪 Запрос админского бейджа: Запрошен",
            f"guid={req.id}",
            f"channel={channel}",
            f"cashier_user_id={user.id}",
            f"cashier_fio={user.full_name}",
            f"cashier_tg_id={getattr(user,'tg_id','') or '—'}",
            f"cashier_max_id={getattr(user,'max_id',None) or '—'}",
            f"stores={store_ids}",
            f"expires_at={timezone.localtime(req.expires_at).isoformat(sep=' ', timespec='seconds')}",
            f"ip={req.ip}",
        ])
    )

    return JsonResponse({
        "status": "ok",
        "guid": str(req.id),
        "stores": store_ids,
        "cashier": {
            "id": user.id,
            "full_name": (user.full_name or "").strip(),
            "tg_id": str(getattr(user, "tg_id", "") or "").strip(),
            "max_id": getattr(user, "max_id", None),
        },
        "expires_at": timezone.localtime(req.expires_at).isoformat(sep=' ', timespec='seconds'),
        "channel": channel,
    })






@csrf_exempt
@require_http_methods(["POST"])
def tg_admin_badge_admins(request):
    """
    ADMINS: {tg_id|max_id, guid, storeid}
    Возвращает админов с полями tg_id/max_id + can_notify_max
    """
    if not _require_bot_token(request):
        return JsonResponse({"status": "error", "message": "FORBIDDEN"}, status=403)

    _expire_old_badge_requests()

    data, err = _json_body_or_400(request)
    if err:
        return err

    tg_id, max_id, channel = _extract_ids(data)
    guid = str(data.get("guid") or "").strip()
    storeid_raw = data.get("storeid")

    if not channel or not guid or storeid_raw is None:
        return JsonResponse({"status": "error", "message": "tg_id/max_id, guid, storeid required"}, status=400)

    try:
        storeid = int(str(storeid_raw).strip())
    except Exception:
        return JsonResponse({"status": "error", "message": "BAD_STOREID"}, status=400)

    try:
        req = AdminBadgeRequest.objects.get(id=guid)
    except Exception:
        return JsonResponse({"status": "error", "message": "GUID_NOT_FOUND"}, status=404)

    if req.status == "EXPIRED" or req.expires_at < timezone.now():
        return JsonResponse({"status": "error", "message": "EXPIRED"}, status=410)

    if not _session_owner_ok(req, channel=channel, tg_id=tg_id, max_id=max_id):
        return JsonResponse({"status": "error", "message": "NOT_YOUR_SESSION"}, status=403)

    allowed_stores = req.store_ids or []
    if storeid not in allowed_stores:
        return JsonResponse({"status": "error", "message": "STORE_NOT_ALLOWED"}, status=403)

    # админы магазина: roleid 11 или 13
    admin_user_ids = list(
        UKMUser.objects.filter(storeid=storeid, roleid__in=[11, 13])
        .values_list("user_id", flat=True)
        .distinct()
    )
    admin_user_ids = [int(x) for x in admin_user_ids if str(x).isdigit()]

    admins = []
    if admin_user_ids:
        qs = User.objects.filter(id__in=admin_user_ids).order_by("full_name")
        for u in qs:
            tg_admin = str(getattr(u, "tg_id", "") or "").strip()
            mx_admin = getattr(u, "max_id", None)
            admins.append({
                "id": u.id,
                "tg_id": tg_admin,
                "max_id": mx_admin,
                "full_name": (u.full_name or "").strip(),
                "can_notify": bool(tg_admin),                 # как раньше (TG)
                "can_notify_max": bool(mx_admin),             # ✅ для MAX
            })

    req.storeid = storeid
    req.status = "STORE_SELECTED"
    req.meta = {**(req.meta or {}), "admins_payload": data, "admins_count": len(admins)}
    req.save(update_fields=["storeid", "status", "meta"])

    return JsonResponse({
        "status": "ok",
        "guid": str(req.id),
        "storeid": storeid,
        "admins": admins,
        "channel": channel,
    })





@csrf_exempt
@require_http_methods(["POST"])
def tg_admin_badge_request(request):
    """
    REQUEST: {tg_id|max_id, guid, storeid, admin_id}
    TG-канал: отправляет админу Telegram сообщение (как раньше)
    MAX-канал: НЕ отправляет Telegram, возвращает admin.max_id (бот MAX сам отправит сообщение админу)
    """
    if not _require_bot_token(request):
        return JsonResponse({"status": "error", "message": "FORBIDDEN"}, status=403)

    _expire_old_badge_requests()

    data, err = _json_body_or_400(request)
    if err:
        return err

    tg_id, max_id, channel = _extract_ids(data)
    guid = str(data.get("guid") or "").strip()
    storeid_raw = data.get("storeid")
    admin_id_raw = data.get("admin_id")

    if not channel or not guid or storeid_raw is None or admin_id_raw is None:
        return JsonResponse({"status": "error", "message": "tg_id/max_id,guid,storeid,admin_id required"}, status=400)

    try:
        storeid = int(str(storeid_raw).strip())
        admin_id = int(str(admin_id_raw).strip())
    except Exception:
        return JsonResponse({"status": "error", "message": "BAD_STOREID_OR_ADMIN_ID"}, status=400)

    try:
        req = AdminBadgeRequest.objects.get(id=guid)
    except Exception:
        return JsonResponse({"status": "error", "message": "GUID_NOT_FOUND"}, status=404)

    if req.status == "EXPIRED" or req.expires_at < timezone.now():
        return JsonResponse({"status": "error", "message": "EXPIRED"}, status=410)

    if not _session_owner_ok(req, channel=channel, tg_id=tg_id, max_id=max_id):
        return JsonResponse({"status": "error", "message": "NOT_YOUR_SESSION"}, status=403)

    if not req.storeid or int(req.storeid) != storeid:
        return JsonResponse({"status": "error", "message": "STORE_NOT_SELECTED_OR_MISMATCH"}, status=400)

    is_admin_here = UKMUser.objects.filter(storeid=storeid, roleid__in=[11, 13], user_id=admin_id).exists()
    if not is_admin_here:
        return JsonResponse({"status": "error", "message": "ADMIN_NOT_IN_STORE_OR_BAD_ROLE"}, status=403)

    admin_user = User.objects.filter(id=admin_id).first()
    if not admin_user:
        return JsonResponse({"status": "error", "message": "ADMIN_USER_NOT_FOUND"}, status=404)

    # сохраняем админа в сессию
    req.admin_user_id = admin_user.id
    req.admin_full_name = (admin_user.full_name or "").strip()
    req.status = "PENDING_ADMIN"
    meta = req.meta or {}

    if channel == "tg":
        admin_tg = str(getattr(admin_user, "tg_id", "") or "").strip()
        if not admin_tg:
            return JsonResponse({"status": "error", "message": "ADMIN_HAS_NO_TG_ID"}, status=409)

        req.admin_tg_id = admin_tg
        req.meta = {**meta, "request_payload": data, "admin_channel": "tg"}
        req.save(update_fields=["admin_user_id", "admin_tg_id", "admin_full_name", "status", "meta"])

        # как раньше — шлём в Telegram
        store_name = ""
        try:
            st = Store.objects.filter(ukm4store=storeid).first()
            store_name = (st.name or "").strip() if st else ""
        except Exception:
            pass

        msg = "\n".join([
            "🪪 Запрос бейджа администратора",
            "",
            f"Кассир: {req.cashier_full_name or '—'} (user_id={req.cashier_user_id})",
            f"Магазин: {storeid}" + (f" — {store_name}" if store_name else ""),
            "",
            f"GUID: {req.id}",
            "",
            "Разрешить выдачу бейджа?",
        ])

        reply_markup = {
            "inline_keyboard": [
                [
                    {"text": "✅ Разрешить", "callback_data": f"admin_badge:accept:{req.id}"},
                    {"text": "⛔ Запретить", "callback_data": f"admin_badge:reject:{req.id}"},
                ]
            ]
        }
        ok_send = _tg_send_message(admin_tg, msg, reply_markup=reply_markup)

        return JsonResponse({
            "status": "ok",
            "guid": str(req.id),
            "message": "WAIT_ADMIN",
            "admin": {"id": admin_user.id, "tg_id": admin_tg, "full_name": req.admin_full_name, "can_notify": bool(admin_tg)},
            "telegram_send_ok": ok_send,
            "channel": "tg",
        })

    # ===== MAX channel =====
    admin_max = getattr(admin_user, "max_id", None)
    if not admin_max:
        return JsonResponse({"status": "error", "message": "ADMIN_HAS_NO_MAX_ID"}, status=409)

    req.admin_tg_id = ""  # не используем TG
    req.meta = {**meta, "request_payload": data, "admin_channel": "max", "admin_max_id": str(admin_max)}
    req.save(update_fields=["admin_user_id", "admin_tg_id", "admin_full_name", "status", "meta"])

    return JsonResponse({
        "status": "ok",
        "guid": str(req.id),
        "message": "WAIT_ADMIN",
        "admin": {"id": admin_user.id, "max_id": admin_max, "full_name": req.admin_full_name, "can_notify_max": True},
        "channel": "max",
    })





@csrf_exempt
@require_http_methods(["POST"])
def tg_admin_badge_decision(request):
    """
    DECISION: { guid, decision, admin_id, (admin_tg_id?), (admin_max_id?) }
    Возвращает send_to_tg_id (если есть) и send_to_max_id (если MAX-сессия).
    """
    if not _require_bot_token(request):
        return JsonResponse({"status": "error", "message": "FORBIDDEN"}, status=403)

    _expire_old_badge_requests()

    data, err = _json_body_or_400(request)
    if err:
        return err

    guid = str(data.get("guid") or "").strip()
    decision = str(data.get("decision") or "").strip().lower()
    admin_id_raw = data.get("admin_id")
    admin_tg_id = str(data.get("admin_tg_id") or "").strip()
    admin_max_id = str(data.get("admin_max_id") or "").strip()

    if not guid or decision not in ("accept", "reject") or admin_id_raw is None:
        return JsonResponse({"status": "error", "message": "guid, decision(accept/reject), admin_id required"}, status=400)

    try:
        admin_id = int(str(admin_id_raw).strip())
    except Exception:
        return JsonResponse({"status": "error", "message": "BAD_ADMIN_ID"}, status=400)

    try:
        req = AdminBadgeRequest.objects.get(id=guid)
    except Exception:
        return JsonResponse({"status": "error", "message": "GUID_NOT_FOUND"}, status=404)

    if req.status == "EXPIRED" or req.expires_at < timezone.now():
        return JsonResponse({"status": "error", "message": "EXPIRED"}, status=410)

    if not req.admin_user_id or int(req.admin_user_id) != int(admin_id):
        return JsonResponse({"status": "error", "message": "ADMIN_MISMATCH"}, status=403)

    # сверка tg_id (для TG-канала)
    if admin_tg_id and req.admin_tg_id and str(req.admin_tg_id) != admin_tg_id:
        return JsonResponse({"status": "error", "message": "ADMIN_TG_MISMATCH"}, status=403)

    # сверка max_id (для MAX-канала)
    meta = req.meta or {}
    if admin_max_id and meta.get("admin_max_id") and str(meta.get("admin_max_id")) != admin_max_id:
        return JsonResponse({"status": "error", "message": "ADMIN_MAX_MISMATCH"}, status=403)

    now = timezone.now()
    req.decision = decision
    req.decided_at = now
    req.status = "ACCEPTED" if decision == "accept" else "REJECTED"
    req.meta = {**meta, "decision_payload": data}
    req.save(update_fields=["decision", "decided_at", "status", "meta"])

    send_to_tg_id = str(req.cashier_tg_id or "").strip()
    send_to_max_id = str(meta.get("cashier_max_id") or "").strip()

    if decision == "reject":
        return JsonResponse({
            "status": "ok",
            "guid": str(req.id),
            "decision": "reject",
            "send_to_tg_id": send_to_tg_id,
            "send_to_max_id": send_to_max_id,
            "message_to_cashier": "Администратор отклонил запрос бейджа.",
        })

    password, open_username, open_row_id = _get_existing_open_password(user_id=admin_id, system_id=9)
    if not password:
        return JsonResponse({"status": "error", "message": "ADMIN_PASSWORD_NOT_FOUND"}, status=404)

    return JsonResponse({
        "status": "ok",
        "guid": str(req.id),
        "decision": "accept",
        "send_to_tg_id": send_to_tg_id,
        "send_to_max_id": send_to_max_id,
        "password": password,
        "admin_open_in_system": {"id": open_row_id, "username": open_username, "system_id": 9},
    })

























































@require_http_methods(["GET"])
@never_cache
def bitrix_inactive_users_ui(request):
    """
    UI: список пользователей Bitrix, фильтры по "не заходил N дней",
    показ ACTIVE / IS_ONLINE / LAST_LOGIN / LAST_ACTIVITY_DATE,
    + массовая/точечная блокировка через отдельный POST endpoint.
    """
    # фильтры
    try:
        days = int((request.GET.get("days") or "30").strip())
    except Exception:
        days = 30
    days = max(1, min(days, 3650))

    mode = (request.GET.get("mode") or "login").strip().lower()
    if mode not in ("login", "activity", "both"):
        mode = "login"

    active_filter = (request.GET.get("active") or "all").strip().lower()
    if active_filter not in ("all", "active", "inactive"):
        active_filter = "all"

    online_filter = (request.GET.get("online") or "all").strip().lower()
    if online_filter not in ("all", "online", "offline"):
        online_filter = "all"

    q = (request.GET.get("q") or "").strip().lower()

    dept_raw = (request.GET.get("dept") or "").strip()
    include_sub = (request.GET.get("sub") or "") in ("1", "true", "on", "yes", "y")

    only_overdue = (request.GET.get("only_overdue") or "1") in ("1", "true", "on", "yes", "y")

    dept_id = int(dept_raw) if dept_raw.isdigit() else None

    now = timezone.now()
    cutoff = now - timezone.timedelta(days=days)

    # departments 
    try:
        depts = bitrix_get_departments()
        by_id, children = _dept_index(depts)
    except Exception:
        depts = []
        by_id, children = {}, {}

    allowed_dept_ids_set = None
    if dept_id:
        if include_sub and children:
            allowed_dept_ids_set = set(_dept_descendants([dept_id], children))
        else:
            allowed_dept_ids_set = {dept_id}

    # users 
    select_fields = [
        "ID", "NAME", "LAST_NAME", "SECOND_NAME",
        "EMAIL", "WORK_POSITION",
        "ACTIVE", "IS_ONLINE",
        "UF_DEPARTMENT",
        "LAST_LOGIN",
        "LAST_ACTIVITY_DATE",
        "DATE_REGISTER",
    ]

    all_users = bitrix_user_get_all(filter_dict={}, select_list=select_fields)

    rows = []
    total = 0
    overdue_cnt = 0

    for u in all_users:
        total += 1

        try:
            uid = int(u.get("ID"))
        except Exception:
            continue

        fio = " ".join([x for x in [u.get("LAST_NAME"), u.get("NAME"), u.get("SECOND_NAME")] if x]).strip()
        email = (u.get("EMAIL") or "").strip()
        position = (u.get("WORK_POSITION") or "").strip()

        is_active = _bx_is_active(u)          # твоя функция (как ранее)
        is_online = _bx_bool(u.get("IS_ONLINE"))

        last_login_dt = _parse_bx_dt(u.get("LAST_LOGIN"))
        last_act_dt = _parse_bx_dt(u.get("LAST_ACTIVITY_DATE"))
        reg_dt = _parse_bx_dt(u.get("DATE_REGISTER"))

        raw_depts = u.get("UF_DEPARTMENT")
        dept_ids = []
        if isinstance(raw_depts, list):
            for x in raw_depts:
                if str(x).isdigit():
                    dept_ids.append(int(x))
        elif str(raw_depts).isdigit():
            dept_ids = [int(raw_depts)]

        if allowed_dept_ids_set is not None:
            if not set(dept_ids).intersection(allowed_dept_ids_set):
                continue

        if active_filter == "active" and not is_active:
            continue
        if active_filter == "inactive" and is_active:
            continue

        if online_filter == "online" and is_online is not True:
            continue
        if online_filter == "offline" and is_online is not False:
            continue

        if q:
            hay = " ".join([str(uid), fio, email, position]).lower()
            if q not in hay:
                continue

        stale_login = (last_login_dt is None) or (last_login_dt < cutoff)
        stale_act = (last_act_dt is None) or (last_act_dt < cutoff)

        if mode == "login":
            overdue = stale_login
        elif mode == "activity":
            overdue = stale_act
        else:
            overdue = stale_login and stale_act

        if overdue:
            overdue_cnt += 1
        if only_overdue and not overdue:
            continue

        def _days_ago(dt):
            if not dt:
                return None
            try:
                return (now - dt).days
            except Exception:
                return None

        dept_names = []
        for did in dept_ids:
            d = by_id.get(did)
            dept_names.append(_dept_name(d) if d else str(did))

        rows.append({
            "id": uid,
            "fio": fio,
            "email": email,
            "position": position,
            "active": is_active,
            "online": is_online,
            "last_login": last_login_dt,
            "last_activity": last_act_dt,
            "date_register": reg_dt,
            "days_login": _days_ago(last_login_dt),
            "days_activity": _days_ago(last_act_dt),
            "dept_ids": dept_ids,
            "dept_names": dept_names,
            "overdue": overdue,
        })

    # ✅ сортировка: больше дней -> выше, None -> вниз
    def _sort_days(v):
        return v if isinstance(v, int) else -1

    rows.sort(key=lambda r: _sort_days(r["days_login"]), reverse=True)

    dept_options = [{"id": did, "name": _dept_name(d)} for did, d in sorted(by_id.items(), key=lambda x: x[0])]

    return render(request, "frostapp/bitrix_inactive_users.html", {
        "rows": rows,
        "total": total,
        "overdue_cnt": overdue_cnt,
        "days": days,
        "mode": mode,
        "active_filter": active_filter,
        "online_filter": online_filter,
        "only_overdue": only_overdue,
        "q": q,
        "dept_id": dept_id,
        "include_sub": include_sub,
        "dept_options": dept_options,
        "csrf": get_token(request),
    })








@require_http_methods(["POST"])
@csrf_protect
def bitrix_users_toggle_active(request):
    """
    POST:
      user_id=123
      user_id=1&user_id=2...
      action=block|unblock

    Делает user.update ACTIVE и ОБЯЗАТЕЛЬНО проверяет, что Bitrix реально поменял значение.
    """
    action = (request.POST.get("action") or "").strip().lower()
    if action not in ("block", "unblock"):
        return JsonResponse({"ok": False, "error": "BAD_ACTION"}, status=400)

    ids = request.POST.getlist("user_id")
    user_ids: list[int] = [int(x) for x in ids if str(x).isdigit()]
    if not user_ids:
        return JsonResponse({"ok": False, "error": "NO_USER_IDS"}, status=400)

    desired_bool = (action == "unblock")

    ok_ids: list[int] = []
    bad: list[dict] = []

    for uid in user_ids:
        try:
            bitrix_set_user_active_strict(uid, desired_bool)
            ok_ids.append(uid)
        except Exception as e:
            bad.append({"id": uid, "error": str(e)})

    try:
        err_preview = "; ".join([f"{x['id']}:{x['error']}" for x in bad[:5]])
        _send_telegram_log_async(
            "\n".join([
                "👥 BITRIX USERS TOGGLE ACTIVE",
                f"by_django_user={getattr(request.user, 'username', 'unknown')}",
                f"action={action} desired_ACTIVE={('Y' if desired_bool else 'N')}",
                f"ok={len(ok_ids)} bad={len(bad)}",
                f"errors_top5={err_preview or '—'}",
                f"ip={_client_ip(request)}",
                f"time={timezone.localtime(timezone.now()).isoformat(sep=' ', timespec='seconds')}",
            ])
        )
    except Exception:
        pass

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"ok": True, "updated": ok_ids, "errors": bad})

    back = request.META.get("HTTP_REFERER") or reverse("bitrix_inactive_users_ui")
    sep = "&" if "?" in back else "?"
    return redirect(f"{back}{sep}updated_ok={len(ok_ids)}&updated_bad={len(bad)}")








_ORA_USERNAME_RE = re.compile(r"^[A-Z][A-Z0-9_$#]{0,127}$")  # безопасный whitelist

def _sm_ui_services_list() -> list[str]:
    # только те базы, которые реально описаны в ORACLE_TNS_MAP
    return sorted(list(ORACLE_TNS_MAP.keys()))

def _parse_int(v, default: int, mn: int, mx: int) -> int:
    try:
        x = int(str(v).strip())
    except Exception:
        x = default
    return max(mn, min(mx, x))

def _client_ip(request) -> str:
    xff = (request.META.get("HTTP_X_FORWARDED_FOR") or "").split(",")[0].strip()
    return xff or (request.META.get("REMOTE_ADDR") or "")

def _oracle_dt_to_str(dt):
    if dt is None:
        return ""
    if isinstance(dt, (datetime.datetime, datetime.date)):
        try:
            if isinstance(dt, datetime.date) and not isinstance(dt, datetime.datetime):
                dt = datetime.datetime(dt.year, dt.month, dt.day)
            return dt.isoformat(sep=" ", timespec="seconds")
        except Exception:
            return str(dt)
    return str(dt)

def _sm_fetch_stale_users_in_db(db: str, days: int, only_enabled: bool, q: str, exclude_users: set[str]):
    """
    Возвращает список dict’ов по одной базе:
      username, last_login, created, has_role, staff_id, fio, serverlogin, userenabled, inn, reason
    """
    conn = cur = None
    try:
        conn = _connect_oracle_service(db)
        cur = conn.cursor()

        staff_cols = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")
        user_cols  = _oracle_get_table_columns_set(cur, owner="SYS", table="DBA_USERS")

        def has_staff(col: str) -> bool:
            return col.upper() in staff_cols

        def has_user(col: str) -> bool:
            return col.upper() in user_cols

        # критично для связки
        if not has_staff("SERVERLOGIN"):
            return [], "SMSTAFF.SERVERLOGIN отсутствует — нельзя связать с DBA_USERS"
        if not user_cols:
            return [], "Нет доступа/описания SYS.DBA_USERS"

        if not has_user("CREATED"):
            return [], "В SYS.DBA_USERS не найдено поле CREATED — нельзя посчитать 'старше N дней'"

        # определяем источник LAST_LOGIN
        last_login_expr = None
        audit_join_sql = ""
        binds = {"b_days": int(days)}

        if has_user("LAST_LOGIN"):
            last_login_expr = "du.last_login"
        else:
            # fallback на аудит
            audit_cols = _oracle_get_table_columns_set(cur, owner="SYS", table="DBA_AUDIT_SESSION")
            ts_col = None
            for cand in ("EXTENDED_TIMESTAMP", "TIMESTAMP", "NTIMESTAMP#", "TIMESTAMP#"):
                if cand in audit_cols:
                    ts_col = cand
                    break

            has_audit_username = ("USERNAME" in audit_cols)

            # доп. условия (если есть колонки)
            action_col = "ACTION_NAME" if "ACTION_NAME" in audit_cols else None
            ret_col = None
            for cand in ("RETURN_CODE", "RETURNCODE"):
                if cand in audit_cols:
                    ret_col = cand
                    break

            if ts_col and has_audit_username:
                lookback_days = int(os.getenv("SM_AUDIT_LOOKBACK_DAYS", "3650"))
                binds["b_lookback"] = lookback_days

                where_a = [f"{ts_col} >= (SYSTIMESTAMP - NUMTODSINTERVAL(:b_lookback, 'DAY'))"]
                if action_col:
                    where_a.append(f"{action_col} = 'LOGON'")
                if ret_col:
                    where_a.append(f"{ret_col} = 0")

                audit_join_sql = f"""
                    LEFT JOIN (
                        SELECT
                            username,
                            MAX({ts_col}) AS last_login
                        FROM sys.dba_audit_session
                        WHERE {" AND ".join(where_a)}
                        GROUP BY username
                    ) al
                      ON UPPER(al.username) = UPPER(du.username)
                """
                last_login_expr = "al.last_login"
            else:
                # last_login нигде не достать — деградируем до NULL
                last_login_expr = "NULL"

        # роль SUPERMAG_USER 
        role_cols = _oracle_get_table_columns_set(cur, owner="SYS", table="DBA_ROLE_PRIVS")
        has_role_view = ("GRANTEE" in role_cols and "GRANTED_ROLE" in role_cols)

        role_join_sql = ""
        role_expr = "0"
        if has_role_view:
            role_join_sql = """
                LEFT JOIN (
                    SELECT grantee
                    FROM sys.dba_role_privs
                    WHERE UPPER(granted_role) = 'SUPERMAG_USER'
                ) drp
                  ON UPPER(du.username) = UPPER(drp.grantee)
            """
            role_expr = "NVL2(drp.grantee, 1, 0)"

        # поля SMSTAFF 
        sel_staff_id   = "ss.id AS staff_id" if has_staff("ID") else "NULL AS staff_id"
        sel_surname    = "ss.surname AS surname" if has_staff("SURNAME") else "NULL AS surname"
        sel_name       = "ss.name AS name" if has_staff("NAME") else "NULL AS name"
        sel_patronymic = "ss.patronymic AS patronymic" if has_staff("PATRONYMIC") else "NULL AS patronymic"
        sel_userenabled= "ss.userenabled AS userenabled" if has_staff("USERENABLED") else "NULL AS userenabled"
        sel_inn        = "ss.inn AS inn" if has_staff("INN") else "NULL AS inn"

        # фильтры 
        q = (q or "").strip().lower()
        where = ["ss.id > 0"]
        # stale predicate
        if last_login_expr != "NULL":
            stale_predicate = f"""
                (
                    ({last_login_expr} IS NULL AND du.created < (TRUNC(SYSDATE) - :b_days))
                    OR
                    ({last_login_expr} IS NOT NULL AND TRUNC({last_login_expr}) < (TRUNC(SYSDATE) - :b_days))
                )
            """
            order_expr = f"NVL({last_login_expr}, du.created)"
        else:
            stale_predicate = "du.created < (TRUNC(SYSDATE) - :b_days)"
            order_expr = "du.created"

        where.append(stale_predicate)

        # exclude users
        if exclude_users:
            for i, u in enumerate(sorted(exclude_users)):
                binds[f"b_ex{i}"] = u.upper()
                where.append(f"UPPER(du.username) != :b_ex{i}")

        # only_enabled (только если колонка есть)
        if only_enabled and has_staff("USERENABLED"):
            where.append("(ss.userenabled = 1 OR ss.userenabled = '1')")

        # search
        if q:
            binds["b_q"] = f"%{q}%"
            or_parts = ["LOWER(du.username) LIKE :b_q"]
            if has_staff("SERVERLOGIN"):
                or_parts.append("LOWER(ss.serverlogin) LIKE :b_q")
            if has_staff("SURNAME"):
                or_parts.append("LOWER(ss.surname) LIKE :b_q")
            if has_staff("NAME"):
                or_parts.append("LOWER(ss.name) LIKE :b_q")
            if has_staff("PATRONYMIC"):
                or_parts.append("LOWER(ss.patronymic) LIKE :b_q")
            if has_staff("INN"):
                or_parts.append("LOWER(TRIM(ss.inn)) LIKE :b_q")
            where.append("(" + " OR ".join(or_parts) + ")")

        where_sql = " AND ".join(where)

        sql = f"""
            SELECT
                du.username                 AS username,
                {last_login_expr}           AS last_login,
                du.created                  AS created,
                {role_expr}                 AS has_role,
                {sel_staff_id},
                {sel_surname},
                {sel_name},
                {sel_patronymic},
                ss.serverlogin              AS serverlogin,
                {sel_userenabled},
                {sel_inn}
            FROM supermag.smstaff ss
            JOIN sys.dba_users du
              ON UPPER(ss.serverlogin) = UPPER(du.username)
            {audit_join_sql}
            {role_join_sql}
            WHERE {where_sql}
            ORDER BY {order_expr} ASC
        """

        cur.execute(sql, binds)
        rows = cur.fetchall()

        out = []
        for (username, last_login, created, has_role,
             staff_id, surname, name, patronymic,
             serverlogin, userenabled, inn) in rows:

            fio = " ".join([x for x in [surname, name, patronymic] if x]).strip()
            # reason оставляем совместимым с твоим шаблоном
            reason = "never_logged_in" if last_login is None else "stale_login"

            out.append({
                "db": db,
                "username": (username or "").strip(),
                "last_login": _oracle_dt_to_str(last_login),
                "created": _oracle_dt_to_str(created),
                "has_role": int(has_role or 0),
                "staff_id": int(staff_id) if staff_id is not None else None,
                "fio": fio,
                "serverlogin": (serverlogin or "").strip(),
                "userenabled": str(userenabled).strip() if userenabled is not None else "",
                "inn": (inn or "").strip(),  # ✅ INN
                "reason": reason,
            })

        return out, None

    except Exception as e:
        logger.exception(f"[SM_LAST_LOGIN] fetch failed db={db}: {e}")
        return [], str(e)

    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass


def _sm_block_users_in_db(db: str, usernames: list[str], dry_run: bool):
    """
    Делает:
      1) UPDATE supermag.smstaff SET userenabled=0 WHERE upper(serverlogin)=upper(username)
      2) REVOKE SUPERMAG_USER FROM <username>
    Возвращает: ok(list), bad(list of {username,error}), details(list)
    """
    conn = cur = None
    ok, bad, details = [], [], []

    # нормализация+валидация имён (чтобы безопасно вставлять в DDL)
    cleaned = []
    for u in usernames:
        u2 = (u or "").strip().upper()
        if not u2:
            continue
        if not _ORA_USERNAME_RE.match(u2):
            bad.append({"username": u2, "error": "BAD_USERNAME_FORMAT"})
            continue
        cleaned.append(u2)

    cleaned = sorted(set(cleaned))

    try:
        conn = _connect_oracle_service(db)
        cur = conn.cursor()

        for u in cleaned:
            if dry_run:
                details.append({"db": db, "username": u, "smstaff_disabled": None, "revoke_role": None, "dry_run": True})
                ok.append(u)
                continue

            try:
                # 1) disable in SMSTAFF
                cur.execute("""
                    UPDATE supermag.smstaff
                       SET userenabled = 0
                     WHERE UPPER(serverlogin) = :b_u
                """, b_u=u)
                smstaff_rows = int(cur.rowcount or 0)

                # 2) revoke role (если роли нет — Oracle может ругнуться; ловим и считаем как "не критично")
                revoked = False
                try:
                    cur.execute(f"REVOKE SUPERMAG_USER FROM {u}")
                    revoked = True
                except Exception as re_err:
                    # бывает ORA-01952 / ORA-01951 и т.п. — роли нет/не выдана
                    revoked = False
                    details.append({"db": db, "username": u, "warn": f"REVOKE_FAILED: {re_err}"})

                details.append({
                    "db": db,
                    "username": u,
                    "smstaff_disabled_rows": smstaff_rows,
                    "revoke_role_ok": revoked,
                    "dry_run": False
                })
                ok.append(u)

            except Exception as one_err:
                bad.append({"username": u, "error": str(one_err)})

        if not dry_run:
            conn.commit()

        return ok, bad, details

    except Exception as e:
        logger.exception(f"[SM_LAST_LOGIN] block failed db={db}: {e}")
        return [], [{"username": "*", "error": str(e)}], details

    finally:
        try:
            if cur: cur.close()
            if conn: conn.close()
        except Exception:
            pass


@staff_member_required
@require_http_methods(["GET"])
@never_cache
def sm_oracle_inactive_users_ui(request):
    """
    UI: список "не заходил N дней" по SYS.DBA_USERS.LAST_LOGIN (в разрезе баз).
    GET /ui/sm/oracle-inactive-users/?db=BINUU00|ALL&days=30&q=...&only_enabled=1
    """
    services = _sm_ui_services_list()

    db = (request.GET.get("db") or "BINUU00").strip().upper()
    if db != "ALL" and db not in ORACLE_TNS_MAP:
        db = "BINUU00"

    days = _parse_int(request.GET.get("days"), default=30, mn=1, mx=3650)
    q = (request.GET.get("q") or "").strip()
    only_enabled = (request.GET.get("only_enabled") or "1").strip().lower() in ("1", "true", "yes", "on")

    # можно расширить список исключений через env
    exclude_env = (os.getenv("SM_LASTLOGIN_EXCLUDE_USERS", "S_BUDAYEV,SUPERMAG,SADMIN,SYSTEM,SYS") or "")
    exclude_users = {x.strip().upper() for x in exclude_env.split(",") if x.strip()}

    rows = []
    db_errors = {}

    target_dbs = services if db == "ALL" else [db]

    for d in target_dbs:
        items, err = _sm_fetch_stale_users_in_db(
            db=d,
            days=days,
            only_enabled=only_enabled,
            q=q,
            exclude_users=exclude_users
        )
        rows.extend(items)
        if err:
            db_errors[d] = err

    # сортировка: сначала те, у кого самый старый last_login/created
    def _sort_key(r):
        # last_login пустой — используем created
        x = r.get("last_login") or r.get("created") or ""
        return x

    rows.sort(key=_sort_key)

    return render(request, "frostapp/sm_oracle_inactive_users.html", {
        "services": services,
        "db": db,
        "days": days,
        "q": q,
        "only_enabled": only_enabled,
        "rows": rows,
        "db_errors": db_errors,
        "total": len(rows),
        "csrf": get_token(request),
    })


@staff_member_required
@require_http_methods(["POST"])
@csrf_protect
def sm_oracle_users_block(request):
    """
    POST /ui/sm/oracle-inactive-users/block/
    Поля:
      action = test | apply
      item = "BINUU00|USERNAME" (может быть несколько)
    """
    action = (request.POST.get("action") or "").strip().lower()
    if action not in ("test", "apply"):
        return JsonResponse({"ok": False, "error": "BAD_ACTION"}, status=400)

    dry_run = (action == "test")

    items = request.POST.getlist("item")
    pairs = []
    for it in items:
        it = (it or "").strip()
        if "|" not in it:
            continue
        db, username = it.split("|", 1)
        db = (db or "").strip().upper()
        username = (username or "").strip().upper()
        if db and username and db in ORACLE_TNS_MAP:
            pairs.append((db, username))

    if not pairs:
        return JsonResponse({"ok": False, "error": "NO_ITEMS"}, status=400)

    # группируем по базе
    by_db = defaultdict(list)
    for db, u in pairs:
        by_db[db].append(u)

    all_ok = {}
    all_bad = {}
    all_details = []

    for db, usernames in by_db.items():
        ok, bad, details = _sm_block_users_in_db(db=db, usernames=usernames, dry_run=dry_run)
        all_ok[db] = ok
        all_bad[db] = bad
        all_details.extend(details)

    # телега-лог (если у тебя есть send_telegram_log)
    try:
        err_cnt = sum(len(v) for v in all_bad.values())
        ok_cnt = sum(len(v) for v in all_ok.values())
        top_errs = []
        for db, lst in all_bad.items():
            for e in lst[:3]:
                top_errs.append(f"{db}:{e.get('username')}:{e.get('error')}")
            if len(top_errs) >= 5:
                break

        msg = "\n".join([
            "🧾 SUPERMAG LAST_LOGIN BLOCK",
            f"by_django_user={getattr(request.user, 'username', 'unknown')}",
            f"action={'dry_run' if dry_run else 'apply'}",
            f"ok={ok_cnt} bad={err_cnt}",
            f"errors_top5={'; '.join(top_errs) if top_errs else '—'}",
            f"ip={_client_ip(request)}",
            f"time={timezone.now().isoformat(sep=' ', timespec='seconds')}",
        ])
        try:
            send_telegram_log(msg)
        except Exception:
            pass
    except Exception:
        pass

    # AJAX
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({
            "ok": True,
            "dry_run": dry_run,
            "updated_ok": all_ok,
            "errors": all_bad,
            "details": all_details,
        }, json_dumps_params={"ensure_ascii": False, "indent": 2})

    # обычный POST -> редирект обратно на список с параметрами
    back = request.META.get("HTTP_REFERER") or reverse("sm_oracle_inactive_users_ui")
    sep = "&" if "?" in back else "?"
    return redirect(f"{back}{sep}done={'test' if dry_run else 'apply'}")








































































































def _require_report_token(request) -> bool:
    """
    Минимальная защёлка:
    - заголовок: X-REPORT-TOKEN: <token>
    - или query: ?token=<token>
    """
    if not INACTIVE_REPORT_TOKEN:
        # если токен не задан — лучше НЕ открывать эндпойнт
        return False
    got = (request.headers.get("X-REPORT-TOKEN") or request.GET.get("token") or "").strip()
    return hmac.compare_digest(got, INACTIVE_REPORT_TOKEN)


def _safe_int(v, default: int, mn: int, mx: int) -> int:
    try:
        x = int(str(v).strip())
    except Exception:
        x = default
    return max(mn, min(mx, x))


def _dt_iso(dt) -> str:
    if not dt:
        return ""
    try:
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        return dt.isoformat()
    except Exception:
        return str(dt)

def _dt_pretty(dt: Optional[datetime.datetime], fmt: str = "%d.%m.%Y %H:%M:%S") -> str:
    """
    Делает человеко-читаемую дату:
    - приводит к aware (если naive)
    - переводит в локальный TZ Django (timezone.localtime)
    - убирает микросекунды/ISO-T/offset (за счёт strftime)
    """
    if not dt:
        return ""
    try:
        if isinstance(dt, datetime.date) and not isinstance(dt, datetime.datetime):
            dt = datetime.datetime(dt.year, dt.month, dt.day)
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        dt = timezone.localtime(dt)  # важно: будет в TIME_ZONE проекта
        return dt.strftime(fmt)
    except Exception:
        return str(dt)


def _dt_pretty_any(v: Any, fmt: str = "%d.%m.%Y %H:%M:%S") -> str:
    """
    Принимает datetime/строку ISO и возвращает красиво.
    Если строка не парсится — вернёт как есть.
    """
    if not v:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return _dt_pretty(v if isinstance(v, datetime.datetime) else None, fmt=fmt) if isinstance(v, datetime.datetime) else _dt_pretty(datetime.datetime(v.year, v.month, v.day), fmt=fmt)

    try:
        s = str(v).strip()
        if not s:
            return ""
        # поддержка "Z"
        s = s.replace("Z", "+00:00")
        # иногда бывает пробел вместо T
        if "T" not in s and " " in s and "-" in s:
            s = s.replace(" ", "T")
        dt = datetime.datetime.fromisoformat(s)
        return _dt_pretty(dt, fmt=fmt)
    except Exception:
        return str(v)





def _parse_ad_filetime(v) -> Optional[datetime.datetime]:
    """
    lastLogonTimestamp в AD — Windows FILETIME:
    кол-во 100-нс интервалов с 1601-01-01.
    """
    if not v:
        return None
    try:
        if isinstance(v, (list, tuple)):
            v = v[0] if v else None
        if v is None:
            return None
        if isinstance(v, bytes):
            v = v.decode("utf-8", "ignore")
        n = int(str(v).strip())
        if n <= 0:
            return None
        base = datetime.datetime(1601, 1, 1, tzinfo=datetime.timezone.utc)
        dt = base + datetime.timedelta(microseconds=n / 10)
        return dt.astimezone(datetime.timezone.utc)
    except Exception:
        return None


def _ldap_decode_first(entry: dict, key: str) -> str:
    try:
        arr = entry.get(key)
        if not arr:
            return ""
        v = arr[0]
        if isinstance(v, bytes):
            return v.decode("utf-8", "ignore")
        return str(v)
    except Exception:
        return ""


def _ad_fetch_inactive_users(days: int = 60, include_never: bool = True) -> list[dict]:
    """
    Возвращает пользователей AD, которые НЕ входили >= days (lastLogonTimestamp),
    по умолчанию включает тех, у кого lastLogonTimestamp пустой (никогда/нет данных).
    Берём только ENABLED (не disabled).
    """
    now = timezone.now()
    cutoff = now - timezone.timedelta(days=days)

    bind_user = AD_USERNAME
    if "@" not in bind_user and "\\" not in bind_user:
        bind_user = f"{AD_DOMAIN}\\{AD_USERNAME}"

    l = ldap.initialize(f"ldap://{AD_IP}")
    l.set_option(ldap.OPT_REFERRALS, 0)
    l.protocol_version = 3
    l.simple_bind_s(bind_user, AD_PASSWORD)

    # userAccountControl disabled flag = 2
    # Фильтр: люди-пользователи
    search_filter = "(&(objectCategory=person)(objectClass=user))"
    attrs = [
        "sAMAccountName", "displayName", "mail", "userPrincipalName",
        "employeeID", "department", "title",
        "lastLogonTimestamp", "userAccountControl",
    ]

    page_size = 1000
    cookie = b""
    ctrl = SimplePagedResultsControl(True, size=page_size, cookie=cookie)

    out = []
    while True:
        msgid = l.search_ext(
            AD_SEARCH_BASE,
            ldap.SCOPE_SUBTREE,
            search_filter,
            attrlist=attrs,
            serverctrls=[ctrl],
        )
        rtype, rdata, rmsgid, serverctrls = l.result3(msgid)

        for dn, entry in (rdata or []):
            if not isinstance(entry, dict):
                continue

            sam = _ldap_decode_first(entry, "sAMAccountName")
            display = _ldap_decode_first(entry, "displayName")
            mail = _ldap_decode_first(entry, "mail")
            upn = _ldap_decode_first(entry, "userPrincipalName")
            employee_id = _ldap_decode_first(entry, "employeeID")
            department = _ldap_decode_first(entry, "department")
            title = _ldap_decode_first(entry, "title")

            uac_raw = _ldap_decode_first(entry, "userAccountControl")
            try:
                uac = int(uac_raw) if uac_raw else 0
            except Exception:
                uac = 0
            disabled = bool(uac & 2)
            if disabled:
                # по ТЗ вы хотите “активных” — значит disabled исключаем
                continue

            last_ts = _parse_ad_filetime(entry.get("lastLogonTimestamp"))
            last_any = last_ts

            if last_any is None and not include_never:
                continue

            if last_any is None:
                is_inactive = True
                days_since = None
            else:
                is_inactive = last_any < cutoff
                days_since = (now - last_any).days

            if not is_inactive:
                continue

            out.append({
                "sam": sam,
                "display": display,
                "mail": mail,
                "upn": upn,
                "employeeID": employee_id,  # часто у вас = ИНН
                "department": department,
                "title": title,
                "disabled": disabled,
                "lastLogonTimestamp": _dt_iso(last_ts),
                "daysSince": days_since,
            })

        # paging cookie
        paged_ctrls = [c for c in serverctrls if c.controlType == SimplePagedResultsControl.controlType]
        if not paged_ctrls:
            break
        cookie = paged_ctrls[0].cookie
        if not cookie:
            break
        ctrl.cookie = cookie

    try:
        l.unbind_s()
    except Exception:
        pass

    # сортировка: самые “древние” наверх
    def _k(r):
        v = r.get("daysSince")
        return v if isinstance(v, int) else 10**9

    out.sort(key=_k, reverse=True)
    return out


def _sm_fetch_inactive_users_binu00(days: int = 100, only_enabled: bool = True) -> list[dict]:
    """
    BINUU00: SYS.DBA_USERS.LAST_LOGIN + SUPERMAG.SMSTAFF.SERVERLOGIN.
    Берём тех, кто не входил > days (или никогда и создан давно).
    """
    now = timezone.now()
    exclude_env = (os.getenv("SM_LASTLOGIN_EXCLUDE_USERS", "S_BUDAYEV,SUPERMAG,SADMIN,SYSTEM,SYS") or "")
    exclude_users = {x.strip().upper() for x in exclude_env.split(",") if x.strip()}

    conn = cur = None
    out = []
    try:
        conn = connect_oracle_supermag()
        cur = conn.cursor()

        where = []
        binds = {"b_days": int(days)}

        # stale condition
        where.append("""
            (
                (du.last_login IS NULL AND du.created < (TRUNC(SYSDATE) - :b_days))
                OR
                (du.last_login IS NOT NULL AND TRUNC(du.last_login) < (TRUNC(SYSDATE) - :b_days))
            )
        """)

        if only_enabled:
            where.append("(ss.userenabled = 1 OR ss.userenabled = '1')")

        for i, u in enumerate(sorted(exclude_users)):
            binds[f"b_ex{i}"] = u
            where.append(f"UPPER(du.username) != :b_ex{i}")

        sql = f"""
            SELECT
                du.username,
                du.last_login,
                du.created,
                ss.id,
                ss.surname,
                ss.name,
                ss.patronymic,
                ss.serverlogin,
                ss.userenabled,
                ss.inn
            FROM supermag.smstaff ss
            JOIN sys.dba_users du
              ON UPPER(ss.serverlogin) = UPPER(du.username)
            WHERE ss.id > 0
              AND {" AND ".join(where)}
            ORDER BY NVL(du.last_login, du.created) ASC
        """

        cur.execute(sql, binds)
        rows = cur.fetchall()

        for (username, last_login, created, staff_id, surname, name, patronymic,
             serverlogin, userenabled, inn) in rows:
            fio = " ".join([x for x in [surname, name, patronymic] if x]).strip()
            out.append({
                "db": "BINUU00",
                "username": (username or "").strip(),
                "fio": fio,
                "inn": (inn or "").strip(),
                "serverlogin": (serverlogin or "").strip(),
                "staff_id": int(staff_id) if staff_id is not None else None,
                "userenabled": str(userenabled).strip() if userenabled is not None else "",
                "last_login": _dt_iso(last_login),
                "created": _dt_iso(created),
            })

        return out
    except Exception as e:
        logger.exception(f"[REPORT][SM] failed: {e}")
        return out
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass


def _bx_parse_dt_any(v: Any) -> Optional[datetime.datetime]:
    """
    Bitrix обычно отдаёт ISO-строки. Делаем мягкий парсер.
    """
    if not v:
        return None
    try:
        s = str(v).strip()
        if not s:
            return None
        # поддержка "2026-02-05T10:20:30+03:00" и "2026-02-05 10:20:30"
        s = s.replace(" ", "T") if "T" not in s and "-" in s else s
        dt = datetime.datetime.fromisoformat(s)
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        return dt
    except Exception:
        return None


def _bitrix_user_get_all_for_report(select_fields: list[str]) -> list[dict]:
    """
    Пагинация Bitrix user.get через параметр start.
    Используем ваш BITRIX_USER_GET_URL.
    """
    all_users = []
    start = 0
    while True:
        payload = {
            "start": start,
            "filter": {},
            "select": select_fields,
        }
        js = _bitrix_call(BITRIX_USER_GET_URL, json_payload=payload, timeout=60)
        res = js.get("result") or []
        if isinstance(res, dict) and "result" in res:
            # на всякий случай
            res = res.get("result") or []
        if not isinstance(res, list):
            break
        all_users.extend(res)

        # Bitrix может возвращать next или total/start
        nxt = js.get("next")
        if nxt is None:
            # иногда next лежит в result
            nxt = (js.get("result") or {}).get("next") if isinstance(js.get("result"), dict) else None
        if nxt is None:
            # если ничего не знаем — прекращаем
            break
        try:
            start = int(nxt)
        except Exception:
            break
        if start <= 0:
            break

    return all_users


def _bitrix_fetch_inactive_users(days: int = 90) -> list[dict]:
    now = timezone.now()
    cutoff = now - timezone.timedelta(days=days)

    select_fields = [
        "ID", "NAME", "LAST_NAME", "SECOND_NAME",
        "EMAIL", "WORK_POSITION",
        "ACTIVE", "IS_ONLINE",
        "LAST_LOGIN", "LAST_ACTIVITY_DATE",
        BITRIX_INN_FIELD,  # ИНН
    ]
    users = _bitrix_user_get_all_for_report(select_fields=select_fields)

    out = []
    for u in users:
        # только активные по ТЗ
        if not _bx_is_active(u):
            continue

        uid = str(u.get("ID") or "").strip()
        fio = " ".join([x for x in [u.get("LAST_NAME"), u.get("NAME"), u.get("SECOND_NAME")] if x]).strip()
        email = (u.get("EMAIL") or "").strip()
        position = (u.get("WORK_POSITION") or "").strip()

        inn = (u.get(BITRIX_INN_FIELD) or "").strip()

        last_login_dt = _bx_parse_dt_any(u.get("LAST_LOGIN"))
        last_act_dt = _bx_parse_dt_any(u.get("LAST_ACTIVITY_DATE"))

        stale_login = (last_login_dt is None) or (last_login_dt < cutoff)
        stale_act = (last_act_dt is None) or (last_act_dt < cutoff)

        # Режим “и login, и activity давно” — меньше ложных срабатываний
        overdue = stale_login and stale_act
        if not overdue:
            continue

        def _days_ago(dt):
            if not dt:
                return None
            return (now - dt).days

        out.append({
            "id": uid,
            "fio": fio,
            "email": email,
            "position": position,
            "inn": inn,
            "last_login": _dt_iso(last_login_dt),
            "last_activity": _dt_iso(last_act_dt),
            "days_login": _days_ago(last_login_dt),
            "days_activity": _days_ago(last_act_dt),
        })

    # сортировка: кто дольше не заходил — выше
    def _k(r):
        v = r.get("days_login")
        return v if isinstance(v, int) else -1

    out.sort(key=_k, reverse=True)
    return out


def _xlsx_autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(60, max(12, max_len + 2))


def _build_report_xlsx(now: datetime.datetime,
                       ad_rows: list[dict],
                       sm_rows: list[dict],
                       bx_rows: list[dict],
                       days_ad: int, days_sm: int, days_bx: int) -> bytes:
    wb = Workbook()
    # удаляем дефолтный лист
    wb.remove(wb.active)

    # SUMMARY
    ws = wb.create_sheet("Общая статистика")
    ws.append(["Дата формирования", _dt_pretty(now)])
    ws.append(["Active directory дней без входа", days_ad])
    ws.append(["Супермаг центральная база дней без входа", days_sm])
    ws.append(["Битрикс дней без входа", days_bx])
    ws.append([])
    ws.append(["Найдено в AD", len(ad_rows)])
    ws.append(["Найдено в Супермаг центральный", len(sm_rows)])
    ws.append(["Найдено в Битрикс", len(bx_rows)])
    ws["A1"].font = Font(bold=True)
    _xlsx_autofit(ws)

    # AD
    ws = wb.create_sheet("AD")
    ws.append(["sam", "display", "mail", "upn", "employeeID(INN)", "department", "title", "lastLogonTimestamp", "daysSince"])
    for r in ad_rows:
        ws.append([
            r.get("sam"), r.get("display"), r.get("mail"), r.get("upn"),
            r.get("employeeID"), r.get("department"), r.get("title"),
            _dt_pretty_any(r.get("lastLogonTimestamp")), r.get("daysSince"),
        ])
    ws.freeze_panes = "A2"
    for c in ws[1]:
        c.font = Font(bold=True)
    _xlsx_autofit(ws)

    # SuperMag
    ws = wb.create_sheet("Супермаг центральный")
    ws.append(["db", "username", "fio", "inn", "serverlogin", "staff_id", "userenabled", "last_login", "created"])
    for r in sm_rows:
        ws.append([
            r.get("db"), r.get("username"), r.get("fio"), r.get("inn"),
            r.get("serverlogin"), r.get("staff_id"), r.get("userenabled"),
            r.get("last_login"), r.get("created"),
        ])
    ws.freeze_panes = "A2"
    for c in ws[1]:
        c.font = Font(bold=True)
    _xlsx_autofit(ws)

    # Bitrix
    ws = wb.create_sheet("Битрикс24")
    ws.append(["id", "fio", "email", "position", "inn", "last_login", "last_activity", "days_login", "days_activity"])
    for r in bx_rows:
        ws.append([
            r.get("id"), r.get("fio"), r.get("email"), r.get("position"), r.get("inn"),
            _dt_pretty_any(r.get("last_login")), _dt_pretty_any(r.get("last_activity")),
            r.get("days_login"), r.get("days_activity"),
        ])
    ws.freeze_panes = "A2"
    for c in ws[1]:
        c.font = Font(bold=True)
    _xlsx_autofit(ws)

    # CROSS_BY_INN
    ws = wb.create_sheet("Сводка по ИНН")
    ws.append(["inn", "fio_best", "ad_sam", "ad_last", "sm_username", "sm_last", "bitrix_id", "bitrix_last"])
    by_inn = {}

    def put(system: str, inn: str, payload: dict):
        if not inn:
            return
        inn = inn.strip()
        if inn not in by_inn:
            by_inn[inn] = {"inn": inn}
        by_inn[inn][system] = payload

    for r in ad_rows:
        put("ad", r.get("employeeID") or "", r)
    for r in sm_rows:
        put("sm", r.get("inn") or "", r)
    for r in bx_rows:
        put("bx", r.get("inn") or "", r)

    for inn, obj in sorted(by_inn.items(), key=lambda x: x[0]):
        ad = obj.get("ad") or {}
        sm = obj.get("sm") or {}
        bx = obj.get("bx") or {}
        fio_best = (sm.get("fio") or bx.get("fio") or ad.get("display") or "")
        ws.append([
            inn,
            fio_best,
            ad.get("sam"),
            _dt_pretty_any(ad.get("lastLogonTimestamp")),
            sm.get("username"),
            sm.get("last_login"),
            bx.get("id"),
            _dt_pretty_any(bx.get("last_login")),
        ])

    ws.freeze_panes = "A2"
    for c in ws[1]:
        c.font = Font(bold=True)
    _xlsx_autofit(ws)

    bio_out = io.BytesIO()
    wb.save(bio_out)
    return bio_out.getvalue()


def _bitrix_resolve_chat_id(chat_id: Optional[int], user_id: Optional[int], dialog_id: Optional[str]) -> int:
    """
    Нужен CHAT_ID для im.disk.*.
    Если передали user_id — получаем чат через im.dialog.get (DIALOG_ID=user_id).
    """
    if chat_id:
        return int(chat_id)

    did = None
    if dialog_id:
        did = str(dialog_id).strip()
    elif user_id:
        did = str(int(user_id))

    if not did:
        raise RuntimeError("chat_id/user_id/dialog_id is required")

    js = _bitrix_call(BITRIX_IM_DIALOG_GET_URL, json_payload={"DIALOG_ID": did}, timeout=30)
    r = js.get("result") or {}
    for k in ("chat_id", "CHAT_ID"):
        if k in r and str(r[k]).isdigit():
            return int(r[k])
    # иногда result.id бывает вида "chat3956"
    _id = str(r.get("id") or "")
    m = re.search(r"(\d+)$", _id)
    if m:
        return int(m.group(1))

    raise RuntimeError(f"Cannot resolve CHAT_ID via im.dialog.get for DIALOG_ID={did}")


def _bitrix_send_file_to_chat(chat_id: int, filename: str, content: bytes, message: str) -> dict:
    """
    1) im.disk.folder.get -> folder ID
    2) disk.folder.uploadfile -> uploadUrl
    3) POST multipart 'file' -> получаем DISK file ID
    4) im.disk.file.commit -> отправляем в чат
    """
    # 1) folder
    folder_js = _bitrix_call(BITRIX_IM_DISK_FOLDER_GET_URL, json_payload={"CHAT_ID": int(chat_id)}, timeout=30)
    folder = folder_js.get("result") or {}
    folder_id = folder.get("ID") or folder.get("id")
    if not folder_id:
        raise RuntimeError("im.disk.folder.get: no folder ID returned")

    # 2) uploadUrl
    up_js = _bitrix_call(
        BITRIX_DISK_FOLDER_UPLOAD_URL,
        json_payload={"id": int(folder_id), "data": {"NAME": filename}},
        timeout=30
    )
    up = up_js.get("result") or {}
    upload_url = up.get("uploadUrl") or up.get("uploadUrl".lower())
    if not upload_url:
        raise RuntimeError("disk.folder.uploadfile: no uploadUrl returned")

    # 3) upload file (multipart, field name 'file')
    mime = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    r = requests.post(
        upload_url,
        files={"file": (filename, io.BytesIO(content), mime)},
        timeout=120
    )
    r.raise_for_status()
    up_res = r.json() if "application/json" in (r.headers.get("Content-Type") or "") else {}
    res = up_res.get("result") if isinstance(up_res, dict) else None
    # пробуем достать ID максимально “мягко”
    disk_id = None
    if isinstance(res, dict):
        disk_id = res.get("ID") or res.get("id") or res.get("FILE_ID") or res.get("fileId")
    if not disk_id and isinstance(up_res, dict):
        disk_id = up_res.get("ID") or up_res.get("id")
    if not disk_id:
        # иногда файл может быть внутри result["FILE"]
        if isinstance(res, dict) and isinstance(res.get("FILE"), dict):
            disk_id = res["FILE"].get("ID") or res["FILE"].get("id")

    if not disk_id:
        raise RuntimeError(f"Upload done but DISK_ID not found in response: {up_res}")

    # 4) commit
    commit_js = _bitrix_call(
        BITRIX_IM_DISK_FILE_COMMIT_URL,
        json_payload={"CHAT_ID": int(chat_id), "DISK_ID": int(disk_id), "MESSAGE": message},
        timeout=30
    )
    return {"disk_id": int(disk_id), "commit": commit_js}


@csrf_exempt
@require_http_methods(["GET", "POST"])
def inactive_users_report_send_to_bitrix(request):
    """
    GET/POST /api/reports/inactive-users/send/?chat_id=3956
      headers: X-REPORT-TOKEN: <INACTIVE_REPORT_TOKEN>

    Параметры:
      chat_id=...            -> отправить в чат
      user_id=...            -> отправить “в личку” (через im.dialog.get получим CHAT_ID)
      dialog_id=...          -> альтернативно (userId или chatXXX)

      days_ad=60
      days_sm=100
      days_bx=90

      download=1             -> вместо отправки вернуть файл как ответ (тоже под токеном)
    """
    if not _require_report_token(request):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    chat_id = request.GET.get("chat_id") or request.POST.get("chat_id")
    user_id = request.GET.get("user_id") or request.POST.get("user_id")
    dialog_id = request.GET.get("dialog_id") or request.POST.get("dialog_id")

    chat_id_i = int(chat_id) if str(chat_id or "").isdigit() else None
    user_id_i = int(user_id) if str(user_id or "").isdigit() else None

    days_ad = _safe_int(request.GET.get("days_ad") or request.POST.get("days_ad"), 60, 1, 3650)
    days_sm = _safe_int(request.GET.get("days_sm") or request.POST.get("days_sm"), 100, 1, 3650)
    days_bx = _safe_int(request.GET.get("days_bx") or request.POST.get("days_bx"), 90, 1, 3650)

    include_never = (request.GET.get("include_never") or "1").strip().lower() in ("1", "true", "yes", "on")
    download = (request.GET.get("download") or "").strip().lower() in ("1", "true", "yes", "on")

    now = timezone.now()

    # 1) collect
    ad_rows = _ad_fetch_inactive_users(days=days_ad, include_never=include_never)
    # sm_rows = _sm_fetch_inactive_users_binu00(days=days_sm, only_enabled=True)
    exclude_env = (os.getenv("SM_LASTLOGIN_EXCLUDE_USERS", "S_BUDAYEV,SUPERMAG,SADMIN,SYSTEM,SYS") or "")
    exclude_users = {x.strip().upper() for x in exclude_env.split(",") if x.strip()}

    items, err = _sm_fetch_stale_users_in_db(
        db="BINUU00",
        days=days_sm,
        only_enabled=True,   # если хотите видеть всех — поставьте False
        q="",
        exclude_users=exclude_users
    )
    if err:
        logger.warning(f"[REPORT][SM] BINUU00 error: {err}")

    sm_rows = items
    bx_rows = _bitrix_fetch_inactive_users(days=days_bx)

    # 2) build xlsx
    content = _build_report_xlsx(
        now=now,
        ad_rows=ad_rows,
        sm_rows=sm_rows,
        bx_rows=bx_rows,
        days_ad=days_ad,
        days_sm=days_sm,
        days_bx=days_bx,
    )
    fname = f"Неактивные пользователи от {now.strftime('%Y%m%d_%H%M%S')}.xlsx"

    if download:
        resp = HttpResponse(
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

    # 3) send to bitrix
    try:
        resolved_chat_id = _bitrix_resolve_chat_id(chat_id_i, user_id_i, dialog_id)
        msg = (
            "Отчёт по неактивным пользователям:\n"
            f"-Active Directory: не входили ≥ {days_ad} дн (найдено: {len(ad_rows)})\n"
            f"-SuperMag (BINUU00): не входили ≥ {days_sm} дн (найдено: {len(sm_rows)})\n"
            f"-Bitrix24: не входили ≥ {days_bx} дн (найдено: {len(bx_rows)})\n"
            f"Сформировано: {now.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        send_res = _bitrix_send_file_to_chat(resolved_chat_id, fname, content, msg)

        return JsonResponse({
            "ok": True,
            "chat_id": resolved_chat_id,
            "counts": {
                "ad": len(ad_rows),
                "supermag": len(sm_rows),
                "bitrix": len(bx_rows),
            },
            "disk_id": send_res.get("disk_id"),
            "bitrix_commit": send_res.get("commit"),
        })
    except Exception as e:
        logger.exception(f"[REPORT] send failed: {e}")
        return JsonResponse({
            "ok": False,
            "error": str(e),
            "counts": {"ad": len(ad_rows), "supermag": len(sm_rows), "bitrix": len(bx_rows)},
        }, status=500)

















# ЭТО ДЛЯ ЗАПИСИ ИЗ БАЗ СУПЕРМАГА В НАШУ БД POSTGRESQL



# SM SYNC: helpers


_SMSTAFF_COL_CACHE: dict[str, dict[str, str]] = {}  # service_key -> {"inn": "INN", "login": "LOGIN", "store": "STORELOC"}
_ORA_CONN_CACHE: dict[str, Any] = {}               # service_key -> connection
_ORA_CONN_HOST_CACHE: dict[str, str] = {}          # service_key -> used_host




def _oracle_get_storeloc_by_dbname(service_key: str, dbname: str) -> Optional[int]:
    """
    Ищем storeloc по REP.DBNAME = <dbname> в рамках ЭТОЙ базы.
    """
    conn = _oracle_connect_by_service_key(service_key)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT storeloc
            FROM smstoreproperties
            WHERE propid = 'REP.DBNAME'
              AND UPPER(TRIM(propval)) = UPPER(TRIM(:db))
              AND ROWNUM = 1
        """, {"db": dbname})
        row = cur.fetchone()
        if not row or row[0] is None:
            return None
        return int(str(row[0]).strip())
    finally:
        try: cur.close()
        except Exception: pass



def _oracle_get_ukm_storeid_by_storeloc(service_key: str, storeloc: int) -> Optional[int]:
    """
    Берём REP.UKMStoreId для конкретного storeloc (в рамках ЭТОЙ базы).
    """
    conn = _oracle_connect_by_service_key(service_key)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT propval
            FROM smstoreproperties
            WHERE storeloc = :sid
              AND propid = 'REP.UKMStoreId'
              AND ROWNUM = 1
        """, {"sid": int(storeloc)})
        row = cur.fetchone()
        if not row or row[0] is None:
            return None
        return int(str(row[0]).strip())
    finally:
        try: cur.close()
        except Exception: pass






def _oracle_connect_by_service_key(service_key: str):
    """
    Подключение к Oracle по service_key из ORACLE_TNS_MAP.
    Кэшируем коннект на время запроса (в памяти процесса).
    """
    if service_key in _ORA_CONN_CACHE:
        return _ORA_CONN_CACHE[service_key]

    cfg = ORACLE_TNS_MAP.get(service_key)
    if not cfg:
        raise RuntimeError(f"ORACLE_TNS_MAP: неизвестный ключ {service_key!r}")

    ORA_USER = os.getenv("ORACLE_USER", "supermag")
    ORA_PASSWORD = os.getenv("ORACLE_PASSWORD", "qqq")
    port = int(cfg.get("port") or 1521)
    service_name = cfg.get("service_name") or service_key

    hosts = []
    if "hosts" in cfg and isinstance(cfg["hosts"], list):
        hosts = cfg["hosts"]
    else:
        hosts = [cfg.get("host")]

    last_err = None
    for host in hosts:
        if not host:
            continue
        try:
            dsn = cx_Oracle.makedsn(host, port, service_name=service_name)
            conn = _oracle_connect(user=ORA_USER, password=ORA_PASSWORD, dsn=dsn)
            _ORA_CONN_CACHE[service_key] = conn
            _ORA_CONN_HOST_CACHE[service_key] = host
            return conn
        except Exception as e:
            last_err = e

    raise RuntimeError(f"Не удалось подключиться к Oracle {service_key}: {last_err}")


def _oracle_close_cached_conns():
    for k, conn in list(_ORA_CONN_CACHE.items()):
        try:
            conn.close()
        except Exception:
            pass
    _ORA_CONN_CACHE.clear()
    _ORA_CONN_HOST_CACHE.clear()


def _load_smstaff_column_map(service_key: str) -> dict[str, str]:
    """
    Определяем имена колонок (inn/login/storeloc) для SMSTAFF динамически.
    login = serverlogin (и fallback на прочие варианты).
    """
    if service_key in _SMSTAFF_COL_CACHE:
        return _SMSTAFF_COL_CACHE[service_key]

    conn = _oracle_connect_by_service_key(service_key)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT column_name
            FROM all_tab_columns
            WHERE owner = 'SUPERMAG'
              AND table_name = 'SMSTAFF'
        """)
        cols = [str(r[0]).strip() for r in cur.fetchall()]
    finally:
        try: cur.close()
        except Exception: pass

    cols_l = {c.lower(): c for c in cols}

    # INN
    inn_candidates = ["inn", "inn_staff", "taxid", "tax_id"]
    inn_col = None
    for c in inn_candidates:
        if c in cols_l:
            inn_col = cols_l[c]
            break
    if not inn_col:
        for lc, orig in cols_l.items():
            if "inn" in lc:
                inn_col = orig
                break

    # LOGIN (ВАЖНО: serverlogin)
    login_candidates = [
        "serverlogin", "server_login",
        "login", "userlogin", "username", "userid",
        "user_name", "name_login", "usrlogin"
    ]
    login_col = None
    for c in login_candidates:
        if c in cols_l:
            login_col = cols_l[c]
            break
    if not login_col:
        for lc, orig in cols_l.items():
            if "login" in lc:
                login_col = orig
                break

    # STORELOC / STOREID (если вдруг есть в SMSTAFF)
    store_candidates = ["storeloc", "store_loc", "storelocid", "storeid", "store", "store_id"]
    store_col = None
    for c in store_candidates:
        if c in cols_l:
            store_col = cols_l[c]
            break

    if not inn_col:
        raise RuntimeError(f"[{service_key}] В SMSTAFF не найден столбец INN (ищу inn/..).")
    if not login_col:
        raise RuntimeError(f"[{service_key}] В SMSTAFF не найден столбец LOGIN (нужен serverlogin).")

    out = {"inn": inn_col, "login": login_col, "store": store_col or ""}
    _SMSTAFF_COL_CACHE[service_key] = out
    return out


def _find_staff_in_service_by_inn(service_key: str, inn: str) -> Optional[dict]:
    """
    Ищем SMSTAFF по inn в конкретной базе (service_key).
    Возвращаем { service, login, storeloc }.
    ROWNUM=1 оставляем (можно).
    """
    colmap = _load_smstaff_column_map(service_key)
    inn_col = colmap["inn"]
    login_col = colmap["login"]
    store_col = colmap["store"]  # может быть ""

    conn = _oracle_connect_by_service_key(service_key)
    cur = conn.cursor()
    try:
        if store_col:
            sql = f"""
                SELECT {login_col} AS login, {store_col} AS storeloc
                FROM smstaff
                WHERE {inn_col} = :inn
                  AND ROWNUM = 1
            """
        else:
            sql = f"""
                SELECT {login_col} AS login
                FROM smstaff
                WHERE {inn_col} = :inn
                  AND ROWNUM = 1
            """
        cur.execute(sql, {"inn": inn})
        row = cur.fetchone()
        if not row:
            return None

        login_str = ("" if row[0] is None else str(row[0])).strip()
        if not login_str:
            return None

        storeloc_int = None
        if store_col and len(row) > 1 and row[1] is not None:
            try:
                storeloc_int = int(str(row[1]).strip())
            except Exception:
                storeloc_int = None

        return {"service": service_key, "login": login_str, "storeloc": storeloc_int}
    finally:
        try: cur.close()
        except Exception: pass


def _ensure_open_in_system(*, user: User, username: str, system_id: int, dry_run: bool) -> dict:
    exists = OpenInSystem.objects.filter(user_id=user.id, username=username, system_id=system_id).exists()
    if exists:
        return {"ok": True, "created": False, "message": "open_in_system уже есть"}

    if dry_run:
        return {"ok": True, "created": False, "message": "DRY-RUN: создали бы open_in_system"}

    OpenInSystem.objects.create(
        user_id=user.id,
        username=username,
        password="",
        system_id=system_id,
        status=True
    )
    return {"ok": True, "created": True, "message": "open_in_system создан"}


def _ensure_ukm_user(*, user: User, storeid: int, roleid: int, dry_run: bool) -> dict:
    exists = UKMUser.objects.filter(user=user, storeid=int(storeid)).exists()
    if exists:
        return {"ok": True, "created": False, "message": "ukm_users уже есть"}

    if dry_run:
        return {"ok": True, "created": False, "message": "DRY-RUN: создали бы ukm_users"}

    UKMUser.objects.create(
        user=user,
        storeid=int(storeid),
        roleid=int(roleid),
        version=1
    )
    return {"ok": True, "created": True, "message": "ukm_users создан"}


def _iterate_services_all() -> list[str]:
    """
    Все базы из ORACLE_TNS_MAP. Можно оставить приоритет BINUU00 первым,
    но поиск идёт по ВСЕМ без break.
    """
    keys = list(ORACLE_TNS_MAP.keys())
    if "BINUU00" in keys:
        rest = sorted([k for k in keys if k != "BINUU00"])
        return ["BINUU00"] + rest
    return sorted(keys)


def _resolve_ukm_storeid_for_service(service_key: str, found_storeloc: Optional[int]) -> tuple[Optional[int], str]:
    """
    Определяем ukm storeid ДЛЯ ЭТОЙ базы:
      1) если SMSTAFF дал storeloc → пробуем PG stores(smstore=storeloc) → ukm4store/ukm5store
         если нет — берём REP.UKMStoreId по storeloc из Oracle
      2) если storeloc нет → storeloc по REP.DBNAME = service_key → REP.UKMStoreId
    """
    # 1) storeloc есть
    if found_storeloc is not None:
        sloc = int(found_storeloc)
        s = Store.objects.filter(smstore=sloc).first()
        if s and (s.ukm4store or s.ukm5store):
            return int(s.ukm4store or s.ukm5store), f"PG stores(smstore={sloc})"

        ukm_id = _oracle_get_ukm_storeid_by_storeloc(service_key, sloc)
        if ukm_id is not None:
            return int(ukm_id), f"Oracle REP.UKMStoreId(storeloc={sloc})"
        return None, f"storeloc={sloc}, но REP.UKMStoreId не найден"

    # 2) storeloc нет → через REP.DBNAME
    sloc = _oracle_get_storeloc_by_dbname(service_key, service_key)
    if sloc is None:
        return None, f"Не найден storeloc по REP.DBNAME={service_key}"
    ukm_id = _oracle_get_ukm_storeid_by_storeloc(service_key, int(sloc))
    if ukm_id is None:
        return None, f"storeloc={sloc}, но REP.UKMStoreId не найден"
    return int(ukm_id), f"Oracle REP.DBNAME->{sloc} then REP.UKMStoreId"



def _sync_one_user(user: User, dry_run: bool) -> dict:
    """
    ГЛАВНОЕ:
    - ищем ИНН в КАЖДОЙ базе
    - если нашли в базе → создаём ukm_users для storeid этой базы (каждый storeid)
    - BINUU00 не исключение: тоже создаём ukm_users
    """
    inn = (user.employee_id or "").strip()

    result = {
        "user_id": user.id,
        "fio": user.full_name,
        "employee_id": user.employee_id,
        "inn_used": inn,
        "found_any": False,
        "found_services": [],   # список находок по каждой базе
        "notes": [],
        "search_trace": [],
    }

    if not inn:
        result["notes"].append("Пустой employee_id — нечего искать")
        return result

    for service_key in _iterate_services_all():
        trace = {"service": service_key, "status": "try"}
        try:
            found = _find_staff_in_service_by_inn(service_key, inn)
            if not found:
                trace["status"] = "no"
                result["search_trace"].append(trace)
                continue

            trace["status"] = "found"
            result["search_trace"].append(trace)
            result["found_any"] = True

            login = found["login"]
            storeloc = found.get("storeloc")

            # open_in_system: BINUU00=4, остальные=5
            system_id = 4 if service_key == "BINUU00" else 5
            oi = _ensure_open_in_system(
                user=user,
                username=login,
                system_id=system_id,
                dry_run=dry_run
            )

            # storeid (UKM) именно этой базы
            ukm_storeid, store_source = _resolve_ukm_storeid_for_service(service_key, storeloc)

            if ukm_storeid is not None:
                uu = _ensure_ukm_user(user=user, storeid=int(ukm_storeid), roleid=1, dry_run=dry_run)
            else:
                uu = {"ok": False, "created": False, "message": "Не удалось определить UKM storeid"}

            result["found_services"].append({
                "service": service_key,
                "login": login,
                "system_id": system_id,
                "open_in_system": oi,
                "storeloc": storeloc,
                "ukm_storeid": ukm_storeid,
                "store_source": store_source,
                "ukm_users": uu,
            })

        except Exception as e:
            trace["status"] = "error"
            trace["error"] = str(e)
            result["search_trace"].append(trace)
            result["notes"].append(f"[{service_key}] ошибка: {e}")

    if not result["found_any"]:
        result["notes"].append("ИНН не найден ни в одной базе")

    return result


def _sse_pack(event: str, payload: dict) -> str:
    return "event: %s\ndata: %s\n\n" % (event, json.dumps(payload, ensure_ascii=False))



# SM SYNC: UI + Stream + API


@staff_member_required
@never_cache
@require_GET
def sm_sync_staff_ui(request):
    """
    UI-страница с кнопками ТЕСТ / НАСТОЯЩИЙ.
    """
    return render(request, "frostapp/sm_sync_staff.html", {
        "stream_url": reverse("sm_sync_staff_stream"),
    })


@staff_member_required
@never_cache
@require_GET
def sm_sync_staff_stream(request):
    """
    SSE stream:
      GET /sm/sync-staff/stream/?mode=test|real

    mode=real пишет в Postgres.
    """
    mode = (request.GET.get("mode") or "test").strip().lower()
    dry_run = (mode != "real")

    def gen():
        start = tz_now()
        summary = {
            "mode": mode,
            "dry_run": dry_run,
            "users_total": 0,
            "found_total": 0,         # users found in ANY base
            "not_found_total": 0,     # users found in NO bases
            "open_created": 0,        # total created rows in open_in_system
            "ukm_created": 0,         # total created rows in ukm_users
            "errors": 0,
            "started_at": start.isoformat(),
        }

        yield _sse_pack("hello", {
            "message": "Старт синхронизации",
            "mode": mode,
            "dry_run": dry_run,
            "ts": start.isoformat(),
        })

        try:
            qs = User.objects.all().order_by("id")
            summary["users_total"] = qs.count()
            yield _sse_pack("meta", summary)

            for u in qs.iterator(chunk_size=200):
                try:
                    if dry_run:
                        res = _sync_one_user(u, dry_run=True)
                    else:
                        with transaction.atomic():
                            res = _sync_one_user(u, dry_run=False)

                    if res.get("found_any"):
                        summary["found_total"] += 1
                    else:
                        summary["not_found_total"] += 1

                    # считаем созданные записи по всем найденным базам
                    for item in (res.get("found_services") or []):
                        oi = (item.get("open_in_system") or {})
                        if oi.get("created"):
                            summary["open_created"] += 1
                        uu = (item.get("ukm_users") or {})
                        if uu.get("created"):
                            summary["ukm_created"] += 1

                    yield _sse_pack("user", res)
                    yield _sse_pack("meta", summary)

                except Exception as e:
                    summary["errors"] += 1
                    yield _sse_pack("log", {
                        "level": "error",
                        "user_id": u.id,
                        "message": str(e),
                    })
                    yield _sse_pack("meta", summary)

            finish = tz_now()
            summary["finished_at"] = finish.isoformat()
            summary["duration_sec"] = int((finish - start).total_seconds())
            yield _sse_pack("done", summary)

        finally:
            _oracle_close_cached_conns()

    resp = StreamingHttpResponse(gen(), content_type="text/event-stream; charset=utf-8")
    resp["Cache-Control"] = "no-cache"
    resp["X-Accel-Buffering"] = "no"
    return resp


@csrf_exempt
@agent_token_required
@require_POST
def sm_sync_staff_run(request):
    """
    Скриптовый запуск:
      POST /api/sm/sync-staff/run/
      Headers: X-AGENT-TOKEN: ...
      Body: {"mode":"test"|"real"}

    Возвращает JSON с итоговой сводкой + примеры первых N результатов.
    (Полный список можно сделать, но он может быть огромный.)
    """
    try:
        data = json.loads(request.body.decode("utf-8") if request.body else "{}")
    except Exception:
        data = {}

    mode = (data.get("mode") or "test").strip().lower()
    dry_run = (mode != "real")

    start = tz_now()
    summary = {
        "mode": mode,
        "dry_run": dry_run,
        "users_total": 0,
        "found_total": 0,         # users found in ANY base
        "not_found_total": 0,     # users found in NO bases
        "open_created": 0,        # total created rows in open_in_system (across all bases)
        "ukm_created": 0,         # total created rows in ukm_users (across all bases)
        "errors": 0,
        "started_at": start.isoformat(),
        "sample_results": [],
    }

    try:
        qs = User.objects.all().order_by("id")
        summary["users_total"] = qs.count()

        sample_limit = 200

        for u in qs.iterator(chunk_size=200):
            try:
                if dry_run:
                    res = _sync_one_user(u, dry_run=True)
                else:
                    with transaction.atomic():
                        res = _sync_one_user(u, dry_run=False)

                # ВАЖНО: теперь используется found_any + found_services
                if res.get("found_any"):
                    summary["found_total"] += 1
                else:
                    summary["not_found_total"] += 1

                for item in (res.get("found_services") or []):
                    oi = (item.get("open_in_system") or {})
                    if oi.get("created"):
                        summary["open_created"] += 1

                    uu = (item.get("ukm_users") or {})
                    if uu.get("created"):
                        summary["ukm_created"] += 1

                if len(summary["sample_results"]) < sample_limit:
                    summary["sample_results"].append(res)

            except Exception as e:
                summary["errors"] += 1
                logger.exception(f"[SM_SYNC_STAFF_RUN] user_id={u.id} error: {e}")

        finish = tz_now()
        summary["finished_at"] = finish.isoformat()
        summary["duration_sec"] = int((finish - start).total_seconds())

        return JsonResponse(
            {"status": "ok", "summary": summary},
            json_dumps_params={"ensure_ascii": False, "indent": 2}
        )

    finally:
        _oracle_close_cached_conns()






























































def maxbot_api_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        token = request.headers.get("X-MAXBOT-TOKEN") or request.META.get("HTTP_X_MAXBOT_TOKEN")
        if not token or token != MAX_BOT_INTERNAL_TOKEN:
            return JsonResponse({"ok": False, "error": "forbidden"}, status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped


def parse_json_body(request):
    try:
        return json.loads(request.body.decode("utf-8") if request.body else "{}")
    except Exception:
        return {}


def normalize_phone(raw_phone: str) -> str:
    digits = re.sub(r"\D", "", raw_phone or "")
    if len(digits) == 10:
        digits = "7" + digits
    elif len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    return digits


def find_user_by_phone(raw_phone: str) -> Optional[User]:
    phone = normalize_phone(raw_phone)
    if not phone:
        return None

    last10 = phone[-10:]

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT id
            FROM users
            WHERE active = TRUE
              AND (
                    regexp_replace(coalesce(phone, ''), '\D', '', 'g') = %s
                 OR right(regexp_replace(coalesce(phone, ''), '\D', '', 'g'), 10) = %s
              )
            ORDER BY id
            LIMIT 1
            """,
            [phone, last10],
        )
        row = cursor.fetchone()

    if not row:
        return None

    return User.objects.filter(pk=row[0]).first()


def find_user_by_max_id(max_user_id: int) -> Optional[User]:
    return User.objects.filter(max_id=max_user_id, active=True).first()


def bind_max_user_to_phone(max_user_id: int, raw_phone: str) -> Tuple[Optional[User], Optional[str]]:
    user = find_user_by_phone(raw_phone)
    if not user:
        return None, "Пользователь с таким номером телефона не найден в базе."

    if user.max_id and int(user.max_id) != int(max_user_id):
        return None, "Этот номер телефона уже привязан к другому MAX-аккаунту."

    if not user.max_id:
        User.objects.filter(pk=user.id).update(
            max_id=max_user_id,
            updated_at=timezone.now(),
        )
        user.refresh_from_db()

    return user, None


def kb_callback(text: str, payload: str):
    return {"type": "callback", "text": text, "payload": payload}


def kb_contact(text: str):
    return {"type": "request_contact", "text": text}


def build_response(text=None, buttons=None, notification=None, fmt="html"):
    return {
        "ok": True,
        "text": text,
        "buttons": buttons or [],
        "notification": notification,
        "format": fmt,
    }


def build_auth_response():
    return build_response(
        text=(
            "Для входа в бот нажми кнопку <b>«Отправить контакт»</b>.\n\n"
            "Я сверю номер телефона с таблицей <code>users</code> и привяжу твой MAX ID."
        ),
        buttons=[[kb_contact("Отправить контакт")]],
    )


def build_main_menu(user: User):
    return build_response(
        text=(
            f"Привет, <b>{escape(user.full_name)}</b>!\n\n"
            "Нажми <b>«Начать»</b>, чтобы пройти сценарий и сохранить запись в реестр."
        ),
        buttons=[
            [kb_callback("Начать", "start")],
        ],
    )


def generate_request_no() -> str:
    return f"MAX-{timezone.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"


def get_open_request(max_user_id: int) -> Optional[MaxBotRequest]:
    return (
        MaxBotRequest.objects
        .filter(
            max_user_id=max_user_id,
            status__in=[
                "draft",
                "awaiting_role",
                "awaiting_employee",
                "awaiting_vehicle",
                "awaiting_answer",
            ],
        )
        .order_by("-created_at")
        .first()
    )


def cancel_previous_open_requests(max_user_id: int):
    MaxBotRequest.objects.filter(
        max_user_id=max_user_id,
        status__in=[
            "draft",
            "awaiting_role",
            "awaiting_employee",
            "awaiting_vehicle",
            "awaiting_answer",
        ],
    ).update(
        status="cancelled",
        updated_at=timezone.now(),
        finished_at=timezone.now(),
    )


def create_new_request_for_user(user: User, max_user_id: int, max_chat_id: Optional[int], raw_event=None) -> MaxBotRequest:
    cancel_previous_open_requests(max_user_id)

    req = MaxBotRequest.objects.create(
        request_no=generate_request_no(),
        applicant_user=user,
        applicant_full_name=user.full_name,
        max_user_id=max_user_id,
        max_chat_id=max_chat_id,
        status="awaiting_role",
        raw_last_event=raw_event or {},
        created_at=timezone.now(),
        updated_at=timezone.now(),
    )
    return req


def get_active_roles():
    return MaxBotRole.objects.filter(is_active=True).order_by("sort_order", "name")


def get_active_scenario_for_role(role: MaxBotRole) -> Optional[MaxBotScenario]:
    return (
        MaxBotScenario.objects
        .filter(role=role, is_active=True)
        .order_by("id")
        .first()
    )


def ask_role(req: MaxBotRequest):
    roles = list(get_active_roles())
    buttons = [[kb_callback(role.name, f"role:{role.id}")] for role in roles]
    buttons.append([kb_callback("Отмена", "cancel")])

    return build_response(
        text="Выберите должность:",
        buttons=buttons,
    )


def ask_employee(req: MaxBotRequest, role: MaxBotRole):
    employees = list(
        MaxBotEmployee.objects.filter(role=role, is_active=True).order_by("sort_order", "full_name")
    )
    if not employees:
        return build_response(
            text="Для этой должности пока нет сотрудников в справочнике.",
            buttons=[[kb_callback("В главное меню", "main_menu")]],
        )

    buttons = [[kb_callback(emp.full_name, f"employee:{emp.id}")] for emp in employees]
    buttons.append([kb_callback("Отмена", "cancel")])

    return build_response(
        text="Выберите ФИО:",
        buttons=buttons,
    )


def ask_vehicle(req: MaxBotRequest, role: MaxBotRole, employee: Optional[MaxBotEmployee] = None):
    qs = MaxBotVehicle.objects.filter(role=role, is_active=True).order_by("sort_order", "reg_number")
    if employee:
        filtered = qs.filter(Q(employee__isnull=True) | Q(employee=employee))
    else:
        filtered = qs

    vehicles = list(filtered)
    if not vehicles:
        return build_response(
            text="Для этой должности пока нет автомобилей в справочнике.",
            buttons=[[kb_callback("В главное меню", "main_menu")]],
        )

    buttons = [[kb_callback(v.title or v.reg_number, f"vehicle:{v.id}")] for v in vehicles]
    buttons.append([kb_callback("Отмена", "cancel")])

    return build_response(
        text="Выберите автомобиль:",
        buttons=buttons,
    )


def ask_question(req: MaxBotRequest, question: MaxBotQuestion, notification: Optional[str] = None):
    req.current_question = question
    req.status = "awaiting_answer"
    req.updated_at = timezone.now()
    req.save(update_fields=["current_question", "status", "updated_at"])

    if question.question_type == "single_choice":
        options = list(question.options.filter(is_active=True).order_by("sort_order", "id"))
        buttons = [[kb_callback(opt.text, f"answer:{opt.id}")] for opt in options]
        buttons.append([kb_callback("Отмена", "cancel")])
        return build_response(text=question.text, buttons=buttons, notification=notification)

    if question.question_type == "photo":
        return build_response(
            text=question.text,
            buttons=[[kb_callback("Отмена", "cancel")]],
            notification=notification,
        )

    if question.question_type in ("text", "number"):
        return build_response(
            text=question.text,
            buttons=[[kb_callback("Отмена", "cancel")]],
            notification=notification,
        )

    return build_response(
        text="Тип вопроса не поддерживается.",
        buttons=[[kb_callback("В главное меню", "main_menu")]],
    )


def rebuild_request_summary(req: MaxBotRequest) -> str:
    lines = []
    answers = req.answers.order_by("created_at", "id")

    for ans in answers:
        if ans.registry_action_text:
            lines.append(ans.registry_action_text)
            continue

        if ans.option_text:
            lines.append(f"{ans.question_text}: {ans.option_text}")
            continue

        if ans.answer_text:
            lines.append(f"{ans.question_text}: {ans.answer_text}")
            continue

        if ans.answer_number is not None:
            lines.append(f"{ans.question_text}: {ans.answer_number}")
            continue

        if ans.photo_file or ans.photo_url:
            lines.append(f"{ans.question_text}: приложено фото")
            continue

    return "\n".join(lines)


def finish_request(req: MaxBotRequest, status: str = "completed", notification: Optional[str] = None):
    req.status = status
    req.current_question = None
    req.summary = rebuild_request_summary(req)
    req.finished_at = timezone.now()
    req.updated_at = timezone.now()
    req.save(update_fields=["status", "current_question", "summary", "finished_at", "updated_at"])

    if status == "emergency_stop":
        text = (
            f"Заявка <b>{req.request_no}</b> сохранена.\n\n"
            f"<b>СТОП!</b> Машину не выгонять из гаража, ставим на ремонт.\n\n"
            f"<b>Итог:</b>\n{escape(req.summary or '-')}"
        )
    else:
        text = (
            f"Готово. Заявка <b>{req.request_no}</b> сохранена в реестр.\n\n"
            f"<b>Итог:</b>\n{escape(req.summary or '-')}"
        )

    return build_response(
        text=text,
        buttons=[[kb_callback("Начать заново", "start")]],
        notification=notification,
    )


def save_photo_to_answer(answer: MaxBotRequestAnswer, photo_url: str):
    answer.photo_url = photo_url
    if not photo_url:
        answer.save(update_fields=["photo_url"])
        return

    try:
        resp = requests.get(photo_url, timeout=20)
        resp.raise_for_status()

        content_type = (resp.headers.get("Content-Type") or "").split(";")[0].strip()
        ext = mimetypes.guess_extension(content_type) or ".jpg"
        filename = f"{uuid.uuid4().hex}{ext}"

        answer.photo_file.save(filename, ContentFile(resp.content), save=False)
    except Exception:
        logger.exception("Не удалось скачать фото из MAX URL: %s", photo_url)

    answer.save()


def start_scenario_after_directory(req: MaxBotRequest):
    if not req.scenario or not req.scenario.first_question:
        return build_response(
            text="Для выбранной должности сценарий ещё не настроен.",
            buttons=[[kb_callback("В главное меню", "main_menu")]],
        )

    return ask_question(req, req.scenario.first_question)


def handle_role_selection(req: MaxBotRequest, role_id: int):
    role = MaxBotRole.objects.filter(id=role_id, is_active=True).first()
    if not role:
        return build_response(text="Должность не найдена.", buttons=[[kb_callback("В главное меню", "main_menu")]])

    scenario = get_active_scenario_for_role(role)
    req.role = role
    req.scenario = scenario
    req.updated_at = timezone.now()

    if role.requires_employee:
        req.status = "awaiting_employee"
        req.save(update_fields=["role", "scenario", "status", "updated_at"])
        return ask_employee(req, role)

    if role.requires_vehicle:
        req.status = "awaiting_vehicle"
        req.save(update_fields=["role", "scenario", "status", "updated_at"])
        return ask_vehicle(req, role, None)

    req.save(update_fields=["role", "scenario", "updated_at"])
    return start_scenario_after_directory(req)


def handle_employee_selection(req: MaxBotRequest, employee_id: int):
    if not req.role:
        return build_response(text="Сначала выберите должность.", buttons=[[kb_callback("В главное меню", "main_menu")]])

    employee = MaxBotEmployee.objects.filter(id=employee_id, role=req.role, is_active=True).first()
    if not employee:
        return build_response(text="Сотрудник не найден.", buttons=[[kb_callback("В главное меню", "main_menu")]])

    req.employee = employee
    req.updated_at = timezone.now()

    if req.role.requires_vehicle:
        req.status = "awaiting_vehicle"
        req.save(update_fields=["employee", "status", "updated_at"])
        return ask_vehicle(req, req.role, employee)

    req.save(update_fields=["employee", "updated_at"])
    return start_scenario_after_directory(req)


def handle_vehicle_selection(req: MaxBotRequest, vehicle_id: int):
    if not req.role:
        return build_response(text="Сначала выберите должность.", buttons=[[kb_callback("В главное меню", "main_menu")]])

    vehicle_qs = MaxBotVehicle.objects.filter(id=vehicle_id, role=req.role, is_active=True)
    if req.employee:
        vehicle_qs = vehicle_qs.filter(Q(employee__isnull=True) | Q(employee=req.employee))

    vehicle = vehicle_qs.first()
    if not vehicle:
        return build_response(text="Автомобиль не найден.", buttons=[[kb_callback("В главное меню", "main_menu")]])

    req.vehicle = vehicle
    req.updated_at = timezone.now()
    req.save(update_fields=["vehicle", "updated_at"])

    return start_scenario_after_directory(req)


def save_option_answer(req: MaxBotRequest, option: MaxBotQuestionOption, raw_payload=None):
    question = req.current_question
    if not question:
        return build_response(text="Нет активного вопроса.", buttons=[[kb_callback("В главное меню", "main_menu")]])

    if option.question_id != question.id:
        return ask_question(req, question, notification="Этот ответ уже не актуален. Повтори выбор.")

    MaxBotRequestAnswer.objects.create(
        request=req,
        question=question,
        option=option,
        question_text=question.text,
        option_text=option.text,
        answer_number=option.numeric_value,
        registry_action_text=option.registry_action_text,
        is_emergency=option.is_emergency,
        raw_payload=raw_payload or {},
        created_at=timezone.now(),
    )

    if option.is_emergency:
        req.emergency_flag = True

    next_question = option.next_question or question.default_next_question
    req.updated_at = timezone.now()
    req.raw_last_event = raw_payload or {}

    if option.request_status_on_select == "emergency_stop":
        req.save(update_fields=["emergency_flag", "updated_at", "raw_last_event"])
        return finish_request(req, status="emergency_stop", notification=option.notification_text)

    if option.is_finish and not next_question:
        req.save(update_fields=["emergency_flag", "updated_at", "raw_last_event"])
        return finish_request(req, status="completed", notification=option.notification_text)

    req.save(update_fields=["emergency_flag", "updated_at", "raw_last_event"])

    if next_question:
        return ask_question(req, next_question, notification=option.notification_text)

    return finish_request(req, status="completed", notification=option.notification_text)


def save_text_or_number_answer(req: MaxBotRequest, text_value: str, raw_payload=None):
    question = req.current_question
    if not question:
        return build_response(text="Нет активного вопроса.", buttons=[[kb_callback("В главное меню", "main_menu")]])

    answer_number = None
    answer_text = text_value

    if question.question_type == "number":
        try:
            normalized = (text_value or "").replace(",", ".").strip()
            answer_number = Decimal(normalized)
            answer_text = None
        except (InvalidOperation, AttributeError):
            return build_response(
                text="Нужно отправить число. Например: 1 или 1.5",
                buttons=[[kb_callback("Отмена", "cancel")]],
            )

    MaxBotRequestAnswer.objects.create(
        request=req,
        question=question,
        question_text=question.text,
        answer_text=answer_text,
        answer_number=answer_number,
        raw_payload=raw_payload or {},
        created_at=timezone.now(),
    )

    req.updated_at = timezone.now()
    req.raw_last_event = raw_payload or {}
    req.save(update_fields=["updated_at", "raw_last_event"])

    next_question = question.default_next_question
    if next_question:
        return ask_question(req, next_question)

    return finish_request(req, status="completed")


def save_photo_answer(req: MaxBotRequest, photo_url: str, raw_payload=None):
    question = req.current_question
    if not question:
        return build_response(text="Нет активного вопроса.", buttons=[[kb_callback("В главное меню", "main_menu")]])

    if question.question_type != "photo":
        return build_response(text="Сейчас бот не ждёт фото.", buttons=[[kb_callback("В главное меню", "main_menu")]])

    ans = MaxBotRequestAnswer.objects.create(
        request=req,
        question=question,
        question_text=question.text,
        registry_action_text="Приложено фото",
        raw_payload=raw_payload or {},
        created_at=timezone.now(),
    )
    save_photo_to_answer(ans, photo_url)

    req.updated_at = timezone.now()
    req.raw_last_event = raw_payload or {}
    req.save(update_fields=["updated_at", "raw_last_event"])

    next_question = question.default_next_question
    if next_question:
        return ask_question(req, next_question)

    return finish_request(req, status="completed")


def handle_callback_event(user: User, payload: str, max_user_id: int, max_chat_id: Optional[int], raw_data: dict):
    payload = (payload or "").strip()

    if payload == "main_menu":
        return build_main_menu(user)

    if payload == "start":
        req = create_new_request_for_user(user, max_user_id, max_chat_id, raw_event=raw_data)
        return ask_role(req)

    if payload == "cancel":
        req = get_open_request(max_user_id)
        if req:
            req.status = "cancelled"
            req.updated_at = timezone.now()
            req.finished_at = timezone.now()
            req.save(update_fields=["status", "updated_at", "finished_at"])
        return build_response(
            text="Текущий сценарий отменён.",
            buttons=[[kb_callback("Начать заново", "start")]],
            notification="Отменено",
        )

    req = get_open_request(max_user_id)
    if not req:
        return build_response(
            text="Нет активного сценария. Нажми «Начать».",
            buttons=[[kb_callback("Начать", "start")]],
        )

    if payload.startswith("role:"):
        role_id = int(payload.split(":", 1)[1])
        return handle_role_selection(req, role_id)

    if payload.startswith("employee:"):
        employee_id = int(payload.split(":", 1)[1])
        return handle_employee_selection(req, employee_id)

    if payload.startswith("vehicle:"):
        vehicle_id = int(payload.split(":", 1)[1])
        return handle_vehicle_selection(req, vehicle_id)

    if payload.startswith("answer:"):
        option_id = int(payload.split(":", 1)[1])
        option = MaxBotQuestionOption.objects.filter(id=option_id, is_active=True).first()
        if not option:
            return build_response(
                text="Вариант ответа не найден.",
                buttons=[[kb_callback("В главное меню", "main_menu")]],
            )
        return save_option_answer(req, option, raw_payload=raw_data)

    return build_response(
        text="Неизвестная команда.",
        buttons=[[kb_callback("В главное меню", "main_menu")]],
    )


@csrf_exempt
@require_POST
@maxbot_api_required
def maxbot_process_update(request):
    data = parse_json_body(request)

    event_type = (data.get("event_type") or "").strip()
    max_user_id = data.get("max_user_id")
    max_chat_id = data.get("chat_id")
    text = (data.get("text") or "").strip()
    phone = (data.get("phone") or "").strip()
    payload = (data.get("payload") or "").strip()
    photo_url = (data.get("photo_url") or "").strip()
    raw_update = data.get("raw_update") or data

    if not max_user_id:
        return JsonResponse({"ok": False, "error": "max_user_id is required"}, status=400)

    max_user_id = int(max_user_id)
    if max_chat_id is not None:
        try:
            max_chat_id = int(max_chat_id)
        except Exception:
            max_chat_id = None

    user = find_user_by_max_id(max_user_id)

    # 1. Пользователь ещё не привязан
    if not user:
        if event_type == "contact_shared":
            user, error = bind_max_user_to_phone(max_user_id, phone)
            if error:
                return JsonResponse(build_response(
                    text=(
                        f"{escape(error)}\n\n"
                        "Нажми кнопку «Отправить контакт» ещё раз или обратись к администратору."
                    ),
                    buttons=[[kb_contact("Отправить контакт")]],
                ))

            return JsonResponse(build_main_menu(user))

        return JsonResponse(build_auth_response())

    # 2. Уже привязан
    if event_type == "bot_started":
        return JsonResponse(build_main_menu(user))

    if event_type == "contact_shared":
        return JsonResponse(build_main_menu(user))

    if event_type == "callback":
        result = handle_callback_event(user, payload, max_user_id, max_chat_id, raw_update)
        return JsonResponse(result)

    req = get_open_request(max_user_id)

    if event_type == "photo_message":
        if not req:
            return JsonResponse(build_response(
                text="Нет активного сценария. Нажми «Начать».",
                buttons=[[kb_callback("Начать", "start")]],
            ))

        return JsonResponse(save_photo_answer(req, photo_url, raw_payload=raw_update))

    if event_type == "message":
        if text.lower() in {"/start", "start", "начать"}:
            return JsonResponse(build_main_menu(user))

        if not req:
            return JsonResponse(build_main_menu(user))

        if not req.current_question:
            return JsonResponse(build_main_menu(user))

        if req.current_question.question_type in ("text", "number"):
            return JsonResponse(save_text_or_number_answer(req, text, raw_payload=raw_update))

        if req.current_question.question_type == "photo":
            return JsonResponse(build_response(
                text="Сейчас нужно отправить именно фото.",
                buttons=[[kb_callback("Отмена", "cancel")]],
            ))

        return JsonResponse(build_response(
            text="Пожалуйста, используй кнопки под сообщением.",
            buttons=[[kb_callback("Отмена", "cancel")]],
        ))

    return JsonResponse(build_main_menu(user))



































MAXBOT_REQUEST_STATUS_LABELS = {
    "draft": "Черновик",
    "awaiting_role": "Ожидание выбора должности",
    "awaiting_employee": "Ожидание выбора сотрудника",
    "awaiting_vehicle": "Ожидание выбора машины",
    "awaiting_answer": "Ожидание ответа",
    "completed": "Завершено",
    "emergency_stop": "Аварийный стоп",
    "cancelled": "Отменено",
}


def _maxbot_stamp_and_save(obj):
    now = timezone.now()
    if hasattr(obj, "created_at") and not getattr(obj, "created_at", None):
        obj.created_at = now
    if hasattr(obj, "updated_at"):
        obj.updated_at = now
    obj.save()
    return obj


def _maxbot_paginate(request, queryset, per_page=25):
    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get("page")
    return paginator.get_page(page_number)


@staff_member_required
def maxbot_dashboard(request):
    total_roles = MaxBotRole.objects.count()
    total_employees = MaxBotEmployee.objects.count()
    total_vehicles = MaxBotVehicle.objects.count()
    total_scenarios = MaxBotScenario.objects.count()

    total_requests = MaxBotRequest.objects.count()
    emergency_requests = MaxBotRequest.objects.filter(emergency_flag=True).count()
    active_requests = MaxBotRequest.objects.filter(
        status__in=["draft", "awaiting_role", "awaiting_employee", "awaiting_vehicle", "awaiting_answer"]
    ).count()
    completed_requests = MaxBotRequest.objects.filter(status="completed").count()

    latest_requests = (
        MaxBotRequest.objects
        .select_related("role", "employee", "vehicle")
        .order_by("-created_at")[:10]
    )

    for req in latest_requests:
        req.status_label = MAXBOT_REQUEST_STATUS_LABELS.get(req.status, req.status)

    return render(request, "maxbot/dashboard.html", {
        "total_roles": total_roles,
        "total_employees": total_employees,
        "total_vehicles": total_vehicles,
        "total_scenarios": total_scenarios,
        "total_requests": total_requests,
        "emergency_requests": emergency_requests,
        "active_requests": active_requests,
        "completed_requests": completed_requests,
        "latest_requests": latest_requests,
    })


# ---------- Roles ----------

@staff_member_required
def maxbot_role_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = MaxBotRole.objects.all().order_by("sort_order", "name")
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))

    page_obj = _maxbot_paginate(request, qs, per_page=30)

    return render(request, "maxbot/role_list.html", {
        "page_obj": page_obj,
        "q": q,
    })


@staff_member_required
def maxbot_role_create(request):
    if request.method == "POST":
        form = MaxBotRoleForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_role_list")
    else:
        form = MaxBotRoleForm()

    return render(request, "maxbot/object_form.html", {
        "title": "Новая должность",
        "form": form,
        "back_url": reverse("maxbot_role_list"),
    })


@staff_member_required
def maxbot_role_edit(request, pk):
    obj = get_object_or_404(MaxBotRole, pk=pk)

    if request.method == "POST":
        form = MaxBotRoleForm(request.POST, instance=obj)
        if form.is_valid():
            obj = form.save(commit=False)
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_role_list")
    else:
        form = MaxBotRoleForm(instance=obj)

    return render(request, "maxbot/object_form.html", {
        "title": f"Редактирование должности: {obj.name}",
        "form": form,
        "back_url": reverse("maxbot_role_list"),
    })


@staff_member_required
def maxbot_role_delete(request, pk):
    obj = get_object_or_404(MaxBotRole, pk=pk)

    if request.method == "POST":
        obj.delete()
        return redirect("maxbot_role_list")

    return render(request, "maxbot/confirm_delete.html", {
        "title": "Удаление должности",
        "object_name": obj.name,
        "back_url": reverse("maxbot_role_list"),
    })


# ---------- Employees ----------

@staff_member_required
def maxbot_employee_list(request):
    q = (request.GET.get("q") or "").strip()
    role_id = (request.GET.get("role_id") or "").strip()

    qs = (
        MaxBotEmployee.objects
        .select_related("role", "user")
        .order_by("role__sort_order", "sort_order", "full_name")
    )

    if q:
        qs = qs.filter(Q(full_name__icontains=q) | Q(phone__icontains=q))
    if role_id:
        qs = qs.filter(role_id=role_id)

    page_obj = _maxbot_paginate(request, qs, per_page=30)

    return render(request, "maxbot/employee_list.html", {
        "page_obj": page_obj,
        "q": q,
        "role_id": role_id,
        "roles": MaxBotRole.objects.all().order_by("sort_order", "name"),
    })


@staff_member_required
def maxbot_employee_create(request):
    if request.method == "POST":
        form = MaxBotEmployeeForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_employee_list")
    else:
        form = MaxBotEmployeeForm()

    return render(request, "maxbot/object_form.html", {
        "title": "Новый сотрудник",
        "form": form,
        "back_url_name": "maxbot_employee_list",
    })


@staff_member_required
def maxbot_employee_edit(request, pk):
    obj = get_object_or_404(MaxBotEmployee, pk=pk)

    if request.method == "POST":
        form = MaxBotEmployeeForm(request.POST, instance=obj)
        if form.is_valid():
            obj = form.save(commit=False)
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_employee_list")
    else:
        form = MaxBotEmployeeForm(instance=obj)

    return render(request, "maxbot/object_form.html", {
        "title": f"Редактирование сотрудника: {obj.full_name}",
        "form": form,
        "back_url_name": "maxbot_employee_list",
    })


@staff_member_required
def maxbot_employee_delete(request, pk):
    obj = get_object_or_404(MaxBotEmployee, pk=pk)

    if request.method == "POST":
        obj.delete()
        return redirect("maxbot_employee_list")

    return render(request, "maxbot/confirm_delete.html", {
        "title": "Удаление сотрудника",
        "object_name": obj.full_name,
        "back_url_name": "maxbot_employee_list",
    })


# ---------- Vehicles ----------

@staff_member_required
def maxbot_vehicle_list(request):
    q = (request.GET.get("q") or "").strip()
    role_id = (request.GET.get("role_id") or "").strip()

    qs = (
        MaxBotVehicle.objects
        .select_related("role", "employee")
        .order_by("role__sort_order", "sort_order", "reg_number")
    )

    if q:
        qs = qs.filter(Q(reg_number__icontains=q) | Q(title__icontains=q))
    if role_id:
        qs = qs.filter(role_id=role_id)

    page_obj = _maxbot_paginate(request, qs, per_page=30)

    return render(request, "maxbot/vehicle_list.html", {
        "page_obj": page_obj,
        "q": q,
        "role_id": role_id,
        "roles": MaxBotRole.objects.all().order_by("sort_order", "name"),
    })


@staff_member_required
def maxbot_vehicle_create(request):
    if request.method == "POST":
        form = MaxBotVehicleForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_vehicle_list")
    else:
        form = MaxBotVehicleForm()

    return render(request, "maxbot/object_form.html", {
        "title": "Новая машина",
        "form": form,
        "back_url_name": "maxbot_vehicle_list",
    })


@staff_member_required
def maxbot_vehicle_edit(request, pk):
    obj = get_object_or_404(MaxBotVehicle, pk=pk)

    if request.method == "POST":
        form = MaxBotVehicleForm(request.POST, instance=obj)
        if form.is_valid():
            obj = form.save(commit=False)
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_vehicle_list")
    else:
        form = MaxBotVehicleForm(instance=obj)

    return render(request, "maxbot/object_form.html", {
        "title": f"Редактирование машины: {obj}",
        "form": form,
        "back_url_name": "maxbot_vehicle_list",
    })


@staff_member_required
def maxbot_vehicle_delete(request, pk):
    obj = get_object_or_404(MaxBotVehicle, pk=pk)

    if request.method == "POST":
        obj.delete()
        return redirect("maxbot_vehicle_list")

    return render(request, "maxbot/confirm_delete.html", {
        "title": "Удаление машины",
        "object_name": str(obj),
        "back_url_name": "maxbot_vehicle_list",
    })


# ---------- Scenarios ----------

@staff_member_required
def maxbot_scenario_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = (
        MaxBotScenario.objects
        .select_related("role", "first_question")
        .order_by("role__sort_order", "name")
    )

    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q) | Q(role__name__icontains=q))

    page_obj = _maxbot_paginate(request, qs, per_page=30)

    return render(request, "maxbot/scenario_list.html", {
        "page_obj": page_obj,
        "q": q,
    })


@staff_member_required
def maxbot_scenario_create(request):
    if request.method == "POST":
        form = MaxBotScenarioForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_scenario_list")
    else:
        form = MaxBotScenarioForm()

    return render(request, "maxbot/object_form.html", {
        "title": "Новый сценарий",
        "form": form,
        "back_url_name": "maxbot_scenario_list",
    })


@staff_member_required
def maxbot_scenario_edit(request, pk):
    obj = get_object_or_404(MaxBotScenario, pk=pk)

    if request.method == "POST":
        form = MaxBotScenarioForm(request.POST, instance=obj)
        if form.is_valid():
            obj = form.save(commit=False)
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_scenario_list")
    else:
        form = MaxBotScenarioForm(instance=obj)

    return render(request, "maxbot/object_form.html", {
        "title": f"Редактирование сценария: {obj.name}",
        "form": form,
        "back_url_name": "maxbot_scenario_list",
    })


@staff_member_required
def maxbot_scenario_delete(request, pk):
    obj = get_object_or_404(MaxBotScenario, pk=pk)

    if request.method == "POST":
        obj.delete()
        return redirect("maxbot_scenario_list")

    return render(request, "maxbot/confirm_delete.html", {
        "title": "Удаление сценария",
        "object_name": obj.name,
        "back_url_name": "maxbot_scenario_list",
    })


# ---------- Questions ----------

@staff_member_required
def maxbot_scenario_questions(request, scenario_id):
    scenario = get_object_or_404(
        MaxBotScenario.objects.select_related("role", "first_question"),
        pk=scenario_id,
    )

    questions = scenario.questions.order_by("sort_order", "id").prefetch_related("options")
    page_obj = _maxbot_paginate(request, questions, per_page=50)

    return render(request, "maxbot/question_list.html", {
        "scenario": scenario,
        "page_obj": page_obj,
    })


@staff_member_required
def maxbot_question_create(request, scenario_id):
    scenario = get_object_or_404(MaxBotScenario, pk=scenario_id)

    if request.method == "POST":
        form = MaxBotQuestionForm(request.POST, scenario=scenario)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.scenario = scenario
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_scenario_questions", scenario_id=scenario.id)
    else:
        form = MaxBotQuestionForm(scenario=scenario)

    return render(request, "maxbot/object_form.html", {
        "title": f"Новый вопрос для сценария: {scenario.name}",
        "form": form,
        "back_url_name": "maxbot_scenario_questions",
        "back_url_kwargs": {"scenario_id": scenario.id},
    })


@staff_member_required
def maxbot_question_edit(request, pk):
    obj = get_object_or_404(MaxBotQuestion.objects.select_related("scenario"), pk=pk)

    if request.method == "POST":
        form = MaxBotQuestionForm(request.POST, instance=obj, scenario=obj.scenario)
        if form.is_valid():
            obj = form.save(commit=False)
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_scenario_questions", scenario_id=obj.scenario_id)
    else:
        form = MaxBotQuestionForm(instance=obj, scenario=obj.scenario)

    return render(request, "maxbot/object_form.html", {
        "title": f"Редактирование вопроса: {obj.code}",
        "form": form,
        "back_url_name": "maxbot_scenario_questions",
        "back_url_kwargs": {"scenario_id": obj.scenario_id},
    })


@staff_member_required
def maxbot_question_delete(request, pk):
    obj = get_object_or_404(MaxBotQuestion.objects.select_related("scenario"), pk=pk)

    if request.method == "POST":
        scenario_id = obj.scenario_id
        obj.delete()
        return redirect("maxbot_scenario_questions", scenario_id=scenario_id)

    return render(request, "maxbot/confirm_delete.html", {
        "title": "Удаление вопроса",
        "object_name": obj.text,
        "back_url_name": "maxbot_scenario_questions",
        "back_url_kwargs": {"scenario_id": obj.scenario_id},
    })


@staff_member_required
def maxbot_question_set_first(request, scenario_id, question_id):
    scenario = get_object_or_404(MaxBotScenario, pk=scenario_id)
    question = get_object_or_404(MaxBotQuestion, pk=question_id, scenario=scenario)

    if request.method == "POST":
        scenario.first_question = question
        scenario.updated_at = timezone.now()
        scenario.save(update_fields=["first_question", "updated_at"])
        return redirect("maxbot_scenario_questions", scenario_id=scenario.id)

    return render(request, "maxbot/confirm_delete.html", {
        "title": "Сделать вопрос первым",
        "object_name": question.text,
        "submit_text": "Сделать первым",
        "back_url_name": "maxbot_scenario_questions",
        "back_url_kwargs": {"scenario_id": scenario.id},
    })


# ---------- Options ----------

@staff_member_required
def maxbot_question_options(request, question_id):
    question = get_object_or_404(
        MaxBotQuestion.objects.select_related("scenario", "scenario__role"),
        pk=question_id,
    )

    options = question.options.order_by("sort_order", "id")
    page_obj = _maxbot_paginate(request, options, per_page=50)

    return render(request, "maxbot/option_list.html", {
        "question": question,
        "page_obj": page_obj,
    })


@staff_member_required
def maxbot_option_create(request, question_id):
    question = get_object_or_404(MaxBotQuestion.objects.select_related("scenario"), pk=question_id)

    if request.method == "POST":
        form = MaxBotQuestionOptionForm(request.POST, question=question)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.question = question
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_question_options", question_id=question.id)
    else:
        form = MaxBotQuestionOptionForm(question=question)

    return render(request, "maxbot/object_form.html", {
        "title": f"Новый вариант ответа для вопроса: {question.code}",
        "form": form,
        "back_url_name": "maxbot_question_options",
        "back_url_kwargs": {"question_id": question.id},
    })


@staff_member_required
def maxbot_option_edit(request, pk):
    obj = get_object_or_404(MaxBotQuestionOption.objects.select_related("question", "question__scenario"), pk=pk)

    if request.method == "POST":
        form = MaxBotQuestionOptionForm(request.POST, instance=obj, question=obj.question)
        if form.is_valid():
            obj = form.save(commit=False)
            _maxbot_stamp_and_save(obj)
            return redirect("maxbot_question_options", question_id=obj.question_id)
    else:
        form = MaxBotQuestionOptionForm(instance=obj, question=obj.question)

    return render(request, "maxbot/object_form.html", {
        "title": f"Редактирование варианта ответа: {obj.code}",
        "form": form,
        "back_url_name": "maxbot_question_options",
        "back_url_kwargs": {"question_id": obj.question_id},
    })


@staff_member_required
def maxbot_option_delete(request, pk):
    obj = get_object_or_404(MaxBotQuestionOption.objects.select_related("question"), pk=pk)

    if request.method == "POST":
        question_id = obj.question_id
        obj.delete()
        return redirect("maxbot_question_options", question_id=question_id)

    return render(request, "maxbot/confirm_delete.html", {
        "title": "Удаление варианта ответа",
        "object_name": obj.text,
        "back_url_name": "maxbot_question_options",
        "back_url_kwargs": {"question_id": obj.question_id},
    })


# ---------- Requests / Registry ----------

@staff_member_required
def maxbot_request_list(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    emergency = (request.GET.get("emergency") or "").strip()

    qs = (
        MaxBotRequest.objects
        .select_related("role", "employee", "vehicle", "scenario", "applicant_user")
        .order_by("-created_at")
    )

    if q:
        qs = qs.filter(
            Q(request_no__icontains=q) |
            Q(applicant_full_name__icontains=q) |
            Q(employee__full_name__icontains=q) |
            Q(vehicle__reg_number__icontains=q) |
            Q(summary__icontains=q)
        )

    if status:
        qs = qs.filter(status=status)

    if emergency == "1":
        qs = qs.filter(emergency_flag=True)
    elif emergency == "0":
        qs = qs.filter(emergency_flag=False)

    page_obj = _maxbot_paginate(request, qs, per_page=40)

    for req in page_obj.object_list:
        req.status_label = MAXBOT_REQUEST_STATUS_LABELS.get(req.status, req.status)

    return render(request, "maxbot/request_list.html", {
        "page_obj": page_obj,
        "q": q,
        "status": status,
        "emergency": emergency,
        "status_choices": MAXBOT_REQUEST_STATUS_LABELS,
    })


@staff_member_required
def maxbot_request_detail(request, pk):
    req = get_object_or_404(
        MaxBotRequest.objects.select_related("role", "employee", "vehicle", "scenario", "applicant_user"),
        pk=pk,
    )

    answers = (
        req.answers
        .select_related("question", "option")
        .order_by("created_at", "id")
    )

    req.status_label = MAXBOT_REQUEST_STATUS_LABELS.get(req.status, req.status)

    return render(request, "maxbot/request_detail.html", {
        "req": req,
        "answers": answers,
    })












































def _safe_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _compose_full_name(item: dict) -> str:
    parts = [
        _safe_text(item.get("Фамилия")),
        _safe_text(item.get("Имя")),
        _safe_text(item.get("Отчество")),
    ]
    return " ".join([p for p in parts if p]).strip()


def _generate_ks_password() -> str:
    # Всего 40 символов: KS + 38 букв
    return "KS" + "".join(random.choices(string.ascii_letters, k=38))


def _encrypt_inn_value(inn: str) -> str:
    # Хеш через HMAC SHA256, длина помещается в 128 символов
    secret = settings.SECRET_KEY.encode("utf-8")
    return hmac.new(secret, inn.encode("utf-8"), hashlib.sha256).hexdigest()


def _fetch_working_employees_from_1c() -> list[dict]:
    auth = None
    if ONEC_WORKING_EMPLOYEES_AUTH_USER and ONEC_WORKING_EMPLOYEES_AUTH_PASSWORD:
        auth = (ONEC_WORKING_EMPLOYEES_AUTH_USER, ONEC_WORKING_EMPLOYEES_AUTH_PASSWORD)

    response = requests.get(
        ONEC_WORKING_EMPLOYEES_URL,
        auth=auth,
        timeout=(20, 240),
    )
    response.raise_for_status()

    payload = response.json()
    if not isinstance(payload, list):
        raise ValueError("Get_WorkingEmployees вернул не список")

    return payload



TARGET_SUPERMAG_POSITIONS = {
    "Администратор магазина",
    "Приемщик товара",
    "Директор магазина",
}

SUPERMAG_CREATE_POSITIONS = {
    "Администратор магазина",
    "Приемщик товара",
}

SUPERMAG_DIRECTOR_POSITION = "Директор магазина"
SUPERMAG_OFFINDEX_ADMIN = 118
SUPERMAG_ORAROLE_ADMIN = "OPERATOR_ADMIN"
WORKING_EMPLOYEES_EXPORT_SUBDIR = "working_employees_sync_exports"


def _chunked(seq, size: int):
    seq = list(seq)
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def _translit_to_login_part(text: str) -> str:
    text = _safe_text(text)
    out = []

    for ch in text:
        upper = ch.upper()
        if upper in _TRANSLIT:
            out.append(_TRANSLIT[upper].lower())
        elif ch.isascii() and ch.isalnum():
            out.append(ch.lower())
        elif ch in (" ", "-", ".", "_"):
            out.append("_")
        # остальное игнорируем

    s = "".join(out)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _build_sm_login_base(first_name: str, last_name: str, inn: str) -> str:
    first_part = _translit_to_login_part(first_name)
    last_part = _translit_to_login_part(last_name)

    if first_part and last_part:
        base = f"{first_part}_{last_part}"
    elif first_part:
        base = first_part
    elif last_part:
        base = last_part
    else:
        base = f"user_{(inn or '')[-4:]}" if inn else "user"

    base = re.sub(r"_+", "_", base).strip("._-")
    if not base:
        base = f"user_{(inn or '')[-4:]}" if inn else "user"

    return base[:64]


def _fetch_dbnames_for_smstores(smstore_ids: set[int]) -> dict[int, str]:
    """
    Берём REP.DBNAME для списка исходных SMSTORE (из 1С ИдМагазина).
    """
    result: dict[int, str] = {}
    if not smstore_ids:
        return result

    conn = cur = None
    try:
        conn = connect_oracle_supermag()
        cur = conn.cursor()

        for chunk in _chunked(sorted(smstore_ids), 500):
            binds = {}
            placeholders = []
            for i, sid in enumerate(chunk):
                key = f"b{i}"
                binds[key] = sid
                placeholders.append(f":{key}")

            sql = f"""
                SELECT storeloc, propval
                FROM smstoreproperties
                WHERE propid = 'REP.DBNAME'
                  AND storeloc IN ({", ".join(placeholders)})
            """
            cur.execute(sql, binds)

            for storeloc, propval in cur.fetchall():
                if storeloc is not None and propval is not None:
                    result[int(storeloc)] = str(propval).strip()

        return result

    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass


def _oracle_fetch_smstaff_by_inns(cur, inns: set[str]) -> dict[str, list[dict]]:
    """
    Возвращает словарь:
      {
        "123456789012": [row1, row2, ...],
        ...
      }
    """
    result: dict[str, list[dict]] = defaultdict(list)

    inns = {str(x).strip() for x in inns if str(x).strip()}
    if not inns:
        return result

    cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")
    if "INN" not in cols_set:
        return result

    def HAS(col: str) -> bool:
        return col.upper() in cols_set

    select_parts = [
        "s.id AS id" if HAS("id") else "NULL AS id",
        _smstaff_afio_sql("s", cols_set),
        "s.serverlogin AS serverlogin" if HAS("serverlogin") else "NULL AS serverlogin",
        "s.inn AS inn" if HAS("inn") else "NULL AS inn",
        "s.userenabled AS userenabled" if HAS("userenabled") else "NULL AS userenabled",
        "s.offindex AS offindex" if HAS("offindex") else "NULL AS offindex",
    ]

    if HAS("offindex"):
        select_parts.append("""
            (
                SELECT c.title
                FROM smoffcfg c
                WHERE c.id = s.offindex
                  AND ROWNUM = 1
            ) AS off_title
        """)
    else:
        select_parts.append("NULL AS off_title")

    for chunk in _chunked(sorted(inns), 500):
        binds = {}
        placeholders = []
        for i, inn in enumerate(chunk):
            key = f"b{i}"
            binds[key] = inn
            placeholders.append(f":{key}")

        sql = f"""
            SELECT {", ".join(select_parts)}
            FROM smstaff s
            WHERE TRIM(s.inn) IN ({", ".join(placeholders)})
            ORDER BY s.id
        """
        cur.execute(sql, binds)
        rows = _oracle_rows_to_jsonable(cur)

        for row in rows:
            inn_val = _safe_text(row.get("inn"))
            if inn_val:
                result[inn_val].append(row)

    return result


def _oracle_login_exists(cur, login: str) -> bool:
    login = _safe_text(login).lower()
    if not login:
        return False

    cols_set = _oracle_get_table_columns_set(cur, owner="SUPERMAG", table="SMSTAFF")
    if "SERVERLOGIN" not in cols_set:
        return False

    cur.execute(
        """
        SELECT COUNT(*)
        FROM smstaff
        WHERE LOWER(TRIM(serverlogin)) = :b_login
        """,
        b_login=login,
    )
    row = cur.fetchone()
    return bool(row and int(row[0] or 0) > 0)


def _resolve_available_sm_login(cur, base_login: str, reserved_logins: set[str]) -> str:
    base_login = _safe_text(base_login).lower()
    if not base_login:
        base_login = "user"

    base_login = re.sub(r"[^a-z0-9_]+", "_", base_login)
    base_login = re.sub(r"_+", "_", base_login).strip("_")[:64]
    if not base_login:
        base_login = "user"

    for n in range(0, 1000):
        if n == 0:
            candidate = base_login
        else:
            suffix = f"_{n+1}"
            candidate = f"{base_login[:64 - len(suffix)]}{suffix}"

        candidate = candidate.lower()
        if candidate in reserved_logins:
            continue

        if not _oracle_login_exists(cur, candidate):
            reserved_logins.add(candidate)
            return candidate

    raise RuntimeError(f"Не удалось подобрать свободный логин для базы '{base_login}'")


def _autosize_openpyxl_columns(ws, max_width: int = 60):
    for col_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > length:
                length = len(val)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 12), max_width)


def _export_working_employees_result(operation_slug: str, summary: dict, rows: list[dict], meta: dict) -> str:
    export_dir = Path(settings.MEDIA_ROOT) / WORKING_EMPLOYEES_EXPORT_SUBDIR
    export_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{operation_slug}_{timezone.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = export_dir / filename

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary.append(["Параметр", "Значение"])

    for k, v in meta.items():
        ws_summary.append([k, v])

    ws_summary.append([])

    for k, v in summary.items():
        ws_summary.append([k, v])

    ws_rows = wb.create_sheet("Rows")
    headers = [
        "№",
        "ИНН",
        "ФИО",
        "Должность",
        "Подразделение",
        "ИдМагазина (1С)",
        "storeid (ukm4store)",
        "DBNAME",
        "user_id",
        "user_exists",
        "ukm_link_exists",
        "oracle_exists",
        "sm_login",
        "sm_password",
        "status_code",
        "status",
        "actions",
    ]
    ws_rows.append(headers)

    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
    for cell in ws_rows[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws_rows.append([
            row.get("index"),
            row.get("employee_id"),
            row.get("full_name"),
            row.get("position"),
            row.get("department"),
            row.get("source_storeid"),
            row.get("target_storeid"),
            row.get("sm_db"),
            row.get("user_id"),
            row.get("user_exists"),
            row.get("ukm_link_exists"),
            row.get("oracle_exists"),
            row.get("sm_login"),
            row.get("sm_password"),
            row.get("status_code"),
            row.get("status"),
            "\n".join(row.get("actions") or []),
        ])

    _autosize_openpyxl_columns(ws_summary)
    _autosize_openpyxl_columns(ws_rows)

    wb.save(filepath)

    return f"{settings.MEDIA_URL.rstrip('/')}/{WORKING_EMPLOYEES_EXPORT_SUBDIR}/{filename}"






def _build_status_row(
    idx: int,
    inn: str,
    full_name: str,
    source_storeid: str,
    target_storeid: Optional[int],
    status_code: str,
    status: str,
    actions: Optional[list[str]] = None,
    position: str = "",
    department: str = "",
    **extra,
) -> dict:
    row = {
        "index": idx,
        "employee_id": inn,
        "full_name": full_name,
        "source_storeid": source_storeid,
        "target_storeid": target_storeid,
        "status_code": status_code,   # success / info / warning / error / skipped
        "status": status,
        "actions": actions or [],
        "position": position,
        "department": department,
    }
    row.update(extra)
    return row


def _sync_working_employees() -> dict:
    started_dt = timezone.now()
    started_ts = time.time()

    employees = _fetch_working_employees_from_1c()

    total_received = len(employees)
    smstore_ids = set()
    inns = set()

    for item in employees:
        raw_store = _safe_text(item.get("ИдМагазина"))
        inn = _safe_text(item.get("ИНН"))

        if raw_store:
            try:
                smstore_ids.add(int(raw_store))
            except ValueError:
                pass

        if inn:
            inns.add(inn)

    stores_map = {}
    for row in Store.objects.filter(smstore__in=smstore_ids).values("smstore", "ukm4store", "name"):
        smstore = row.get("smstore")
        if smstore is not None and smstore not in stores_map:
            stores_map[smstore] = row

    existing_users_qs = User.objects.filter(employee_id__in=inns).only(
        "id", "employee_id", "full_name", "mail", "phone"
    ).order_by("id")

    users_map = {}
    duplicate_user_employee_ids = set()
    existing_user_ids = []

    for user in existing_users_qs:
        if user.employee_id in users_map:
            duplicate_user_employee_ids.add(user.employee_id)
            continue
        users_map[user.employee_id] = user
        existing_user_ids.append(user.id)

    ukm_lookup = set(
        UKMUser.objects.filter(user_id__in=existing_user_ids).values_list("user_id", "storeid")
    )
    open_lookup = set(
        OpenInSystem.objects.filter(user_id__in=existing_user_ids, system_id=9).values_list("user_id", "system_id")
    )

    rows = []

    new_users_by_inn = {}
    pending_ukm_by_inn = defaultdict(set)
    pending_open_by_inn = set()

    existing_ukm_to_create = []
    existing_open_to_create = []

    summary = Counter()
    summary["total_received"] = total_received
    summary["unique_inn"] = len(inns)

    for idx, item in enumerate(employees, start=1):
        inn = _safe_text(item.get("ИНН"))
        raw_store = _safe_text(item.get("ИдМагазина"))
        full_name = _compose_full_name(item)
        position = _safe_text(item.get("Должность"))
        department = _safe_text(item.get("Подразделение"))

        if not raw_store:
            summary["skipped_no_storeid"] += 1
            rows.append(
                _build_status_row(
                    idx=idx,
                    inn=inn,
                    full_name=full_name,
                    source_storeid=raw_store,
                    target_storeid=None,
                    status_code="skipped",
                    status="пропущен: пустой ИдМагазина",
                    position=position,
                    department=department,
                )
            )
            continue

        summary["with_storeid"] += 1

        try:
            source_smstore = int(raw_store)
        except ValueError:
            summary["warning_rows"] += 1
            rows.append(
                _build_status_row(
                    idx=idx,
                    inn=inn,
                    full_name=full_name,
                    source_storeid=raw_store,
                    target_storeid=None,
                    status_code="warning",
                    status="неверный формат ИдМагазина",
                    position=position,
                    department=department,
                )
            )
            continue

        store_row = stores_map.get(source_smstore)
        if not store_row or not store_row.get("ukm4store"):
            summary["skipped_store_not_mapped"] += 1
            rows.append(
                _build_status_row(
                    idx=idx,
                    inn=inn,
                    full_name=full_name,
                    source_storeid=raw_store,
                    target_storeid=None,
                    status_code="warning",
                    status="не найден stores.smstore или пустой stores.ukm4store",
                    position=position,
                    department=department,
                )
            )
            continue

        target_storeid = int(store_row["ukm4store"])

        if not inn:
            summary["skipped_no_inn"] += 1
            rows.append(
                _build_status_row(
                    idx=idx,
                    inn=inn,
                    full_name=full_name,
                    source_storeid=raw_store,
                    target_storeid=target_storeid,
                    status_code="warning",
                    status="пропущен: пустой ИНН",
                    position=position,
                    department=department,
                )
            )
            continue

        existing_user = users_map.get(inn)
        if existing_user:
            actions = []

            if inn in duplicate_user_employee_ids:
                actions.append("в users найдено несколько записей с таким employee_id, использована первая")

            if (existing_user.id, target_storeid) not in ukm_lookup:
                existing_ukm_to_create.append(
                    UKMUser(
                        user_id=existing_user.id,
                        roleid=1,
                        storeid=target_storeid,
                        version=1,
                    )
                )
                ukm_lookup.add((existing_user.id, target_storeid))
                actions.append(f"добавлен ukm_users(storeid={target_storeid})")

            if (existing_user.id, 9) not in open_lookup:
                existing_open_to_create.append(
                    OpenInSystem(
                        user_id=existing_user.id,
                        username=(full_name or existing_user.full_name or inn)[:255],
                        password=_generate_ks_password(),
                        system_id=9,
                        status=True,
                    )
                )
                open_lookup.add((existing_user.id, 9))
                actions.append("добавлен open_in_system(system_id=9)")

            if actions:
                summary["success_rows"] += 1
                rows.append(
                    _build_status_row(
                        idx=idx,
                        inn=inn,
                        full_name=full_name,
                        source_storeid=raw_store,
                        target_storeid=target_storeid,
                        status_code="success",
                        status="обновлён",
                        actions=actions,
                        position=position,
                        department=department,
                    )
                )
            else:
                summary["info_rows"] += 1
                rows.append(
                    _build_status_row(
                        idx=idx,
                        inn=inn,
                        full_name=full_name,
                        source_storeid=raw_store,
                        target_storeid=target_storeid,
                        status_code="info",
                        status="без изменений",
                        position=position,
                        department=department,
                    )
                )
            continue

        first_new_user = inn not in new_users_by_inn
        if first_new_user:
            new_users_by_inn[inn] = User(
                employee_id=inn,
                max_id=None,
                encrypted_inn=_encrypt_inn_value(inn),
                full_name=(full_name or inn)[:255],
                mail=_safe_text(item.get("Почта"))[:255],
                phone=_safe_text(item.get("НомерТелефона"))[:50],
                department_id=311,
                position_id=5,
                active=True,
                tg_status=False,
                tg_id=None,
                created_at=started_dt,
                updated_at=started_dt,
            )
            pending_open_by_inn.add(inn)

        store_relation_is_new = target_storeid not in pending_ukm_by_inn[inn]
        if store_relation_is_new:
            pending_ukm_by_inn[inn].add(target_storeid)

        actions = []
        if first_new_user:
            actions.append("создан user")
            actions.append("добавлен open_in_system(system_id=9)")
        if store_relation_is_new:
            actions.append(f"добавлен ukm_users(storeid={target_storeid})")

        if actions:
            summary["success_rows"] += 1
            rows.append(
                _build_status_row(
                    idx=idx,
                    inn=inn,
                    full_name=full_name,
                    source_storeid=raw_store,
                    target_storeid=target_storeid,
                    status_code="success",
                    status="будет создан",
                    actions=actions,
                    position=position,
                    department=department,
                )
            )
        else:
            summary["info_rows"] += 1
            rows.append(
                _build_status_row(
                    idx=idx,
                    inn=inn,
                    full_name=full_name,
                    source_storeid=raw_store,
                    target_storeid=target_storeid,
                    status_code="info",
                    status="повтор в выгрузке, действий не требуется",
                    position=position,
                    department=department,
                )
            )

    new_users_list = list(new_users_by_inn.values())
    new_ukm_to_create = []
    new_open_to_create = []

    try:
        with transaction.atomic():
            if new_users_list:
                User.objects.bulk_create(new_users_list, batch_size=500)

            created_users_map = {}
            if new_users_by_inn:
                for user in User.objects.filter(employee_id__in=list(new_users_by_inn.keys())).order_by("id"):
                    created_users_map.setdefault(user.employee_id, user)

            for inn, target_store_ids in pending_ukm_by_inn.items():
                created_user = created_users_map.get(inn)
                if not created_user:
                    raise RuntimeError(f"Не удалось повторно получить созданного пользователя по employee_id={inn}")

                for target_storeid in sorted(target_store_ids):
                    new_ukm_to_create.append(
                        UKMUser(
                            user_id=created_user.id,
                            roleid=1,
                            storeid=target_storeid,
                            version=1,
                        )
                    )

                if inn in pending_open_by_inn:
                    new_open_to_create.append(
                        OpenInSystem(
                            user_id=created_user.id,
                            username=(created_user.full_name or inn)[:255],
                            password=_generate_ks_password(),
                            system_id=9,
                            status=True,
                        )
                    )

            if existing_ukm_to_create:
                UKMUser.objects.bulk_create(existing_ukm_to_create, batch_size=1000)

            if existing_open_to_create:
                OpenInSystem.objects.bulk_create(existing_open_to_create, batch_size=1000)

            if new_ukm_to_create:
                UKMUser.objects.bulk_create(new_ukm_to_create, batch_size=1000)

            if new_open_to_create:
                OpenInSystem.objects.bulk_create(new_open_to_create, batch_size=1000)

    except Exception:
        logger.exception("Ошибка при синхронизации рабочих сотрудников из 1С")
        raise

    summary["new_users_created"] = len(new_users_list)
    summary["ukm_users_created"] = len(existing_ukm_to_create) + len(new_ukm_to_create)
    summary["open_in_system_created"] = len(existing_open_to_create) + len(new_open_to_create)
    summary["rows_total"] = len(rows)

    # Дополнительно пересчитываем, если что-то не попало
    summary["success_rows"] = sum(1 for r in rows if r["status_code"] == "success")
    summary["info_rows"] = sum(1 for r in rows if r["status_code"] == "info")
    summary["warning_rows"] = sum(1 for r in rows if r["status_code"] == "warning")
    summary["error_rows"] = sum(1 for r in rows if r["status_code"] == "error")
    summary["skipped_rows"] = sum(1 for r in rows if r["status_code"] == "skipped")

    duration_sec = round(time.time() - started_ts, 2)

    logger.info(
        "[WORKING_EMPLOYEES_SYNC] total=%s with_storeid=%s new_users=%s new_ukm=%s new_open=%s duration=%s",
        summary["total_received"],
        summary["with_storeid"],
        summary["new_users_created"],
        summary["ukm_users_created"],
        summary["open_in_system_created"],
        duration_sec,
    )

    meta = {
        "started_at": started_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "duration_sec": duration_sec,
        "default_page_size": SYNC_DEFAULT_PAGE_SIZE,
        "mode": "sync_users_ukm_open9",
    }

    export_url = _export_working_employees_result(
        operation_slug="working_employees_sync",
        summary=dict(summary),
        rows=rows,
        meta=meta,
    )

    return {
        "ok": True,
        "summary": dict(summary),
        "rows": rows,
        "meta": meta,
        "export_url": export_url,
        "result_title": "Синхронизация users / ukm_users / open_in_system",
    }



def _run_working_employees_supermag(dry_run: bool) -> dict:
    started_dt = timezone.now()
    started_ts = time.time()

    employees = _fetch_working_employees_from_1c()
    filtered = [
        item for item in employees
        if _safe_text(item.get("Должность")) in TARGET_SUPERMAG_POSITIONS
    ]

    summary = Counter()
    summary["total_received"] = len(employees)
    summary["target_positions"] = len(filtered)
    summary["create_positions"] = sum(
        1 for x in filtered if _safe_text(x.get("Должность")) in SUPERMAG_CREATE_POSITIONS
    )
    summary["director_positions"] = sum(
        1 for x in filtered if _safe_text(x.get("Должность")) == SUPERMAG_DIRECTOR_POSITION
    )
    summary["dry_run"] = 1 if dry_run else 0

    source_smstore_ids = set()
    inns = set()

    for item in filtered:
        raw_store = _safe_text(item.get("ИдМагазина"))
        inn = _safe_text(item.get("ИНН"))

        if raw_store:
            try:
                source_smstore_ids.add(int(raw_store))
            except ValueError:
                pass

        if inn:
            inns.add(inn)

    stores_map = {}
    for row in Store.objects.filter(smstore__in=source_smstore_ids).values("smstore", "ukm4store", "name"):
        smstore = row.get("smstore")
        if smstore is not None and smstore not in stores_map:
            stores_map[int(smstore)] = row

    dbname_map = _fetch_dbnames_for_smstores(source_smstore_ids)

    existing_users_qs = User.objects.filter(employee_id__in=inns).only(
        "id", "employee_id", "full_name", "mail", "phone"
    ).order_by("id")

    users_map = {}
    duplicate_user_employee_ids = set()
    existing_user_ids = []

    for user in existing_users_qs:
        if user.employee_id in users_map:
            duplicate_user_employee_ids.add(user.employee_id)
            continue
        users_map[user.employee_id] = user
        existing_user_ids.append(user.id)

    ukm_lookup = set(
        UKMUser.objects.filter(user_id__in=existing_user_ids).values_list("user_id", "storeid")
    )
    open4_user_ids = set(
        OpenInSystem.objects.filter(user_id__in=existing_user_ids, system_id=4).values_list("user_id", flat=True)
    )

    eligible_by_db = defaultdict(list)
    results_by_index = {}

    for idx, item in enumerate(filtered, start=1):
        inn = _safe_text(item.get("ИНН"))
        raw_store = _safe_text(item.get("ИдМагазина"))
        full_name = _compose_full_name(item)
        position = _safe_text(item.get("Должность"))
        department = _safe_text(item.get("Подразделение"))
        first_name = _safe_text(item.get("Имя"))
        last_name = _safe_text(item.get("Фамилия"))

        actions = []

        if not raw_store:
            summary["skipped_no_storeid"] += 1
            results_by_index[idx] = _build_status_row(
                idx=idx,
                inn=inn,
                full_name=full_name,
                source_storeid=raw_store,
                target_storeid=None,
                status_code="skipped",
                status="пропущен: пустой ИдМагазина",
                actions=actions,
                position=position,
                department=department,
                sm_db="",
                user_id=None,
                user_exists=False,
                ukm_link_exists=False,
                oracle_exists=False,
                sm_login="",
                sm_password="",
            )
            continue

        try:
            source_smstore = int(raw_store)
        except ValueError:
            summary["warning_bad_storeid"] += 1
            results_by_index[idx] = _build_status_row(
                idx=idx,
                inn=inn,
                full_name=full_name,
                source_storeid=raw_store,
                target_storeid=None,
                status_code="warning",
                status="неверный формат ИдМагазина",
                actions=actions,
                position=position,
                department=department,
                sm_db="",
                user_id=None,
                user_exists=False,
                ukm_link_exists=False,
                oracle_exists=False,
                sm_login="",
                sm_password="",
            )
            continue

        store_row = stores_map.get(source_smstore)
        if not store_row or not store_row.get("ukm4store"):
            summary["warning_store_not_mapped"] += 1
            results_by_index[idx] = _build_status_row(
                idx=idx,
                inn=inn,
                full_name=full_name,
                source_storeid=raw_store,
                target_storeid=None,
                status_code="warning",
                status="не найден stores.smstore или пустой stores.ukm4store",
                actions=actions,
                position=position,
                department=department,
                sm_db="",
                user_id=None,
                user_exists=False,
                ukm_link_exists=False,
                oracle_exists=False,
                sm_login="",
                sm_password="",
            )
            continue

        target_storeid = int(store_row["ukm4store"])

        dbname = _safe_text(dbname_map.get(source_smstore))
        if not dbname:
            summary["warning_dbname_not_found"] += 1
            results_by_index[idx] = _build_status_row(
                idx=idx,
                inn=inn,
                full_name=full_name,
                source_storeid=raw_store,
                target_storeid=target_storeid,
                status_code="warning",
                status="не найден REP.DBNAME для магазина",
                actions=actions,
                position=position,
                department=department,
                sm_db="",
                user_id=None,
                user_exists=False,
                ukm_link_exists=False,
                oracle_exists=False,
                sm_login="",
                sm_password="",
            )
            continue

        if dbname.startswith("BINCH"):
            summary["skipped_binch"] += 1
            results_by_index[idx] = _build_status_row(
                idx=idx,
                inn=inn,
                full_name=full_name,
                source_storeid=raw_store,
                target_storeid=target_storeid,
                status_code="skipped",
                status=f"база {dbname} пропущена (BINCH пока не обрабатываем)",
                actions=actions,
                position=position,
                department=department,
                sm_db=dbname,
                user_id=None,
                user_exists=False,
                ukm_link_exists=False,
                oracle_exists=False,
                sm_login="",
                sm_password="",
            )
            continue

        if not dbname.startswith("BINUU"):
            summary["warning_bad_db_prefix"] += 1
            results_by_index[idx] = _build_status_row(
                idx=idx,
                inn=inn,
                full_name=full_name,
                source_storeid=raw_store,
                target_storeid=target_storeid,
                status_code="warning",
                status=f"неподдерживаемая база {dbname}",
                actions=actions,
                position=position,
                department=department,
                sm_db=dbname,
                user_id=None,
                user_exists=False,
                ukm_link_exists=False,
                oracle_exists=False,
                sm_login="",
                sm_password="",
            )
            continue

        if dbname not in ORACLE_TNS_MAP:
            summary["warning_db_not_configured"] += 1
            results_by_index[idx] = _build_status_row(
                idx=idx,
                inn=inn,
                full_name=full_name,
                source_storeid=raw_store,
                target_storeid=target_storeid,
                status_code="warning",
                status=f"база {dbname} не описана в ORACLE_TNS_MAP",
                actions=actions,
                position=position,
                department=department,
                sm_db=dbname,
                user_id=None,
                user_exists=False,
                ukm_link_exists=False,
                oracle_exists=False,
                sm_login="",
                sm_password="",
            )
            continue

        if not inn:
            summary["warning_no_inn"] += 1
            results_by_index[idx] = _build_status_row(
                idx=idx,
                inn=inn,
                full_name=full_name,
                source_storeid=raw_store,
                target_storeid=target_storeid,
                status_code="warning",
                status="пропущен: пустой ИНН",
                actions=actions,
                position=position,
                department=department,
                sm_db=dbname,
                user_id=None,
                user_exists=False,
                ukm_link_exists=False,
                oracle_exists=False,
                sm_login="",
                sm_password="",
            )
            continue

        user = users_map.get(inn)
        if not user:
            summary["warning_user_not_found"] += 1
            actions.append("в users запись не найдена")
            results_by_index[idx] = _build_status_row(
                idx=idx,
                inn=inn,
                full_name=full_name,
                source_storeid=raw_store,
                target_storeid=target_storeid,
                status_code="warning",
                status="нет записи в users",
                actions=actions,
                position=position,
                department=department,
                sm_db=dbname,
                user_id=None,
                user_exists=False,
                ukm_link_exists=False,
                oracle_exists=False,
                sm_login="",
                sm_password="",
            )
            continue

        if inn in duplicate_user_employee_ids:
            actions.append("в users найдено несколько записей с таким employee_id, использована первая")

        has_ukm_link = (user.id, target_storeid) in ukm_lookup
        if not has_ukm_link:
            summary["warning_ukm_link_not_found"] += 1
            actions.append(f"нет ukm_users для user_id={user.id} и storeid={target_storeid}")
            results_by_index[idx] = _build_status_row(
                idx=idx,
                inn=inn,
                full_name=full_name,
                source_storeid=raw_store,
                target_storeid=target_storeid,
                status_code="warning",
                status="нет привязки ukm_users к нужному storeid",
                actions=actions,
                position=position,
                department=department,
                sm_db=dbname,
                user_id=user.id,
                user_exists=True,
                ukm_link_exists=False,
                oracle_exists=False,
                sm_login="",
                sm_password="",
            )
            continue

        eligible_by_db[dbname].append({
            "index": idx,
            "item": item,
            "inn": inn,
            "full_name": full_name,
            "position": position,
            "department": department,
            "first_name": first_name,
            "last_name": last_name,
            "source_storeid": raw_store,
            "target_storeid": target_storeid,
            "sm_db": dbname,
            "user": user,
            "duplicate_user": inn in duplicate_user_employee_ids,
        })

    for dbname, items in eligible_by_db.items():
        conn = cur = None
        try:
            conn = _connect_oracle_service(dbname)
            cur = conn.cursor()

            existing_sm_by_inn = _oracle_fetch_smstaff_by_inns(cur, {x["inn"] for x in items})
            reserved_logins = set()
            processed_inn_in_run = set()

            for item in items:
                idx = item["index"]
                inn = item["inn"]
                full_name = item["full_name"]
                position = item["position"]
                department = item["department"]
                raw_store = item["source_storeid"]
                target_storeid = item["target_storeid"]
                user = item["user"]

                actions = []
                if item["duplicate_user"]:
                    actions.append("в users найдено несколько записей с таким employee_id, использована первая")

                actions.append(f"user_id={user.id}")
                actions.append(f"найден ukm_users(storeid={target_storeid})")

                existing_rows = existing_sm_by_inn.get(inn, [])
                existing_row = existing_rows[0] if existing_rows else None

                if existing_row:
                    if len(existing_rows) > 1:
                        actions.append("в SMSTAFF найдено несколько записей с таким ИНН, показана первая")

                    login_existing = _safe_text(existing_row.get("serverlogin"))
                    if login_existing:
                        reserved_logins.add(login_existing.lower())
                        actions.append(f"логин в Супермаге: {login_existing}")

                    if position == SUPERMAG_DIRECTOR_POSITION:
                        summary["director_exists"] += 1
                        status_text = "директор уже есть в Супермаге"
                    else:
                        summary["sm_exists"] += 1
                        status_text = "учётная запись уже есть в Супермаге"

                    results_by_index[idx] = _build_status_row(
                        idx=idx,
                        inn=inn,
                        full_name=full_name,
                        source_storeid=raw_store,
                        target_storeid=target_storeid,
                        status_code="info",
                        status=status_text,
                        actions=actions,
                        position=position,
                        department=department,
                        sm_db=dbname,
                        user_id=user.id,
                        user_exists=True,
                        ukm_link_exists=True,
                        oracle_exists=True,
                        sm_login=login_existing,
                        sm_password="",
                    )
                    processed_inn_in_run.add(inn)
                    continue

                if inn in processed_inn_in_run:
                    summary["duplicate_feed_rows"] += 1
                    results_by_index[idx] = _build_status_row(
                        idx=idx,
                        inn=inn,
                        full_name=full_name,
                        source_storeid=raw_store,
                        target_storeid=target_storeid,
                        status_code="info",
                        status="дубликат в текущей выгрузке для этой базы, повторное создание не требуется",
                        actions=actions,
                        position=position,
                        department=department,
                        sm_db=dbname,
                        user_id=user.id,
                        user_exists=True,
                        ukm_link_exists=True,
                        oracle_exists=False,
                        sm_login="",
                        sm_password="",
                    )
                    continue

                if position == SUPERMAG_DIRECTOR_POSITION:
                    summary["director_missing"] += 1
                    actions.append("для директоров создание не выполняется")
                    results_by_index[idx] = _build_status_row(
                        idx=idx,
                        inn=inn,
                        full_name=full_name,
                        source_storeid=raw_store,
                        target_storeid=target_storeid,
                        status_code="warning",
                        status="для директора УЗ в Супермаге не найдена, создание не выполняется",
                        actions=actions,
                        position=position,
                        department=department,
                        sm_db=dbname,
                        user_id=user.id,
                        user_exists=True,
                        ukm_link_exists=True,
                        oracle_exists=False,
                        sm_login="",
                        sm_password="",
                    )
                    processed_inn_in_run.add(inn)
                    continue

                login_base = _build_sm_login_base(
                    first_name=item["first_name"],
                    last_name=item["last_name"],
                    inn=inn,
                )
                login_new = _resolve_available_sm_login(cur, login_base, reserved_logins)
                pin_code = _generate_pin_code()

                actions.append(f"логин: {login_new}")
                actions.append(f"пароль: {pin_code}")
                actions.append(f"AFIO: {full_name}")
                actions.append(f"ADOL: {SUPERMAG_ORAROLE_ADMIN}")
                actions.append(f"OFFINDEX: {SUPERMAG_OFFINDEX_ADMIN}")

                if dry_run:
                    summary["sm_will_create"] += 1
                    if user.id in open4_user_ids:
                        summary["open4_will_update"] += 1
                        actions.append("будет обновлён open_in_system(system_id=4)")
                    else:
                        summary["open4_will_create"] += 1
                        actions.append("будет создан open_in_system(system_id=4)")

                    results_by_index[idx] = _build_status_row(
                        idx=idx,
                        inn=inn,
                        full_name=full_name,
                        source_storeid=raw_store,
                        target_storeid=target_storeid,
                        status_code="success",
                        status="будет создан в Супермаге (тест)",
                        actions=actions,
                        position=position,
                        department=department,
                        sm_db=dbname,
                        user_id=user.id,
                        user_exists=True,
                        ukm_link_exists=True,
                        oracle_exists=False,
                        sm_login=login_new,
                        sm_password=pin_code,
                    )
                    processed_inn_in_run.add(inn)
                    continue

                proc_name = os.getenv("SM_BIN_CREATEUSER_PROC", "bin_createuser").strip()
                if not re.fullmatch(r"[A-Za-z0-9_.$]+", proc_name):
                    raise ValueError("Некорректное имя процедуры в SM_BIN_CREATEUSER_PROC.")

                try:
                    cur.execute(
                        f"""
                        BEGIN
                            {proc_name}(
                                AUSER => :p_auser,
                                APASS => :p_apass,
                                ADOL => :p_adol,
                                ACHECK => :p_acheck,
                                AUSERENABLED => :p_auserenabled,
                                AINN => :p_ainn,
                                AFIO => :p_afio
                            );
                        END;
                        """,
                        p_auser=login_new,
                        p_apass=pin_code,
                        p_adol=SUPERMAG_ORAROLE_ADMIN,
                        p_acheck=0,
                        p_auserenabled=1,
                        p_ainn=inn,
                        p_afio=full_name,
                    )

                    updated_smstaff_rows = _oracle_update_smstaff_after_create(
                        cur=cur,
                        login=login_new,
                        inn=inn,
                        afio=full_name,
                        offindex=SUPERMAG_OFFINDEX_ADMIN,
                    )

                    conn.commit()

                except Exception as oracle_exc:
                    try:
                        conn.rollback()
                    except Exception:
                        pass

                    summary["oracle_create_errors"] += 1
                    results_by_index[idx] = _build_status_row(
                        idx=idx,
                        inn=inn,
                        full_name=full_name,
                        source_storeid=raw_store,
                        target_storeid=target_storeid,
                        status_code="error",
                        status=f"ошибка создания в Супермаге: {oracle_exc}",
                        actions=actions,
                        position=position,
                        department=department,
                        sm_db=dbname,
                        user_id=user.id,
                        user_exists=True,
                        ukm_link_exists=True,
                        oracle_exists=False,
                        sm_login=login_new,
                        sm_password=pin_code,
                    )
                    continue

                row_status_code = "success"
                row_status_text = "создан в Супермаге"

                try:
                    qs = OpenInSystem.objects.filter(user_id=user.id, system_id=4)
                    if qs.exists():
                        updated_cnt = qs.update(
                            username=login_new,
                            password=pin_code,
                            status=True,
                        )
                        summary["open4_updated"] += 1
                        actions.append(f"обновлён open_in_system(system_id=4), строк: {updated_cnt}")
                    else:
                        OpenInSystem.objects.create(
                            user_id=user.id,
                            username=login_new,
                            password=pin_code,
                            system_id=4,
                            status=True,
                        )
                        open4_user_ids.add(user.id)
                        summary["open4_created"] += 1
                        actions.append("создан open_in_system(system_id=4)")
                except Exception as pg_exc:
                    summary["open4_errors"] += 1
                    row_status_code = "warning"
                    row_status_text = "создан в Супермаге, но ошибка записи в open_in_system"
                    actions.append(f"ОШИБКА open_in_system: {pg_exc}")

                if updated_smstaff_rows:
                    actions.append(f"обновлено строк в SMSTAFF: {updated_smstaff_rows}")

                summary["sm_created"] += 1
                processed_inn_in_run.add(inn)

                existing_sm_by_inn[inn] = [{
                    "serverlogin": login_new,
                    "inn": inn,
                    "afio": full_name,
                    "userenabled": 1,
                    "offindex": SUPERMAG_OFFINDEX_ADMIN,
                    "off_title": "Администратор",
                }]

                results_by_index[idx] = _build_status_row(
                    idx=idx,
                    inn=inn,
                    full_name=full_name,
                    source_storeid=raw_store,
                    target_storeid=target_storeid,
                    status_code=row_status_code,
                    status=row_status_text,
                    actions=actions,
                    position=position,
                    department=department,
                    sm_db=dbname,
                    user_id=user.id,
                    user_exists=True,
                    ukm_link_exists=True,
                    oracle_exists=True,
                    sm_login=login_new,
                    sm_password=pin_code,
                )

        except Exception as db_exc:
            logger.exception("Ошибка обработки базы %s", dbname)
            for item in items:
                idx = item["index"]
                results_by_index[idx] = _build_status_row(
                    idx=idx,
                    inn=item["inn"],
                    full_name=item["full_name"],
                    source_storeid=item["source_storeid"],
                    target_storeid=item["target_storeid"],
                    status_code="error",
                    status=f"ошибка подключения/обработки базы {dbname}: {db_exc}",
                    actions=[],
                    position=item["position"],
                    department=item["department"],
                    sm_db=dbname,
                    user_id=item["user"].id if item.get("user") else None,
                    user_exists=bool(item.get("user")),
                    ukm_link_exists=True,
                    oracle_exists=False,
                    sm_login="",
                    sm_password="",
                )
                summary["db_errors"] += 1

        finally:
            try:
                if cur:
                    cur.close()
                if conn:
                    conn.close()
            except Exception:
                pass

    rows = [results_by_index[i] for i in range(1, len(filtered) + 1)]

    summary["rows_total"] = len(rows)
    summary["success_rows"] = sum(1 for r in rows if r["status_code"] == "success")
    summary["info_rows"] = sum(1 for r in rows if r["status_code"] == "info")
    summary["warning_rows"] = sum(1 for r in rows if r["status_code"] == "warning")
    summary["error_rows"] = sum(1 for r in rows if r["status_code"] == "error")
    summary["skipped_rows"] = sum(1 for r in rows if r["status_code"] == "skipped")

    duration_sec = round(time.time() - started_ts, 2)
    mode_name = "supermag_test" if dry_run else "supermag_run"

    meta = {
        "started_at": started_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "duration_sec": duration_sec,
        "mode": mode_name,
        "default_page_size": SYNC_DEFAULT_PAGE_SIZE,
    }

    export_url = _export_working_employees_result(
        operation_slug=mode_name,
        summary=dict(summary),
        rows=rows,
        meta=meta,
    )

    logger.info(
        "[WORKING_EMPLOYEES_SUPERMAG] mode=%s total=%s target=%s sm_created=%s sm_will_create=%s warnings=%s errors=%s duration=%s",
        mode_name,
        summary["total_received"],
        summary["target_positions"],
        summary.get("sm_created", 0),
        summary.get("sm_will_create", 0),
        summary["warning_rows"],
        summary["error_rows"],
        duration_sec,
    )

    return {
        "ok": True,
        "summary": dict(summary),
        "rows": rows,
        "meta": meta,
        "export_url": export_url,
        "result_title": "Создание в Супермаге (ТЕСТ)" if dry_run else "Создание в Супермаге",
    }










@never_cache
@require_GET
def working_employees_sync_page(request):
    return render(request, "frostapp/working_employees_sync.html")


@never_cache
@require_POST
def working_employees_sync_run(request):
    try:
        result = _sync_working_employees()
        return JsonResponse(result, json_dumps_params={"ensure_ascii": False})
    except RequestException as exc:
        logger.exception("Ошибка запроса к 1С Get_WorkingEmployees")
        return JsonResponse(
            {
                "ok": False,
                "error": f"Ошибка запроса к 1С: {str(exc)}",
            },
            status=502,
            json_dumps_params={"ensure_ascii": False},
        )
    except Exception as exc:
        logger.exception("Внутренняя ошибка синхронизации сотрудников")
        return JsonResponse(
            {
                "ok": False,
                "error": f"Внутренняя ошибка: {str(exc)}",
            },
            status=500,
            json_dumps_params={"ensure_ascii": False},
        )








@never_cache
@require_POST
def working_employees_sync_supermag_run(request):
    try:
        result = _run_working_employees_supermag(dry_run=False)
        return JsonResponse(result, json_dumps_params={"ensure_ascii": False})
    except RequestException as exc:
        logger.exception("Ошибка запроса к 1С Get_WorkingEmployees для Супермага")
        return JsonResponse(
            {
                "ok": False,
                "error": f"Ошибка запроса к 1С: {str(exc)}",
            },
            status=502,
            json_dumps_params={"ensure_ascii": False},
        )
    except Exception as exc:
        logger.exception("Внутренняя ошибка создания пользователей в Супермаге")
        return JsonResponse(
            {
                "ok": False,
                "error": f"Внутренняя ошибка: {str(exc)}",
            },
            status=500,
            json_dumps_params={"ensure_ascii": False},
        )


@never_cache
@require_POST
def working_employees_sync_supermag_test_run(request):
    try:
        result = _run_working_employees_supermag(dry_run=True)
        return JsonResponse(result, json_dumps_params={"ensure_ascii": False})
    except RequestException as exc:
        logger.exception("Ошибка запроса к 1С Get_WorkingEmployees для теста Супермага")
        return JsonResponse(
            {
                "ok": False,
                "error": f"Ошибка запроса к 1С: {str(exc)}",
            },
            status=502,
            json_dumps_params={"ensure_ascii": False},
        )
    except Exception as exc:
        logger.exception("Внутренняя ошибка тестового создания пользователей в Супермаге")
        return JsonResponse(
            {
                "ok": False,
                "error": f"Внутренняя ошибка: {str(exc)}",
            },
            status=500,
            json_dumps_params={"ensure_ascii": False},
        )
