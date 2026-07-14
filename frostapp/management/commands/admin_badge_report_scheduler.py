import os
import time
import datetime
import tempfile
from pathlib import Path
from zoneinfo import ZoneInfo

from django.core.management.base import BaseCommand
from django.db.models import Q
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from frostapp.models import AdminBadgeRequest, QRIssueLog, Store


def _read_simple_env_file_value(name: str, env_path: str = "/app/.env") -> str | None:
    """
    Простое чтение .env без python-dotenv.

    Поддерживает:
      KEY=value
      export KEY=value
      KEY="value"
      KEY='value'
    """
    try:
        if not os.path.exists(env_path):
            return None

        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = (line or "").strip()

                if not line or line.startswith("#"):
                    continue

                if line.startswith("export "):
                    line = line[len("export "):].strip()

                if "=" not in line:
                    continue

                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip()

                if key != name:
                    continue

                if (
                    len(value) >= 2
                    and (
                        (value[0] == value[-1] == '"')
                        or (value[0] == value[-1] == "'")
                    )
                ):
                    value = value[1:-1]

                return value

    except Exception:
        return None

    return None


def _env_value(name: str, default: str = "", env_path: str = "/app/.env") -> str:
    """
    Сначала читает настоящие переменные окружения Docker,
    потом файл /app/.env.
    """
    value = os.getenv(name)

    if value is not None and str(value).strip() != "":
        return str(value).strip()

    value_from_file = _read_simple_env_file_value(name, env_path)

    if value_from_file is not None and str(value_from_file).strip() != "":
        return str(value_from_file).strip()

    return default


def _int_env(name: str, default: int) -> int:
    try:
        return int(_env_value(name, str(default)))
    except Exception:
        return default


def _get_report_tz() -> ZoneInfo:
    tz_name = (
        _env_value(
            "BADGE_WEEKLY_REPORT_TZ",
            _env_value("BADGE_REPORT_TIMEZONE", "Asia/Irkutsk"),
        )
        or "Asia/Irkutsk"
    ).strip()

    try:
        return ZoneInfo(tz_name)
    except Exception:
        return ZoneInfo("Asia/Irkutsk")


def _local_dt(dt, tz: ZoneInfo):
    if not dt:
        dt = timezone.now()

    return timezone.localtime(dt, tz)


def _date_time_parts(dt, tz: ZoneInfo) -> tuple[str, str]:
    local_dt = _local_dt(dt, tz)
    return local_dt.strftime("%d.%m.%Y"), local_dt.strftime("%H:%M:%S")


def _period_for_run_date(run_date: datetime.date) -> tuple[datetime.date, datetime.date]:
    """
    Если запуск 09.07.2026, период будет:
      02.07.2026 - 08.07.2026

    То есть 7 календарных дней до дня запуска.
    """
    start_date = run_date - datetime.timedelta(days=7)
    end_date = run_date - datetime.timedelta(days=1)
    return start_date, end_date


def _period_datetimes(start_date: datetime.date, end_date: datetime.date, tz: ZoneInfo):
    """
    Возвращает границы периода:
      start_date 00:00:00 inclusive
      end_date + 1 день 00:00:00 exclusive
    """
    start_dt = datetime.datetime.combine(start_date, datetime.time.min, tzinfo=tz)
    end_exclusive_dt = datetime.datetime.combine(
        end_date + datetime.timedelta(days=1),
        datetime.time.min,
        tzinfo=tz,
    )

    return start_dt, end_exclusive_dt


def _filename_for_period(start_date: datetime.date, end_date: datetime.date) -> str:
    return f"{start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}.xlsx"


def _to_int_or_none(value):
    try:
        if value is None:
            return None

        s = str(value).strip()
        if not s:
            return None

        return int(s)
    except Exception:
        return None


def _requested_smstore_from_raw_request(raw_request) -> int | None:
    """
    В qr_issue_logs.raw_request лежит JSON из 1С.
    Там ожидаем storeId или storeid.
    """
    if not isinstance(raw_request, dict):
        return None

    value = (
        raw_request.get("storeId")
        or raw_request.get("storeid")
        or raw_request.get("smstore")
    )

    return _to_int_or_none(value)


def _load_store_names_by_ukm4(store_ids: list[int]) -> dict[int, str]:
    if not store_ids:
        return {}

    result = {}

    qs = Store.objects.filter(ukm4store__in=store_ids).only("ukm4store", "name")

    for store in qs:
        if store.ukm4store is None:
            continue

        name = str(store.name or "").strip()
        result[int(store.ukm4store)] = name or f"Магазин {store.ukm4store}"

    return result


def _load_store_names_by_smstore(sm_store_ids: list[int]) -> dict[int, str]:
    if not sm_store_ids:
        return {}

    result = {}

    qs = Store.objects.filter(smstore__in=sm_store_ids).only("smstore", "name")

    for store in qs:
        if store.smstore is None:
            continue

        name = str(store.name or "").strip()
        result[int(store.smstore)] = name or f"Магазин {store.smstore}"

    return result


def _apply_header_style(ws, headers: list[str]):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_num, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    return border


def _apply_body_style(ws, border):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _finalize_sheet(ws, widths: dict[str, int]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _fill_admin_badge_transfer_sheet(wb: Workbook, start_dt, end_exclusive_dt, tz: ZoneInfo) -> int:
    """
    Лист 1:
    Передача бейджа администратора через MAX.
    Источник: admin_badge_requests.
    """
    ws = wb.active
    ws.title = "Передача бейджа"

    headers = ["Магазин", "Дата", "Время", "Кассир", "Администратор"]
    ws.append(headers)

    rows = list(
        AdminBadgeRequest.objects
        .filter(
            Q(status="ACCEPTED") | Q(decision__iexact="accept"),
            decided_at__isnull=False,
            decided_at__gte=start_dt,
            decided_at__lt=end_exclusive_dt,
        )
        .order_by("decided_at")
    )

    store_ids = sorted({
        int(x.storeid)
        for x in rows
        if x.storeid is not None
    })

    stores_map = _load_store_names_by_ukm4(store_ids)

    border = _apply_header_style(ws, headers)

    for item in rows:
        if item.storeid is not None:
            store_name = stores_map.get(int(item.storeid), f"Магазин {item.storeid}")
        else:
            store_name = "—"

        date_text, time_text = _date_time_parts(item.decided_at, tz)

        ws.append([
            store_name,
            date_text,
            time_text,
            item.cashier_full_name or "—",
            item.admin_full_name or "—",
        ])

    _apply_body_style(ws, border)

    _finalize_sheet(ws, {
        "A": 35,
        "B": 14,
        "C": 12,
        "D": 35,
        "E": 35,
    })

    return len(rows)


def _fill_1c_admin_badge_print_sheet(wb: Workbook, start_dt, end_exclusive_dt, tz: ZoneInfo) -> int:
    """
    Лист 2:
    Печать бейджа администратора из 1С.
    Источник: qr_issue_logs, записи get_qr_code_by_employee_id.

    Важно:
    get_qr_code_by_employee_id сейчас пишет несколько строк по магазинам пользователя.
    Поэтому берём только строку, где sm_store_id совпадает с storeId из raw_request.
    """
    ws = wb.create_sheet("Печать бейджа из 1С")

    headers = ["Магазин", "Дата", "Время", "Администратор"]
    ws.append(headers)

    raw_rows = list(
        QRIssueLog.objects
        .filter(
            endpoint="get_qr_code_by_employee_id",
            method="BY_INN",
            status="ok",
            created_at__gte=start_dt,
            created_at__lt=end_exclusive_dt,
        )
        .select_related("user")
        .order_by("created_at", "id")
    )

    filtered_rows = []

    for item in raw_rows:
        requested_smstore = _requested_smstore_from_raw_request(item.raw_request)

        # Если raw_request есть и там есть storeId,
        # оставляем только лог именно по запрошенному магазину.
        if requested_smstore is not None:
            if item.sm_store_id is None:
                continue

            if int(item.sm_store_id) != int(requested_smstore):
                continue

        filtered_rows.append(item)

    sm_store_ids = sorted({
        int(x.sm_store_id)
        for x in filtered_rows
        if x.sm_store_id is not None
    })

    stores_map_by_sm = _load_store_names_by_smstore(sm_store_ids)

    border = _apply_header_style(ws, headers)

    for item in filtered_rows:
        if item.sm_store_id is not None:
            store_name = stores_map_by_sm.get(
                int(item.sm_store_id),
                f"Магазин {item.sm_store_id}",
            )
        else:
            store_name = "—"

        date_text, time_text = _date_time_parts(item.created_at, tz)

        admin_name = (
            str(item.employee_fio or "").strip()
            or str(getattr(item.user, "full_name", "") or "").strip()
            or "—"
        )

        ws.append([
            store_name,
            date_text,
            time_text,
            admin_name,
        ])

    _apply_body_style(ws, border)

    _finalize_sheet(ws, {
        "A": 35,
        "B": 14,
        "C": 12,
        "D": 35,
    })

    return len(filtered_rows)


def build_admin_badge_report_file(
    *,
    start_date: datetime.date,
    end_date: datetime.date,
    output_path: Path,
    tz: ZoneInfo,
) -> dict:
    """
    Формирует Excel-отчёт и сохраняет его в output_path.

    Лист 1: Передача бейджа
    Лист 2: Печать бейджа администратора из 1С
    """
    start_dt, end_exclusive_dt = _period_datetimes(start_date, end_date, tz)

    wb = Workbook()

    transfer_rows_count = _fill_admin_badge_transfer_sheet(
        wb,
        start_dt,
        end_exclusive_dt,
        tz,
    )

    onec_print_rows_count = _fill_1c_admin_badge_print_sheet(
        wb,
        start_dt,
        end_exclusive_dt,
        tz,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Пишем сначала во временный файл, потом атомарно переименовываем.
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        delete=False,
        dir=str(output_path.parent),
    ) as tmp:
        tmp_path = Path(tmp.name)

    try:
        wb.save(tmp_path)
        os.replace(tmp_path, output_path)
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            pass

    return {
        "transfer_rows": transfer_rows_count,
        "onec_print_rows": onec_print_rows_count,
    }


class Command(BaseCommand):
    help = "Еженедельный Excel-отчёт по передаче бейджа администратора"

    def add_arguments(self, parser):
        parser.add_argument(
            "--once",
            action="store_true",
            help="Сформировать отчёт один раз и выйти",
        )
        parser.add_argument(
            "--date",
            type=str,
            default="",
            help="Дата запуска в формате YYYY-MM-DD. Нужна для теста. Например: 2026-07-09",
        )

    def handle(self, *args, **options):
        tz = _get_report_tz()

        report_dir = Path(
            _env_value(
                "BADGE_WEEKLY_REPORT_DIR",
                "/app/badge_reports_shared",
            )
        )

        schedule_weekday = _int_env("BADGE_WEEKLY_REPORT_WEEKDAY", 2)
        schedule_hour = _int_env("BADGE_WEEKLY_REPORT_HOUR", 8)
        schedule_minute = _int_env("BADGE_WEEKLY_REPORT_MINUTE", 0)
        check_sec = _int_env("BADGE_WEEKLY_REPORT_CHECK_SEC", 60)

        once = bool(options.get("once"))
        date_raw = str(options.get("date") or "").strip()

        if once:
            if date_raw:
                run_date = datetime.date.fromisoformat(date_raw)
            else:
                run_date = timezone.now().astimezone(tz).date()

            self._generate_for_run_date(
                run_date=run_date,
                report_dir=report_dir,
                tz=tz,
                force=True,
            )
            return

        self.stdout.write(
            self.style.SUCCESS(
                "[BADGE WEEKLY REPORT] scheduler started. "
                f"dir={report_dir} tz={tz.key} "
                f"weekday={schedule_weekday} time={schedule_hour:02d}:{schedule_minute:02d}"
            )
        )

        while True:
            try:
                now_local = timezone.now().astimezone(tz)

                is_schedule_day = now_local.weekday() == schedule_weekday
                is_after_schedule_time = (
                    now_local.hour > schedule_hour
                    or (
                        now_local.hour == schedule_hour
                        and now_local.minute >= schedule_minute
                    )
                )

                if is_schedule_day and is_after_schedule_time:
                    self._generate_for_run_date(
                        run_date=now_local.date(),
                        report_dir=report_dir,
                        tz=tz,
                        force=False,
                    )

            except Exception as e:
                self.stderr.write(
                    self.style.ERROR(f"[BADGE WEEKLY REPORT] scheduler error: {e}")
                )

            time.sleep(max(10, check_sec))

    def _generate_for_run_date(
        self,
        *,
        run_date: datetime.date,
        report_dir: Path,
        tz: ZoneInfo,
        force: bool,
    ):
        start_date, end_date = _period_for_run_date(run_date)
        filename = _filename_for_period(start_date, end_date)
        output_path = report_dir / filename

        if output_path.exists() and not force:
            self.stdout.write(
                f"[BADGE WEEKLY REPORT] already exists, skip: {output_path}"
            )
            return

        result = build_admin_badge_report_file(
            start_date=start_date,
            end_date=end_date,
            output_path=output_path,
            tz=tz,
        )

        self.stdout.write(
            self.style.SUCCESS(
                "[BADGE WEEKLY REPORT] created "
                f"{output_path} "
                f"transfer_rows={result['transfer_rows']} "
                f"onec_print_rows={result['onec_print_rows']} "
                f"period={start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}"
            )
        )














# import os
# import time
# import datetime
# import tempfile
# from pathlib import Path
# from zoneinfo import ZoneInfo

# from django.core.management.base import BaseCommand
# from django.db.models import Q
# from django.utils import timezone

# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# from frostapp.models import AdminBadgeRequest, Store


# def _read_simple_env_file_value(name: str, env_path: str = "/app/.env") -> str | None:
#     """
#     Простое чтение .env без python-dotenv.

#     Поддерживает:
#       KEY=value
#       export KEY=value
#       KEY="value"
#       KEY='value'
#     """
#     try:
#         if not os.path.exists(env_path):
#             return None

#         with open(env_path, "r", encoding="utf-8") as f:
#             for line in f:
#                 line = (line or "").strip()

#                 if not line or line.startswith("#"):
#                     continue

#                 if line.startswith("export "):
#                     line = line[len("export "):].strip()

#                 if "=" not in line:
#                     continue

#                 key, value = line.split("=", 1)
#                 key = key.strip()
#                 value = value.strip()

#                 if key != name:
#                     continue

#                 if (
#                     len(value) >= 2
#                     and (
#                         (value[0] == value[-1] == '"')
#                         or (value[0] == value[-1] == "'")
#                     )
#                 ):
#                     value = value[1:-1]

#                 return value

#     except Exception:
#         return None

#     return None


# def _env_value(name: str, default: str = "", env_path: str = "/app/.env") -> str:
#     """
#     Сначала читает настоящие переменные окружения Docker,
#     потом файл /app/.env.
#     """
#     value = os.getenv(name)

#     if value is not None and str(value).strip() != "":
#         return str(value).strip()

#     value_from_file = _read_simple_env_file_value(name, env_path)

#     if value_from_file is not None and str(value_from_file).strip() != "":
#         return str(value_from_file).strip()

#     return default


# def _int_env(name: str, default: int) -> int:
#     try:
#         return int(_env_value(name, str(default)))
#     except Exception:
#         return default


# def _get_report_tz() -> ZoneInfo:
#     tz_name = (
#         _env_value(
#             "BADGE_WEEKLY_REPORT_TZ",
#             _env_value("BADGE_REPORT_TIMEZONE", "Asia/Irkutsk"),
#         )
#         or "Asia/Irkutsk"
#     ).strip()

#     try:
#         return ZoneInfo(tz_name)
#     except Exception:
#         return ZoneInfo("Asia/Irkutsk")


# def _local_dt(dt, tz: ZoneInfo):
#     if not dt:
#         dt = timezone.now()

#     return timezone.localtime(dt, tz)


# def _date_time_parts(dt, tz: ZoneInfo) -> tuple[str, str]:
#     local_dt = _local_dt(dt, tz)
#     return local_dt.strftime("%d.%m.%Y"), local_dt.strftime("%H:%M:%S")


# def _period_for_run_date(run_date: datetime.date) -> tuple[datetime.date, datetime.date]:
#     """
#     Если запуск 09.07.2026, период будет:
#       02.07.2026 - 08.07.2026

#     То есть 7 календарных дней до дня запуска.
#     """
#     start_date = run_date - datetime.timedelta(days=7)
#     end_date = run_date - datetime.timedelta(days=1)
#     return start_date, end_date


# def _period_datetimes(start_date: datetime.date, end_date: datetime.date, tz: ZoneInfo):
#     """
#     Возвращает границы периода:
#       start_date 00:00:00 inclusive
#       end_date + 1 день 00:00:00 exclusive
#     """
#     start_dt = datetime.datetime.combine(start_date, datetime.time.min, tzinfo=tz)
#     end_exclusive_dt = datetime.datetime.combine(
#         end_date + datetime.timedelta(days=1),
#         datetime.time.min,
#         tzinfo=tz,
#     )

#     return start_dt, end_exclusive_dt


# def _filename_for_period(start_date: datetime.date, end_date: datetime.date) -> str:
#     return f"{start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}.xlsx"


# def _load_store_names(store_ids: list[int]) -> dict[int, str]:
#     if not store_ids:
#         return {}

#     result = {}

#     qs = Store.objects.filter(ukm4store__in=store_ids).only("ukm4store", "name")

#     for store in qs:
#         if store.ukm4store is None:
#             continue

#         name = str(store.name or "").strip()
#         result[int(store.ukm4store)] = name or f"Магазин {store.ukm4store}"

#     return result


# def build_admin_badge_report_file(
#     *,
#     start_date: datetime.date,
#     end_date: datetime.date,
#     output_path: Path,
#     tz: ZoneInfo,
# ) -> int:
#     """
#     Формирует Excel-отчёт и сохраняет его в output_path.
#     Возвращает количество строк отчёта без заголовка.
#     """
#     start_dt, end_exclusive_dt = _period_datetimes(start_date, end_date, tz)

#     qs = (
#         AdminBadgeRequest.objects
#         .filter(
#             Q(status="ACCEPTED") | Q(decision__iexact="accept"),
#             decided_at__isnull=False,
#             decided_at__gte=start_dt,
#             decided_at__lt=end_exclusive_dt,
#         )
#         .order_by("decided_at")
#     )

#     rows = list(qs)

#     store_ids = sorted({
#         int(x.storeid)
#         for x in rows
#         if x.storeid is not None
#     })

#     stores_map = _load_store_names(store_ids)

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Передача бейджа"

#     headers = ["Магазин", "Дата", "Время", "Кассир", "Администратор"]
#     ws.append(headers)

#     header_fill = PatternFill("solid", fgColor="D9EAF7")
#     thin = Side(style="thin", color="B7B7B7")
#     border = Border(left=thin, right=thin, top=thin, bottom=thin)

#     for col_num, title in enumerate(headers, start=1):
#         cell = ws.cell(row=1, column=col_num)
#         cell.font = Font(bold=True)
#         cell.fill = header_fill
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.border = border

#     for item in rows:
#         if item.storeid is not None:
#             store_name = stores_map.get(int(item.storeid), f"Магазин {item.storeid}")
#         else:
#             store_name = "—"

#         date_text, time_text = _date_time_parts(item.decided_at, tz)

#         ws.append([
#             store_name,
#             date_text,
#             time_text,
#             item.cashier_full_name or "—",
#             item.admin_full_name or "—",
#         ])

#     for row in ws.iter_rows(min_row=2):
#         for cell in row:
#             cell.border = border
#             cell.alignment = Alignment(vertical="center", wrap_text=True)

#     widths = {
#         "A": 35,
#         "B": 14,
#         "C": 12,
#         "D": 35,
#         "E": 35,
#     }

#     for col, width in widths.items():
#         ws.column_dimensions[col].width = width

#     ws.freeze_panes = "A2"
#     ws.auto_filter.ref = ws.dimensions

#     output_path.parent.mkdir(parents=True, exist_ok=True)

#     # Пишем сначала во временный файл, потом атомарно переименовываем.
#     with tempfile.NamedTemporaryFile(
#         suffix=".xlsx",
#         delete=False,
#         dir=str(output_path.parent),
#     ) as tmp:
#         tmp_path = Path(tmp.name)

#     try:
#         wb.save(tmp_path)
#         os.replace(tmp_path, output_path)
#     finally:
#         try:
#             if tmp_path.exists():
#                 tmp_path.unlink()
#         except Exception:
#             pass

#     return len(rows)


# class Command(BaseCommand):
#     help = "Еженедельный Excel-отчёт по передаче бейджа администратора"

#     def add_arguments(self, parser):
#         parser.add_argument(
#             "--once",
#             action="store_true",
#             help="Сформировать отчёт один раз и выйти",
#         )
#         parser.add_argument(
#             "--date",
#             type=str,
#             default="",
#             help="Дата запуска в формате YYYY-MM-DD. Нужна для теста. Например: 2026-07-09",
#         )

#     def handle(self, *args, **options):
#         tz = _get_report_tz()

#         report_dir = Path(
#             _env_value(
#                 "BADGE_WEEKLY_REPORT_DIR",
#                 "/app/badge_reports_shared",
#             )
#         )

#         schedule_weekday = _int_env("BADGE_WEEKLY_REPORT_WEEKDAY", 2)
#         schedule_hour = _int_env("BADGE_WEEKLY_REPORT_HOUR", 8)
#         schedule_minute = _int_env("BADGE_WEEKLY_REPORT_MINUTE", 0)
#         check_sec = _int_env("BADGE_WEEKLY_REPORT_CHECK_SEC", 60)

#         once = bool(options.get("once"))
#         date_raw = str(options.get("date") or "").strip()

#         if once:
#             if date_raw:
#                 run_date = datetime.date.fromisoformat(date_raw)
#             else:
#                 run_date = timezone.now().astimezone(tz).date()

#             self._generate_for_run_date(
#                 run_date=run_date,
#                 report_dir=report_dir,
#                 tz=tz,
#                 force=True,
#             )
#             return

#         self.stdout.write(
#             self.style.SUCCESS(
#                 "[BADGE WEEKLY REPORT] scheduler started. "
#                 f"dir={report_dir} tz={tz.key} "
#                 f"weekday={schedule_weekday} time={schedule_hour:02d}:{schedule_minute:02d}"
#             )
#         )

#         while True:
#             try:
#                 now_local = timezone.now().astimezone(tz)

#                 is_schedule_day = now_local.weekday() == schedule_weekday
#                 is_after_schedule_time = (
#                     now_local.hour > schedule_hour
#                     or (
#                         now_local.hour == schedule_hour
#                         and now_local.minute >= schedule_minute
#                     )
#                 )

#                 if is_schedule_day and is_after_schedule_time:
#                     self._generate_for_run_date(
#                         run_date=now_local.date(),
#                         report_dir=report_dir,
#                         tz=tz,
#                         force=False,
#                     )

#             except Exception as e:
#                 self.stderr.write(
#                     self.style.ERROR(f"[BADGE WEEKLY REPORT] scheduler error: {e}")
#                 )

#             time.sleep(max(10, check_sec))

#     def _generate_for_run_date(
#         self,
#         *,
#         run_date: datetime.date,
#         report_dir: Path,
#         tz: ZoneInfo,
#         force: bool,
#     ):
#         start_date, end_date = _period_for_run_date(run_date)
#         filename = _filename_for_period(start_date, end_date)
#         output_path = report_dir / filename

#         if output_path.exists() and not force:
#             self.stdout.write(
#                 f"[BADGE WEEKLY REPORT] already exists, skip: {output_path}"
#             )
#             return

#         rows_count = build_admin_badge_report_file(
#             start_date=start_date,
#             end_date=end_date,
#             output_path=output_path,
#             tz=tz,
#         )

#         self.stdout.write(
#             self.style.SUCCESS(
#                 "[BADGE WEEKLY REPORT] created "
#                 f"{output_path} rows={rows_count} "
#                 f"period={start_date.strftime('%d.%m.%Y')}-{end_date.strftime('%d.%m.%Y')}"
#             )
#         )
