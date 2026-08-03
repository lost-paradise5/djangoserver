from __future__ import annotations

import datetime as dt
import os
import time
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4

import pymysql
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from frostapp.models import Store


TARGET_STORE_IDS: tuple[int, ...] = (
    1003, 1004, 1005, 1006,
    2007, 2010, 2011, 2013,
    3002, 3003, 3011, 3012, 3013,
    4002, 4004, 4011, 4015,
    5001, 5006, 5007, 5009,
    6010, 6011, 6012, 6014, 6015, 6017,
    7002, 7012, 7013,
    8002, 8003, 8005, 8007, 8012,
    9001, 9004, 9006, 9008, 9010, 9014, 9016,
    11004, 11006, 11008, 11015,
    12005, 12008, 12009,
    14005, 14006, 14015, 14021,
)

EXPORT_DIR = Path(
    os.getenv("TRM_USERS_EXPORT_DIR", "/tmp/trm_users_exports")
)
EXPORT_TTL_SECONDS = int(
    os.getenv("TRM_USERS_EXPORT_TTL_SECONDS", "86400")
)
EXPORT_MAX_WORKERS = max(
    1,
    int(os.getenv("TRM_USERS_EXPORT_MAX_WORKERS", "6")),
)

# Отдельный сервер КСО. Пароль лучше передавать через .env.
# Если KSO_DB_USER/KSO_DB_PASSWORD не заданы, используем те же
# учётные данные, что и для обычных кассовых серверов UKM.
KSO_DB_HOST = os.getenv("KSO_DB_HOST", "192.168.17.38")
KSO_DB_PORT = int(os.getenv("KSO_DB_PORT", "3306"))
KSO_DB_USER = os.getenv(
    "KSO_DB_USER",
    os.getenv("UKMSERVER_USER", "ukminfo"),
)
KSO_DB_PASSWORD = os.getenv(
    "KSO_DB_PASSWORD",
    os.getenv("UKMSERVER_PASSWORD", ""),
)
KSO_DB_NAME = os.getenv("KSO_DB_NAME", "srvdata")
KSO_EXPORT_BATCH_SIZE = max(
    100,
    int(os.getenv("KSO_EXPORT_BATCH_SIZE", "5000")),
)
KSO_CONNECT_TIMEOUT = int(os.getenv("KSO_DB_CONNECT_TIMEOUT", "10"))
KSO_READ_TIMEOUT = int(os.getenv("KSO_DB_READ_TIMEOUT", "600"))
KSO_WRITE_TIMEOUT = int(os.getenv("KSO_DB_WRITE_TIMEOUT", "30"))


_COLUMN_CANDIDATES: dict[str, tuple[str, ...]] = {
    "store_id": ("store_id", "storeid", "store", "shop_id", "shop"),
    "name": ("name", "fio", "full_name"),
    "user_inn": ("user_inn", "inn"),
    "role_id": ("role_id", "roleid"),
    "start_date": ("start_date",),
    "end_date": ("end_date",),
    "deleted": ("deleted", "is_deleted"),
}


class TRMExportError(RuntimeError):
    pass


def _mysql_connection(host: str):
    return pymysql.connect(
        host=host,
        port=int(os.getenv("UKMSERVER_PORT", "3306")),
        user=os.getenv("UKMSERVER_USER", "ukminfo"),
        password=os.getenv("UKMSERVER_PASSWORD", ""),
        database=os.getenv("UKMSERVER_DATABASE", "ukmserver"),
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=int(os.getenv("TRM_USERS_CONNECT_TIMEOUT", "5")),
        read_timeout=int(os.getenv("TRM_USERS_READ_TIMEOUT", "30")),
        write_timeout=int(os.getenv("TRM_USERS_WRITE_TIMEOUT", "10")),
        autocommit=True,
    )


def _pick_column(columns: set[str], logical_name: str) -> str | None:
    for candidate in _COLUMN_CANDIDATES[logical_name]:
        if candidate in columns:
            return candidate
    return None


def _safe_int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _query_one_host(host: str, store_ids: list[int]) -> dict[str, Any]:
    conn = cur = None

    try:
        conn = _mysql_connection(host)
        cur = conn.cursor()

        cur.execute("SHOW COLUMNS FROM `trm_in_users`")
        schema_rows = cur.fetchall() or []
        columns = {
            str(row.get("Field") or "").strip()
            for row in schema_rows
            if row.get("Field")
        }

        selected = {
            logical_name: _pick_column(columns, logical_name)
            for logical_name in _COLUMN_CANDIDATES
        }

        required = ("store_id", "name", "user_inn", "role_id", "deleted")
        missing_required = [
            name for name in required if not selected.get(name)
        ]

        if missing_required:
            raise TRMExportError(
                "В trm_in_users отсутствуют обязательные колонки: "
                + ", ".join(missing_required)
                + ". Фактические колонки: "
                + ", ".join(sorted(columns))
            )

        optional_missing = [
            name
            for name in ("start_date", "end_date")
            if not selected.get(name)
        ]

        def expr(logical_name: str, alias: str) -> str:
            actual = selected.get(logical_name)
            if actual:
                return f"`{actual}` AS `{alias}`"
            return f"NULL AS `{alias}`"

        placeholders = ", ".join(["%s"] * len(store_ids))
        store_col = selected["store_id"]
        deleted_col = selected["deleted"]
        name_col = selected["name"]

        sql = f"""
            SELECT
                {expr('store_id', 'store_id')},
                {expr('name', 'name')},
                {expr('user_inn', 'user_inn')},
                {expr('role_id', 'role_id')},
                {expr('start_date', 'start_date')},
                {expr('end_date', 'end_date')}
            FROM `trm_in_users`
            WHERE `{store_col}` IN ({placeholders})
              AND COALESCE(`{deleted_col}`, 0) = 0
            ORDER BY `{store_col}`, `{name_col}`
        """

        cur.execute(sql, tuple(store_ids))
        rows = list(cur.fetchall() or [])

        normalized_rows: list[dict[str, Any]] = []
        for row in rows:
            sid = _safe_int(row.get("store_id"))
            if sid is None:
                continue

            normalized_rows.append({
                "store_id": sid,
                "name": str(row.get("name") or "").strip(),
                "user_inn": str(row.get("user_inn") or "").strip(),
                "role_id": row.get("role_id"),
                "start_date": row.get("start_date"),
                "end_date": row.get("end_date"),
                "source_host": host,
            })

        return {
            "ok": True,
            "host": host,
            "store_ids": list(store_ids),
            "rows": normalized_rows,
            "warning": (
                "Нет колонок: " + ", ".join(optional_missing)
                if optional_missing
                else ""
            ),
            "error": "",
        }

    except Exception as exc:
        return {
            "ok": False,
            "host": host,
            "store_ids": list(store_ids),
            "rows": [],
            "warning": "",
            "error": str(exc),
        }

    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass


def _load_store_map(store_ids: Iterable[int]) -> dict[int, dict[str, Any]]:
    store_ids = [int(value) for value in store_ids]

    rows = list(
        Store.objects
        .filter(ukm4store__in=store_ids)
        .values("ukm4store", "name", "ukm4ip")
        .order_by("ukm4store", "id")
    )

    result: dict[int, dict[str, Any]] = {}

    for row in rows:
        sid = _safe_int(row.get("ukm4store"))
        if sid is None:
            continue

        candidate = {
            "store_id": sid,
            "store_name": str(row.get("name") or "").strip(),
            "host": str(row.get("ukm4ip") or "").strip(),
        }

        current = result.get(sid)

        # Если есть дубликаты stores, предпочитаем запись с заполненным IP.
        if current is None or (not current.get("host") and candidate.get("host")):
            result[sid] = candidate

    return result


def _cleanup_old_exports() -> None:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    if EXPORT_TTL_SECONDS <= 0:
        return

    cutoff = time.time() - EXPORT_TTL_SECONDS

    for file_path in EXPORT_DIR.glob("*_users_*.xlsx"):
        try:
            if file_path.stat().st_mtime < cutoff:
                file_path.unlink(missing_ok=True)
        except Exception:
            pass


def _excel_value(value: Any) -> Any:
    """
    Приводит значение к формату, который поддерживает openpyxl.

    Excel не умеет сохранять datetime/time с tzinfo. Для aware-datetime
    сначала сохраняем локальное время Django, затем удаляем только tzinfo.
    """
    if value is None:
        return None

    if isinstance(value, dt.datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.replace(tzinfo=None)

    if isinstance(value, dt.time):
        return value.replace(tzinfo=None)

    if isinstance(value, (dt.date, int, float, bool)):
        return value

    return str(value)


def _apply_sheet_style(ws, *, freeze_cell: str = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def _set_widths(ws, widths: dict[int, float]) -> None:
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width


def _build_workbook(
    employee_rows: list[dict[str, Any]],
    report_rows: list[dict[str, Any]],
    generated_at,
) -> Workbook:
    wb = Workbook()

    employees_ws = wb.active
    employees_ws.title = "Сотрудники"

    employee_headers = [
        "ID магазина",
        "Название магазина",
        "ФИО",
        "ИНН",
        "Роль",
        "Дата начала",
        "Дата окончания",
    ]
    employees_ws.append(employee_headers)

    for item in employee_rows:
        employees_ws.append([
            item["store_id"],
            item.get("store_name") or "",
            item.get("name") or "",
            item.get("user_inn") or "",
            _excel_value(item.get("role_id")),
            _excel_value(item.get("start_date")),
            _excel_value(item.get("end_date")),
        ])

    for row_number in range(2, employees_ws.max_row + 1):
        employees_ws.cell(row=row_number, column=4).number_format = "@"
        employees_ws.cell(row=row_number, column=6).number_format = "DD.MM.YYYY HH:MM:SS"
        employees_ws.cell(row=row_number, column=7).number_format = "DD.MM.YYYY HH:MM:SS"

    _apply_sheet_style(employees_ws)
    _set_widths(employees_ws, {
        1: 14,
        2: 34,
        3: 36,
        4: 18,
        5: 12,
        6: 21,
        7: 21,
    })

    report_ws = wb.create_sheet("Отчёт")
    report_headers = [
        "ID магазина",
        "Название магазина",
        "Кассовый сервер",
        "Статус",
        "Найдено записей",
        "Предупреждение",
        "Ошибка",
    ]
    report_ws.append(report_headers)

    for item in report_rows:
        report_ws.append([
            item["store_id"],
            item.get("store_name") or "",
            item.get("host") or "",
            item.get("status") or "",
            int(item.get("row_count") or 0),
            item.get("warning") or "",
            item.get("error") or "",
        ])

    _apply_sheet_style(report_ws)
    _set_widths(report_ws, {
        1: 14,
        2: 34,
        3: 20,
        4: 20,
        5: 18,
        6: 36,
        7: 70,
    })

    summary_ws = wb.create_sheet("Сводка")
    success_count = sum(
        1 for row in report_rows if row.get("status") in {"ok", "empty"}
    )
    warning_count = sum(
        1 for row in report_rows if row.get("status") == "warning"
    )
    error_count = sum(
        1 for row in report_rows
        if row.get("status") in {"error", "missing_store", "missing_ip"}
    )

    summary_ws.append(["Показатель", "Значение"])
    summary_ws.append(["Дата формирования", _excel_value(generated_at)])
    summary_ws.append(["Целевых магазинов", len(report_rows)])
    summary_ws.append(["Успешно обработано", success_count])
    summary_ws.append(["С предупреждениями", warning_count])
    summary_ws.append(["С ошибками", error_count])
    summary_ws.append(["Всего выгружено сотрудников", len(employee_rows)])
    summary_ws.append([
        "Уникальных кассовых серверов",
        len({row.get("host") for row in report_rows if row.get("host")}),
    ])

    _apply_sheet_style(summary_ws)
    _set_widths(summary_ws, {1: 34, 2: 24})
    summary_ws["B2"].number_format = "DD.MM.YYYY HH:MM:SS"

    return wb


def build_trm_users_export() -> dict[str, Any]:
    started_at = time.monotonic()
    generated_at = timezone.localtime()
    target_ids = list(TARGET_STORE_IDS)
    store_map = _load_store_map(target_ids)

    report_by_store: dict[int, dict[str, Any]] = {}
    host_groups: dict[str, list[int]] = defaultdict(list)

    for store_id in target_ids:
        store_info = store_map.get(store_id)

        if not store_info:
            report_by_store[store_id] = {
                "store_id": store_id,
                "store_name": "",
                "host": "",
                "status": "missing_store",
                "row_count": 0,
                "warning": "",
                "error": "Магазин не найден в PostgreSQL stores по ukm4store",
            }
            continue

        host = str(store_info.get("host") or "").strip()
        report_by_store[store_id] = {
            "store_id": store_id,
            "store_name": store_info.get("store_name") or "",
            "host": host,
            "status": "pending",
            "row_count": 0,
            "warning": "",
            "error": "",
        }

        if not host:
            report_by_store[store_id].update({
                "status": "missing_ip",
                "error": "В stores.ukm4ip не указан адрес кассового сервера",
            })
            continue

        host_groups[host].append(store_id)

    employee_rows: list[dict[str, Any]] = []

    if host_groups:
        max_workers = min(EXPORT_MAX_WORKERS, len(host_groups))

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_map = {
                executor.submit(
                    _query_one_host,
                    host,
                    sorted(store_ids),
                ): (host, sorted(store_ids))
                for host, store_ids in host_groups.items()
            }

            for future in as_completed(future_map):
                host, store_ids = future_map[future]

                try:
                    host_result = future.result()
                except Exception as exc:
                    host_result = {
                        "ok": False,
                        "host": host,
                        "store_ids": store_ids,
                        "rows": [],
                        "warning": "",
                        "error": str(exc),
                    }

                if not host_result.get("ok"):
                    for store_id in store_ids:
                        report_by_store[store_id].update({
                            "status": "error",
                            "error": host_result.get("error") or "Неизвестная ошибка",
                        })
                    continue

                rows = host_result.get("rows") or []
                counts = Counter(
                    int(row["store_id"])
                    for row in rows
                    if _safe_int(row.get("store_id")) is not None
                )
                warning = str(host_result.get("warning") or "")

                for row in rows:
                    store_id = _safe_int(row.get("store_id"))
                    if store_id not in report_by_store:
                        continue

                    store_info = report_by_store[store_id]
                    employee_rows.append({
                        **row,
                        "store_name": store_info.get("store_name") or "",
                    })

                for store_id in store_ids:
                    count = int(counts.get(store_id, 0))
                    report_by_store[store_id].update({
                        "status": (
                            "warning" if warning
                            else ("ok" if count > 0 else "empty")
                        ),
                        "row_count": count,
                        "warning": warning,
                        "error": "",
                    })

    employee_rows.sort(key=lambda row: (
        int(row.get("store_id") or 0),
        str(row.get("name") or "").casefold(),
        str(row.get("user_inn") or ""),
        str(row.get("role_id") or ""),
    ))

    report_rows = [report_by_store[store_id] for store_id in target_ids]

    _cleanup_old_exports()
    token = uuid4().hex
    filename = (
        "trm_users_"
        + generated_at.strftime("%Y%m%d_%H%M%S")
        + f"_{token[:8]}.xlsx"
    )
    file_path = EXPORT_DIR / filename

    workbook = _build_workbook(
        employee_rows=employee_rows,
        report_rows=report_rows,
        generated_at=generated_at,
    )
    try:
        workbook.save(file_path)
    except Exception:
        file_path.unlink(missing_ok=True)
        raise

    elapsed_sec = round(time.monotonic() - started_at, 2)

    success_stores = sum(
        1 for row in report_rows if row["status"] in {"ok", "empty"}
    )
    warning_stores = sum(
        1 for row in report_rows if row["status"] == "warning"
    )
    failed_stores = len(report_rows) - success_stores - warning_stores

    return {
        "export_type": "trm",
        "result_title": "Выгрузка сотрудников кассовых серверов",
        "token": token,
        "filename": filename,
        "file_path": str(file_path),
        "generated_at": generated_at.isoformat(),
        "elapsed_sec": elapsed_sec,
        "target_store_count": len(target_ids),
        "unique_host_count": len(host_groups),
        "employee_count": len(employee_rows),
        "success_store_count": success_stores,
        "warning_store_count": warning_stores,
        "failed_store_count": failed_stores,
        "summary_labels": {
            "stores": "Магазинов",
            "hosts": "Кассовых серверов",
            "rows": "Записей сотрудников",
            "elapsed": "Время обработки",
        },
        "details_text": (
            f"Успешно: {success_stores}; "
            f"с предупреждениями: {warning_stores}; "
            f"с ошибками: {failed_stores}."
        ),
        "stores": report_rows,
    }


def _kso_mysql_connection():
    if not KSO_DB_PASSWORD:
        raise TRMExportError(
            "Не задан пароль КСО. Добавьте KSO_DB_PASSWORD "
            "или UKMSERVER_PASSWORD в .env и передайте его контейнеру."
        )

    return pymysql.connect(
        host=KSO_DB_HOST,
        port=KSO_DB_PORT,
        user=KSO_DB_USER,
        password=KSO_DB_PASSWORD,
        database=KSO_DB_NAME,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.SSDictCursor,
        connect_timeout=KSO_CONNECT_TIMEOUT,
        read_timeout=KSO_READ_TIMEOUT,
        write_timeout=KSO_WRITE_TIMEOUT,
        autocommit=True,
    )


def _validate_kso_schema(cur) -> None:
    required_columns = {
        "store": {"id", "name"},
        "user_role": {"id", "name"},
        "user": {"store_id", "id", "name", "role_id", "inn"},
        "user_card": {
            "store_id",
            "number",
            "user_id",
            "date_from",
            "date_till",
            "active",
        },
    }

    for table_name, required in required_columns.items():
        cur.execute(f"SHOW COLUMNS FROM `{table_name}`")
        actual = {
            str(row.get("Field") or "").strip()
            for row in (cur.fetchall() or [])
            if row.get("Field")
        }
        missing = sorted(required - actual)
        if missing:
            raise TRMExportError(
                f"В таблице {table_name} отсутствуют колонки: "
                + ", ".join(missing)
            )


def _write_only_cell(
    ws,
    value: Any,
    *,
    header: bool = False,
    number_format: str | None = None,
):
    cell = WriteOnlyCell(ws, value=_excel_value(value))
    thin = Side(style="thin", color="D9E2F3")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(vertical="top", wrap_text=True)

    if header:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    if number_format:
        cell.number_format = number_format

    return cell


def _append_kso_header(ws) -> None:
    ws.append([
        _write_only_cell(ws, "Магазин", header=True),
        _write_only_cell(ws, "ФИО", header=True),
        _write_only_cell(ws, "ИНН", header=True),
        _write_only_cell(ws, "Роль", header=True),
        _write_only_cell(ws, "Дата начала", header=True),
        _write_only_cell(ws, "Дата окончания", header=True),
    ])


def build_kso_users_export() -> dict[str, Any]:
    """
    Формирует потоковую выгрузку всех активных записей user_card КСО.

    Одна строка Excel соответствует одной записи user_card с active=1.
    LEFT JOIN сохраняет карты, у которых уже нет строки в таблице user.
    Пользователь связывается одновременно по user.id и user.store_id,
    чтобы не сопоставить одинаковый user_id из другого магазина.
    """
    started_at = time.monotonic()
    generated_at = timezone.localtime()
    _cleanup_old_exports()

    token = uuid4().hex
    filename = (
        "kso_users_"
        + generated_at.strftime("%Y%m%d_%H%M%S")
        + f"_{token[:8]}.xlsx"
    )
    file_path = EXPORT_DIR / filename
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook(write_only=True)
    data_ws = wb.create_sheet("КСО")
    data_ws.freeze_panes = "A2"
    data_ws.column_dimensions["A"].width = 38
    data_ws.column_dimensions["B"].width = 42
    data_ws.column_dimensions["C"].width = 18
    data_ws.column_dimensions["D"].width = 32
    data_ws.column_dimensions["E"].width = 21
    data_ws.column_dimensions["F"].width = 21
    _append_kso_header(data_ws)

    total_rows = 0
    linked_users = 0
    missing_users = 0
    missing_roles = 0
    missing_stores = 0
    store_stats: dict[int, dict[str, Any]] = {}

    conn = cur = None
    try:
        conn = _kso_mysql_connection()
        cur = conn.cursor()
        _validate_kso_schema(cur)

        sql = """
            SELECT
                uc.`store_id` AS `store_id`,
                uc.`number` AS `card_number`,
                uc.`user_id` AS `card_user_id`,
                uc.`date_from` AS `date_from`,
                uc.`date_till` AS `date_till`,
                s.`name` AS `store_name`,
                u.`id` AS `matched_user_id`,
                u.`name` AS `user_name`,
                u.`inn` AS `user_inn`,
                u.`role_id` AS `role_id`,
                ur.`name` AS `role_name`
            FROM `user_card` AS uc
            LEFT JOIN `store` AS s
              ON s.`id` = uc.`store_id`
            LEFT JOIN `user` AS u
              ON u.`store_id` = uc.`store_id`
             AND u.`id` = uc.`user_id`
            LEFT JOIN `user_role` AS ur
              ON ur.`id` = u.`role_id`
            WHERE uc.`active` = 1
            ORDER BY
                uc.`store_id`,
                COALESCE(u.`name`, ''),
                uc.`user_id`,
                uc.`number`
        """
        cur.execute(sql)

        while True:
            batch = cur.fetchmany(KSO_EXPORT_BATCH_SIZE)
            if not batch:
                break

            for row in batch:
                total_rows += 1
                store_id = _safe_int(row.get("store_id"))
                if store_id is None:
                    store_id = 0

                raw_store_name = str(row.get("store_name") or "").strip()
                store_missing = not bool(raw_store_name)
                store_name = (
                    raw_store_name
                    if raw_store_name
                    else f"[МАГАЗИН НЕ НАЙДЕН] store_id={store_id}"
                )

                matched_user_id = row.get("matched_user_id")
                user_missing = matched_user_id is None
                card_user_id = row.get("card_user_id")
                card_number = str(row.get("card_number") or "").strip()

                if user_missing:
                    missing_users += 1
                    fio = (
                        f"[ПОЛЬЗОВАТЕЛЬ НЕ НАЙДЕН] "
                        f"user_id={card_user_id}"
                    )
                    if card_number:
                        fio += f", карта={card_number}"
                    user_inn = ""
                    role_name = "[ПОЛЬЗОВАТЕЛЬ НЕ НАЙДЕН]"
                else:
                    linked_users += 1
                    fio = str(row.get("user_name") or "").strip()
                    if not fio:
                        fio = f"[ФИО НЕ УКАЗАНО] user_id={matched_user_id}"

                    user_inn = str(row.get("user_inn") or "").strip()
                    raw_role_name = str(row.get("role_name") or "").strip()
                    if raw_role_name:
                        role_name = raw_role_name
                    else:
                        missing_roles += 1
                        role_id = row.get("role_id")
                        role_name = (
                            f"[РОЛЬ НЕ НАЙДЕНА] role_id={role_id}"
                            if role_id not in (None, "")
                            else "[РОЛЬ НЕ УКАЗАНА]"
                        )

                if store_missing:
                    missing_stores += 1

                stat = store_stats.setdefault(store_id, {
                    "store_id": store_id,
                    "store_name": store_name,
                    "host": f"{KSO_DB_HOST}:{KSO_DB_PORT}/{KSO_DB_NAME}",
                    "row_count": 0,
                    "missing_user_count": 0,
                    "missing_role_count": 0,
                    "missing_store_count": 0,
                })
                stat["row_count"] += 1
                if user_missing:
                    stat["missing_user_count"] += 1
                elif not str(row.get("role_name") or "").strip():
                    stat["missing_role_count"] += 1
                if store_missing:
                    stat["missing_store_count"] += 1

                data_ws.append([
                    _write_only_cell(data_ws, store_name),
                    _write_only_cell(data_ws, fio),
                    _write_only_cell(data_ws, user_inn, number_format="@"),
                    _write_only_cell(data_ws, role_name),
                    _write_only_cell(
                        data_ws,
                        row.get("date_from"),
                        number_format="DD.MM.YYYY HH:MM:SS",
                    ),
                    _write_only_cell(
                        data_ws,
                        row.get("date_till"),
                        number_format="DD.MM.YYYY HH:MM:SS",
                    ),
                ])

        data_ws.auto_filter.ref = f"A1:F{max(1, total_rows + 1)}"

    except Exception:
        file_path.unlink(missing_ok=True)
        raise
    finally:
        try:
            if cur:
                cur.close()
            if conn:
                conn.close()
        except Exception:
            pass

    report_rows: list[dict[str, Any]] = []
    for store_id in sorted(store_stats):
        stat = store_stats[store_id]
        warning_parts = []
        if stat["missing_user_count"]:
            warning_parts.append(
                f"без записи user: {stat['missing_user_count']}"
            )
        if stat["missing_role_count"]:
            warning_parts.append(
                f"роль не найдена: {stat['missing_role_count']}"
            )
        if stat["missing_store_count"]:
            warning_parts.append(
                f"магазин не найден: {stat['missing_store_count']}"
            )

        warning = "; ".join(warning_parts)
        report_rows.append({
            "store_id": store_id,
            "store_name": stat["store_name"],
            "host": stat["host"],
            "status": "warning" if warning else "ok",
            "row_count": stat["row_count"],
            "warning": warning,
            "error": "",
        })

    summary_ws = wb.create_sheet("Сводка")
    summary_rows = [
        ["Показатель", "Значение"],
        ["Дата формирования", _excel_value(generated_at)],
        ["Сервер", f"{KSO_DB_HOST}:{KSO_DB_PORT}"],
        ["База данных", KSO_DB_NAME],
        ["Активных записей user_card", total_rows],
        ["С найденным пользователем", linked_users],
        ["Без записи в user", missing_users],
        ["Без найденной роли", missing_roles],
        ["Без записи в store", missing_stores],
        ["Магазинов в выгрузке", len(store_stats)],
        [
            "Принцип выгрузки",
            "Одна строка Excel = одна запись user_card с active=1",
        ],
    ]
    for row_index, values in enumerate(summary_rows):
        summary_ws.append([
            _write_only_cell(
                summary_ws,
                values[0],
                header=(row_index == 0),
            ),
            _write_only_cell(
                summary_ws,
                values[1],
                header=(row_index == 0),
                number_format=(
                    "DD.MM.YYYY HH:MM:SS"
                    if row_index == 1
                    else None
                ),
            ),
        ])
    summary_ws.freeze_panes = "A2"
    summary_ws.column_dimensions["A"].width = 36
    summary_ws.column_dimensions["B"].width = 58

    try:
        wb.save(file_path)
    except Exception:
        file_path.unlink(missing_ok=True)
        raise

    elapsed_sec = round(time.monotonic() - started_at, 2)
    warning_stores = sum(
        1 for row in report_rows if row.get("status") == "warning"
    )
    success_stores = len(report_rows) - warning_stores

    return {
        "export_type": "kso",
        "result_title": "Выгрузка активных карт КСО",
        "token": token,
        "filename": filename,
        "file_path": str(file_path),
        "generated_at": generated_at.isoformat(),
        "elapsed_sec": elapsed_sec,
        "target_store_count": len(store_stats),
        "unique_host_count": 1,
        "employee_count": total_rows,
        "success_store_count": success_stores,
        "warning_store_count": warning_stores,
        "failed_store_count": 0,
        "summary_labels": {
            "stores": "Магазинов в выгрузке",
            "hosts": "Серверов КСО",
            "rows": "Активных карт",
            "elapsed": "Время обработки",
        },
        "details_text": (
            f"Активных карт: {total_rows}; "
            f"с найденным user: {linked_users}; "
            f"без user: {missing_users}; "
            f"без роли: {missing_roles}; "
            f"без магазина: {missing_stores}."
        ),
        "stores": report_rows,
    }


def get_export_file_path(filename: str) -> Path | None:
    filename = str(filename or "").strip()

    allowed_prefixes = ("trm_users_", "kso_users_")
    if (
        not filename.startswith(allowed_prefixes)
        or not filename.endswith(".xlsx")
    ):
        return None

    if Path(filename).name != filename:
        return None

    candidate = EXPORT_DIR / filename
    return candidate if candidate.is_file() else None
